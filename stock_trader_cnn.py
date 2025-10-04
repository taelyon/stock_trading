import sys
import ctypes
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QTimer, pyqtSignal, QProcess, QObject, QThread, Qt, pyqtSlot, QRunnable, QThreadPool, QEventLoop
from PyQt5.QtGui import QIcon, QPainter, QFont
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from datetime import datetime, timedelta
import pandas as pd
import win32com.client
import time
import numpy as np
from scipy.stats import entropy

import pickle
import re
import mplfinance as mpf
import sqlite3
import os
import psutil
import configparser
import pygetwindow as gw
import requests
from io import BytesIO
import warnings
import logging
from openpyxl import Workbook
import json
from slacker import Slacker
import threading
import queue
import copy
import talib
import win32file
import struct
import pywintypes
import traceback
import uuid
import pyautogui

# pip install scikit-learn --only-binary :all:
# pip install .\ta_lib-0.6.3-cp39-cp39-win32.whl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001)

# matplotlib의 폰트 관련 설정
plt.rcParams['font.family'] = 'Malgun Gothic'  # '맑은 고딕'으로 설정
plt.rcParams['axes.unicode_minus'] = False  # 마이너스 기호 깨짐 방지

# PLUS 공통 OBJECT
cpCodeMgr = win32com.client.Dispatch('CpUtil.CpCodeMgr')
cpStatus = win32com.client.Dispatch('CpUtil.CpCybos')
cpTrade = win32com.client.Dispatch('CpTrade.CpTdUtil')
cpBalance = win32com.client.Dispatch('CpTrade.CpTd6033')
cpCash = win32com.client.Dispatch('CpTrade.CpTdNew5331A')
cpOrder = win32com.client.Dispatch('CpTrade.CpTd0311')
cpStock = win32com.client.Dispatch('DsCbo1.StockMst')

def init_plus_check():
    if not ctypes.windll.shell32.IsUserAnAdmin():
        logging.error(f"오류: 일반권한으로 실행됨. 관리자 권한으로 실행해 주세요")
        return False
    if (cpStatus.IsConnect == 0):
        logging.error(f"PLUS가 정상적으로 연결되지 않음")
        return False
    if (cpTrade.TradeInit(0) != 0):
        logging.error(f"주문 초기화 실패")
        return False
    return True

def setup_logging():
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)  # 루트 로거의 레벨을 DEBUG로 설정
    logging.getLogger('matplotlib').setLevel(logging.WARNING)

    log_dir = os.path.join(os.getcwd(), 'log')
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # 1.1 FileHandler 설정 (모든 로그를 파일에 기록)
    log_path = os.path.join(log_dir, f"trading_{datetime.now().strftime('%Y%m%d')}.log")
    file_handler = logging.FileHandler(log_path, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)

    # 1.2 StreamHandler 설정 (WARNING 이상의 로그를 콘솔에 출력)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.WARNING)
    console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)

def send_slack_message(login_handler, channel, message):
    """Slack 메시지를 비동기로 전송한다."""
    if login_handler is None or login_handler.slack is None:
        logging.warning("Slack 설정이 되어 있지 않습니다.")
        return
    sender = SlackMessageSender(login_handler, channel, message)
    QThreadPool.globalInstance().start(sender)

class SlackMessageSender(QRunnable):
    def __init__(self, login_handler, channel, message):
        super().__init__()
        self.login_handler = login_handler
        self.channel = channel
        self.message = message

    def run(self):
        try:
            if self.login_handler.slack:
                self.login_handler.slack.chat.post_message(self.channel, self.message)
        except Exception as ex:
            logging.error(f"Slack 메시지 전송 실패: {ex}")

class CpEvent:
    def __init__(self):
        self.last_update_time = None  # 실시간 데이터 업데이트 시각 추적

    def set_params(self, client, name, caller):
        self.client = client
        self.name = name
        self.caller = caller
        self.dic = {ord('1'): "종목별 VI", ord('2'): "배분정보", ord('3'): "기준가결정", ord('4'): "임의종료", ord('5'): "종목정보공개", ord('6'): "종목조치", ord('7'): "시장조치"}

    def OnReceived(self):
        if self.name == '9619s':  # 시장조치사항
            time_num = self.client.GetHeaderValue(0)  # 시간
            flag = self.client.GetHeaderValue(1)  # 조치 구분
            time_str = datetime.strptime(f"{time_num:06d}", '%H%M%S')
            combined_datetime = datetime.now().replace(hour=time_str.hour, minute=time_str.minute, second=time_str.second)
            time = combined_datetime.strftime('%m/%d %H:%M:%S')

            if self.dic.get(flag) == "종목별 VI":
                code = self.client.GetHeaderValue(3)  # 종목코드
                event = self.client.GetHeaderValue(5)  # 조치내용
                event2 = self.client.GetHeaderValue(6)  # 변경사항 (괴리율 포함)
                match1 = re.search(r'^A\d{6}$', code)  # 종목코드 형식 확인
                match2 = re.search(r"괴리율:(-?\d+\.\d+)%", event2)  # 괴리율 추출

                # 조건: 정상 종목 + 주권 + 상승 괴리율 + 정적 VI + 장중
                if (cpCodeMgr.GetStockControlKind(code) == 0 and  # 정상 종목
                    cpCodeMgr.GetStockSectionKind(code) == 1 and  # 주권
                    match1 and match2 and "정적" in event):
                        gap_rate = float(match2.group(1))
                        if gap_rate > 0 and combined_datetime < combined_datetime.replace(hour=15, minute=15, second=0):
                            logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> VI 발동")
                            self.caller.monitor_vi(time, code, event2)
            return
        if self.name == 'stockcur':
            code = self.client.GetHeaderValue(0)  # 종목코드
            timess = self.client.GetHeaderValue(18)  # 시간
            exFlag = self.client.GetHeaderValue(19)  # 예상체결 플래그
            cprice = self.client.GetHeaderValue(13)  # 현재가
            cVol = self.client.GetHeaderValue(17)  # 순간체결수량

            if (exFlag == ord('2')):
                item = {'code': code, 'time': timess, 'cur': cprice, 'vol': cVol}
                self.caller.updateCurData(item)
            return
        if self.name == 'cssalert':
            stgid = self.client.GetHeaderValue(0)
            stgmonid = self.client.GetHeaderValue(1)
            code = self.client.GetHeaderValue(2)
            inoutflag = self.client.GetHeaderValue(3)
            stgtime = self.client.GetHeaderValue(4)
            stgprice = self.client.GetHeaderValue(5)
            time_str = datetime.strptime(stgtime.zfill(6), '%H%M%S')
            combined_datetime = datetime.now().replace(hour=time_str.hour, minute=time_str.minute, second=time_str.second)
            time = combined_datetime.strftime('%m/%d %H:%M:%S')
            if inoutflag == ord('1') and combined_datetime < combined_datetime.replace(hour=15, minute=15, second=0): # 진입
                self.caller.checkRealtimeStg(stgid, stgmonid, code, stgprice, time)
            return
        if self.name == 'conclusion':
            conflag = self.client.GetHeaderValue(14) # 체결 플래그
            ordernum = self.client.GetHeaderValue(5) # 주문번호
            qty = self.client.GetHeaderValue(3) # 체결수량
            price = self.client.GetHeaderValue(4) # 체결가격
            code = self.client.GetHeaderValue(9) # 종목코드
            bs = self.client.GetHeaderValue(12)  # 매수/매도 구분
            buyprice = self.client.GetHeaderValue(21) # 장부가
            balance = self.client.GetHeaderValue(23) # 체결 후 잔고 수량
            conflags = { "1": "체결", "2": "확인", "3": "거부", "4": "접수" }.get(conflag, "")
            self.caller.monitorOrderStatus(code, ordernum, conflags, price, qty, bs, balance, buyprice)

class CpPublish:
    def __init__(self, name, service_id):
        self.name = name
        self.obj = win32com.client.Dispatch(service_id)
        self.bIsSB = False

    def Subscribe(self, var, caller):
        if self.bIsSB:
            self.Unsubscribe()

        if len(var) > 0:
            self.obj.SetInputValue(0, var)
        handler = win32com.client.WithEvents(self.obj, CpEvent)
        handler.set_params(self.obj, self.name, caller)
        self.obj.Subscribe()
        self.bIsSB = True

    def Unsubscribe(self):
        if self.bIsSB:
            self.obj.Unsubscribe()
            self.bIsSB = False

class CpPBConclusion(CpPublish):
    def __init__(self):
        super().__init__('conclusion', 'DsCbo1.CpConclusion')

class CpPBStockCur(CpPublish):
    def __init__(self):
        super().__init__('stockcur', 'DsCbo1.StockCur')

class CpPB9619(CpPublish): # 시장조치사항
    def __init__(self):
        super().__init__('9619s', 'CpSysDib.CpSvr9619s')

class CpPBCssAlert(CpPublish): # 종목검색 전략
    def __init__(self):
        super().__init__('cssalert', 'CpSysDib.CssAlert')

class CpRequest:
    def __init__(self):
        self.result = None
        self.loop = QEventLoop()
        self.is_finished = False
        self.client = None
        self.name = None
        self.caller = None
        self.params = {}
        self.is_requesting = False  # 전역 요청 플래그

    def set_params(self, client, name, caller, params=None):
        self.client = client
        self.name = name
        self.caller = caller
        self.params = params if params else {}

    def wait(self):
        if not self.is_finished:
            self.loop.exec_()
        return self.result

    def OnReceived(self):
        try:
            if self.name == 'order':
                rqStatus = self.client.GetDibStatus()
                if rqStatus != 0:
                    rqRet = self.client.GetDibMsg1()
                    logging.warning(f"{self.params.get('stock_name')}({self.params.get('code')}) 주문 요청 오류, {rqRet}")
                    self.result = False
                    return
                self.result = True

        except Exception as ex:
            logging.error(f"OnReceived -> {ex}")
            self.result = False

        finally:
            self.is_finished = True
            if self.loop.isRunning():
                self.loop.quit()
            # logging.debug(f"{self.name} 요청 완료")
            if self.caller and hasattr(self.caller, 'cp_request'):
                self.caller.cp_request.is_requesting = False

class CpStrategy:
    def __init__(self, trader):
        self.monList = {}
        self.trader = trader
        self.stgname = {}
        self.objpb = CpPBCssAlert()

    def requestList(self):
        retStgList = {}
        objRq = win32com.client.Dispatch("CpSysDib.CssStgList")
        objRq.SetInputValue(0, ord('1'))  # '0' : 예제전략, '1': 나의전략
        objRq.BlockRequest2(1)

        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"나의전략 조회실패, {rqStatus}, {rqRet}")
            return (False, retStgList)

        cnt = objRq.GetHeaderValue(0)  # 0 - (long) 전략 목록 수
        flag = objRq.GetHeaderValue(1)  # 1 - (char) 요청구분

        for i in range(cnt):
            item = {}
            item['전략명'] = objRq.GetDataValue(0, i)
            item['ID'] = objRq.GetDataValue(1, i)
            # item['전략등록일시'] = objRq.GetDataValue(2, i)
            # item['작성자필명'] = objRq.GetDataValue(3, i)
            # item['평균종목수'] = objRq.GetDataValue(4, i)
            # item['평균승률'] = objRq.GetDataValue(5, i)
            item['평균수익률'] = objRq.GetDataValue(6, i)
            retStgList[item['전략명']] = item
        return retStgList

    def requestStgID(self, id):
        retStgstockList = []
        objRq = None
        objRq = win32com.client.Dispatch("CpSysDib.CssStgFind")
        objRq.SetInputValue(0, id)  # 전략 id 요청
        objRq.BlockRequest2(1)
        # 통신 및 통신 에러 처리
        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"전략ID 조회실패, {rqStatus}, {rqRet}")
            return (False, retStgstockList)

        cnt = objRq.GetHeaderValue(0)  # 0 - (long) 검색된 결과 종목 수
        totcnt = objRq.GetHeaderValue(1)  # 1 - (long) 총 검색 종목 수
        stime = objRq.GetHeaderValue(2)  # 2 - (string) 검색시간
        # print('검색된 종목수:', cnt, '전체종목수:', totcnt, '검색시간:', stime)

        for i in range(cnt):
            item = {}
            item['code'] = objRq.GetDataValue(0, i)
            item['종목명'] = cpCodeMgr.CodeToName(item['code'])
            retStgstockList.append(item)

        return (True, retStgstockList)

    def requestMonitorID(self, id):
        objRq = win32com.client.Dispatch("CpSysDib.CssWatchStgSubscribe")
        objRq.SetInputValue(0, id)  # 전략 id 요청
        objRq.BlockRequest2(1)

        # 통신 및 통신 에러 처리
        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"감시번호 조회실패, {rqStatus}, {rqRet}")
            return (False, 0)

        monID = objRq.GetHeaderValue(0)
        if monID == 0:
            logging.warning(f"감시 일련번호 구하기 실패")
            return (False, 0)

        # monID - 전략 감시를 위한 일련번호를 구해온다.
        # 현재 감시되는 전략이 없다면 감시일련번호로 1을 리턴하고,
        # 현재 감시되는 전략이 있다면 각 통신 ID에 대응되는 새로운 일련번호를 리턴한다.
        return (True, monID)

    def requestStgControl(self, id, monID, bStart, stgname):
        objRq = win32com.client.Dispatch("CpSysDib.CssWatchStgControl")
        objRq.SetInputValue(0, id)  # 전략 id 요청
        objRq.SetInputValue(1, monID)  # 감시일련번호

        if bStart == True:
            objRq.SetInputValue(2, ord('1'))  # 감시시작
            self.stgname[id] = stgname
        else:
            objRq.SetInputValue(2, ord('3'))  # 감시취소
            if id in self.stgname:
                del self.stgname[id]
        objRq.BlockRequest2(1)

        # 통신 및 통신 에러 처리
        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"감시시작 실패, {rqStatus}, {rqRet}")
            return (False, '')

        status = objRq.GetHeaderValue(0)

        # if status == 0 :
        #     print('전략 감시 초기화')
        # if status == 1:
        #     logging.info(f"{stgname} 전략 감시 시작")
        # elif status == 2:
        #     logging.info(f"{stgname} 전략 감시 중단")
        # elif status == 3:
        #     logging.info(f"{stgname} 전략 등록 취소")

        self.objpb.Subscribe('', self)

        # 진행 중인 전략들 저장
        if bStart == True:
            self.monList[id] = monID
        else:
            if id in self.monList:
                del self.monList[id]

        return (True, status)

    def checkRealtimeStg(self, stgid, stgmonid, code, stgprice, time):
        # 감시중인 전략인 경우만 체크
        if stgid not in self.monList:
            return
        if (stgmonid != self.monList[stgid]):
            return
        remain_time0 = cpStatus.GetLimitRemainTime(0)
        remain_time1 = cpStatus.GetLimitRemainTime(1)
        if remain_time0 != 0 or remain_time1 != 0:
            return
        if datetime.now() < datetime.now().replace(hour=9, minute=3, second=0, microsecond=0):
            return
        if code not in self.trader.monistock_set:
            if self.trader.daydata.select_code(code) and self.trader.tickdata.monitor_code(code) and self.trader.mindata.monitor_code(code):
                stgname = self.stgname.get(stgid, '')
                if stgname in ['급등주']:
                    self.trader.starting_time[code] = datetime.now().strftime('%m/%d 09:00:00')
                    self.trader.starting_price[code] = stgprice
                else:
                    if code not in self.trader.starting_time:
                        self.trader.starting_time[code] = datetime.now().strftime('%m/%d 09:00:00')
                self.trader.monistock_set.add(code)
                self.trader.stock_added_to_monitor.emit(code)
                self.trader.save_list_db(code, self.trader.starting_time[code], self.trader.starting_price[code], 1)
                logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 종목 추가")
            else:
                self.trader.daydata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)

    def Clear(self):
        delitem = []
        for id, monId in self.monList.items():
            delitem.append((id, monId))

        for item in delitem:
            self.requestStgControl(item[0], item[1], False, "Unknown")

        self.objpb.Unsubscribe()

# stock_trader_cnn.py의 CpIndicators 클래스 교체
class CpIndicators:
    def __init__(self, chart_type):
        self.chart_type = chart_type
        self.params = self._get_default_params()
    
    def _get_default_params(self):
        """지표별 기본 파라미터 정의"""
        if self.chart_type == 'T':
            return {
                'MA_PERIODS': [5, 20, 60, 120],
                'RSI_PERIOD': 12,
                'RSI_SIGNAL_PERIOD': 12,
                'MACD': (12, 26, 9),
                'STOCH': (7, 3, 3),
                'ATR_PERIOD': 10,
                'CCI_PERIOD': 12,
                'BB_PERIOD': 20,
                'BB_STD': 2
            }
        elif self.chart_type == 'm':
            return {
                'MA_PERIODS': [5, 10, 20],
                'RSI_PERIOD': 9,
                'RSI_SIGNAL_PERIOD': 9,
                'MACD': (12, 26, 9),
                'STOCH': (7, 3, 3),
                'ATR_PERIOD': 10,
                'CCI_PERIOD': 12,
                'BB_PERIOD': 20,
                'BB_STD': 2
            }
        elif self.chart_type == 'D':
            return {
                'MA_PERIODS': [5, 10],
                'MACD': (12, 26, 9)
            }
        return {}
    
    def _validate_data(self, chart_data, min_length):
        """데이터 유효성 검증"""
        required_keys = ['C', 'H', 'L', 'V', 'D', 'T']
        for key in required_keys:
            if key not in chart_data:
                return False
            if len(chart_data[key]) < min_length:
                return False
        return True
    
    def _fill_nan(self, data, method='smart'):
        """스마트한 NaN 처리"""
        if method == 'smart':
            # pandas를 사용한 forward fill → backward fill → 0
            series = pd.Series(data)
            filled = series.fillna(method='ffill').fillna(method='bfill').fillna(0)
            return filled.tolist()
        else:
            # 단순 0 치환
            return np.nan_to_num(data, nan=0.0).tolist()
    
    def _get_default_result(self, indicator_type, length):
        """데이터 부족 시 기본값 반환"""
        default_value = [0] * length
        
        if indicator_type == 'MA':
            if self.chart_type == 'T':
                return {
                    'MAT5': default_value, 'MAT20': default_value,
                    'MAT60': default_value, 'MAT120': default_value,
                    'MAT5_MAT20_DIFF': default_value, 'MAT20_MAT60_DIFF': default_value,
                    'MAT60_MAT120_DIFF': default_value, 'C_MAT5_DIFF': default_value,
                    'MAT5_CHANGE': default_value, 'MAT20_CHANGE': default_value,
                    'MAT60_CHANGE': default_value, 'MAT120_CHANGE': default_value
                }
            elif self.chart_type == 'm':
                return {
                    'MAM5': default_value, 'MAM10': default_value, 'MAM20': default_value,
                    'MAM5_MAM10_DIFF': default_value, 'MAM10_MAM20_DIFF': default_value,
                    'C_MAM5_DIFF': default_value, 'C_ABOVE_MAM5': default_value
                }
            elif self.chart_type == 'D':
                return {'MAD5': default_value, 'MAD10': default_value}
        
        elif indicator_type == 'MACD':
            if self.chart_type == 'T':
                return {
                    'MACDT': default_value, 'MACDT_SIGNAL': default_value,
                    'OSCT': default_value
                }
            else:
                return {
                    'MACD': default_value, 'MACD_SIGNAL': default_value,
                    'OSC': default_value
                }
        
        elif indicator_type == 'RSI':
            if self.chart_type == 'T':
                return {'RSIT': default_value, 'RSIT_SIGNAL': default_value}
            else:
                return {'RSI': default_value, 'RSI_SIGNAL': default_value}
        
        elif indicator_type == 'STOCH':
            return {'STOCHK': default_value, 'STOCHD': default_value}
        
        elif indicator_type == 'ATR':
            return {'ATR': default_value}
        
        elif indicator_type == 'CCI':
            return {'CCI': default_value}
        
        elif indicator_type == 'BBANDS':
            return {
                'BB_UPPER': default_value, 'BB_MIDDLE': default_value,
                'BB_LOWER': default_value, 'BB_POSITION': default_value,
                'BB_BANDWIDTH': default_value
            }
        
        elif indicator_type == 'VWAP':
            return {'VWAP': default_value}
        
        return {}

    def make_indicator(self, indicator_type, code, chart_data):
        try:
            result = {}
            closes = np.array(chart_data['C'], dtype=np.float64)
            highs = np.array(chart_data['H'], dtype=np.float64)
            lows = np.array(chart_data['L'], dtype=np.float64)
            volumes = np.array(chart_data['V'], dtype=np.float64)
            dates = np.array(chart_data['D'], dtype=np.int64)
            
            desired_length = len(closes)
            
            # 지표별 최소 데이터 길이 정의
            min_lengths = {
                'MA': max(self.params.get('MA_PERIODS', [5])),
                'MACD': 35,  # 26 + 9
                'RSI': self.params.get('RSI_PERIOD', 14),
                'STOCH': 14,
                'ATR': self.params.get('ATR_PERIOD', 14),
                'CCI': self.params.get('CCI_PERIOD', 14),
                'BBANDS': self.params.get('BB_PERIOD', 20),
                'VWAP': 1
            }
            
            min_required = min_lengths.get(indicator_type, 20)
            
            # 데이터 유효성 검증
            if not self._validate_data(chart_data, min_required):
                logging.debug(f"{code}: {indicator_type} 데이터 부족 ({len(closes)} < {min_required})")
                return self._get_default_result(indicator_type, desired_length)
            
            # 지표 계산
            if indicator_type == 'MA':
                if self.chart_type == 'T':
                    terms = self.params['MA_PERIODS']
                    index_names = ["MAT5", "MAT20", "MAT60", "MAT120"]
                elif self.chart_type == 'm':
                    terms = self.params['MA_PERIODS']
                    index_names = ["MAM5", "MAM10", "MAM20"]
                elif self.chart_type == 'D':
                    terms = self.params['MA_PERIODS']
                    index_names = ["MAD5", "MAD10"]
                
                # 이동평균 계산
                for term, name in zip(terms, index_names):
                    sma = talib.SMA(closes, timeperiod=term)
                    result[name] = self._fill_nan(sma)
                
                # 이동평균 간 차이 및 변화율 계산
                if self.chart_type == 'T':
                    # 이동평균 간 차이
                    result['MAT5_MAT20_DIFF'] = [
                        (result['MAT5'][i] - result['MAT20'][i]) / result['MAT20'][i] 
                        if result['MAT20'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['MAT20_MAT60_DIFF'] = [
                        (result['MAT20'][i] - result['MAT60'][i]) / result['MAT60'][i] 
                        if result['MAT60'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['MAT60_MAT120_DIFF'] = [
                        (result['MAT60'][i] - result['MAT120'][i]) / result['MAT120'][i] 
                        if result['MAT120'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['C_MAT5_DIFF'] = [
                        (closes[i] - result['MAT5'][i]) / result['MAT5'][i] 
                        if result['MAT5'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    
                    # 이동평균 변화율 (전봉 대비)
                    for name in ['MAT5', 'MAT20', 'MAT60', 'MAT120']:
                        ma_values = result[name]
                        changes = [0] + [
                            (ma_values[i] - ma_values[i-1]) / ma_values[i-1]
                            if ma_values[i-1] != 0 else 0
                            for i in range(1, len(ma_values))
                        ]
                        result[f'{name}_CHANGE'] = changes
                
                elif self.chart_type == 'm':
                    result['MAM5_MAM10_DIFF'] = [
                        (result['MAM5'][i] - result['MAM10'][i]) / result['MAM10'][i] 
                        if result['MAM10'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['MAM10_MAM20_DIFF'] = [
                        (result['MAM10'][i] - result['MAM20'][i]) / result['MAM20'][i] 
                        if result['MAM20'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['C_MAM5_DIFF'] = [
                        (closes[i] - result['MAM5'][i]) / result['MAM5'][i] 
                        if result['MAM5'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['C_ABOVE_MAM5'] = [
                        1 if closes[i] > result['MAM5'][i] else 0 
                        for i in range(desired_length)
                    ]

            elif indicator_type == 'MACD':
                fast, slow, signal = self.params['MACD']
                macd_line, signal_line, macd_hist = talib.MACD(
                    closes, fastperiod=fast, slowperiod=slow, signalperiod=signal
                )
                
                if self.chart_type == 'T':
                    result['MACDT'] = self._fill_nan(macd_line)
                    result['MACDT_SIGNAL'] = self._fill_nan(signal_line)
                    result['OSCT'] = self._fill_nan(macd_hist)
                else:
                    result['MACD'] = self._fill_nan(macd_line)
                    result['MACD_SIGNAL'] = self._fill_nan(signal_line)
                    result['OSC'] = self._fill_nan(macd_hist)

            elif indicator_type == 'RSI':
                rsi_period = self.params['RSI_PERIOD']
                rsi_signal_period = self.params['RSI_SIGNAL_PERIOD']
                
                rsi_values = talib.RSI(closes, timeperiod=rsi_period)
                rsi_signal = talib.SMA(rsi_values, timeperiod=rsi_signal_period)
                
                if self.chart_type == 'T':
                    result['RSIT'] = self._fill_nan(rsi_values)
                    result['RSIT_SIGNAL'] = self._fill_nan(rsi_signal)
                else:
                    result['RSI'] = self._fill_nan(rsi_values)
                    result['RSI_SIGNAL'] = self._fill_nan(rsi_signal)

            elif indicator_type == 'STOCH':
                fastk_period, slowk_period, slowd_period = self.params['STOCH']
                stoch_k, stoch_d = talib.STOCH(
                    highs, lows, closes,
                    fastk_period=fastk_period,
                    slowk_period=slowk_period,
                    slowd_period=slowd_period
                )
                result['STOCHK'] = self._fill_nan(stoch_k)
                result['STOCHD'] = self._fill_nan(stoch_d)

            elif indicator_type == 'ATR':
                atr_period = self.params['ATR_PERIOD']
                atr = talib.ATR(highs, lows, closes, timeperiod=atr_period)
                result['ATR'] = self._fill_nan(atr)

            elif indicator_type == 'CCI':
                cci_period = self.params['CCI_PERIOD']
                cci = talib.CCI(highs, lows, closes, timeperiod=cci_period)
                result['CCI'] = self._fill_nan(cci)

            elif indicator_type == 'BBANDS':
                bb_period = self.params['BB_PERIOD']
                bb_std = self.params['BB_STD']
                
                upper, middle, lower = talib.BBANDS(
                    closes, timeperiod=bb_period,
                    nbdevup=bb_std, nbdevdn=bb_std
                )
                
                result['BB_UPPER'] = self._fill_nan(upper)
                result['BB_MIDDLE'] = self._fill_nan(middle)
                result['BB_LOWER'] = self._fill_nan(lower)
                
                # BB Position 계산 (개선)
                bandwidth = upper - lower
                bb_position = np.where(
                    bandwidth > 1e-6,
                    (closes - middle) / bandwidth,
                    0.5  # 밴드폭이 매우 좁을 때는 중립
                )
                bb_position = np.clip(bb_position, -2, 2)  # 극단값 제한
                result['BB_POSITION'] = bb_position.tolist()
                
                # BB Bandwidth 계산
                bb_bandwidth = np.where(
                    middle > 1e-6,
                    bandwidth / middle,
                    0
                )
                result['BB_BANDWIDTH'] = bb_bandwidth.tolist()

            elif indicator_type == 'VWAP':
                # 개선된 VWAP 계산
                vwap = np.zeros_like(closes)
                
                if len(dates) == len(closes):
                    unique_dates = np.unique(dates)
                    
                    for d in unique_dates:
                        mask = dates == d
                        day_closes = closes[mask]
                        day_volumes = volumes[mask]
                        
                        # 장 시작부터 누적 계산
                        cumsum_pv = np.cumsum(day_closes * day_volumes)
                        cumsum_v = np.cumsum(day_volumes)
                        
                        # 0으로 나누기 방지
                        day_vwap = np.divide(
                            cumsum_pv, cumsum_v,
                            out=np.zeros_like(cumsum_pv),
                            where=cumsum_v != 0
                        )
                        vwap[mask] = day_vwap
                
                result['VWAP'] = vwap.tolist()

            else:
                logging.error(f"알 수 없는 지표 유형: {indicator_type}")
                return self._get_default_result(indicator_type, desired_length)

            return result

        except Exception as ex:
            logging.error(f"make_indicator -> {code}, {indicator_type}{self.chart_type} {ex}\n{traceback.format_exc()}")
            return self._get_default_result(indicator_type, len(chart_data.get('C', [])))
       
class CpData(QObject):
    def __init__(self, interval, chart_type, number, trader):
        super().__init__()
        self.interval = interval
        self.number = number
        self.chart_type = chart_type
        self.objCur = {}
        self.stockdata = {}
        self.objIndicators = {}
        self.code = ''
        self.LASTTIME = 1530
        self.trader = trader
        self.is_updating = {}
        self.is_initial_loaded = {}
        self.stockdata_lock = threading.Lock()
        self.last_update_time = {}
        
        # 실시간 최적화 관련 추가
        self.last_indicator_update = {}
        self.indicator_update_interval = 1.0  # 1초마다 지표 재계산
        self.latest_snapshot = {}  # 읽기 전용 스냅샷

        now = time.localtime()
        self.todayDate = now.tm_year * 10000 + now.tm_mon * 100 + now.tm_mday

        self.update_data_timer = QTimer()
        self.update_data_timer.timeout.connect(self.periodic_update_data)
        self.update_data_timer.start(10000)  # 10초마다 호출

    def periodic_update_data(self):
        try:
            current_time = time.time()
            with self.stockdata_lock:
                codes = list(self.stockdata.keys())
            
            for code in codes:
                if (code in self.trader.vistock_set and code not in self.trader.monistock_set 
                    and code not in self.trader.bought_set):
                    continue
                
                interval = 10 if code in self.trader.bought_set else 15

                last_time = self.last_update_time.get(code, 0)
                if current_time - last_time < interval:
                    continue

                with self.stockdata_lock:
                    if code not in self.stockdata:
                        logging.debug(f"{code}: stockdata에서 제거됨, 스킵")
                        continue
                    if self.is_updating.get(code, False):
                        logging.debug(f"{code}: 데이터 업데이트 진행 중, 스킵")
                        continue
                
                self.update_chart_data(code, self.interval, self.number)
                self.last_update_time[code] = current_time

                with self.stockdata_lock:
                    if code not in self.stockdata:
                        logging.debug(f"{code}: 업데이트 후 stockdata에서 제거됨, 스킵")
                        continue
                    if code not in self.objIndicators:
                        self.objIndicators[code] = CpIndicators(self.chart_type)
                        results = [
                            self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                            for ind in ["MA", "MACD", "RSI", "STOCH", "ATR", "CCI", "BBANDS", "VWAP"]
                        ]
                        if all(results):
                            for result in results:
                                self.stockdata[code].update(result)
                            self._update_snapshot(code)
                        else:
                            logging.warning(f"{code}: 지표 생성 실패")
        except Exception as ex:
            logging.error(f"periodic_update_data -> {ex}")

    def select_code(self, code):
        try:
            if code in self.stockdata:
                return True
            
            self.stockdata[code] = {
                'D': [], 'T': [], 'O': [], 'H': [], 'L': [], 'C': [], 'V': [], 'TV': [], 
                'MAD5': [], 'MAD10': [], 'VWAP': []
            }
            
            self.update_chart_data(code, self.interval, self.number)
            self.is_initial_loaded[code] = False
            
            if code not in self.objIndicators:
                self.objIndicators[code] = CpIndicators(self.chart_type)
                result = self.objIndicators[code].make_indicator("MA", code, self.stockdata[code])

                if result:
                    self.stockdata[code].update(result)
                    self._update_snapshot(code)
                    return True
                else:
                    return False
            return True
        except Exception as ex:
            logging.error(f"select_code -> {ex}")
            return False

    def monitor_code(self, code):
        try:
            if code in self.stockdata:
                return True
            
            self.stockdata[code] = {
                'D': [], 'T': [], 'O': [], 'H': [], 'L': [], 'C': [], 'V': [], 'TV': [], 
                'MAT5': [], 'MAT20': [], 'MAT60': [], 'MAT120': [], 
                'MAM5': [], 'MAM10': [], 'MAM20': [], 'MACDT': [], 'MACDT_SIGNAL': [], 'OSCT': [], 
                'MACD': [], 'MACD_SIGNAL': [], 'OSC': [], 
                'RSIT': [], 'RSIT_SIGNAL': [], 'RSI': [], 'RSI_SIGNAL': [], 
                'STOCHK': [], 'STOCHD': [], 'ATR': [], 'CCI': [],  
                'BB_UPPER': [], 'BB_MIDDLE': [], 'BB_LOWER': [], 'BB_POSITION': [], 'BB_BANDWIDTH': [], 
                'TICKS': [], 'MAT5_MAT20_DIFF': [], 'MAT20_MAT60_DIFF': [], 'MAT60_MAT120_DIFF': [], 
                'C_MAT5_DIFF': [], 'MAM5_MAM10_DIFF': [], 'MAM10_MAM20_DIFF': [], 
                'C_MAM5_DIFF': [], 'C_ABOVE_MAM5': [], 'VWAP': [], 
                'MAT5_CHANGE': [], 'MAT20_CHANGE': [], 'MAT60_CHANGE': [], 'MAT120_CHANGE': []
            }

            success = self.update_chart_data_from_market_open(code)
            
            if not success:
                logging.warning(f"{code}: 장 시작 데이터 로드 실패, 개수 기준으로 폴백")
                self.update_chart_data(code, self.interval, self.number)
                self.is_initial_loaded[code] = False
            else:
                self.is_initial_loaded[code] = True

            with self.stockdata_lock:
                if code not in self.objIndicators:
                    self.objIndicators[code] = CpIndicators(self.chart_type)
                    results = [
                        self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                        for ind in ["MA", "MACD", "RSI", "STOCH", "ATR", "CCI", "BBANDS", "VWAP"]
                    ]
                    if all(results):
                        for result in results:
                            self.stockdata[code].update(result)
                        self._update_snapshot(code)
                        if code not in self.objCur:
                            self.objCur[code] = CpPBStockCur()
                            self.objCur[code].Subscribe(code, self)
                        return True
                    else:
                        return False
            return True
        except Exception as ex:
            logging.error(f"monitor_code -> {code}, {ex}")
            return False

    def monitor_stop(self, code):
        try:
            if self.is_updating.get(code, False):
                logging.debug(f"{code}: 데이터 업데이트 진행 중, 1초 후 재시도")
                QTimer.singleShot(1000, lambda: self.monitor_stop(code))
                return
            with self.stockdata_lock:
                if code in self.objCur:
                    self.objCur[code].Unsubscribe()
                    del self.objCur[code]
                if code in self.stockdata:
                    del self.stockdata[code]
                if code in self.objIndicators:
                    del self.objIndicators[code]
                if code in self.is_updating:
                    del self.is_updating[code]
                if code in self.is_initial_loaded:
                    del self.is_initial_loaded[code]
                if code in self.last_indicator_update:
                    del self.last_indicator_update[code]
                if code in self.latest_snapshot:
                    del self.latest_snapshot[code]
        except Exception as ex:
            logging.error(f"monitor_stop -> {code}, {ex}")
            return False

    def _request_chart_data(self, code, request_type='count', count=None, start_date=None, end_date=None):
        """공통 차트 데이터 요청 로직"""
        try:
            objRq = win32com.client.Dispatch("CpSysDib.StockChart")
            objRq.SetInputValue(0, code)
            
            if request_type == 'count':
                objRq.SetInputValue(1, ord('2'))
                objRq.SetInputValue(4, count)
            elif request_type == 'period':
                objRq.SetInputValue(1, ord('1'))
                objRq.SetInputValue(2, end_date)
                objRq.SetInputValue(3, start_date)
            else:
                logging.error(f"잘못된 request_type: {request_type}")
                return None
            
            objRq.SetInputValue(5, [0, 1, 2, 3, 4, 5, 8, 9, 13])
            objRq.SetInputValue(6, ord(self.chart_type))
            objRq.SetInputValue(7, self.interval)
            objRq.SetInputValue(9, ord('1'))
            objRq.BlockRequest2(1)
            
            rqStatus = objRq.GetDibStatus()
            if rqStatus != 0:
                rqRet = objRq.GetDibMsg1()
                logging.warning(f"{code} 데이터 조회 실패, {rqStatus}, {rqRet}")
                return None
            
            len_data = objRq.GetHeaderValue(3)
            lastCount = objRq.GetHeaderValue(4)
            
            new_data = {
                'D': [], 'T': [], 'O': [], 'H': [], 'L': [], 'C': [], 'V': [], 'TV': [], 'TICKS': []
            }
            
            for i in range(len_data):
                new_data['D'].append(objRq.GetDataValue(0, i))
                new_data['T'].append(objRq.GetDataValue(1, i))
                new_data['O'].append(objRq.GetDataValue(2, i))
                new_data['H'].append(objRq.GetDataValue(3, i))
                new_data['L'].append(objRq.GetDataValue(4, i))
                new_data['C'].append(objRq.GetDataValue(5, i))
                new_data['V'].append(objRq.GetDataValue(6, i))
                new_data['TV'].append(objRq.GetDataValue(7, i))
                
                if self.chart_type == 'T':
                    if i == (len_data - 1):
                        new_data['TICKS'].append(lastCount)
                    else:
                        new_data['TICKS'].append(self.interval)
            
            for key in ['D', 'T', 'O', 'H', 'L', 'C', 'V', 'TV']:
                new_data[key].reverse()
            if self.chart_type == 'T':
                new_data['TICKS'].reverse()
            
            return new_data
            
        except Exception as ex:
            logging.error(f"_request_chart_data -> {code}, {ex}")
            return None

    def update_chart_data(self, code, interval, number):
        """실시간 증분 업데이트 (주기적 호출)"""
        try:
            self.is_updating[code] = True
            
            new_data = self._request_chart_data(code, request_type='count', count=number)
            
            if new_data is None:
                self.is_updating[code] = False
                return False
            
            with self.stockdata_lock:
                if code not in self.stockdata:
                    logging.debug(f"{code}: stockdata에 없음, 업데이트 중단")
                    self.is_updating[code] = False
                    return False
                
                if self.is_initial_loaded.get(code, False):
                    if self.stockdata[code].get('T') and len(self.stockdata[code]['T']) > 0:
                        last_time = self.stockdata[code]['T'][-1]
                        last_date = self.stockdata[code]['D'][-1]
                        
                        new_indices = [
                            i for i in range(len(new_data['T']))
                            if (new_data['D'][i] > last_date) or 
                               (new_data['D'][i] == last_date and new_data['T'][i] > last_time)
                        ]
                        
                        if new_indices:
                            for key in new_data:
                                filtered_data = [new_data[key][i] for i in new_indices]
                                self.stockdata[code][key].extend(filtered_data)
                            
                            if self.chart_type == 'T':
                                max_length = 400
                            elif self.chart_type == 'm':
                                max_length = 150
                            else:
                                max_length = 50
                            
                            for key in self.stockdata[code]:
                                if isinstance(self.stockdata[code][key], list):
                                    self.stockdata[code][key] = self.stockdata[code][key][-max_length:]
                    else:
                        for key in new_data:
                            self.stockdata[code][key] = new_data[key]
                else:
                    for key in new_data:
                        self.stockdata[code][key] = new_data[key]
            
            self.is_updating[code] = False
            return True
            
        except Exception as ex:
            logging.error(f"update_chart_data -> {ex}")
            self.is_updating[code] = False
            return False

    def update_chart_data_from_market_open(self, code):
        """장 시작부터 전체 로딩 (초기 1회)"""
        try:
            self.is_updating[code] = True
            
            today = datetime.now().strftime('%Y%m%d')
            
            new_data = self._request_chart_data(
                code, 
                request_type='period',
                start_date=today,
                end_date=today
            )
            
            if new_data is None:
                self.is_updating[code] = False
                return False
            
            with self.stockdata_lock:
                if code not in self.stockdata:
                    logging.debug(f"{code}: stockdata에 없음, 업데이트 중단")
                    self.is_updating[code] = False
                    return False
                
                for key in new_data:
                    self.stockdata[code][key] = new_data[key]
            
            logging.info(f"{code}: 장 시작부터 데이터 로드 완료 ({len(new_data['D'])}개, {self.chart_type})")
            self.is_updating[code] = False
            return True
            
        except Exception as ex:
            logging.error(f"update_chart_data_from_market_open -> {ex}")
            self.is_updating[code] = False
            return False

    def verify_data_coverage(self, code):
        """데이터가 장 시작부터 커버하는지 확인"""
        try:
            with self.stockdata_lock:
                if code not in self.stockdata:
                    return False
                
                stock_data = self.stockdata[code]
                if not stock_data.get('T') or len(stock_data['T']) == 0:
                    return False
                
                first_time = stock_data['T'][0]
                market_open = 900
                
                if first_time > market_open + 30:
                    logging.warning(
                        f"{code}: {self.chart_type} 데이터가 {first_time}부터 시작 "
                        f"(장 시작 데이터 부족, 총 {len(stock_data['T'])}개)"
                    )
                    return False
                
                logging.info(
                    f"{code}: {self.chart_type} 데이터 커버리지 양호 "
                    f"({first_time}부터 시작, 총 {len(stock_data['T'])}개)"
                )
                return True
                
        except Exception as ex:
            logging.error(f"verify_data_coverage -> {ex}")
            return False

    def _update_snapshot(self, code):
        """읽기 전용 스냅샷 업데이트 (락 내부에서 호출)"""
        try:
            if code not in self.stockdata:
                return
            
            data = self.stockdata[code]
            
            # 틱 데이터용 스냅샷
            if self.chart_type == 'T':
                self.latest_snapshot[code] = {
                    'C': data.get('C', [0])[-1] if data.get('C') else 0,
                    'O': data.get('O', [0])[-1] if data.get('O') else 0,
                    'H': data.get('H', [0])[-1] if data.get('H') else 0,
                    'L': data.get('L', [0])[-1] if data.get('L') else 0,
                    'V': data.get('V', [0])[-1] if data.get('V') else 0,
                    'MAT5': data.get('MAT5', [0])[-1] if data.get('MAT5') else 0,
                    'MAT20': data.get('MAT20', [0])[-1] if data.get('MAT20') else 0,
                    'MAT60': data.get('MAT60', [0])[-1] if data.get('MAT60') else 0,
                    'MAT120': data.get('MAT120', [0])[-1] if data.get('MAT120') else 0,
                    'RSIT': data.get('RSIT', [0])[-1] if data.get('RSIT') else 0,
                    'RSIT_SIGNAL': data.get('RSIT_SIGNAL', [0])[-1] if data.get('RSIT_SIGNAL') else 0,
                    'MACDT': data.get('MACDT', [0])[-1] if data.get('MACDT') else 0,
                    'MACDT_SIGNAL': data.get('MACDT_SIGNAL', [0])[-1] if data.get('MACDT_SIGNAL') else 0,
                    'OSCT': data.get('OSCT', [0])[-1] if data.get('OSCT') else 0,
                    'STOCHK': data.get('STOCHK', [0])[-1] if data.get('STOCHK') else 0,
                    'STOCHD': data.get('STOCHD', [0])[-1] if data.get('STOCHD') else 0,
                    'ATR': data.get('ATR', [0])[-1] if data.get('ATR') else 0,
                    'CCI': data.get('CCI', [0])[-1] if data.get('CCI') else 0,
                    'BB_UPPER': data.get('BB_UPPER', [0])[-1] if data.get('BB_UPPER') else 0,
                    'BB_MIDDLE': data.get('BB_MIDDLE', [0])[-1] if data.get('BB_MIDDLE') else 0,
                    'BB_LOWER': data.get('BB_LOWER', [0])[-1] if data.get('BB_LOWER') else 0,
                    'BB_POSITION': data.get('BB_POSITION', [0])[-1] if data.get('BB_POSITION') else 0,
                    'BB_BANDWIDTH': data.get('BB_BANDWIDTH', [0])[-1] if data.get('BB_BANDWIDTH') else 0,
                    'VWAP': data.get('VWAP', [0])[-1] if data.get('VWAP') else 0,
                    # 최근 N개 값도 포함 (전략 평가용)
                    'C_recent': data.get('C', [0])[-3:] if data.get('C') else [0, 0, 0],
                    'H_recent': data.get('H', [0])[-3:] if data.get('H') else [0, 0, 0],
                    'L_recent': data.get('L', [0])[-3:] if data.get('L') else [0, 0, 0],
                }
            
            # 분 데이터용 스냅샷
            elif self.chart_type == 'm':
                self.latest_snapshot[code] = {
                    'C': data.get('C', [0])[-1] if data.get('C') else 0,
                    'O': data.get('O', [0])[-1] if data.get('O') else 0,
                    'H': data.get('H', [0])[-1] if data.get('H') else 0,
                    'L': data.get('L', [0])[-1] if data.get('L') else 0,
                    'V': data.get('V', [0])[-1] if data.get('V') else 0,
                    'MAM5': data.get('MAM5', [0])[-1] if data.get('MAM5') else 0,
                    'MAM10': data.get('MAM10', [0])[-1] if data.get('MAM10') else 0,
                    'MAM20': data.get('MAM20', [0])[-1] if data.get('MAM20') else 0,
                    'RSI': data.get('RSI', [0])[-1] if data.get('RSI') else 0,
                    'RSI_SIGNAL': data.get('RSI_SIGNAL', [0])[-1] if data.get('RSI_SIGNAL') else 0,
                    'MACD': data.get('MACD', [0])[-1] if data.get('MACD') else 0,
                    'MACD_SIGNAL': data.get('MACD_SIGNAL', [0])[-1] if data.get('MACD_SIGNAL') else 0,
                    'OSC': data.get('OSC', [0])[-1] if data.get('OSC') else 0,
                    'STOCHK': data.get('STOCHK', [0])[-1] if data.get('STOCHK') else 0,
                    'STOCHD': data.get('STOCHD', [0])[-1] if data.get('STOCHD') else 0,
                    'CCI': data.get('CCI', [0])[-1] if data.get('CCI') else 0,
                    'VWAP': data.get('VWAP', [0])[-1] if data.get('VWAP') else 0,
                    # 최근 N개 값도 포함
                    'C_recent': data.get('C', [0])[-2:] if data.get('C') else [0, 0],
                    'O_recent': data.get('O', [0])[-2:] if data.get('O') else [0, 0],
                    'H_recent': data.get('H', [0])[-2:] if data.get('H') else [0, 0],
                    'L_recent': data.get('L', [0])[-2:] if data.get('L') else [0, 0],
                }
            
            # 일 데이터용 스냅샷
            elif self.chart_type == 'D':
                self.latest_snapshot[code] = {
                    'C': data.get('C', [0])[-1] if data.get('C') else 0,
                    'V': data.get('V', [0])[-1] if data.get('V') else 0,
                    'MAD5': data.get('MAD5', [0])[-1] if data.get('MAD5') else 0,
                    'MAD10': data.get('MAD10', [0])[-1] if data.get('MAD10') else 0,
                    'VWAP': data.get('VWAP', [0])[-1] if data.get('VWAP') else 0,
                }
                
        except Exception as ex:
            logging.error(f"_update_snapshot -> {code}, {ex}")

    def get_latest_data(self, code):
        """빠른 읽기 - 스냅샷 반환 (락 불필요)"""
        return self.latest_snapshot.get(code, {})
    
    def get_full_data(self, code):
        """전체 데이터 읽기 (락 필요, 차트 그리기용)"""
        with self.stockdata_lock:
            return copy.deepcopy(self.stockdata.get(code, {}))
    
    def get_recent_data(self, code, count=10):
        """최근 N개 데이터 읽기 (락 필요)"""
        with self.stockdata_lock:
            if code not in self.stockdata:
                return {}
            
            data = self.stockdata[code]
            result = {}
            for key, values in data.items():
                if isinstance(values, list) and len(values) > 0:
                    result[key] = values[-count:] if len(values) >= count else values
                else:
                    result[key] = values
            return result

    def updateCurData(self, item):
        """실시간 체결 데이터 업데이트"""
        try:
            if self.is_updating.get(item['code'], False):
                return
            
            code = item['code']
            time_val = item['time']
            cur = item['cur']
            vol = item['vol']
            current_time = time.time()

            with self.stockdata_lock:
                if self.chart_type == 'T':
                    hh, mm = divmod(time_val, 10000)
                    mm, tt = divmod(mm, 100)
                    if mm == 60:
                        hh += 1
                        mm = 0
                    lCurTime = hh * 100 + mm

                    bFind = False
                    if lCurTime > self.LASTTIME:
                        lCurTime = self.LASTTIME

                    if code in self.stockdata:
                        if len(self.stockdata[code]['T']) > 0:
                            lastCount = self.stockdata[code]['TICKS'][-1]
                            if 1 <= lastCount < 60:
                                bFind = True
                                self.stockdata[code]['T'][-1] = lCurTime
                                self.stockdata[code]['C'][-1] = cur
                                if self.stockdata[code]['H'][-1] < cur:
                                    self.stockdata[code]['H'][-1] = cur
                                if self.stockdata[code]['L'][-1] > cur:
                                    self.stockdata[code]['L'][-1] = cur
                                self.stockdata[code]['V'][-1] += vol
                                self.stockdata[code]['TICKS'][-1] += 1

                        if not bFind:
                            self.stockdata[code]['D'].append(self.todayDate)
                            self.stockdata[code]['T'].append(lCurTime)
                            self.stockdata[code]['O'].append(cur)
                            self.stockdata[code]['H'].append(cur)
                            self.stockdata[code]['L'].append(cur)
                            self.stockdata[code]['C'].append(cur)
                            self.stockdata[code]['V'].append(vol)
                            self.stockdata[code]['TICKS'].append(1)

                        # 데이터 길이 유지
                        desired_length = 400
                        for key in self.stockdata[code]:
                            self.stockdata[code][key] = self.stockdata[code][key][-desired_length:]

                        # 스냅샷 즉시 업데이트 (가격/거래량)
                        self._update_snapshot(code)

                        # 지표는 1초마다만 재계산
                        last_update = self.last_indicator_update.get(code, 0)
                        if current_time - last_update >= self.indicator_update_interval:
                            if code in self.objIndicators:
                                results = [
                                    self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                                    for ind in ["MA", "RSI", "MACD", "STOCH", "ATR", "CCI", "BBANDS", "VWAP"]
                                ]
                                if all(results):
                                    for result in results:
                                        self.stockdata[code].update(result)
                                    # 지표 업데이트 후 스냅샷 갱신
                                    self._update_snapshot(code)
                            
                            self.last_indicator_update[code] = current_time
                    
                elif self.chart_type == 'm':
                    hh, mm = divmod(time_val, 10000)
                    mm, tt = divmod(mm, 100)
                    convertedMintime = hh * 60 + mm

                    bFind = False
                    a, b = divmod(convertedMintime, self.interval)
                    intervaltime = a * self.interval
                    lChartTime = intervaltime + self.interval
                    hour, minute = divmod(lChartTime, 60)
                    lCurTime = hour * 100 + minute

                    if lCurTime > self.LASTTIME:
                        lCurTime = self.LASTTIME

                    if code in self.stockdata:
                        if len(self.stockdata[code]['T']) > 0:
                            lLastTime = self.stockdata[code]['T'][-1]
                            if lLastTime == lCurTime:
                                bFind = True
                                self.stockdata[code]['C'][-1] = cur
                                if self.stockdata[code]['H'][-1] < cur:
                                    self.stockdata[code]['H'][-1] = cur
                                if self.stockdata[code]['L'][-1] > cur:
                                    self.stockdata[code]['L'][-1] = cur
                                self.stockdata[code]['V'][-1] += vol
        
                        if not bFind:
                            self.stockdata[code]['D'].append(self.todayDate)
                            self.stockdata[code]['T'].append(lCurTime)
                            self.stockdata[code]['O'].append(cur)
                            self.stockdata[code]['H'].append(cur)
                            self.stockdata[code]['L'].append(cur)
                            self.stockdata[code]['C'].append(cur)
                            self.stockdata[code]['V'].append(vol)

                        # 데이터 길이 유지
                        desired_length = 150
                        for key in self.stockdata[code]:
                            self.stockdata[code][key] = self.stockdata[code][key][-desired_length:]

                        # 스냅샷 즉시 업데이트
                        self._update_snapshot(code)

                        # 지표는 1초마다만 재계산
                        last_update = self.last_indicator_update.get(code, 0)
                        if current_time - last_update >= self.indicator_update_interval:
                            if code in self.objIndicators:
                                results = [
                                    self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                                    for ind in ["MA", "MACD", "RSI", "STOCH", "ATR", "CCI", "BBANDS", "VWAP"]
                                ]
                                if all(results):
                                    for result in results:
                                        self.stockdata[code].update(result)
                                    self._update_snapshot(code)
                            
                            self.last_indicator_update[code] = current_time
                
        except Exception as ex:
            logging.error(f"updateCurData -> {ex}")

class DatabaseWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, db_name, queue, tickdata, mindata, parent=None):
        super().__init__(parent)
        self.db_name = db_name
        self.queue = queue  # queue.Queue 인스턴스
        self.tickdata = tickdata
        self.mindata = mindata
        self.running = True
        self.last_saved_timestamp = {}

    def run(self):
        while self.running:
            try:
                # 큐에서 항목을 비차단 방식으로 가져오기
                item = self.queue.get(timeout=1)
                if item is None:
                    self.queue.task_done()
                    break
                code, tick_data_copy, min_data_copy, start_date, start_hhmm = item
                try:
                    self.save_vi_data(code, tick_data_copy, min_data_copy, start_date, start_hhmm)
                except Exception as ex:
                    error_msg = f"DB Worker 오류 ({code}): {ex}"
                    logging.error(error_msg)
                    self.error.emit(error_msg)
                finally:
                    self.queue.task_done()  # 작업 완료 표시
            except queue.Empty:
                continue
            except Exception as ex:
                logging.error(f"Queue processing error: {ex}")
        self.finished.emit()

    def stop(self):
        self.running = False
        self.queue.put(None)

    def save_vi_data(self, code, tick_data, min_data, start_date, start_hhmm):
        try:
            with sqlite3.connect(self.db_name, timeout=10) as conn:
                conn.execute("PRAGMA journal_mode=WAL")
                c = conn.cursor()
                c.execute("BEGIN TRANSACTION")

                # last_saved_timestamp 초기화
                if code not in self.last_saved_timestamp:
                    c.execute("SELECT MAX(date), MAX(time) FROM tick_data WHERE code = ?", (code,))
                    result = c.fetchone()
                    self.last_saved_timestamp[code] = result if result[0] else ("00000000", "0000")
                    logging.debug(f"{code}: last_saved_timestamp 초기화 - {self.last_saved_timestamp[code]}")

                last_date, last_time = self.last_saved_timestamp.get(code, ("00000000", "0000"))
                
                # 장 시작부터 저장 (start_hhmm은 항상 0900)
                start_hhmm = "0900"  # 항상 장 시작부터
                end_hhmm = "1515"

                # 틱 데이터 저장
                if tick_data and tick_data["T"]:
                    dates, times = tick_data["D"], tick_data["T"]
                    values = [tick_data[key] for key in ["C", "V", "MAT5", "MAT20", "MAT60", "MAT120", "RSIT",
                            "RSIT_SIGNAL", "MACDT", "MACDT_SIGNAL", "OSCT", "STOCHK", "STOCHD", "ATR",
                            "CCI", "BB_UPPER", "BB_MIDDLE", "BB_LOWER", "BB_POSITION", "BB_BANDWIDTH",
                            "MAT5_MAT20_DIFF", "MAT20_MAT60_DIFF", "MAT60_MAT120_DIFF",
                            "C_MAT5_DIFF", "VWAP"]]
                    
                    new_data = [
                        (i, str(date), f"{int(time_val):04d}")
                        for i, (date, time_val) in enumerate(zip(dates, times))
                        if len(str(date)) == 8 and len(f"{int(time_val):04d}") == 4
                        and str(date) == start_date and start_hhmm <= f"{int(time_val):04d}" <= end_hhmm
                        and (str(date) > last_date or (str(date) == last_date and f"{int(time_val):04d}" >= last_time))
                    ]

                    if new_data:
                        unique_date_times = set((date, time) for _, date, time in new_data)
                        for date, time in unique_date_times:
                            c.execute("DELETE FROM tick_data WHERE code = ? AND date = ? AND time = ?", (code, date, time))
                        
                        inserted_count = 0
                        for date, time in unique_date_times:
                            for seq, (i, _, _) in enumerate(sorted([(i, d, t) for i, d, t in new_data if d == date and t == time], key=lambda x: x[2])):
                                c.execute("""INSERT INTO tick_data
                                            (code, date, time, sequence, C, V, MAT5, MAT20, MAT60, MAT120, RSIT,
                                            RSIT_SIGNAL, MACDT, MACDT_SIGNAL, OSCT, STOCHK, STOCHD, ATR, CCI,
                                            BB_UPPER, BB_MIDDLE, BB_LOWER, BB_POSITION, BB_BANDWIDTH, MAT5_MAT20_DIFF,
                                            MAT20_MAT60_DIFF, MAT60_MAT120_DIFF, C_MAT5_DIFF, VWAP)
                                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                        (code, date, time, seq, *[val[i] for val in values]))
                                inserted_count += 1

                        if inserted_count:
                            latest_date = max(date for _, date, _ in new_data)
                            latest_time = max(time for _, date, time in new_data if date == latest_date)
                            self.last_saved_timestamp[code] = (latest_date, latest_time)
                            logging.debug(f"{code}: {inserted_count}개 틱 데이터 저장 (장시작부터)")

                # 분 데이터 저장 (동일하게 start_hhmm = "0900")
                if min_data and min_data["T"]:
                    c.execute("SELECT MAX(date), MAX(time) FROM min_data WHERE code = ?", (code,))
                    result = c.fetchone()
                    last_min_date, last_min_time = result if result[0] else ("00000000", "0000")

                    dates, times = min_data["D"], min_data["T"]
                    values = [min_data[key] for key in ["C", "V", "MAM5", "MAM10", "MAM20", "RSI", "RSI_SIGNAL",
                            "MACD", "MACD_SIGNAL", "OSC", "STOCHK", "STOCHD", "CCI",
                            "MAM5_MAM10_DIFF", "MAM10_MAM20_DIFF", "C_MAM5_DIFF", "C_ABOVE_MAM5", "VWAP"]]
                    
                    new_indices = [
                        (i, str(date), f"{int(time_val):04d}")
                        for i, (date, time_val) in enumerate(zip(dates, times))
                        if len(str(date)) == 8 and f"{int(time_val):04d}"[2:4].isdigit() and int(f"{int(time_val):04d}"[2:4]) % 3 == 0
                        and str(date) == start_date and start_hhmm <= f"{int(time_val):04d}" <= end_hhmm
                        and (str(date) > last_min_date or (str(date) == last_min_date and f"{int(time_val):04d}" >= last_min_time))
                    ]

                    if new_indices:
                        unique_date_times = set((date, time) for _, date, time in new_indices)
                        for date, time in unique_date_times:
                            c.execute("DELETE FROM min_data WHERE code = ? AND date = ? AND time = ?", (code, date, time))
                        
                        inserted_count = 0
                        for i, date, time in new_indices:
                            c.execute("""INSERT INTO min_data
                                        (code, date, time, sequence, C, V, MAM5, MAM10, MAM20, RSI, RSI_SIGNAL,
                                        MACD, MACD_SIGNAL, OSC, STOCHK, STOCHD, CCI, MAM5_MAM10_DIFF,
                                        MAM10_MAM20_DIFF, C_MAM5_DIFF, C_ABOVE_MAM5, VWAP)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                    (code, date, time, 0, *[val[i] for val in values]))
                            inserted_count += 1

                        if inserted_count:
                            latest_date = max(date for _, date, _ in new_indices)
                            latest_time = max(time for _, date, time in new_indices if date == latest_date)
                            self.last_saved_timestamp[code] = (latest_date, latest_time)
                            logging.debug(f"{code}: {inserted_count}개 분 데이터 저장 (장시작부터)")

                conn.commit()

        except sqlite3.OperationalError as e:
            logging.error(f"{code}: DB 저장 오류 (OperationalError) - {e}")
            conn.rollback()
        except Exception as ex:
            logging.error(f"{code}: VI 데이터 저장 오류 - {ex}")
            conn.rollback()

class CTrader(QObject):
    stock_added_to_monitor = pyqtSignal(str)
    stock_bought = pyqtSignal(str)
    stock_sold = pyqtSignal(str)

    def __init__(self, obj_cp_trade, cp_balance, cpCodeMgr, cp_cash, cp_order, cp_stock, buycount, window):
        super().__init__()
        self.cpTrade = obj_cp_trade
        self.cpBalance = cp_balance
        self.cpCodeMgr = cpCodeMgr
        self.cpCash = cp_cash
        self.cpOrder = cp_order
        self.cpStock = cp_stock
        self.target_buy_count = buycount

        self.bought_set = set()
        self.database_set = set()
        self.monistock_set = set()
        self.vistock_set = set()
        self.sell_half_set = set()
        self.buyorder_set = set()
        self.buyordering_set = set()
        self.sellorder_set = set()
        
        self.starting_price = {}
        self.starting_time = {}
        self.highest_price = {}
        self.buy_price = {}
        self.buy_qty = {}
        self.buyorder_qty = {}
        self.sell_half_qty = {}
        self.balance_qty = {}
        self.sell_amount = {}
        self.buy_amount = 0
        self.conclusion = CpPBConclusion()
        self.conclusion.Subscribe('', self)
        self.window = window

        self.cp_request = CpRequest()

        self.daydata = CpData(1, 'D', 50, self)
        self.mindata = CpData(3, 'm', 150, self)   # 110 → 150 (450분 = 7.5시간)
        self.tickdata = CpData(60, 'T', 400, self) # 300 → 400 (24,000틱, 충분한 여유)        

        self.last_saved_timestamp = {}
        self.db_name = 'vi_stock_data.db'
        # self.init_database()

        self.db_queue = queue.Queue()
        self.db_thread = QThread()
        self.db_worker = DatabaseWorker(self.db_name, self.db_queue, self.tickdata, self.mindata)
        self.db_worker.moveToThread(self.db_thread)
        self.db_thread.started.connect(self.db_worker.run)
        self.db_worker.finished.connect(self.db_thread.quit)
        self.db_worker.error.connect(lambda msg: logging.error(msg))
        self.db_thread.start()

        self.save_data_timer = QTimer()
        self.save_data_timer.timeout.connect(self.periodic_save_vi_data)
        self.save_data_timer.start(600000)  # 10분마다 호출

    def init_database(self):
        try:
            if os.path.exists(self.db_name):
                os.remove(self.db_name)
                logging.debug(f"{self.db_name} 삭제 완료")
            else:
                logging.debug(f"{self.db_name} 파일이 이미 존재하지 않음")

            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            # tick_data 테이블
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tick_data (
                    code TEXT, date TEXT, time TEXT, sequence INTEGER, C REAL, V INTEGER,
                    MAT5 REAL, MAT20 REAL, MAT60 REAL, MAT120 REAL, RSIT REAL, RSIT_SIGNAL REAL,
                    MACDT REAL, MACDT_SIGNAL REAL, OSCT REAL, STOCHK REAL, STOCHD REAL, 
                    ATR REAL, CCI REAL, BB_UPPER REAL, BB_MIDDLE REAL, BB_LOWER REAL, BB_POSITION REAL, BB_BANDWIDTH REAL,
                    MAT5_MAT20_DIFF REAL, MAT20_MAT60_DIFF REAL, MAT60_MAT120_DIFF REAL,
                    C_MAT5_DIFF REAL, VWAP REAL,
                    PRIMARY KEY (code, date, time, sequence)
                )
            ''')
            # min_data 테이블
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS min_data (
                    code TEXT, date TEXT, time TEXT, sequence INTEGER, C REAL, V INTEGER,
                    MAM5 REAL, MAM10 REAL, MAM20 REAL, RSI REAL, RSI_SIGNAL REAL,
                    MACD REAL, MACD_SIGNAL REAL, OSC REAL, STOCHK REAL, STOCHD REAL,
                    CCI REAL, MAM5_MAM10_DIFF REAL, MAM10_MAM20_DIFF REAL,
                    C_MAM5_DIFF REAL, C_ABOVE_MAM5 REAL, VWAP REAL,
                    PRIMARY KEY (code, date, time, sequence)
                )
            ''')
            # tick_data 중복 데이터 정리 (재시작 시 단일 데이터 유지)
            cursor.execute('''
                DELETE FROM tick_data
                WHERE rowid NOT IN (
                    SELECT MAX(rowid)
                    FROM tick_data
                    GROUP BY code, date, time, sequence
                )
            ''')
            # min_data 중복 데이터 정리
            cursor.execute('''
                DELETE FROM min_data
                WHERE rowid NOT IN (
                    SELECT MAX(rowid)
                    FROM min_data
                    GROUP BY code, date, time
                )
            ''')
            conn.commit()
            conn.close()
        except Exception as ex:
            logging.error(f"init_database -> {ex}")
            raise

    def save_vi_data(self, code):
        try:
            start_date, start_hhmm = None, None
            try:
                if code in self.starting_time:
                    current_year = datetime.now().year
                    time_str = self.starting_time[code]
                    start_dt = datetime.strptime(f"{current_year}/{time_str}", '%Y/%m/%d %H:%M:%S')
                    start_date = start_dt.strftime('%Y%m%d')
                    start_hhmm = start_dt.strftime('%H%M')
                else:
                    logging.warning(f"{code}: starting_time 없음, 기본값 0900 사용")
                    start_date = datetime.now().strftime('%Y%m%d')
                    start_hhmm = '0900'
            except ValueError as ve:
                logging.error(f"{code}: starting_time 형식 오류 - {self.starting_time.get(code, '없음')}: {ve}")
                start_date = datetime.now().strftime('%Y%m%d')
                start_hhmm = '0900'

            # stockdata 복사본 생성
            with self.tickdata.stockdata_lock:
                tick_data_copy = copy.deepcopy(self.tickdata.stockdata.get(code, {}))
            with self.mindata.stockdata_lock:
                min_data_copy = copy.deepcopy(self.mindata.stockdata.get(code, {}))

            self.db_queue.put((code, tick_data_copy, min_data_copy, start_date, start_hhmm))
            # logging.debug(f"{code}: DB 저장 요청 큐에 추가, 큐 크기={self.db_queue.qsize()}")
        except Exception as ex:
            logging.error(f"save_vi_data -> {code}, {ex}")

    def periodic_save_vi_data(self):
        try:
            with self.tickdata.stockdata_lock:
                codes = list(self.tickdata.stockdata.keys())
            for code in codes:
                # logging.debug(f"Saving VI data for {code}")
                self.save_vi_data(code)
        except Exception as ex:
            logging.error(f"periodic_save_vi_data -> {ex}")

    def get_stock_balance(self, code, func):
        try:
            stock_name = self.cpCodeMgr.CodeToName(code)
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)
            self.cpBalance.SetInputValue(0, acc) # 계좌번호
            self.cpBalance.SetInputValue(1, accFlag[0]) # 상품관리구분코드
            self.cpBalance.SetInputValue(2, 50) # 요청건수
            ret = self.cpBalance.BlockRequest2(1)
            if ret != 0:
                logging.warning(f"{stock_name}({code}) 잔고 조회 실패, {ret}")
                return False
            
            if code == 'START':
                logging.info(f"계좌명 : {str(self.cpBalance.GetHeaderValue(0))}")
                logging.info(f"결제잔고수량 : {str(self.cpBalance.GetHeaderValue(1))}")
                logging.info(f"평가금액 : {str(self.cpBalance.GetHeaderValue(3))}")
                logging.info(f"평가손익 : {str(self.cpBalance.GetHeaderValue(4))}")
                logging.info(f"종목수 : {str(self.cpBalance.GetHeaderValue(7))}")
                return

            stocks = []
            for i in range(self.cpBalance.GetHeaderValue(7)):
                stock_code = self.cpBalance.GetDataValue(12, i)
                stock_name = self.cpBalance.GetDataValue(0, i)
                stock_qty = self.cpBalance.GetDataValue(15, i)
                buy_price = self.cpBalance.GetDataValue(17, i)
                stocks.append({'code': stock_code, 'name': stock_name, 'qty': stock_qty, 'buy_price': buy_price})

            if code == 'ALL':
                logging.debug("잔고 전부 조회 성공")
                return stocks

            for s in stocks:
                if s['code'] == code:
                    logging.debug(f"{s['name']}({s['code']}) 잔고 조회 성공")
                    return s['name'], s['qty'], s['buy_price']             

        except Exception as ex:
            logging.error(f"get_stock_balance({func}) -> {ex}")
            return False

    def init_stock_balance(self):
        try:
            stocks = self.get_stock_balance('ALL', 'init_stock_balance')

            for s in stocks:
                if self.daydata.select_code(s['code']) and self.tickdata.monitor_code(s['code']) and self.mindata.monitor_code(s['code']):
                    if s['code'] not in self.starting_time:
                        self.starting_time[s['code']] = datetime.now().strftime('%m/%d 09:00:00')
                    self.monistock_set.add(s['code'])
                    self.stock_added_to_monitor.emit(s['code'])
                    self.bought_set.add(s['code'])
                    self.stock_bought.emit(s['code'])
                    self.buy_price[s['code']] = s['buy_price']
                    self.buy_qty[s['code']] = s['qty']
                    self.balance_qty[s['code']] = s['qty']

            remaining_count = self.target_buy_count - len(stocks)
            self.buy_percent = 1/remaining_count if remaining_count > 0 else 0
            self.total_cash = self.get_current_cash() * 0.9
            self.buy_amount = int(self.total_cash * self.buy_percent)
            
            logging.info(f"주문 가능 금액 : {self.total_cash}")
            logging.info(f"종목별 주문 비율 : {self.buy_percent}")
            logging.info(f"종목별 주문 금액 : {self.buy_amount}")

        except Exception as ex:
            logging.error(f"init_stock_balance -> {ex}")

    def get_current_cash(self):
        try:
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)
            self.cpCash.SetInputValue(0, acc)  # 계좌번호
            self.cpCash.SetInputValue(1, accFlag[0]) # 상품관리구분코드
            self.cpCash.SetInputValue(5, "Y")
            self.cpCash.BlockRequest2(1)

            rqStatus = self.cpCash.GetDibStatus()
            if rqStatus != 0:
                rqRet = self.cpCash.GetDibMsg1()
                logging.warning(f"주문가능금액 조회실패, {rqStatus}, {rqRet}")
                return (False, '')
            current_cash = self.cpCash.GetHeaderValue(9) # 증거금100%주문가능금액
            return current_cash
        except Exception as ex:
            logging.error(f"get_current_cash -> {ex}")
            return False

    # def get_current_qty(self, code, ask_price, order_style):
    #     try:
    #         acc = self.cpTrade.AccountNumber[0]
    #         accFlag = self.cpTrade.GoodsList(acc, 1)
    #         self.cpCash.SetInputValue(0, acc)  # 계좌번호
    #         self.cpCash.SetInputValue(1, accFlag[0]) # 상품관리구분코드
    #         self.cpCash.SetInputValue(2, code)  # 종목코드
    #         self.cpCash.SetInputValue(3, order_style)
    #         self.cpCash.SetInputValue(4, ask_price)
    #         self.cpCash.SetInputValue(5, 'Y')
    #         self.cpCash.SetInputValue(6, ord('2'))

    #         ret = self.cpCash.BlockRequest2(1)
    #         if ret != 0:
    #             logging.warning(f"{code} 주문가능수량 조회 실패, {ret}")
    #             return False

    #         rqStatus = self.cpCash.GetDibStatus()
    #         if rqStatus != 0:
    #             rqRet = self.cpCash.GetDibMsg1()
    #             logging.warning(f"주문가능수량 조회 오류, {rqStatus}, {rqRet}")
    #             return False
    #         current_qty = self.cpCash.GetHeaderValue(17) # 증거금100%주문가능수량
    #         return current_qty
        
    #     except Exception as ex:
    #         logging.error(f"get_current_qty -> {ex}")
    #         return False

    def get_current_price(self, code):
        try:
            self.cpStock.SetInputValue(0, code)
            ret = self.cpStock.BlockRequest2(1)
            if ret != 0:
                logging.warning(f"{code} 현재가 조회 실패, {ret}")
                return False

            rqStatus = self.cpStock.GetDibStatus()
            if rqStatus != 0:
                rqRet = self.cpStock.GetDibMsg1()
                logging.warning(f"현재가 조회 오류, {rqStatus}, {rqRet}")
                return False
            
            item = {'cur_price': self.cpStock.GetHeaderValue(11), 'ask': self.cpStock.GetHeaderValue(16), 'upper': self.cpStock.GetHeaderValue(8)}
            return item['cur_price'], item['ask'], item['upper']
        
        except Exception as ex:
            logging.error(f"get_current_price -> {ex}")
            return False
    
    def get_trade_profit(self):
        try:
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            self.objRq = win32com.client.Dispatch("CpTrade.CpTd6032")
            self.objRq.SetInputValue(0, acc)
            self.objRq.SetInputValue(1, accFlag[0])
            ret = self.objRq.BlockRequest2(1)
            if ret != 0:
                logging.warning(f"매매실현손익 조회 실패, {ret}")
                return
            
            rqStatus = self.objRq.GetDibStatus()
            if rqStatus != 0:
                rqRet = self.objRq.GetDibMsg1()
                logging.warning(f"매매실현손익 조회 오류, {rqStatus}, {rqRet}")
                return False
            
            profit = self.objRq.GetHeaderValue(2)
            logging.info(f"매매실현손익 : {self.objRq.GetHeaderValue(2)}원")
            send_slack_message(self.window.login_handler, "#stock", f"매매실현손익 : {profit}원")

        except Exception as ex:
            logging.error(f"get_trade_profit -> {ex}")
    
    def update_highest_price(self, code, current_price):
        if code not in self.highest_price:
            self.highest_price[code] = current_price
        elif current_price > self.highest_price[code]:
            self.highest_price[code] = current_price    

    def monitor_vi(self, time, code, event2):
        try:
            # 중복 및 개수 제한 체크
            if code in self.monistock_set or len(self.monistock_set) >= 10 or code in self.bought_set:
                return

            # 1️⃣ 일봉 데이터 로드
            if not self.daydata.select_code(code):
                logging.debug(f"{code}: 일봉 데이터 로드 실패")
                return
            
            # 2️⃣ 추세 확인 (다층 이동평균)
            day_data = self.daydata.stockdata[code]
            if not (day_data['MAD5'][-1] > day_data['MAD10'][-1]):
                logging.debug(f"{code}: MAD5 < MAD10, 추세 미달")
                self.daydata.monitor_stop(code)
                return
            
            # 추세 강도 확인
            recent_ma5 = day_data['MAD5'][-3:]
            if len(recent_ma5) >= 3 and not all(recent_ma5[i] < recent_ma5[i+1] for i in range(2)):
                logging.debug(f"{code}: MAD5 추세 약화")
                self.daydata.monitor_stop(code)
                return
            
            # 3️⃣ 유동성 검증
            if len(day_data['V']) < 30:
                logging.debug(f"{code}: 데이터 부족 ({len(day_data['V'])}일)")
                self.daydata.monitor_stop(code)
                return
            
            Trading_amount = sum(day_data['V'][-30:]) / 30
            Trading_Value = sum(day_data['TV'][-30:]) / 30
            
            MIN_VOLUME = 100000
            MIN_VALUE = 3000000000
            
            if Trading_amount < MIN_VOLUME or Trading_Value < MIN_VALUE:
                logging.debug(
                    f"{code}: 유동성 부족 - "
                    f"거래량 {Trading_amount:.0f}/{MIN_VOLUME}, "
                    f"거래금액 {Trading_Value/100000000:.1f}/{MIN_VALUE/100000000}억"
                )
                self.daydata.monitor_stop(code)
                return
            
            # 4️⃣ 시간대별 거래량 기준
            now = datetime.now()
            elapsed_minutes = (now - now.replace(hour=9, minute=0, second=0)).total_seconds() / 60
            
            if elapsed_minutes <= 120:
                volume_threshold = 0.5
            elif elapsed_minutes <= 240:
                volume_threshold = 0.7
            else:
                volume_threshold = 1.0
            
            current_volume = day_data['V'][-1]
            prev_volume = day_data['V'][-2]
            
            if current_volume < prev_volume * volume_threshold:
                logging.debug(
                    f"{code}: 거래량 부족 - "
                    f"{current_volume}/{prev_volume * volume_threshold:.0f} ({volume_threshold*100}%)"
                )
                self.daydata.monitor_stop(code)
                return
            
            # 5️⃣ 가격 필터
            current_price = day_data['C'][-1]
            if current_price < 1000 or current_price > 500000:
                logging.debug(f"{code}: 가격 {current_price}원 범위 초과")
                self.daydata.monitor_stop(code)
                return
            
            # 6️⃣ VI 괴리율 검증
            match_gap = re.search(r"괴리율:(-?\d+\.\d+)%", event2)
            if match_gap:
                gap_rate = float(match_gap.group(1))
                if gap_rate < 3.0:
                    logging.debug(f"{code}: 괴리율 {gap_rate}% (3% 미만)")
                    self.daydata.monitor_stop(code)
                    return
            
            # 7️⃣ 섹터 분산 체크
            sector = cpCodeMgr.GetStockSectionKind(code)
            sector_count = sum(1 for c in self.monistock_set 
                            if cpCodeMgr.GetStockSectionKind(c) == sector)
            if sector_count >= 2:
                logging.debug(f"{code}: 동일 섹터 종목 {sector_count}개 초과")
                self.daydata.monitor_stop(code)
                return
            
            # ✅ 모든 조건 통과 - 데이터 로드 및 추가
            if not (self.tickdata.monitor_code(code) and self.mindata.monitor_code(code)):
                logging.error(f"{code}: 틱/분 데이터 모니터링 시작 실패")
                self.daydata.monitor_stop(code)
                return
            
            logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 장 시작부터 데이터 로드 시작")
            
            tick_loaded = self.tickdata.update_chart_data_from_market_open(code)
            min_loaded = self.mindata.update_chart_data_from_market_open(code)
            
            if not tick_loaded or not min_loaded:
                logging.warning(f"{code}: 장 시작 데이터 로드 실패, 개수 기준으로 폴백")
                self.tickdata.update_chart_data(code, 60, 400)
                self.mindata.update_chart_data(code, 3, 150)
            
            # VI 발동 시각 저장
            self.starting_time[code] = time
            
            # VI 발동 가격 저장
            match_price = re.search(r"발동가격:\s*(\d+)", event2)
            if match_price:
                self.starting_price[code] = int(match_price.group(1))
            else:
                self.starting_price[code] = current_price
            
            # 모니터링 세트 추가
            self.monistock_set.add(code)
            self.stock_added_to_monitor.emit(code)
            
            logging.info(
                f"{cpCodeMgr.CodeToName(code)}({code}) -> "
                f"투자 대상 추가 (VI: {time}, 발동가: {self.starting_price[code]:.0f}원, "
                f"괴리율: {gap_rate if match_gap else 'N/A'}%)"
            )
            
            self.save_list_db(code, self.starting_time[code], self.starting_price[code], 1)
            self.save_vi_data(code)

        except Exception as ex:
            logging.error(f"monitor_vi -> {code}, {ex}\n{traceback.format_exc()}")

    def download_vi(self):
        try:
            date = datetime.today() - timedelta(days=1)
            while True:
                date_str = date.strftime('%Y%m%d')

                # OTP 생성 URL 및 다운로드 URL 설정
                otp_url = 'http://data.krx.co.kr/comm/fileDn/GenerateOTP/generate.cmd'
                download_url = 'http://data.krx.co.kr/comm/fileDn/download_excel/download.cmd'

                # OTP 요청에 필요한 파라미터 설정
                query_str_params = {
                    'locale': 'ko_KR',
                    'mktId': 'ALL',
                    'inqTpCd1': '01',
                    'viKindCd': '1',
                    'tboxisuCd_finder_stkisu1_0': '전체',
                    'isuCd': 'ALL',
                    'isuCd2': 'ALL',
                    'param1isuCd_finder_stkisu1_0': 'ALL',
                    'prcDetailView': '1',
                    'share': '1',
                    'money': '1',
                    'strtDd': date_str,
                    'endDd': date_str,
                    'csvxls_isNo': 'true',
                    'name': 'fileDown',
                    'url': 'dbms/MDC/STAT/issue/MDCSTAT22401'
                }

                # HTTP 헤더 설정
                headers = {
                    'Referer': 'http://data.krx.co.kr/contents/MDC/MDI/mdiLoader/index.cmd?menuId=MDC02021501',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36'
                }

                with requests.Session() as session:
                    r = session.get(otp_url, params=query_str_params, headers=headers)
                    otp_code = r.content
                    r = session.post(download_url, data={'code': otp_code}, headers=headers)
                    df_vi = pd.read_excel(BytesIO(r.content), engine='openpyxl')

                if df_vi['발동가격_가격'].apply(pd.to_numeric, errors='coerce').notnull().all():
                    break
                date -= timedelta(days=1)

            df_vi = df_vi[df_vi['발동가격_괴리율'] > 0]
            df_vi = df_vi[df_vi['거래량'] > 1000000]
            df_vi = df_vi[df_vi['현재가'] > df_vi['시가']]
            df_vi = df_vi[df_vi['현재가'] > df_vi['발동가격_가격']]
            df_vi = df_vi[df_vi['종목코드'].str.match(r'^\d+$')]
            df_vi['종목코드'] = df_vi['종목코드'].apply(lambda x: f"A{int(x):06d}")
            idx = df_vi.groupby('종목코드')['발동가격_가격'].idxmax()
            df_vi = df_vi.loc[idx].reset_index(drop=True)
            vi_list = df_vi['종목코드'].tolist()

            for code in vi_list:
                if self.daydata.select_code(code):
                    if len(self.daydata.stockdata[code].get('MAD5', [])) > 0 and len(self.daydata.stockdata[code].get('MAD10', [])) > 0:
                        if (self.daydata.stockdata[code]['MAD5'][-1] > self.daydata.stockdata[code]['MAD10'][-1]):
                            if (self.daydata.stockdata[code]['O'][-1] > self.daydata.stockdata[code]['C'][-2] * 0.99):
                                if code not in self.monistock_set:
                                    if self.tickdata.monitor_code(code) == True and self.mindata.monitor_code(code) == True:
                                        if code not in self.starting_time:
                                            self.starting_time[code] = datetime.now().strftime('%m/%d 09:00:00')
                                        self.monistock_set.add(code)
                                        self.stock_added_to_monitor.emit(code)
                                    else:
                                        self.daydata.monitor_stop(code)
                                        self.mindata.monitor_stop(code)
                                        self.tickdata.monitor_stop(code)
                            else:
                                self.daydata.monitor_stop(code)
                        else:
                            self.daydata.monitor_stop(code)
                    else:
                        self.daydata.monitor_stop(code)
                else:
                    self.daydata.monitor_stop(code)
        except Exception as ex:
            logging.error(f"download_vi -> {ex}")

    def save_list_db(self, code, starting_time, starting_price, is_moni=0, db_file='mylist.db'):
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS items (
                codes TEXT PRIMARY KEY,
                starting_time DATETIME,
                starting_price REAL,
                is_moni INTEGER
            )
        ''')
        c.execute('INSERT OR REPLACE INTO items (codes, starting_time, starting_price, is_moni) VALUES (?, ?, ?, ?)', (code, starting_time, starting_price, is_moni))
        conn.commit()
        conn.close()

    def delete_list_db(self, code, db_file='mylist.db'):
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS items (
                codes TEXT PRIMARY KEY,
                starting_time DATETIME,
                starting_price REAL,
                is_moni INTEGER
            )
        ''')

        c.execute("DELETE FROM items WHERE codes = ?", (code,))
        conn.commit()
        conn.close()

    def clear_list_db(self, db_file='mylist.db'):
        try:
            if os.path.exists(db_file):
                os.remove(db_file)
        except Exception as ex:
            logging.error(f"clear_list_db -> {ex}")
        self.monistock_set = set()
        self.vistock_set = set()
        self.database_set = set()
        self.starting_time = {}
        self.starting_price = {}

    def load_from_list_db(self, db_file='mylist.db'):
        if not os.path.exists(db_file):
            self.monistock_set = set()
            self.vistock_set = set()
            self.database_set = set()
            self.stgname = {}
            self.starting_time = {}
            self.starting_price = {}
            return
        
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS items (
                codes TEXT PRIMARY KEY,
                starting_time DATETIME,
                starting_price REAL,
                is_moni INTEGER
            )
        ''')
        c.execute('SELECT codes, starting_time, starting_price, is_moni FROM items')
        rows = c.fetchall()
        conn.close()
        if rows:
            self.vistock_set = {row[0] for row in rows if not row[3]}
            self.database_set = {row[0] for row in rows if row[3]}
            self.starting_time = {row[0]: row[1] for row in rows}
            self.starting_price = {row[0]: row[2] for row in rows}
        # else:
        #     self.monistock_set = set()
        #     self.database_set = set()
        #     self.stgname = {}
        #     self.starting_time = {}
        #     self.starting_price = {}

    @pyqtSlot(str, str, str, str)
    def buy_stock(self, code, buy_message, order_condition, order_style):
        try:
            if code in self.bought_set or code in self.buyorder_set or code in self.buyordering_set:
                return
            self.buyordering_set.add(code) # 초단시간 재매수 방지
            stock_name = self.cpCodeMgr.CodeToName(code)
            cur_price, ask_price, upper_price = self.get_current_price(code)

            if not ask_price:
                if ask_price == 0:
                    self.buyorder_set.add(code)
                    logging.info(f"{stock_name}({code}) 상한가 주문 불가")
                else:
                    logging.error(f"{stock_name}({code}) 현재가 조회 실패")
                return
            if (len(self.buyorder_set) + len(self.bought_set)) >= self.target_buy_count:
                return
            buy_qty = self.buy_amount // ask_price
            total_amount = self.total_cash - self.buy_amount * (len(self.buyorder_set) + len(self.bought_set))
            max_buy_qty = total_amount // upper_price
            if max_buy_qty <= 0 or buy_qty <= 0: 
                # logging.warning(f"{stock_name}({code}) 매수 가능 수량이 부족합니다.")
                return
            self.buyorder_qty[code] = int(min(buy_qty, max_buy_qty))

            if self.buyorder_qty[code] >= 0:
                acc = self.cpTrade.AccountNumber[0]
                accFlag = self.cpTrade.GoodsList(acc, 1)
                if self.cp_request.is_requesting:
                    logging.debug("요청 진행 중, 매수 주문 스킵")
                    return None  # 즉시 반환

                self.cp_request.is_requesting = True
                self.cpOrder.SetInputValue(0, "2") # 2: 매수
                self.cpOrder.SetInputValue(1, acc) # 계좌번호
                self.cpOrder.SetInputValue(2, accFlag[0]) # 상품구분 - 주식 상품 중 첫번째
                self.cpOrder.SetInputValue(3, code) # 종목코드
                self.cpOrder.SetInputValue(4, self.buyorder_qty[code]) # 매수수량
                if buy_message == '발동가':
                    self.cpOrder.SetInputValue(5, self.starting_price[code]) # 주문단가
                    # self.buy_price[code] = self.starting_price[code]
                else:
                    self.cpOrder.SetInputValue(5, ask_price) # 주문단가
                    # self.buy_price[code] = ask_price
                self.cpOrder.SetInputValue(7, order_condition) # 주문 조건 구분 코드, 0: 기본 1: IOC 2:FOK
                self.cpOrder.SetInputValue(8, order_style) # 주문호가 구분코드 - 01: 보통, 03:시장
                
                remain_count0 = cpStatus.GetLimitRemainCount(0) # 0: 주문/계좌 관련 1: 시세 요청 관련 2: 실시간 요청 관련
                if remain_count0 == 0:
                    logging.error(f"거래 요청 제한")
                    return
                
                logging.info(f"{stock_name}({code}), {buy_message} -> 매수 요청({self.buyorder_qty[code]}주)")
                
                handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
                handler.set_params(self.cpOrder, 'order', self, {'code': code, 'stock_name': stock_name})
                self.cpOrder.Request()          
                self.buyorder_set.add(code)
                if code in self.buyordering_set:
                    self.buyordering_set.remove(code)
                result = handler.wait()
                if not result:
                    logging.warning(f"{stock_name}({code}) 매수주문 실패")
                    if code in self.buyorder_set:
                        self.buyorder_set.remove(code)
                    return            

        except Exception as ex:
            logging.error(f"buy_stock -> {code}, {ex}")

    @pyqtSlot(str, str)
    def sell_stock(self, code, message):
        try:
            if code in self.sellorder_set:
                return
            
            stock_name = self.cpCodeMgr.CodeToName(code)
            if code in self.buy_qty:
                stock_qty = self.buy_qty[code]
            else:
                return
            
            sell_order_qty = min(stock_qty, self.balance_qty[code])
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            if self.cp_request.is_requesting:
                logging.debug("요청 진행 중, 매도 주문 스킵")
                return None  # 즉시 반환

            self.cp_request.is_requesting = True
            self.cpOrder.SetInputValue(0, "1")
            self.cpOrder.SetInputValue(1, acc)
            self.cpOrder.SetInputValue(2, accFlag[0])
            self.cpOrder.SetInputValue(3, code)
            self.cpOrder.SetInputValue(4, sell_order_qty)
            self.cpOrder.SetInputValue(7, "0")
            self.cpOrder.SetInputValue(8, "03")
        
            remain_count0 = cpStatus.GetLimitRemainCount(0) # 0: 주문/계좌 관련 1: 시세 요청 관련 2: 실시간 요청 관련
            if remain_count0 == 0:
                logging.error(f"거래 요청 제한")
                return
            
            logging.info(f"{stock_name}({code}), {message} -> 매도 요청({sell_order_qty}주)")
            handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
            handler.set_params(self.cpOrder, 'order', self, {'code': code, 'stock_name': stock_name})
            self.cpOrder.Request()
            self.sellorder_set.add(code) 

            result = handler.wait()
            if not result:
                logging.warning(f"{stock_name}({code}) 매도주문 실패")
                if code in self.sellorder_set:
                    self.sellorder_set.remove(code)
                return
                       
            # order_price = self.cpOrder.GetHeaderValue(5) # 시장가는 0        

        except Exception as ex:
            logging.error(f"sell_stock -> {code}, {ex}")

    @pyqtSlot(str, str)
    def sell_half_stock(self, code, message):
        try:
            if code in self.sellorder_set:
                return
            
            stock_name = self.cpCodeMgr.CodeToName(code)
            if code in self.buy_qty:
                stock_qty = self.buy_qty[code]
            else:
                return
            self.sell_half_qty[code] = stock_qty - ((stock_qty + 1) // 2)
            sell_half_order_qty = min(((stock_qty + 1) // 2), self.balance_qty.get(code, 0))

            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            if self.cp_request.is_requesting:
                logging.debug("요청 진행 중, 분할매도 주문 스킵")
                return None  # 즉시 반환

            self.cp_request.is_requesting = True
            self.cpOrder.SetInputValue(0, "1")
            self.cpOrder.SetInputValue(1, acc)
            self.cpOrder.SetInputValue(2, accFlag[0])
            self.cpOrder.SetInputValue(3, code)
            self.cpOrder.SetInputValue(4, sell_half_order_qty)
            self.cpOrder.SetInputValue(7, "0")
            self.cpOrder.SetInputValue(8, "03")

            remain_count0 = cpStatus.GetLimitRemainCount(0) # 0: 주문/계좌 관련 1: 시세 요청 관련 2: 실시간 요청 관련
            if remain_count0 == 0:
                logging.error(f"거래 요청 제한")
                return
            
            logging.info(f"{stock_name}({code}), {message} -> 분할 매도 요청({sell_half_order_qty}주)")
            handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
            handler.set_params(self.cpOrder, 'order', self, {'code': code, 'stock_name': stock_name})
            self.cpOrder.Request()
            self.sellorder_set.add(code) 

            result = handler.wait()
            if not result:
                logging.warning(f"{stock_name}({code}) 분할매도주문 실패")
                if code in self.sellorder_set:
                    self.sellorder_set.remove(code)
                return
                       
            # order_price = self.cpOrder.GetHeaderValue(5) # 시장가는 0
            
        except Exception as ex:
            logging.error(f"sell_half_stock -> {code}, {ex}")

    @pyqtSlot()
    def sell_all(self):
        try:
            # remain_count0 = cpStatus.GetLimitRemainCount(0)
            # remain_time0 = cpStatus.GetLimitRemainTime(0) # 0:주문/계좌 관련 RQ 요청, 1:시세관련 RQ 요청
            # if remain_count0 == 0 and remain_time0 != 0:
            #     QTimer.singleShot(int(remain_time0), lambda: self.sell_all())
            #     logging.error(f"거래 요청 제한, 대기시간: {remain_time0/1000}초")
            #     return 
            stocks = self.get_stock_balance('ALL', 'sell_all')
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            for s in stocks:
                if s['qty'] > 0:
                    if self.cp_request.is_requesting:
                        logging.debug("요청 진행 중, 전부매도 주문 스킵")
                        return None  # 즉시 반환

                    self.cp_request.is_requesting = True
                    self.cpOrder.SetInputValue(0, "1")
                    self.cpOrder.SetInputValue(1, acc)
                    self.cpOrder.SetInputValue(2, accFlag[0])
                    self.cpOrder.SetInputValue(3, s['code'])
                    self.cpOrder.SetInputValue(4, s['qty'])
                    self.cpOrder.SetInputValue(7, "0")
                    self.cpOrder.SetInputValue(8, "03")

                    logging.info(f"{s['name']}({s['code']}) -> 매도 요청({s['qty']}주)")
                    handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
                    handler.set_params(self.cpOrder, 'order', self, {'code': s['code'], 'stock_name': s['name']})
                    self.cpOrder.Request()
                    self.sellorder_set.add(s['code'])

                    result = handler.wait()
                    if not result:
                        logging.warning(f"{s['name']}({s['code']}) 전부매도주문 실패")
                        if s['code'] in self.sellorder_set:
                            self.sellorder_set.remove(s['code'])
            return True

        except Exception as ex:
            logging.error(f"sell_all -> {ex}")
            return False

    def monitorOrderStatus(self, code, ordernum, conflags, price, qty, bs, balance, buyprice):
        try:
            stock_name = self.cpCodeMgr.CodeToName(code)

            if bs == '1' and conflags == "체결":  # 매도
                logging.debug(f"{stock_name}({code}), {price}원, {qty}주 매도, 잔고: {balance}주")
                self.balance_qty[code] = balance

                if code not in self.sell_amount:
                    self.sell_amount[code] = 0
                self.sell_amount[code] += price * qty

                if code in self.sell_half_qty and balance == self.sell_half_qty[code]: # 분할매도
                    logging.info(f"{stock_name}({code}), 분할 매도 완료")                  
                    
                    stock_profit = self.sell_amount[code] * 0.99835 - self.buy_price[code] * (self.buy_qty[code] - balance) * 1.00015
                    stock_rate = (stock_profit / (self.buy_price[code] * (self.buy_qty[code] - balance))) * 100
                    if stock_profit > 0:
                        logging.info(f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                    else:
                        logging.info(f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                    self.get_trade_profit()
                    if code in self.sellorder_set:
                        self.sellorder_set.remove(code)
                    if code not in self.sell_half_set:
                        self.sell_half_set.add(code)
                    self.buy_qty[code] = balance
                    self.sell_amount[code] = 0
                    if code in self.sell_half_qty:
                        del self.sell_half_qty[code]

                if balance == 0: # 전부매도                    
                    logging.info(f"{stock_name}({code}), 매도 완료")

                    stock_profit = self.sell_amount[code] * 0.99835 - self.buy_price[code] * self.buy_qty[code] * 1.00015
                    stock_rate = (stock_profit / (self.buy_price[code] * self.buy_qty[code] - balance)) * 100
                    if stock_profit > 0:
                        logging.info(f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                    else:
                        logging.info(f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                    self.get_trade_profit()                    
                                        
                    self.stock_sold.emit(code)                    
                    if code in self.bought_set:
                        self.bought_set.remove(code) 
                    if code in self.sell_half_set:
                        self.sell_half_set.remove(code)
                    if code in self.sellorder_set:
                        self.sellorder_set.remove(code)                 
                    if code in self.buy_price:
                        del self.buy_price[code]
                    if code in self.buy_qty:
                        del self.buy_qty[code]
                    if code in self.sell_amount:
                        del self.sell_amount[code]


            elif bs == '2' and conflags == "체결": # 매수
                logging.debug(f"{stock_name}({code}), {qty}주 매수, 잔고: {balance}주")
                self.balance_qty[code] = balance
                
                if code in self.buyorder_qty and balance >= self.buyorder_qty[code]:
                    self.buy_qty[code] = balance
                    self.buy_price[code] = buyprice
                    logging.info(f"{stock_name}({code}), {self.buy_qty[code]}주, 매수 완료({int(self.buy_price[code])}원)")
                    
                    if code not in self.bought_set:
                        self.bought_set.add(code)
                        self.stock_bought.emit(code)
                    if code in self.buyorder_set:
                        self.buyorder_set.remove(code)
                    if code in self.buyorder_qty:
                        del self.buyorder_qty[code]                
                    return

        except Exception as ex:
            logging.error(f"monitorOrderStatus -> {ex}")

class ChartDrawerThread(QThread):
    data_ready = pyqtSignal(dict)

    def __init__(self, trader, code):
        super().__init__()
        self.trader = trader
        self.code = code
        self.is_running = True

    def run(self):
        while self.is_running:
            if self.code:
                # 차트 그리기는 전체 데이터 필요 (복사본)
                tick_data = self.trader.tickdata.get_full_data(self.code)
                min_data = self.trader.mindata.get_full_data(self.code)

                data = {'tick_data': tick_data, 'min_data': min_data, 'code': self.code}
                self.data_ready.emit(data)
                
            self.msleep(2000)

    def stop(self):
        self.is_running = False
        self.quit()
        self.wait()

class ChartDrawer(QObject):
    def __init__(self, fig, canvas, trader, trader_thread, window):
        super().__init__()
        self.fig = fig
        self.canvas = canvas
        self.trader = trader
        self.trader_thread = trader_thread
        self.window = window
        self.code = None
        self.create_initial_chart()
        self.chart_thread = None
        self.stop_loss = {}
        self.last_chart_update_time = None
        self.current_prediction = None

    def set_code(self, code):
        self.code = code
        # 기존 스레드 중지
        if self.chart_thread:
            self.chart_thread.stop()
            self.chart_thread = None

        if code:
            # 새로운 스레드 시작
            self.chart_thread = ChartDrawerThread(self.trader, code)
            self.chart_thread.data_ready.connect(self.update_chart)
            self.chart_thread.start()
        else:
            self.create_initial_chart()

    def create_initial_chart(self):
        self.fig.clear()
        self.tick_axes = [self.fig.add_subplot(18, 1, (1, 4), title='Tick Chart'), self.fig.add_subplot(18, 1, (5, 6)), self.fig.add_subplot(18, 1, (7, 8))]
        self.min_axes = [self.fig.add_subplot(18, 1, (10, 13), title='3 Min Chart'), self.fig.add_subplot(18, 1, 14), self.fig.add_subplot(18, 1, (15, 16)), self.fig.add_subplot(18, 1, (17, 18))]

        for ax in self.tick_axes + self.min_axes:
            ax.clear()
            ax.set_xticklabels([])
            ax.set_yticklabels([])
            ax.grid(True)

        self.fig.subplots_adjust(hspace=0)
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        self.canvas.draw()
        self.canvas.flush_events()

    @pyqtSlot(str, float)
    def update_prediction(self, code, prediction):
        """AutoTraderThread로부터 받은 prediction 값을 저장"""
        if code == self.code:
            self.current_prediction = prediction * 100  # 예측값을 퍼센트로 변환

    @pyqtSlot(dict)
    def update_chart(self, data):
        self.draw_chart(data)
        self.last_chart_update_time = data['tick_data'].get('T', [None])[-1]

    def draw_chart(self, data):
        try:
            if data['code']:
                code_name = cpCodeMgr.CodeToName(data['code'])
                for ax in self.tick_axes:
                    ax.clear()
                for ax in self.min_axes:
                    ax.clear()

                self.draw_chart_data(data['tick_data'], self.tick_axes, data['code'], data_type='tick')
                self.draw_chart_data(data['min_data'], self.min_axes, data['code'], data_type='min')
                self.fig.suptitle(f"Chart for {data['code']}({code_name})", fontsize=15)
                self.fig.subplots_adjust(hspace=0)
                self.canvas.draw()
            else:
                self.create_initial_chart()

        except Exception as ex:
            logging.error(f"draw_chart -> {ex}")

    def draw_chart_data(self, chart_data, axes, code, data_type):
        try:
            current_strategy = self.window.comboStg.currentText()
            all_strategies = self.window.strategies.get(current_strategy, [])
            
            if chart_data:
                if data_type == 'tick':
                    keys_to_keep = ['O', 'H', 'L', 'C', 'V', 'D', 'T', 'MAT5', 'MAT20', 'MAT60', 'MAT120', 'RSIT', 'RSIT_SIGNAL', 'MACDT', 'MACDT_SIGNAL', 'OSCT']
                elif data_type == 'min':
                    keys_to_keep = ['O', 'H', 'L', 'C', 'V', 'D', 'T', 'MAM5', 'MAM10', 'MAM20', 'RSI', 'RSI_SIGNAL', 'MACD', 'MACD_SIGNAL', 'OSC']

                if data_type == 'tick':
                    filtered_data = {key: [x for x in (chart_data[key][-60:] if len(chart_data[key]) >= 60 else chart_data[key][:])] for key in keys_to_keep}

                    df = pd.DataFrame({
                        'Open': filtered_data['O'], 'High': filtered_data['H'], 'Low': filtered_data['L'], 'Close': filtered_data['C'], 'Volume': filtered_data['V'], 
                        'Date': filtered_data['D'], 'Time': filtered_data['T'], 'MAT5': filtered_data['MAT5'], 
                        'MAT20': filtered_data['MAT20'], 'MAT60': filtered_data['MAT60'], 'MAT120': filtered_data['MAT120'], 
                        'RSIT': filtered_data['RSIT'], 'RSIT_SIGNAL': filtered_data['RSIT_SIGNAL'], 'MACDT': filtered_data['MACDT'], 
                        'MACDT_SIGNAL': filtered_data['MACDT_SIGNAL'], 'OSCT': filtered_data['OSCT']
                    })
                    df.index = pd.to_datetime(df['Date'].astype(str) + df['Time'].astype(str), format='%Y%m%d%H%M', errors='coerce')

                    addplots = [
                        mpf.make_addplot(df['MAT5'], color='magenta', label='MAT5', ax=axes[0], width=0.9),
                        mpf.make_addplot(df['MAT20'], color='blue', label='MAT20', ax=axes[0], width=0.9),
                        mpf.make_addplot(df['MAT60'], color='darkorange', label='MAT60', ax=axes[0], width=1.0),                        
                        mpf.make_addplot(df['MAT120'], color='black', label='MAT120', ax=axes[0], width=0.9),
                        mpf.make_addplot(df['RSIT'], color='red', label='RSIT', ax=axes[1], width=0.9),
                        mpf.make_addplot(df['RSIT_SIGNAL'], color='blue', label='RSIT_SIGNAL', ax=axes[1], width=0.9),
                        mpf.make_addplot(df['MACDT'], color='red', label='MACDT', ax=axes[2], width=0.9),
                        mpf.make_addplot(df['MACDT_SIGNAL'], color='blue', label='MACDT_SIGNAL', ax=axes[2], width=0.9),
                        mpf.make_addplot(df['OSCT'], color='purple', type='bar', label='OSCT', ax=axes[2]),
                    ]
                    mpf.plot(df, ax=axes[0], type='candle', style='yahoo', addplot=addplots)

                    # 현재가 레이블 추가
                    current_price = df['Close'].iloc[-1]
                    prediction_text = f"{self.current_prediction:.2f}" if self.current_prediction is not None else ""
                    axes[0].text(
                        1.01, current_price, f'{current_price}({prediction_text})',
                        transform=axes[0].get_yaxis_transform(), color='red',
                        verticalalignment='center', bbox=dict(facecolor='white', edgecolor='none')
                    )
                    axes[0].legend(loc='upper left')

                    x_raw = list(range(len(df['Date'])))
                    axes[1].fill_between(x_raw, df['RSIT'], 20, where=(df['RSIT'] <= 20), color='blue', alpha=0.5)
                    axes[1].fill_between(x_raw, df['RSIT'], 80, where=(df['RSIT'] >= 80), color='red', alpha=0.5)
                    axes[1].set_yticks([20, 50, 80])
                    axes[1].legend(loc='upper left')
                    axes[2].legend(loc='upper left')

                    x_labels = self.create_labels(filtered_data)
                    for index, ax in enumerate(axes):
                        ax.set_xlim(-2, len(filtered_data['D']) + 1)
                        ax.grid(True, axis='y')
                        ax.set_xticks(x_raw)
                        if index == 0 or index == 1:
                            ax.set_xticklabels([])
                        if index == 2:
                            ax.set_xticklabels(x_labels)

                elif data_type == 'min':
                    filtered_data = {key: [x for x in (chart_data[key][-50:] if len(chart_data[key]) >= 50 else chart_data[key][:])] for key in keys_to_keep}

                    df = pd.DataFrame({
                        'Open': filtered_data['O'], 'High': filtered_data['H'], 'Low': filtered_data['L'], 'Close': filtered_data['C'], 'Volume': filtered_data['V'],
                        'Date': filtered_data['D'], 'Time': filtered_data['T'], 'MAM5': filtered_data['MAM5'], 'MAM10': filtered_data['MAM10'], 'MAM20': filtered_data['MAM20'],
                        'RSI': filtered_data['RSI'], 'RSI_SIGNAL': filtered_data['RSI_SIGNAL'], 'MACD': filtered_data['MACD'], 'MACD_SIGNAL': filtered_data['MACD_SIGNAL'], 'OSC': filtered_data['OSC']
                    })
                    df.index = pd.to_datetime(df['Date'].astype(str) + df['Time'].astype(str), format='%Y%m%d%H%M')

                    addplots = [
                        mpf.make_addplot(df['MAM5'], color='magenta', label='MAM5', ax=axes[0], width=0.9),
                        mpf.make_addplot(df['MAM10'], color='blue', label='MAM10', ax=axes[0], width=0.9),
                        mpf.make_addplot(df['MAM20'], color='darkorange', label='MAM20', ax=axes[0], width=1.0),
                        mpf.make_addplot(df['RSI'], color='red', label='RSI', ax=axes[2], width=0.9),
                        mpf.make_addplot(df['RSI_SIGNAL'], color='blue', label='RSI_SIGNAL', ax=axes[2], width=0.9),
                        mpf.make_addplot(df['MACD'], color='red', label='MACD', ax=axes[3], width=0.9),
                        mpf.make_addplot(df['MACD_SIGNAL'], color='blue', label='MACD_SIGNAL', ax=axes[3], width=0.9),
                        mpf.make_addplot(df['OSC'], color='purple', type='bar', label='OSC', ax=axes[3])
                    ]
                    mpf.plot(df, ax=axes[0], type='candle', style='yahoo', volume=axes[1], addplot=addplots)

                    # 현재가 레이블 추가
                    current_price = df['Close'].iloc[-1]
                    axes[0].text(
                        1.01, current_price, f'<{current_price}>',
                        transform=axes[0].get_yaxis_transform(), color='red',
                        verticalalignment='center', bbox=dict(facecolor='white', edgecolor='none')
                    )
                    axes[0].legend(loc='upper left')
                    
                    x_raw = list(range(len(df['Date'])))
                    axes[2].fill_between(x_raw, df['RSI'], 30, where=(df['RSI'] <= 30), color='blue', alpha=0.5)
                    axes[2].fill_between(x_raw, df['RSI'], 70, where=(df['RSI'] >= 70), color='red', alpha=0.5)
                    axes[2].set_yticks([30, 50, 70])
                    axes[2].legend(loc='upper left')
                    axes[3].legend(loc='upper left')

                    x_labels = self.create_labels(filtered_data)
                    for index, ax in enumerate(axes):
                        ax.set_xlim(-2, len(filtered_data['D']) + 1)
                        ax.grid(True, axis='y')
                        ax.set_xticks(x_raw)
                        if index == 0 or index == 1 or index == 2:
                            ax.set_xticklabels([])
                        if index == 3:
                            ax.set_xticklabels(x_labels)
                    axes[1].yaxis.set_label_position("right")
                    axes[1].yaxis.tick_right()
                    axes[1].grid(False)

                # 공통: starting_price_line 표시 (y축 범위 내)
                if code in self.trader.starting_price:
                    starting_price_line = self.trader.starting_price[code]
                    y_min, y_max = axes[0].get_ylim()  # y축 범위 가져오기
                    if data_type == 'tick':
                        if y_min <= starting_price_line <= y_max:
                            axes[0].axhline(y=starting_price_line, color='orangered', linestyle='--', linewidth=1, label='Starting Price')
                    elif data_type == 'min':
                        if starting_price_line < y_min or starting_price_line > y_max:
                            axes[0].set_ylim(min(y_min, starting_price_line), max(y_max, starting_price_line))
                        axes[0].axhline(y=starting_price_line, color='orangered', linestyle='--', linewidth=1, label='Starting Price')

                # 공통: buy_price_line 표시
                if code in self.trader.buy_price:
                    buy_price_line = self.trader.buy_price[code]
                    axes[0].axhline(y=buy_price_line, color='red', linestyle='-', linewidth=1, label='Buy Price')

        except Exception as ex:
            logging.error(f"draw_chart_data -> {ex}\n{traceback.format_exc()}")

    def create_labels(self, filtered_data):
        labels = []
        current_hour = None
        mmm_0_added = False
        mmm_15_added = False
        mmm_30_added = False
        mmm_45_added = False

        for i in range(len(filtered_data['D'])):
            date = filtered_data['D'][i]
            time = filtered_data['T'][i]

            # 날짜 분리
            yy, mmdd = divmod(date, 10000)
            mm, dd = divmod(mmdd, 100)
            formatted_date = f"{mm:02}/{dd:02} "  # 예: "01/01 "

            # 시간 분리
            hhh, mmm = divmod(time, 100)
            formatted_time = f"{hhh:02}:{mmm:02}"  # 예: "12:21"

            # 시간이 변경되었는지 확인하여 플래그 초기화
            if hhh != current_hour:
                current_hour = hhh
                mmm_0_added = False
                mmm_15_added = False
                mmm_30_added = False
                mmm_45_added = False

            # 레이블 초기화
            label = ''

            if i == 0:
                # 첫 번째 요소는 항상 날짜와 시간을 함께 추가
                label = f"{formatted_date}{formatted_time}"
            elif i == len(filtered_data['D']) - 1:
                # 마지막 요소도 항상 날짜와 시간을 함께 추가
                label = f"{formatted_date}{formatted_time}"
            else:
                if mmm == 0:
                    if not mmm_0_added:
                        label = formatted_time
                        mmm_0_added = True
                elif mmm == 15:
                    if not mmm_15_added:
                        label = formatted_time
                        mmm_15_added = True
                elif mmm == 30:
                    if not mmm_30_added:
                        label = formatted_time
                        mmm_30_added = True
                elif mmm == 45:
                    if not mmm_45_added:
                        label = formatted_time
                        mmm_45_added = True

            labels.append(label if label else '')

        return labels

class CnnTrainerThread(QThread):
    """CNN 훈련 스레드 - 개선 버전"""
    
    training_completed = pyqtSignal(bool, str, float)
    training_error = pyqtSignal(str)

    def __init__(self, trader, pipe_handle, pipe_lock, partial_update=False, seq_length=5, parent=None):
        super().__init__(parent=parent)
        
        # 경로 설정
        self.MODEL_DIR = self._get_model_dir()
        
        self.trader = trader
        self.pipe_handle = pipe_handle
        self.pipe_lock = pipe_lock
        self.partial_update = partial_update
        self.running = True
        self.seq_length = seq_length
        
        # 통신 설정
        self.chunk_size = 65536  # 64KB
        self.max_data_size = 52428800  # 50MB
        self.max_partial_update_size = 10485760  # 10MB
        
        # 특징 정의
        self.tick_features = [
            'C', 'V', 'MAT5', 'MAT20', 'MAT60', 'MAT120', 'RSIT', 'RSIT_SIGNAL',
            'MACDT', 'MACDT_SIGNAL', 'OSCT', 'STOCHK', 'STOCHD', 'ATR', 'CCI',
            'BB_UPPER', 'BB_MIDDLE', 'BB_LOWER', 'BB_POSITION', 'BB_BANDWIDTH',
            'MAT5_MAT20_DIFF', 'MAT20_MAT60_DIFF', 'MAT60_MAT120_DIFF',
            'C_MAT5_DIFF', 'MAT5_CHANGE', 'MAT20_CHANGE', 'MAT60_CHANGE', 
            'MAT120_CHANGE', 'VWAP'
        ]
        self.min_features = [
            'MAM5', 'MAM10', 'MAM20', 'RSI', 'RSI_SIGNAL', 'MACD', 'MACD_SIGNAL',
            'OSC', 'STOCHK', 'STOCHD', 'CCI', 'MAM5_MAM10_DIFF', 'MAM10_MAM20_DIFF',
            'C_MAM5_DIFF', 'C_ABOVE_MAM5', 'VWAP'
        ]

    def _get_model_dir(self):
        """모델 디렉토리 동적 설정"""
        # 1순위: 환경변수
        model_dir = os.getenv('CNN_MODEL_DIR')
        if model_dir and os.path.exists(model_dir):
            return model_dir
        
        # 2순위: 현재 작업 디렉토리
        cwd_model_dir = os.path.join(os.getcwd(), 'models')
        if not os.path.exists(cwd_model_dir):
            os.makedirs(cwd_model_dir)
        
        return cwd_model_dir

    def run(self):
        """훈련 메인 루프"""
        try:
            if not self.pipe_handle or not self.running:
                self.training_error.emit("훈련 파이프 연결 없음")
                return

            # 1. 데이터 로드
            df_tick, df_min = self._load_training_data()
            if df_tick is None or df_min is None:
                return
            
            # 2. 시퀀스 생성
            X, y = self._create_training_sequences(df_tick, df_min)
            if X is None or y is None:
                return
            
            # 3. 데이터 검증
            if not self._validate_training_data(X, y):
                return
            
            # 4. 클래스 가중치 계산
            scale_pos_weight = self._calculate_class_weight(y)
            
            # 5. 훈련 데이터 전송
            if not self._send_training_data(X, y, scale_pos_weight):
                return
            
            # 6. 응답 수신
            self._receive_training_response()
            
        except Exception as ex:
            logging.error(f"CnnTrainerThread 오류: {ex}\n{traceback.format_exc()}")
            self.training_error.emit(f"훈련 실패: {str(ex)}")
        finally:
            self.running = False
            self.quit()

    # ==================== 데이터 로드 ====================
    
    def _load_training_data(self):
        """훈련 데이터 로드"""
        try:
            today = datetime.now().strftime('%Y%m%d')
            
            if self.partial_update:
                # 실시간 데이터 로드
                df_tick, df_min = self._load_realtime_data(today)
            else:
                # DB에서 로드
                df_tick, df_min = self._load_from_database(today)
            
            if df_tick is None or df_tick.empty:
                self.training_error.emit("tick_data 없음")
                return None, None
            
            if df_min is None or df_min.empty:
                self.training_error.emit("min_data 없음")
                return None, None
            
            logging.info(f"데이터 로드 완료: tick={len(df_tick)}, min={len(df_min)}")
            return df_tick, df_min
            
        except Exception as ex:
            logging.error(f"_load_training_data 오류: {ex}")
            self.training_error.emit(f"데이터 로드 실패: {str(ex)}")
            return None, None

    def _load_realtime_data(self, today):
        """실시간 데이터 로드 (부분 업데이트)"""
        try:
            df_tick = self._extract_realtime_tick_data(today)
            df_min = self._extract_realtime_min_data(today)
            
            return df_tick, df_min
            
        except Exception as ex:
            logging.error(f"_load_realtime_data 오류: {ex}")
            return None, None

    def _extract_realtime_tick_data(self, today):
        """실시간 틱 데이터 추출"""
        with self.trader.tickdata.stockdata_lock:
            dfs_tick = []
            
            for code, td in self.trader.tickdata.stockdata.items():
                if not td.get('C'):
                    continue
                
                # VI 발동 시각 추출
                start_hhmm = self._get_vi_start_time(code, today)
                if start_hhmm is None:
                    continue
                
                # 유효 인덱스 추출
                indices = self._get_valid_realtime_tick_indices(td, code, today, start_hhmm)
                if not indices:
                    continue
                
                # 데이터프레임 생성
                df = self._create_tick_dataframe(td, code, indices)
                if df is not None:
                    dfs_tick.append(df)
            
            if not dfs_tick:
                return None
            
            return pd.concat(dfs_tick, ignore_index=True)

    def _get_vi_start_time(self, code, today):
        """VI 발동 시각 추출"""
        if code not in self.trader.starting_time:
            return '0900'
        
        try:
            start_dt = datetime.strptime(
                f"{datetime.now().year}/{self.trader.starting_time[code]}", 
                '%Y/%m/%d %H:%M:%S'
            )
            if start_dt.strftime('%Y%m%d') == today:
                return start_dt.strftime('%H%M')
        except ValueError as ve:
            logging.error(f"{code}: starting_time 형식 오류: {ve}")
        
        return '0900'

    def _get_valid_realtime_tick_indices(self, td, code, today, start_hhmm):
        """유효 실시간 틱 인덱스 추출"""
        min_length = min(len(td.get(f, [])) for f in ['C', 'V', 'D', 'T'])
        
        indices = [
            i for i, (d, t) in enumerate(zip(td['D'], td['T']))
            if i < min_length 
            and str(d) == today 
            and start_hhmm <= f"{t:04d}" <= '1515'
            and all(
                f in td and len(td[f]) > i 
                and td[f][i] is not None 
                and not np.isnan(float(td[f][i])) 
                for f in ['C', 'V']
            )
        ]
        
        if len(indices) < self.seq_length + 10:  # 최소 시퀀스 + 미래 데이터
            logging.debug(f"{code}: 데이터 부족 ({len(indices)} < {self.seq_length + 10})")
            return []
        
        return indices

    def _create_tick_dataframe(self, td, code, indices):
        """틱 데이터프레임 생성"""
        try:
            # 시퀀스 매핑
            sequence_map = {}
            current_seq = {}
            for i in indices:
                t = td['T'][i]
                if t not in current_seq:
                    current_seq[t] = 0
                else:
                    current_seq[t] += 1
                sequence_map[i] = current_seq[t]
            
            # 데이터 딕셔너리 생성
            data_dict = {
                'code': code,
                'date': [td['D'][i] for i in indices],
                'time': [td['T'][i] for i in indices],
                'sequence': [sequence_map[i] for i in indices],
            }
            
            # 특징 추가
            for feature in self.tick_features:
                if feature.endswith('_CHANGE'):
                    base_feature = feature.replace('_CHANGE', '')
                    values = np.array([td.get(base_feature, [0]*len(indices))[i] for i in indices])
                    changes = np.diff(values, prepend=values[0])
                    data_dict[feature] = changes
                else:
                    data_dict[feature] = [td.get(feature, [0]*len(indices))[i] for i in indices]
            
            return pd.DataFrame(data_dict)
            
        except Exception as ex:
            logging.error(f"{code}: 틱 데이터프레임 생성 오류: {ex}")
            return None

    def _extract_realtime_min_data(self, today):
        """실시간 분 데이터 추출"""
        with self.trader.mindata.stockdata_lock:
            dfs_min = []
            
            for code, md in self.trader.mindata.stockdata.items():
                if not md.get('C'):
                    continue
                
                start_hhmm = self._get_vi_start_time(code, today)
                if start_hhmm is None:
                    continue
                
                indices = self._get_valid_realtime_min_indices(md, code, today, start_hhmm)
                if not indices:
                    continue
                
                df = self._create_min_dataframe(md, code, indices)
                if df is not None:
                    dfs_min.append(df)
            
            if not dfs_min:
                return None
            
            return pd.concat(dfs_min, ignore_index=True)

    def _get_valid_realtime_min_indices(self, md, code, today, start_hhmm):
        """유효 실시간 분 인덱스 추출"""
        min_length = min(len(md.get(f, [])) for f in self.min_features + ['C', 'V', 'D', 'T'])
        
        indices = [
            i for i, (d, t) in enumerate(zip(md['D'], md['T']))
            if i < min_length 
            and str(d) == today 
            and start_hhmm <= f"{t:04d}" <= '1515'
            and all(
                f in md and len(md[f]) > i 
                and md[f][i] is not None 
                and not np.isnan(float(md[f][i])) 
                for f in ['C', 'V', 'MAM5', 'MAM10', 'MAM20']
            )
        ]
        
        if len(indices) < self.seq_length:
            logging.debug(f"{code}: 분 데이터 부족 ({len(indices)} < {self.seq_length})")
            return []
        
        return indices

    def _create_min_dataframe(self, md, code, indices):
        """분 데이터프레임 생성"""
        try:
            data_dict = {
                'code': code,
                'date': [md['D'][i] for i in indices],
                'time': [md['T'][i] for i in indices],
                'sequence': [0] * len(indices),
            }
            
            for feature in self.min_features:
                data_dict[feature] = [md.get(feature, [0]*len(indices))[i] for i in indices]
            
            return pd.DataFrame(data_dict)
            
        except Exception as ex:
            logging.error(f"{code}: 분 데이터프레임 생성 오류: {ex}")
            return None

    def _load_from_database(self, today):
        """DB에서 데이터 로드"""
        try:
            conn = sqlite3.connect(self.trader.db_name)
            
            df_tick = pd.read_sql_query(
                "SELECT * FROM tick_data WHERE date = ?",
                conn, params=(today,)
            )
            
            df_min = pd.read_sql_query(
                "SELECT * FROM min_data WHERE date = ?",
                conn, params=(today,)
            )
            
            conn.close()
            
            # NaN 처리
            df_tick = df_tick.replace([None], np.nan)
            df_min = df_min.replace([None], np.nan)
            
            # Forward/Backward fill
            for feature in self.tick_features:
                if feature in df_tick.columns and df_tick[feature].isna().any():
                    df_tick[feature] = df_tick.groupby('code')[feature].fillna(method='ffill').fillna(method='bfill').fillna(0)
            
            for feature in self.min_features:
                if feature in df_min.columns and df_min[feature].isna().any():
                    df_min[feature] = df_min.groupby('code')[feature].fillna(method='ffill').fillna(method='bfill').fillna(0)
            
            return df_tick, df_min
            
        except Exception as ex:
            logging.error(f"_load_from_database 오류: {ex}")
            return None, None

    # ==================== 시퀀스 생성 ====================
    
    def _create_training_sequences(self, df_tick, df_min):
        """훈련 시퀀스 생성"""
        try:
            X, y = [], []
            
            for code in df_tick['code'].unique():
                code_tick = df_tick[df_tick['code'] == code].sort_values(['date', 'time', 'sequence']).reset_index(drop=True)
                code_min = df_min[df_min['code'] == code].sort_values(['date', 'time']).reset_index(drop=True)
                
                # 최소 길이 체크
                if len(code_tick) < self.seq_length + 10:  # seq + 미래 데이터
                    logging.debug(f"{code}: 틱 데이터 부족 ({len(code_tick)})")
                    continue
                
                # 부분 업데이트 시 최근 데이터만
                if self.partial_update:
                    code_tick = code_tick.tail(100).reset_index(drop=True)
                    code_min = code_min.tail(100).reset_index(drop=True)
                
                # 시퀀스 생성
                sequences = self._generate_sequences_for_code(code, code_tick, code_min)
                if sequences:
                    X.extend(sequences['X'])
                    y.extend(sequences['y'])
            
            if not X:
                self.training_error.emit("시퀀스 생성 실패")
                return None, None
            
            X = np.array(X)
            y = np.array(y)
            
            logging.info(f"시퀀스 생성 완료: X={X.shape}, y={y.shape}")
            return X, y
            
        except Exception as ex:
            logging.error(f"_create_training_sequences 오류: {ex}")
            self.training_error.emit(f"시퀀스 생성 실패: {str(ex)}")
            return None, None

    def _generate_sequences_for_code(self, code, code_tick, code_min):
        """종목별 시퀀스 생성"""
        try:
            data_list = []
            targets = []
            
            for i in range(len(code_tick) - self.seq_length + 1):
                # 틱 시퀀스
                tick_seq = code_tick.iloc[i:i + self.seq_length]
                tick_data = tick_seq[self.tick_features].values
                
                # NaN 체크
                if np.isnan(tick_data).any():
                    continue
                
                # 분 데이터 매핑
                min_data = self._map_min_data_to_ticks(tick_seq, code_min)
                if min_data is None:
                    continue
                
                # 결합
                combined_data = np.hstack((tick_data, min_data)).flatten()
                
                # 크기 검증
                expected_size = self.seq_length * (len(self.tick_features) + len(self.min_features))
                if len(combined_data) != expected_size:
                    logging.warning(f"{code}: 데이터 크기 불일치 ({len(combined_data)} != {expected_size})")
                    continue
                
                # 레이블 생성 (개선됨)
                target = self._create_label(code_tick, i)
                if target is None:
                    continue
                
                data_list.append(combined_data)
                targets.append(target)
            
            if not data_list:
                return None
            
            return {'X': data_list, 'y': targets}
            
        except Exception as ex:
            logging.error(f"{code}: 시퀀스 생성 오류: {ex}")
            return None

    def _map_min_data_to_ticks(self, tick_seq, code_min):
        """틱에 분 데이터 매핑"""
        try:
            min_data_values = np.zeros((self.seq_length, len(self.min_features)))
            
            # 분 데이터 딕셔너리
            min_dict = {
                row['time']: idx 
                for idx, row in code_min.iterrows()
            }
            
            for idx, tick_time in enumerate(tick_seq['time'].values):
                # 틱 시간을 분봉 시간으로 변환
                hh, mm = divmod(int(tick_time), 100)
                converted_mintime = hh * 60 + mm
                a, _ = divmod(converted_mintime, 3)
                interval_time = a * 3
                chart_time = interval_time + 3
                hour, minute = divmod(chart_time, 60)
                min_time_mapped = hour * 100 + minute
                
                if min_time_mapped > 1515:
                    min_time_mapped = 1515
                
                # 매핑
                if min_time_mapped in min_dict:
                    min_idx = min_dict[min_time_mapped]
                    min_data_values[idx] = code_min.iloc[min_idx][self.min_features].values
                else:
                    # 이전 데이터 사용
                    prev_times = [t for t in min_dict.keys() if t < min_time_mapped]
                    if prev_times:
                        prev_time = max(prev_times)
                        min_idx = min_dict[prev_time]
                        min_data_values[idx] = code_min.iloc[min_idx][self.min_features].values
            
            return min_data_values
            
        except Exception as ex:
            logging.error(f"_map_min_data_to_ticks 오류: {ex}")
            return None

    def _create_label(self, code_tick, index):
        """레이블 생성 (개선됨)"""
        try:
            base_price = code_tick['C'].iloc[index + self.seq_length - 1]
            
            if base_price == 0:
                return None
            
            # ✅ 10틱 후 최고가 확인 (3틱 → 10틱)
            max_future_ticks = min(10, len(code_tick) - (index + self.seq_length))
            
            if max_future_ticks < 5:  # 최소 5틱 필요
                return None
            
            # 미래 가격
            future_prices = code_tick['C'].iloc[
                index + self.seq_length:index + self.seq_length + max_future_ticks
            ]
            
            max_future_price = future_prices.max()
            cumulative_change = (max_future_price - base_price) / base_price * 100
            
            # ✅ 3단계 레이블링
            if cumulative_change >= 1.5:  # 1.5% 이상 상승
                return 2  # 강한 매수
            elif cumulative_change >= 0.5:  # 0.5~1.5% 상승
                return 1  # 약한 매수
            else:
                return 0  # 매도
                
        except Exception as ex:
            logging.error(f"_create_label 오류: {ex}")
            return None

    # ==================== 데이터 검증 ====================
    
    def _validate_training_data(self, X, y):
        """훈련 데이터 검증"""
        try:
            # NaN/Inf 체크
            if np.isnan(X).any():
                nan_count = np.isnan(X).sum()
                self.training_error.emit(f"X에 NaN {nan_count}개 존재")
                return False
            
            if np.isinf(X).any():
                inf_count = np.isinf(X).sum()
                self.training_error.emit(f"X에 Inf {inf_count}개 존재")
                return False
            
            if np.isnan(y).any():
                self.training_error.emit("y에 NaN 존재")
                return False
            
            # 클래스 분포 체크
            unique_classes, counts = np.unique(y, return_counts=True)
            class_dist = dict(zip(unique_classes, counts))
            
            logging.info(f"클래스 분포: {class_dist}")
            
            if len(unique_classes) < 2:
                self.training_error.emit(f"클래스 부족: {unique_classes}")
                return False
            
            # 최소 샘플 수 체크
            min_samples_per_class = 10
            for cls, count in class_dist.items():
                if count < min_samples_per_class:
                    self.training_error.emit(f"클래스 {cls} 샘플 부족 ({count} < {min_samples_per_class})")
                    return False
            
            # 데이터 범위 체크
            x_min, x_max = X.min(), X.max()
            if abs(x_min) > 1e6 or abs(x_max) > 1e6:
                logging.warning(f"X 범위 비정상: [{x_min:.2e}, {x_max:.2e}]")
            
            logging.info(f"데이터 검증 완료: X={X.shape}, y={y.shape}, 클래스={class_dist}")
            return True
            
        except Exception as ex:
            logging.error(f"_validate_training_data 오류: {ex}")
            self.training_error.emit(f"데이터 검증 실패: {str(ex)}")
            return False

    def _calculate_class_weight(self, y):
        """클래스 가중치 계산"""
        try:
            unique_classes, counts = np.unique(y, return_counts=True)
            
            if len(unique_classes) == 2:
                # 이진 분류
                scale_pos_weight = counts[0] / counts[1] if counts[1] > 0 else 1.0
            else:
                # 다중 분류
                total = len(y)
                scale_pos_weight = {
                    int(cls): total / (len(unique_classes) * count)
                    for cls, count in zip(unique_classes, counts)
                }
            
            logging.info(f"클래스 가중치: {scale_pos_weight}")
            return scale_pos_weight
            
        except Exception as ex:
            logging.error(f"_calculate_class_weight 오류: {ex}")
            return 1.0

    # ==================== 데이터 전송/수신 ====================
    
    def _send_training_data(self, X, y, scale_pos_weight):
        """훈련 데이터 전송"""
        try:
            request_id = str(uuid.uuid4())
            
            training_data = {
                'request_id': request_id,
                'X': X,
                'y': y,
                'scale_pos_weight': scale_pos_weight,
                'partial_update': self.partial_update,
                'seq_length': self.seq_length
            }
            
            data_bytes = pickle.dumps(training_data)
            data_len = len(data_bytes)
            
            logging.info(f"훈련 데이터 전송 시작: {data_len} 바이트")
            
            # 크기 제한 체크
            if self.partial_update and data_len > self.max_partial_update_size:
                self.training_error.emit(f"부분 업데이트 데이터 크기 초과: {data_len} > {self.max_partial_update_size}")
                return False
            
            if data_len > self.max_data_size:
                self.training_error.emit(f"데이터 크기 초과: {data_len} > {self.max_data_size}")
                return False
            
            # 청크 전송
            total_chunks = (data_len + self.chunk_size - 1) // self.chunk_size
            
            with self.pipe_lock:
                # 헤더 전송
                win32file.WriteFile(self.pipe_handle, struct.pack('I', data_len))
                win32file.WriteFile(self.pipe_handle, struct.pack('I', total_chunks))
                win32file.WriteFile(self.pipe_handle, b'TRAIN')
                
                # 청크 전송
                for chunk_idx in range(total_chunks):
                    if not self.running:
                        logging.info("훈련 중지 요청")
                        return False
                    
                    start_idx = chunk_idx * self.chunk_size
                    end_idx = min((chunk_idx + 1) * self.chunk_size, data_len)
                    chunk = data_bytes[start_idx:end_idx]
                    
                    chunk_header = struct.pack('I', chunk_idx)
                    win32file.WriteFile(self.pipe_handle, chunk_header + chunk)
                    
                    # 확인 수신
                    result = win32file.ReadFile(self.pipe_handle, 4)
                    if result[0] != 0:
                        self.training_error.emit(f"청크 {chunk_idx} 확인 실패")
                        return False
                    
                    ack_idx = struct.unpack('I', result[1])[0]
                    if ack_idx != chunk_idx:
                        self.training_error.emit(f"청크 인덱스 불일치: {ack_idx} != {chunk_idx}")
                        return False
                
                logging.info("훈련 데이터 전송 완료")
                return True
                
        except pywintypes.error as e:
            if e.winerror == 109:  # ERROR_BROKEN_PIPE
                self.training_error.emit("훈련 파이프 연결 끊김")
            else:
                self.training_error.emit(f"파이프 오류: {e}")
            return False
        except Exception as ex:
            logging.error(f"_send_training_data 오류: {ex}")
            self.training_error.emit(f"데이터 전송 실패: {str(ex)}")
            return False

    def _receive_training_response(self):
        """훈련 응답 수신"""
        try:
            with self.pipe_lock:
                # 응답 길이 수신
                result = win32file.ReadFile(self.pipe_handle, 4)
                if result[0] != 0:
                    self.training_error.emit("응답 길이 수신 실패")
                    return
                
                response_len = struct.unpack('I', result[1])[0]
                
                # 응답 데이터 수신
                response_data = b''
                while len(response_data) < response_len:
                    chunk = win32file.ReadFile(self.pipe_handle, response_len - len(response_data))[1]
                    if not chunk:
                        self.training_error.emit("응답 데이터 수신 중단")
                        return
                    response_data += chunk
                
                # 응답 파싱
                response = pickle.loads(response_data)
                
                if isinstance(response, dict):
                    status = response.get('status')
                    
                    if status == "TRAINING_COMPLETED":
                        best_threshold = response.get('best_threshold', -1.0)
                        f1_score = response.get('f1', 0.0)
                        
                        logging.info(f"훈련 완료: threshold={best_threshold:.3f}, f1={f1_score:.3f}")
                        
                        self.training_completed.emit(
                            True, 
                            f"훈련 완료 (F1: {f1_score:.3f})", 
                            best_threshold
                        )
                    else:
                        error_msg = response.get('status', '알 수 없는 오류')
                        self.training_error.emit(f"훈련 실패: {error_msg}")
                else:
                    self.training_error.emit("응답 형식 오류")
                    
        except pywintypes.error as e:
            if e.winerror == 109:
                self.training_error.emit("훈련 파이프 연결 끊김")
            else:
                self.training_error.emit(f"응답 수신 오류: {e}")
        except Exception as ex:
            logging.error(f"_receive_training_response 오류: {ex}")
            self.training_error.emit(f"응답 처리 실패: {str(ex)}")

    def stop(self):
        """스레드 정지"""
        logging.info("CnnTrainerThread 정지 중...")
        self.running = False
        self.quit()
        self.wait()
        logging.info("CnnTrainerThread 정지 완료")

class AutoTraderThread(QThread):
    """자동매매 스레드"""
    
    # 시그널 정의
    buy_signal = pyqtSignal(str, str, str, str)
    sell_signal = pyqtSignal(str, str)
    sell_half_signal = pyqtSignal(str, str)
    sell_all_signal = pyqtSignal()
    stock_removed_from_monitor = pyqtSignal(str)
    counter_updated = pyqtSignal(int)
    prediction_signal = pyqtSignal(str, float)
    stock_data_updated = pyqtSignal(list)

    def __init__(self, trader, window, seq_length=5, 
                 buy_threshold_base=0.75, sell_threshold_base=0.65, 
                 threshold_weight=0.05):
        super().__init__()
        
        # 핵심 참조
        self.trader = trader
        self.window = window
        
        # 상태 플래그
        self.running = True
        self.sell_all_emitted = False
        self.restart_after_training = False
        self.last_partial_update = False
        
        # 카운터
        self.counter = 0
        
        # 매매 파라미터
        self.buy_threshold_base = buy_threshold_base
        self.sell_threshold_base = sell_threshold_base
        self.threshold_weight = threshold_weight
        self.seq_length = seq_length
        
        # 예측 시스템 초기화
        self._init_prediction_system()
        
        # CNN 프로세스 초기화
        self._init_cnn_process()
        
        # 훈련 스레드
        self.trainer_thread = None

    # ==================== 초기화 메서드 ====================
    
    def _init_prediction_system(self):
        """예측 시스템 초기화"""
        self.prediction_history = {}
        self.prediction_ema = {}
        self.prediction_cache = {}
        self.last_prediction_time = {}
        self.prediction_update_interval = 3.0

    def _init_cnn_process(self):
        """CNN 프로세스 관련 초기화"""
        self.training_pipe_name = r'\\.\pipe\CnnTrainingPipe'
        self.prediction_pipe_name = r'\\.\pipe\CnnPredictionPipe'
        self.training_pipe_handle = None
        self.prediction_pipe_handle = None
        self.training_pipe_lock = threading.Lock()
        self.prediction_pipe_lock = threading.Lock()
        self.cnn_process = None

    # ==================== 스레드 생명주기 ====================
    
    def run(self):
        """메인 루프"""
        while self.running:
            self.autotrade()
            self.msleep(1000)

    def stop(self):
        """스레드 정지"""
        logging.info("AutoTraderThread 정지 시작...")
        self.running = False
        
        # 훈련 스레드 정지
        if self.trainer_thread and self.trainer_thread.isRunning():
            logging.info("CnnTrainerThread 정지 중...")
            self.trainer_thread.stop()
            self.trainer_thread = None
        
        # CNN 프로세스 정지
        self.stop_cnn_process()
        
        # 스레드 종료 대기
        self.quit()
        self.wait()
        logging.info("AutoTraderThread 정지 완료")

    # ==================== 설정 저장 ====================
    
    def save_trade_params(self):
        """거래 파라미터 저장"""
        config = configparser.ConfigParser()
        if os.path.exists('settings.ini'):
            config.read('settings.ini', encoding='utf-8')
        
        if not config.has_section('TRADE_PARAMS'):
            config.add_section('TRADE_PARAMS')
        
        config.set('TRADE_PARAMS', 'buy_threshold_base', str(round(self.buy_threshold_base, 3)))
        config.set('TRADE_PARAMS', 'sell_threshold_base', str(round(self.sell_threshold_base, 3)))
        config.set('TRADE_PARAMS', 'threshold_weight', str(self.threshold_weight))
        
        with open('settings.ini', 'w', encoding='utf-8') as cfg:
            config.write(cfg)

    # ==================== CNN 프로세스 관리 ====================
    
    def start_cnn_process(self):
        """CNN 서버 프로세스 시작"""
        try:
            python64_path = r"C:\MyAPP\day_trading\venv64\Scripts\python.exe"
            script_path = r"C:\MyAPP\day_trading\cnn_server.py"
            
            # 경로 검증
            if not os.path.exists(python64_path):
                logging.error(f"Python 실행 파일 없음: {python64_path}")
                logging.info("가상환경 경로를 확인하세요. 예: .venv/Scripts/python.exe")
                return False
            
            if not os.path.exists(script_path):
                logging.error(f"CNN 서버 스크립트 없음: {script_path}")
                return False
            
            # 기존 프로세스 정리
            if self.cnn_process is not None:
                logging.warning("기존 CNN 프로세스 종료 중...")
                self.stop_cnn_process()
                time.sleep(1)
            
            logging.info("CNN 프로세스 시작 중...")
            
            self.cnn_process = QProcess()
            self.cnn_process.setProcessChannelMode(QProcess.MergedChannels)
            
            # 시그널 연결
            self.cnn_process.readyReadStandardOutput.connect(self._on_cnn_output)
            self.cnn_process.readyReadStandardError.connect(self._on_cnn_error)
            self.cnn_process.errorOccurred.connect(self._on_cnn_process_error)
            self.cnn_process.finished.connect(self._on_cnn_finished)
            
            # 프로세스 시작
            self.cnn_process.start(python64_path, [script_path])
            
            # 타임아웃 30초
            if not self.cnn_process.waitForStarted(30000):
                error_msg = self.cnn_process.errorString()
                logging.error(f"CNN 프로세스 시작 실패: {error_msg}")
                self._cleanup_failed_process()
                return False
            
            logging.info("CNN 프로세스 시작 성공")
            
            # 파이프 연결
            if not self._connect_pipes():
                logging.error("파이프 연결 실패")
                self.stop_cnn_process()
                return False
            
            return True

        except Exception as ex:
            logging.error(f"start_cnn_process 오류: {ex}\n{traceback.format_exc()}")
            self._cleanup_failed_process()
            return False

    def _connect_pipes(self, max_attempts=20, retry_delay=0.5):
        """파이프 연결"""
        for attempt in range(max_attempts):
            try:
                self.training_pipe_handle = win32file.CreateFile(
                    self.training_pipe_name,
                    win32file.GENERIC_READ | win32file.GENERIC_WRITE,
                    0, None,
                    win32file.OPEN_EXISTING,
                    0, None
                )
                self.prediction_pipe_handle = win32file.CreateFile(
                    self.prediction_pipe_name,
                    win32file.GENERIC_READ | win32file.GENERIC_WRITE,
                    0, None,
                    win32file.OPEN_EXISTING,
                    0, None
                )
                logging.info("파이프 연결 성공")
                return True
                
            except Exception as e:
                if attempt < max_attempts - 1:
                    logging.debug(f"파이프 연결 시도 {attempt + 1}/{max_attempts}: {e}")
                    time.sleep(retry_delay)
                else:
                    logging.error(f"파이프 연결 최종 실패: {e}")
                    return False
        
        return False

    def _cleanup_failed_process(self):
        """실패한 프로세스 정리"""
        if self.cnn_process:
            try:
                if self.cnn_process.state() != QProcess.NotRunning:
                    self.cnn_process.terminate()
                    if not self.cnn_process.waitForFinished(3000):
                        self.cnn_process.kill()
            except Exception as ex:
                logging.error(f"프로세스 정리 오류: {ex}")
            finally:
                self.cnn_process = None

    def _on_cnn_output(self):
        """CNN 서버 출력 처리"""
        try:
            data = self.cnn_process.readAllStandardOutput().data().decode('utf-8', errors='ignore')
            for line in data.strip().split('\n'):
                if line:
                    logging.debug(f"[CNN] {line}")
        except Exception as ex:
            logging.error(f"CNN 출력 처리 오류: {ex}")

    def _on_cnn_error(self):
        """CNN 서버 에러 처리"""
        try:
            data = self.cnn_process.readAllStandardError().data().decode('utf-8', errors='ignore')
            for line in data.strip().split('\n'):
                if line:
                    logging.error(f"[CNN Error] {line}")
        except Exception as ex:
            logging.error(f"CNN 에러 처리 오류: {ex}")

    def _on_cnn_process_error(self, error):
        """QProcess 에러 핸들러"""
        error_map = {
            QProcess.FailedToStart: "프로세스 시작 실패 (파일 없음 또는 권한 부족)",
            QProcess.Crashed: "프로세스 비정상 종료",
            QProcess.Timedout: "프로세스 타임아웃",
            QProcess.WriteError: "쓰기 오류",
            QProcess.ReadError: "읽기 오류",
            QProcess.UnknownError: "알 수 없는 오류"
        }
        logging.error(f"CNN 프로세스 오류: {error_map.get(error, '알 수 없음')}")

    def _on_cnn_finished(self, exit_code, exit_status):
        """CNN 프로세스 종료 핸들러"""
        if exit_status == QProcess.CrashExit:
            logging.error(f"CNN 프로세스 비정상 종료 (코드: {exit_code})")
        else:
            logging.info(f"CNN 프로세스 정상 종료 (코드: {exit_code})")

    def stop_cnn_process(self):
        """CNN 프로세스 안전 종료"""
        try:
            # 1. 파이프에 종료 신호 전송
            self._send_stop_signal_to_pipes()
            
            # 2. 파이프 핸들 닫기
            self._close_pipe_handles()
            
            # 3. 프로세스 종료
            self._terminate_cnn_process()
            
            logging.info("CNN 프로세스 종료 완료")
            
        except Exception as ex:
            logging.error(f"stop_cnn_process 오류: {ex}\n{traceback.format_exc()}")

    def _send_stop_signal_to_pipes(self):
        """파이프에 종료 신호 전송"""
        stop_signal = struct.pack('I', 0) + b'STOP'
        
        if self.training_pipe_handle:
            with self.training_pipe_lock:
                try:
                    win32file.WriteFile(self.training_pipe_handle, stop_signal)
                    logging.debug("훈련 파이프에 종료 신호 전송")
                except pywintypes.error as e:
                    # ERROR_INVALID_HANDLE(6), ERROR_BROKEN_PIPE(109) 등은 무시
                    if e.winerror not in (6, 109, 232):
                        logging.warning(f"훈련 파이프 종료 신호 전송 실패: {e}")
                except Exception as e:
                    logging.warning(f"훈련 파이프 종료 신호 오류: {e}")
        
        if self.prediction_pipe_handle:
            with self.prediction_pipe_lock:
                try:
                    win32file.WriteFile(self.prediction_pipe_handle, stop_signal)
                    logging.debug("예측 파이프에 종료 신호 전송")
                except pywintypes.error as e:
                    if e.winerror not in (6, 109, 232):
                        logging.warning(f"예측 파이프 종료 신호 전송 실패: {e}")
                except Exception as e:
                    logging.warning(f"예측 파이프 종료 신호 오류: {e}")

    def _close_pipe_handles(self):
        """파이프 핸들 닫기"""
        if self.training_pipe_handle:
            with self.training_pipe_lock:
                try:
                    win32file.CloseHandle(self.training_pipe_handle)
                    logging.debug("훈련 파이프 핸들 닫기 완료")
                except Exception as e:
                    logging.warning(f"훈련 파이프 핸들 닫기 오류: {e}")
                finally:
                    self.training_pipe_handle = None
        
        if self.prediction_pipe_handle:
            with self.prediction_pipe_lock:
                try:
                    win32file.CloseHandle(self.prediction_pipe_handle)
                    logging.debug("예측 파이프 핸들 닫기 완료")
                except Exception as e:
                    logging.warning(f"예측 파이프 핸들 닫기 오류: {e}")
                finally:
                    self.prediction_pipe_handle = None

    def _terminate_cnn_process(self):
        """CNN 프로세스 종료"""
        if not self.cnn_process:
            return
        
        try:
            state = self.cnn_process.state()
            
            if state == QProcess.NotRunning:
                logging.debug("CNN 프로세스가 이미 종료됨")
                return
            
            # 정상 종료 시도
            self.cnn_process.terminate()
            
            # 3초 대기
            if not self.cnn_process.waitForFinished(3000):
                logging.warning("CNN 프로세스 정상 종료 실패, 강제 종료 시도")
                self.cnn_process.kill()
                
                # 추가 1초 대기
                if not self.cnn_process.waitForFinished(1000):
                    logging.error("CNN 프로세스 강제 종료 실패")
            else:
                logging.debug("CNN 프로세스 정상 종료 완료")
                
        except Exception as e:
            logging.error(f"CNN 프로세스 종료 오류: {e}")
        finally:
            self.cnn_process = None

    # ==================== 훈련 관리 ====================
    
    def start_training(self, partial_update=False):
        """훈련 시작"""
        if self.trainer_thread and self.trainer_thread.isRunning():
            logging.info("이미 훈련 스레드가 실행 중입니다")
            return
        
        self.last_partial_update = partial_update
        self.trainer_thread = CnnTrainerThread(
            self.trader, 
            self.training_pipe_handle, 
            self.training_pipe_lock, 
            partial_update, 
            self.seq_length
        )
        self.trainer_thread.training_completed.connect(self.on_training_completed)
        self.trainer_thread.training_error.connect(self.on_training_error)
        self.trainer_thread.start()

    def on_training_completed(self, success, message, threshold):
        """훈련 완료 핸들러"""
        logging.info(f"{message}")
        self.trainer_thread = None
        
        if threshold and threshold > 0:
            self.buy_threshold_base = float(threshold)
            self.sell_threshold_base = max(0.5, self.buy_threshold_base - 0.05)
            
            # 매수 직후 매도 방지
            if self.sell_threshold_base >= self.buy_threshold_base:
                self.buy_threshold_base = self.sell_threshold_base + 0.05
            
            self.save_trade_params()
            logging.info(f"임계치 업데이트: buy={self.buy_threshold_base:.3f}, sell={self.sell_threshold_base:.3f}")
        
        if success and self.restart_after_training:
            logging.info("모델/스케일러 생성 완료")
            self.restart_after_training = False
        
        if success and not self.last_partial_update:
            self.trader.init_database()

    def on_training_error(self, error_message):
        """훈련 오류 핸들러"""
        logging.error(f"훈련 오류: {error_message}")
        self.trainer_thread = None
        self.restart_after_training = False

    # ==================== 자동매매 메인 로직 ====================
    
    def autotrade(self):
        """자동매매 메인 루프"""
        try:
            t_now = datetime.now()
            
            # 카운터 업데이트
            self.counter += 1
            self.counter_updated.emit(self.counter)
            
            # 주식 데이터 업데이트
            self._update_stock_data_table()
            
            # 시간대별 로직 실행
            if self._is_trading_hours(t_now):
                self._execute_trading_logic(t_now)
            elif self._is_market_close_time(t_now):
                self._handle_market_close()
                
        except Exception as ex:
            logging.error(f"autotrade 오류: {ex}\n{traceback.format_exc()}")

    def _is_trading_hours(self, t_now):
        """거래 시간인지 확인"""
        t_0903 = t_now.replace(hour=9, minute=3, second=0, microsecond=0)
        t_1515 = t_now.replace(hour=15, minute=15, second=0, microsecond=0)
        return t_0903 < t_now <= t_1515

    def _is_market_close_time(self, t_now):
        """장 종료 시간인지 확인"""
        t_1515 = t_now.replace(hour=15, minute=15, second=0, microsecond=0)
        t_exit = t_now.replace(hour=15, minute=20, second=0, microsecond=0)
        return t_1515 < t_now < t_exit and not self.sell_all_emitted

    def _update_stock_data_table(self):
        """주식 데이터 테이블 업데이트"""
        stock_data_list = []
        
        # 스냅샷 복사로 안전성 확보
        monitored_codes = list(self.trader.monistock_set.copy())
        
        for code in monitored_codes:
            # 재확인 (제거되었을 수 있음)
            if code not in self.trader.monistock_set:
                continue
            
            tick_latest = self.trader.tickdata.get_latest_data(code)
            current_price = tick_latest.get('C', 0.0) if tick_latest else 0.0
            upward_prob = self.prediction_ema.get(code, 0.5)
            buy_price = self.trader.buy_price.get(code, 0.0)
            quantity = self.trader.buy_qty.get(code, 0)

            stock_data_list.append({
                'code': code,
                'current_price': float(current_price),
                'upward_probability': float(upward_prob),
                'buy_price': float(buy_price),
                'quantity': quantity
            })
        
        self.stock_data_updated.emit(stock_data_list)

    def _execute_trading_logic(self, t_now):
        """거래 로직 실행"""
        current_strategy = self.window.comboStg.currentText()
        buy_strategies = [
            stg for stg in self.window.strategies.get(current_strategy, []) 
            if stg['key'].startswith('buy')
        ]
        sell_strategies = [
            stg for stg in self.window.strategies.get(current_strategy, []) 
            if stg['key'].startswith('sell')
        ]
        
        if self.restart_after_training or not self.prediction_pipe_handle:
            return
        
        # 스냅샷 복사로 안전성 확보
        monitored_codes = list(self.trader.monistock_set.copy())
        
        for code in monitored_codes:
            # 재확인
            if code not in self.trader.monistock_set:
                continue
            
            try:
                # 매수 로직
                if code not in self.trader.buyorder_set and code not in self.trader.bought_set:
                    self._evaluate_buy_condition(code, t_now, current_strategy, buy_strategies)
                
                # 매도 로직
                elif (code in self.trader.bought_set and 
                      code not in self.trader.buyorder_set and 
                      code not in self.trader.sellorder_set):
                    self._evaluate_sell_condition(code, t_now, current_strategy, sell_strategies)
                    
            except Exception as ex:
                logging.error(f"{code} 거래 로직 오류: {ex}")

    def _handle_market_close(self):
        """장 종료 처리"""
        if self.trader.buyorder_set or self.trader.sellorder_set:
            return
        
        if self.trader.bought_set:
            logging.info("보유 주식 전부 매도")
            self.sell_all_signal.emit()
        
        self.sell_all_emitted = True

    # ==================== 매수 로직 ====================
    
    def _evaluate_buy_condition(self, code, t_now, strategy, buy_strategies):
        """매수 조건 평가"""
        # 데이터 로드
        tick_latest = self.trader.tickdata.get_latest_data(code)
        min_latest = self.trader.mindata.get_latest_data(code)
        
        if not tick_latest or not min_latest:
            return
        
        # 지표 검증
        if tick_latest.get('MAT5', 0) == 0:
            logging.debug(f"{code}: 지표 준비 미완료")
            return
        
        # 투자 대상 제거 조건 체크
        if self._should_remove_from_monitor(code, tick_latest, min_latest, t_now):
            return
        
        # CNN 예측
        recommendation, ema = self.get_cnn_recommendation_cached(code)
        
        # 시간대별 임계값 조정
        buy_threshold = self._get_adjusted_buy_threshold(t_now)
        
        # 매수 조건 평가
        if self._check_buy_conditions(code, strategy, tick_latest, min_latest, 
                                       recommendation, ema, buy_threshold, buy_strategies):
            self.buy_signal.emit(code, f"매수 신호({(ema * 100):.1f}%)", "0", "03")

    def _should_remove_from_monitor(self, code, tick_latest, min_latest, t_now):
        """투자 대상에서 제거해야 하는지 확인"""
        min_close = min_latest.get('C', 0)
        MAM5 = min_latest.get('MAM5', 0)
        MAM10 = min_latest.get('MAM10', 0)
        
        # starting_price 기준 하락 + MAM5 < MAM10
        if code in self.trader.starting_price:
            if min_close < self.trader.starting_price[code] * 0.99 and MAM5 < MAM10:
                try:
                    vi_time = datetime.strptime(
                        f"{datetime.now().year}/{self.trader.starting_time[code]}", 
                        '%Y/%m/%d %H:%M:%S'
                    )
                    if t_now - vi_time > timedelta(hours=1):
                        logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 제거 (하락)")
                        self.stock_removed_from_monitor.emit(code)
                        return True
                except Exception as ex:
                    logging.error(f"{code} VI 시각 파싱 오류: {ex}")
        
        # 상한가 종목 제거
        min_high_recent = min_latest.get('H_recent', [0, 0])
        min_low_recent = min_latest.get('L_recent', [0, 0])
        
        if len(min_high_recent) >= 2 and len(min_low_recent) >= 2:
            if all(h == l for h, l in zip(min_high_recent[-2:], min_low_recent[-2:])):
                logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 제거 (상한가)")
                self.stock_removed_from_monitor.emit(code)
                return True
        
        return False

    def _get_adjusted_buy_threshold(self, t_now):
        """시간대별 매수 임계값 조정"""
        time_factor = 1.0
        
        if t_now > t_now.replace(hour=10, minute=0, second=0):
            time_factor = 1.1  # 10시 이후 10% 상향
        if t_now > t_now.replace(hour=13, minute=0, second=0):
            time_factor = 1.2  # 13시 이후 20% 상향
        
        return self.buy_threshold_base * time_factor

    def _check_buy_conditions(self, code, strategy, tick_latest, min_latest, 
                              recommendation, ema, buy_threshold, buy_strategies):
        """매수 조건 종합 체크"""
        # 기본 조건
        if (recommendation != "BUY" or 
            ema <= buy_threshold or
            len(self.trader.bought_set) >= self.trader.target_buy_count):
            return False
        
        # 최신값 추출
        min_close = min_latest.get('C', 0)
        MAM5 = min_latest.get('MAM5', 0)
        MAM10 = min_latest.get('MAM10', 0)
        MAT5 = tick_latest.get('MAT5', 0)
        MAT20 = tick_latest.get('MAT20', 0)
        MAT60 = tick_latest.get('MAT60', 0)
        MAT120 = tick_latest.get('MAT120', 0)
        
        # 최근 데이터 (트렌드)
        tick_recent = self.trader.tickdata.get_recent_data(code, 3)
        MACDT_SIGNAL = tick_recent.get('MACDT_SIGNAL', [0, 0, 0])
        OSCT = tick_recent.get('OSCT', [0, 0, 0])
        
        macdt_increasing = (len(MACDT_SIGNAL) >= 2 and 
                           MACDT_SIGNAL[-1] > MACDT_SIGNAL[-2])
        osct_positive = all(x > 0 for x in OSCT[-2:])
        
        # positive_candle
        min_close_recent = min_latest.get('C_recent', [0, 0])
        min_open_recent = min_latest.get('O_recent', [0, 0])
        positive_candle = all(
            min_close_recent[i] > min_open_recent[i] 
            for i in range(min(2, len(min_close_recent), len(min_open_recent)))
        )
        
        # VI 발동 전략 핵심 조건
        if strategy == "VI 발동":
            return (min_close > MAM5 > MAM10 and
                    MAT5 > MAT20 and
                    MAT60 > MAT120 and
                    macdt_increasing and
                    osct_positive and
                    positive_candle)
        
        # 기타 전략 조건 평가
        return self._evaluate_strategy_conditions(code, buy_strategies, tick_latest, min_latest)

    def _evaluate_strategy_conditions(self, code, strategies, tick_latest, min_latest):
        """전략별 조건 평가"""
        if not strategies:
            return False
        
        # 전체 데이터 로드 (전략 평가용)
        tick_data_full = self.trader.tickdata.get_recent_data(code, 10)
        min_data_full = self.trader.mindata.get_recent_data(code, 10)
        
        # 변수 설정
        tick_close_price = tick_data_full.get('C', [0])
        MAT5 = tick_data_full.get('MAT5', [0])
        MAT20 = tick_data_full.get('MAT20', [0])
        MAT60 = tick_data_full.get('MAT60', [0])
        RSIT = tick_data_full.get('RSIT', [0])
        OSCT = tick_data_full.get('OSCT', [0])
        bb_upper = tick_data_full.get('BB_UPPER', [0])
        
        min_close_price = min_data_full.get('C', [0])
        MAM5 = min_data_full.get('MAM5', [0])
        MAM10 = min_data_full.get('MAM10', [0])
        RSI = min_data_full.get('RSI', [0])
        OSC = min_data_full.get('OSC', [0])
        VWAP = min_data_full.get('VWAP', [0])
        
        # positive_candle
        min_close_recent = min_data_full.get('C', [0, 0])[-2:]
        min_open_recent = min_data_full.get('O', [0, 0])[-2:]
        positive_candle = all(
            min_close_recent[i] > min_open_recent[i] 
            for i in range(min(2, len(min_close_recent), len(min_open_recent)))
        )
        
        # 전략 평가
        for strategy in strategies:
            try:
                condition = strategy.get('content', '')
                if eval(condition):
                    logging.debug(f"{code}: {strategy.get('name')} 조건 만족")
                    return True
            except Exception as ex:
                logging.error(f"{code} 전략 평가 오류: {ex}")
        
        return False

    # ==================== 매도 로직 ====================
    
    def _evaluate_sell_condition(self, code, t_now, strategy, sell_strategies):
        """매도 조건 평가"""
        # 데이터 로드
        tick_latest = self.trader.tickdata.get_latest_data(code)
        min_latest = self.trader.mindata.get_latest_data(code)
        
        if not tick_latest or not min_latest:
            return
        
        tick_close = tick_latest.get('C', 0)
        
        # 최고가 업데이트
        self.trader.update_highest_price(code, tick_close)
        
        # 수익률 계산
        buy_price = self.trader.buy_price.get(code, 0)
        if buy_price == 0:
            return
        
        self.stock_rate = (tick_close / buy_price - 1) * 100
        
        # 손절매
        if self.stock_rate < -0.5:
            logging.info(f"{cpCodeMgr.CodeToName(code)}({code}): 손절매")
            self.sell_signal.emit(code, "손절매")
            return
        
        # 분할 매도
        if self.stock_rate > 1.0 and code not in self.trader.sell_half_set:
            logging.info(f"{cpCodeMgr.CodeToName(code)}({code}): 분할 매도")
            self.sell_half_signal.emit(code, "1.0% 초과")
            return
        
        # CNN 예측
        recommendation, ema = self.get_cnn_recommendation_cached(code)
        
        # VI 발동 전략: CNN 매도 신호
        if strategy == "VI 발동" and recommendation == "SELL":
            self.sell_signal.emit(code, f"CNN 매도({(ema * 100):.1f}%)")
            return
        
        # 전략별 매도 조건 평가
        if self._evaluate_strategy_conditions(code, sell_strategies, tick_latest, min_latest):
            self.sell_signal.emit(code, f"전략 매도({(ema * 100):.1f}%)")

    # ==================== CNN 예측 시스템 ====================
    
    def get_cnn_recommendation_cached(self, code):
        """CNN 예측 (캐싱 적용)"""
        current_time = time.time()
        
        # 캐시 확인
        if code in self.prediction_cache:
            cached_rec, cached_ema, cached_time = self.prediction_cache[code]
            
            # 3초 이내면 캐시 사용
            if current_time - cached_time < self.prediction_update_interval:
                return cached_rec, cached_ema
        
        # 새로 예측
        recommendation, ema = self.get_cnn_recommendation(code)
        
        # 캐시 업데이트
        if recommendation is not None:
            self.prediction_cache[code] = (recommendation, ema, current_time)
        
        return recommendation if recommendation is not None else "HOLD", ema if ema is not None else 0.5

    def get_cnn_recommendation(self, code):
        """CNN 예측 메인"""
        try:
            # 1. 데이터 준비
            tick_data, min_data = self._prepare_prediction_data(code)
            if tick_data is None or min_data is None:
                return None, 0.5
            
            # 2. 시퀀스 생성
            combined_data = self._create_prediction_sequence(code, tick_data, min_data)
            if combined_data is None:
                return None, 0.5
            
            # 3. 예측 요청
            prediction = self._request_prediction(code, combined_data)
            if prediction is None:
                return None, 0.5
            
            # 4. 예측값 후처리
            recommendation, ema = self._post_process_prediction(code, prediction)
            
            # 5. 시그널 발송
            self.prediction_signal.emit(code, ema)
            
            return recommendation, ema
            
        except Exception as ex:
            logging.error(f"get_cnn_recommendation({code}) 오류: {ex}")
            return None, 0.5

    def _prepare_prediction_data(self, code):
        """예측용 데이터 준비"""
        today = datetime.now().strftime('%Y%m%d')
        
        # VI 발동 시각 추출
        start_hhmm = self._get_vi_start_time(code, today)
        
        # Tick 데이터 추출
        tick_data = self._extract_tick_data(code, today, start_hhmm)
        if tick_data is None:
            return None, None
        
        # Min 데이터 추출
        min_data = self._extract_min_data(code, today, start_hhmm)
        
        return tick_data, min_data

    def _get_vi_start_time(self, code, today):
        """VI 발동 시각 추출"""
        if code in self.trader.starting_time:
            try:
                start_dt = datetime.strptime(
                    f"{datetime.now().year}/{self.trader.starting_time[code]}", 
                    '%Y/%m/%d %H:%M:%S'
                )
                if start_dt.strftime('%Y%m%d') == today:
                    return start_dt.strftime('%H%M')
            except ValueError as ve:
                logging.error(f"{code}: starting_time 형식 오류: {ve}")
        
        return '0900'  # 기본값

    def _extract_tick_data(self, code, today, start_hhmm):
        """Tick 데이터 추출"""
        tick_features = [
            'C', 'V', 'MAT5', 'MAT20', 'MAT60', 'MAT120', 'RSIT', 'RSIT_SIGNAL',
            'MACDT', 'MACDT_SIGNAL', 'OSCT', 'STOCHK', 'STOCHD', 'ATR', 'CCI',
            'BB_UPPER', 'BB_MIDDLE', 'BB_LOWER', 'BB_POSITION', 'BB_BANDWIDTH',
            'MAT5_MAT20_DIFF', 'MAT20_MAT60_DIFF', 'MAT60_MAT120_DIFF',
            'C_MAT5_DIFF', 'MAT5_CHANGE', 'MAT20_CHANGE', 'MAT60_CHANGE', 
            'MAT120_CHANGE', 'VWAP'
        ]
        
        with self.trader.tickdata.stockdata_lock:
            tick_data = self.trader.tickdata.stockdata.get(code, {})
            
            # 검증
            if not self._validate_tick_data(tick_data, tick_features):
                return None
            
            # 유효 인덱스 추출
            valid_indices = self._get_valid_tick_indices(tick_data, today, start_hhmm, tick_features)
            if len(valid_indices) < self.seq_length:
                logging.debug(f"{code}: 유효 틱 데이터 부족 ({len(valid_indices)} < {self.seq_length})")
                return None
            
            # 최신 seq_length개 선택
            indices = self._select_latest_indices(tick_data, valid_indices)
            
            # 데이터 추출
            return self._build_tick_array(tick_data, indices, tick_features)

    def _validate_tick_data(self, tick_data, features):
        """Tick 데이터 유효성 검증"""
        if not tick_data or not any(tick_data.get(f) for f in ['C', 'V', 'D', 'T']):
            return False
        
        dates = tick_data.get('D', [])
        timestamps = tick_data.get('T', [])
        
        if not dates or not timestamps or len(dates) != len(timestamps):
            return False
        
        # 누락 피처 보완
        for f in features:
            if f not in tick_data or len(tick_data[f]) < max(len(dates), self.seq_length):
                tick_data[f] = [0] * max(len(dates), self.seq_length)
        
        return True

    def _get_valid_tick_indices(self, tick_data, today, start_hhmm, features):
        """유효 Tick 인덱스 추출"""
        dates = tick_data['D']
        timestamps = tick_data['T']
        min_length = min(len(tick_data.get(f, [])) for f in ['C', 'V', 'D', 'T'])
        
        return [
            i for i in range(min_length)
            if str(dates[i]) == today 
            and start_hhmm <= f"{timestamps[i]:04d}" <= '1515'
            and all(
                len(tick_data.get(f, [])) > i 
                and tick_data.get(f, [0])[i] is not None 
                and not np.isnan(float(tick_data.get(f, [0])[i]))
                for f in ['C', 'V', 'MAT5', 'RSIT', 'OSCT']  # 핵심 피처만 체크
            )
        ]

    def _select_latest_indices(self, tick_data, valid_indices):
        """최신 seq_length개 인덱스 선택"""
        dates = tick_data['D']
        timestamps = tick_data['T']
        
        # (날짜, 시간, 인덱스) 튜플 생성 후 정렬
        items = [(str(dates[i]), f"{timestamps[i]:04d}", i) for i in valid_indices]
        sorted_items = sorted(items, key=lambda x: (x[0], x[1], -x[2]), reverse=True)
        
        return [item[2] for item in sorted_items[:self.seq_length]]

    def _build_tick_array(self, tick_data, indices, features):
        """Tick 배열 생성"""
        result = np.zeros((self.seq_length, len(features)))
        
        for j, feature in enumerate(features):
            data_list = tick_data.get(feature, [])
            
            # MAT 변화율 계산
            if feature.endswith('_CHANGE'):
                base_feature = feature.replace('_CHANGE', '')
                base_data = tick_data.get(base_feature, [0] * len(indices))
                values = np.array([base_data[i] for i in indices])
                changes = [(values[k] - values[k-1]) if k > 0 else 0 for k in range(len(values))]
                result[:, j] = changes
            else:
                result[:, j] = [data_list[i] for i in indices]
        
        return result

    def _extract_min_data(self, code, today, start_hhmm):
        """Min 데이터 추출"""
        min_features = [
            'MAM5', 'MAM10', 'MAM20', 'RSI', 'RSI_SIGNAL', 'MACD', 'MACD_SIGNAL',
            'OSC', 'STOCHK', 'STOCHD', 'CCI', 'MAM5_MAM10_DIFF', 'MAM10_MAM20_DIFF',
            'C_MAM5_DIFF', 'C_ABOVE_MAM5', 'VWAP'
        ]
        
        with self.trader.mindata.stockdata_lock:
            min_data = self.trader.mindata.stockdata.get(code, {})
            
            if not min_data or not any(min_data.get(f) for f in min_features + ['D', 'T']):
                logging.debug(f"{code}: min_data 없음")
                return np.zeros((self.seq_length, len(min_features)))
            
            dates = min_data.get('D', [])
            timestamps = min_data.get('T', [])
            
            if not dates or not timestamps:
                return np.zeros((self.seq_length, len(min_features)))
            
            # 유효 인덱스 딕셔너리 생성
            min_data_dict = {
                t: i for i, t in enumerate(timestamps) 
                if str(dates[i]) == today and start_hhmm <= f"{t:04d}" <= '1515'
            }
            
            return min_data_dict, min_data, min_features

    def _create_prediction_sequence(self, code, tick_data, min_data_info):
        """예측 시퀀스 생성"""
        try:
            min_data_dict, min_data, min_features = min_data_info
            
            # Tick 시간에 대응하는 Min 데이터 매핑
            tick_times = self.trader.tickdata.stockdata[code]['T']
            with self.trader.tickdata.stockdata_lock:
                tick_indices = self._select_latest_indices(
                    self.trader.tickdata.stockdata[code],
                    self._get_valid_tick_indices(
                        self.trader.tickdata.stockdata[code],
                        datetime.now().strftime('%Y%m%d'),
                        self._get_vi_start_time(code, datetime.now().strftime('%Y%m%d')),
                        ['C', 'V']
                    )
                )
            
            tick_times_selected = [tick_times[i] for i in tick_indices]
            
            # Min 데이터 매핑
            min_array = np.zeros((self.seq_length, len(min_features)))
            
            for idx, tick_time in enumerate(tick_times_selected):
                # 틱 시간을 분봉 시간으로 변환
                hh, mm = divmod(int(tick_time), 100)
                converted_mintime = hh * 60 + mm
                a, _ = divmod(converted_mintime, 3)
                interval_time = a * 3
                chart_time = interval_time + 3
                hour, minute = divmod(chart_time, 60)
                min_time_mapped = hour * 100 + minute
                
                if min_time_mapped > 1515:
                    min_time_mapped = 1515
                
                # 매핑된 시간 또는 이전 시간 데이터 찾기
                if min_time_mapped in min_data_dict:
                    min_idx = min_data_dict[min_time_mapped]
                else:
                    prev_times = [t for t in min_data_dict.keys() if t < min_time_mapped]
                    if prev_times:
                        min_time_mapped = max(prev_times)
                        min_idx = min_data_dict[min_time_mapped]
                    else:
                        continue
                
                # 데이터 추출
                for j, feature in enumerate(min_features):
                    values = min_data.get(feature, [])
                    if len(values) > min_idx:
                        min_array[idx, j] = values[min_idx]
            
            # 결합
            combined_data = np.hstack((tick_data, min_array)).flatten().reshape(1, -1)
            
            # 검증
            expected_size = (1, self.seq_length * (len(tick_data[0]) + len(min_features)))
            if combined_data.shape != expected_size:
                logging.error(f"{code}: 데이터 크기 불일치 - 예상 {expected_size}, 실제 {combined_data.shape}")
                return None
            
            return combined_data
            
        except Exception as ex:
            logging.error(f"_create_prediction_sequence({code}) 오류: {ex}")
            return None

    def _request_prediction(self, code, combined_data):
        """예측 요청"""
        try:
            request_id = str(uuid.uuid4())
            combined_data_with_id = {
                'request_id': request_id, 
                'data': combined_data
            }
            
            with self.prediction_pipe_lock:
                data_bytes = pickle.dumps(combined_data_with_id)
                data_len = len(data_bytes)
                
                # 전송
                win32file.WriteFile(self.prediction_pipe_handle, struct.pack('I', data_len))
                win32file.WriteFile(self.prediction_pipe_handle, b'PREDI')
                win32file.WriteFile(self.prediction_pipe_handle, data_bytes)

                # 수신
                result = win32file.ReadFile(self.prediction_pipe_handle, 4)
                if result[0] != 0:
                    logging.error(f"{code}: 예측 파이프 읽기 실패")
                    return None
                
                prediction_len = struct.unpack('I', result[1])[0]
                prediction_data = b''
                
                while len(prediction_data) < prediction_len:
                    chunk = win32file.ReadFile(
                        self.prediction_pipe_handle, 
                        prediction_len - len(prediction_data)
                    )[1]
                    if not chunk:
                        logging.error(f"{code}: 예측 데이터 수신 중 손실")
                        break
                    prediction_data += chunk
                
                if len(prediction_data) != prediction_len:
                    logging.error(f"{code}: 수신 데이터 길이 불일치")
                    return None
                
                response = pickle.loads(prediction_data)
                
                # 응답 검증
                if isinstance(response, dict):
                    if 'error' in response and response['error']:
                        error_msg = response['error']
                        logging.error(f"{code}: 서버 오류: {error_msg}")
                        
                        # Feature count mismatch 처리
                        if 'Feature count mismatch' in error_msg:
                            self._handle_feature_mismatch(error_msg)
                        
                        return None
                    
                    if response.get('request_id') == request_id:
                        return response['prediction']
                    else:
                        logging.error(f"{code}: 요청 ID 불일치")
                        return None
                else:
                    logging.error(f"{code}: 응답 형식 오류")
                    return None
                    
        except pywintypes.error as e:
            if e.winerror == 109:  # ERROR_BROKEN_PIPE
                logging.error(f"{code}: 예측 파이프 연결 끊김")
            else:
                logging.error(f"{code}: 예측 파이프 오류: {e}")
            return None
        except Exception as e:
            logging.error(f"{code}: 예측 요청 오류: {e}")
            return None

    def _handle_feature_mismatch(self, error_msg):
        """Feature count mismatch 처리"""
        try:
            # "expected 145, got 130" 형식에서 숫자 추출
            expected = int(error_msg.split('expected')[1].split(',')[0].strip())
            tick_features_count = 29
            min_features_count = 16
            features_per_step = tick_features_count + min_features_count
            
            new_seq = expected // features_per_step
            
            if new_seq > 0 and new_seq != self.seq_length:
                logging.warning(f"seq_length {self.seq_length} -> {new_seq} (서버 요구)")
                self.seq_length = new_seq
        except Exception as parse_err:
            logging.warning(f"seq_length 파싱 실패: {parse_err}")

    def _post_process_prediction(self, code, prediction):
        """예측값 후처리"""
        try:
            # 예측값 검증
            prediction = float(prediction)
            if not 0 <= prediction <= 1:
                logging.warning(f"{code}: 비정상 예측값 {prediction}, 기본값 0.5 사용")
                prediction = 0.5
        except (ValueError, TypeError) as e:
            logging.error(f"{code}: 예측값 변환 실패: {e}")
            prediction = 0.5
        
        # 예측 이력 저장
        if code not in self.prediction_history:
            self.prediction_history[code] = []
        
        ema_window = 3
        self.prediction_history[code].append(prediction)
        self.prediction_history[code] = self.prediction_history[code][-ema_window*2:]
        
        # EMA 계산
        if len(self.prediction_history[code]) >= ema_window:
            ema_array = talib.EMA(
                np.array(self.prediction_history[code], dtype=np.float64), 
                timeperiod=ema_window
            )
            ema = ema_array[-1]
        else:
            ema = np.mean(self.prediction_history[code])
        
        self.prediction_ema[code] = ema
        
        # ATR 정규화
        atr_values = self.trader.tickdata.stockdata.get(code, {}).get('ATR', [0])[-ema_window:]
        if atr_values and max(atr_values) > min(atr_values):
            atr_normalized = (atr_values[-1] - min(atr_values)) / (max(atr_values) - min(atr_values))
        else:
            atr_normalized = 0.0
        
        # 엔트로피 계산
        hist, _ = np.histogram(self.prediction_history[code], bins=10, range=(0, 1), density=True)
        pred_entropy = entropy(hist + 1e-6) / np.log(10)
        
        # 동적 임계값 계산
        buy_threshold = self.buy_threshold_base + self.threshold_weight * (atr_normalized + pred_entropy)
        sell_threshold = self.sell_threshold_base - self.threshold_weight * (atr_normalized + pred_entropy)
        
        # 추천 결정
        if ema > buy_threshold:
            recommendation = "BUY"
        elif ema < sell_threshold:
            recommendation = "SELL"
        else:
            recommendation = "HOLD"
        
        return recommendation, ema
    
class LoginHandler:
    def __init__(self, parent_window):
        self.parent = parent_window
        self.config = configparser.ConfigParser()
        self.config_file = 'settings.ini'
        self.process = None
        self.slack = None
        self.slack_channel = '#stock'

    def load_settings(self):
        if os.path.exists(self.config_file):
            self.config.read(self.config_file, encoding='utf-8')
            self.parent.loginEdit.setText(self.config.get('LOGIN', 'username', fallback=''))
            self.parent.passwordEdit.setText(self.config.get('LOGIN', 'password', fallback=''))
            self.parent.certpasswordEdit.setText(self.config.get('LOGIN', 'certpassword', fallback=''))
            self.parent.autoLoginCheckBox.setChecked(self.config.getboolean('LOGIN', 'autologin', fallback=False))

            self.parent.buycountEdit.setText(self.config.get('BUYCOUNT', 'target_buy_count', fallback='3'))

            if self.config.has_section('SLACK'):
                token = self.config.get('SLACK', 'token', fallback='')
                self.slack_channel = self.config.get('SLACK', 'channel', fallback='#stock')
                if token:
                    self.slack = Slacker(token)

            if self.config.has_section('TRADE_PARAMS'):
                self.parent.buy_threshold_base = float(self.config.get('TRADE_PARAMS', 'buy_threshold_base', fallback='0.75'))
                self.parent.sell_threshold_base = float(self.config.get('TRADE_PARAMS', 'sell_threshold_base', fallback='0.65'))
                self.parent.threshold_weight = float(self.config.get('TRADE_PARAMS', 'threshold_weight', fallback='0.05'))
            else:
                self.parent.buy_threshold_base = 0.75
                self.parent.sell_threshold_base = 0.65
                self.parent.threshold_weight = 0.05

    def save_settings(self):
        if not self.config.has_section('LOGIN'):
            self.config.add_section('LOGIN')
        self.config.set('LOGIN', 'username', self.parent.loginEdit.text())
        self.config.set('LOGIN', 'password', self.parent.passwordEdit.text())
        self.config.set('LOGIN', 'certpassword', self.parent.certpasswordEdit.text())
        self.config.set('LOGIN', 'autologin', str(self.parent.autoLoginCheckBox.isChecked()))
        with open(self.config_file, 'w', encoding='utf-8') as configfile:
            self.config.write(configfile)

    def attempt_auto_login(self):
        if self.parent.autoLoginCheckBox.isChecked() and self.parent.loginEdit.text() and self.parent.passwordEdit.text() and self.parent.certpasswordEdit.text():
            self.handle_login()

    def handle_login(self):
        username = self.parent.loginEdit.text()
        password = self.parent.passwordEdit.text()
        certpassword = self.parent.certpasswordEdit.text()

        self.save_settings()
        self.parent.clean_up_processes()

        self.process = QProcess(self.parent)
        creon_path = 'C:\\CREON\\STARTER\\coStarter.exe'
        args = ["/prj:cp", f"/id:{username}", f"/pwd:{password}", f"/pwdcert:{certpassword}"]
        if self.parent.autoLoginCheckBox.isChecked():
            args.append('/autostart')
        self.process.start(creon_path, args)
        self.process.finished.connect(self.init_plus_check_and_continue)
       
    def buycount_setting(self):
        if not self.config.has_section('BUYCOUNT'):
            self.config.add_section('BUYCOUNT')
        self.config.set('BUYCOUNT', 'target_buy_count', self.parent.buycountEdit.text())

        with open(self.config_file, 'w', encoding='utf-8') as configfile:
            self.config.write(configfile)

        logging.info(f"최대투자 종목수가 업데이트되었습니다.")

    def init_plus_check_and_continue(self):
        if not init_plus_check():
            exit()
        # self.auto_select_creon_popup()
        self.parent.post_login_setup()

    def auto_select_creon_popup(self):       
        # time.sleep(10)

        try:
            button_x, button_y = 960, 500
            pyautogui.moveTo(button_x, button_y, duration=0.5)
            pyautogui.click()
            
            logging.info("모의투자 접속 버튼 클릭 완료")
        except Exception as e:
            logging.error(f"모의투자 접속 버튼 클릭 실패: {e}")

class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.is_loading_strategy = False
        self.market_close_emitted = False
        self.login_handler = LoginHandler(self)
        self.init_ui()
        self.login_handler.load_settings()
        self.login_handler.attempt_auto_login()
        self.cnn_train_timer = None
        self.update_chart_status_timer = None

    def __del__(self):
        self.objstg.Clear()

    def post_login_setup(self):
        logger = logging.getLogger()
        if not any(isinstance(handler, QTextEditLogger) for handler in logger.handlers):
            text_edit_logger = QTextEditLogger(self.terminalOutput)
            text_edit_logger.setLevel(logging.INFO)
            logger.addHandler(text_edit_logger)
        buycount = int(self.buycountEdit.text())
        self.trader = CTrader(cpTrade, cpBalance, cpCodeMgr, cpCash, cpOrder, cpStock, buycount, self)
        self.objstg = CpStrategy(self.trader)
        self.trader_thread = AutoTraderThread(
            self.trader,
            self,
            buy_threshold_base=getattr(self, 'buy_threshold_base', 0.75),
            sell_threshold_base=getattr(self, 'sell_threshold_base', 0.65),
            threshold_weight=getattr(self, 'threshold_weight', 0.05),
        )

        self.chartdrawer = ChartDrawer(self.fig, self.canvas, self.trader, self.trader_thread, self)

        self.code = ''
        self.stocks = []
        self.counter = 0

        self.trader.get_stock_balance('START', 'post_login_setup')
        logging.info(f"시작 시간 : {datetime.now().strftime('%m/%d %H:%M:%S')}")

        self.trader.stock_added_to_monitor.connect(self.on_stock_added)
        self.trader.stock_bought.connect(self.on_stock_bought)
        self.trader.stock_sold.connect(self.on_stock_sold)

        self.close_external_popup()         

        self.load_strategy()

        self.start_timers()
        self.trader_thread.buy_signal.connect(self.trader.buy_stock)
        self.trader_thread.sell_signal.connect(self.trader.sell_stock)
        self.trader_thread.sell_half_signal.connect(self.trader.sell_half_stock)
        self.trader_thread.sell_all_signal.connect(self.trader.sell_all)
        self.trader_thread.stock_removed_from_monitor.connect(self.on_stock_removed)
        self.trader_thread.counter_updated.connect(self.update_counter_label)
        self.trader_thread.prediction_signal.connect(self.chartdrawer.update_prediction)
        self.trader_thread.stock_data_updated.connect(self.update_stock_table)
        self.trader_thread.start()

    def start_timers(self):
        now = datetime.now()
        start_time = now.replace(hour=9, minute=0, second=0, microsecond=0)
        end_time = now.replace(hour=15, minute=20, second=0, microsecond=0)
        today = datetime.today().weekday()

        if today == 5 or today == 6:
            logging.info(f"Today is {'Saturday.' if today == 5 else 'Sunday.'}")
            logging.info(f"오늘은 장이 쉽니다.")
            return
        
        if now < start_time:
            logging.info(f"자동 매매 시작 대기")
            self.trader.init_database()
            QTimer.singleShot(int((start_time - now).total_seconds() * 1000) + 1000, self.start_timers)

        elif start_time <= now < end_time:
            logging.info(f"자동 매매 시작")
            send_slack_message(self.login_handler, "#stock", f"자동 매매 시작")

            self.MODEL_DIR = r"C:\MyAPP\stock_trading"
            if not os.path.exists(self.MODEL_DIR):
                os.makedirs(self.MODEL_DIR)
            if not os.access(self.MODEL_DIR, os.W_OK):
                raise PermissionError(f"No write access to {self.MODEL_DIR}")

            model_path = os.path.join(self.MODEL_DIR, 'cnn_model.keras')
            scaler_path = os.path.join(self.MODEL_DIR, 'cnn_scaler.pkl')
            if not os.path.exists(model_path) or not os.path.exists(scaler_path):
                logging.info("모델 또는 스케일러 파일이 없음")
                self.trader_thread.restart_after_training = True
                self.cnn_train_timer = QTimer()
                self.cnn_train_timer.timeout.connect(self.periodic_train_cnn)
                self.cnn_train_timer.start(300000)  # 5분마다 호출
                self.periodic_train_cnn()

            self.update_chart_status_timer = QTimer(self)
            self.update_chart_status_timer.timeout.connect(self.update_chart_status_label)
            self.update_chart_status_timer.start(2000)  # 2초마다 갱신
            
            QTimer.singleShot(int((end_time - now).total_seconds() * 1000) + 1000, self.start_timers)
            
        elif end_time <= now and not self.market_close_emitted:
            self.trader_thread.start_training(partial_update=False)
            
            if self.cnn_train_timer is not None:
                self.cnn_train_timer.stop()
            if self.trader.save_data_timer is not None:
                self.trader.save_data_timer.stop()
            if self.trader.tickdata is not None:
                self.trader.tickdata.update_data_timer.stop()
            if self.trader.mindata is not None:
                self.trader.mindata.update_data_timer.stop()
            if self.trader.daydata is not None:
                self.trader.daydata.update_data_timer.stop()
            if self.update_chart_status_timer is not None:
                self.update_chart_status_timer.stop()
                
            for code in list(self.trader.monistock_set):
                if code not in self.trader.bought_set:
                    self.on_stock_removed(code)
                    self.trader.delete_list_db(code)
            for code in list(self.trader.vistock_set):
                if code not in self.trader.monistock_set and code not in self.trader.bought_set:
                    self.trader.delete_list_db(code)

            self.market_close_emitted = True
            logging.info(f"자동 매매 종료")
            send_slack_message(self.login_handler, "#stock", f"자동 매매 종료")
            

    def periodic_train_cnn(self):
        if self.trader.monistock_set:
            self.trader_thread.start_training(partial_update=True)

    def update_chart_status_label(self):
        if hasattr(self, 'chartdrawer') and self.chartdrawer.last_chart_update_time:
            chart_age = int(datetime.now().strftime("%H%M")) - self.chartdrawer.last_chart_update_time
            if chart_age < 2:
                chart_color = "green"
            else:
                chart_color = "red"
            self.chart_status_label.setText(f"Chart: {chart_age}m ago")
            self.chart_status_label.setStyleSheet(f"color: {chart_color}")
        else:
            self.chart_status_label.setText("Chart: None")
            self.chart_status_label.setStyleSheet("color: red")

    def update_stock_table(self, stock_data_list):
        """monistock_set 데이터를 기반으로 표 업데이트"""
        self.stock_table.setRowCount(len(stock_data_list))
        for row, stock_data in enumerate(stock_data_list):
            code = stock_data.get('code', '')
            current_price = stock_data.get('current_price', 0.0)
            upward_prob = stock_data.get('upward_probability', 0.0) * 100  # 0~1을 %로 변환
            buy_price = stock_data.get('buy_price', 0.0)
            quantity = stock_data.get('quantity', 0)
            profit_loss = (current_price - buy_price) * quantity
            return_pct = ((current_price - buy_price) / buy_price * 100) if buy_price != 0 else 0.0

            code_item = QTableWidgetItem(code)
            code_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.stock_table.setItem(row, 0, code_item)

            current_price_item = QTableWidgetItem(f"{current_price:,.0f}")
            current_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 1, current_price_item)

            upward_prob_item = QTableWidgetItem(f"{upward_prob:.2f}")
            upward_prob_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 2, upward_prob_item)

            buy_price_item = QTableWidgetItem(f"{buy_price:,.0f}")
            buy_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 3, buy_price_item)

            profit_loss_item = QTableWidgetItem(f"{profit_loss:,.0f}")
            profit_loss_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 4, profit_loss_item)
            profit_loss_item = self.stock_table.item(row, 4)
            if profit_loss > 0:
                profit_loss_item.setForeground(Qt.green)
            elif profit_loss < 0:
                profit_loss_item.setForeground(Qt.red)
            else:
                profit_loss_item.setForeground(Qt.black)

            return_item = QTableWidgetItem(f"{return_pct:.2f}")
            return_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 5, return_item)
            return_item = self.stock_table.item(row, 5)
            if return_pct > 0:
                return_item.setForeground(Qt.green)
            elif return_pct < 0:
                return_item.setForeground(Qt.red)
            else:
                return_item.setForeground(Qt.black)

    @pyqtSlot(str)
    def on_stock_added(self, code):
        if code not in [self.firstListBox.item(i).text() for i in range(self.firstListBox.count())]:
            self.firstListBox.addItem(code)

    @pyqtSlot(str)
    def on_stock_removed(self, code):
        if code in self.trader.vistock_set:
            self.trader.vistock_set.remove(code)
        if code in self.trader.monistock_set:
            self.trader.monistock_set.remove(code)
            self.trader.tickdata.monitor_stop(code)
            self.trader.mindata.monitor_stop(code)
            self.trader.daydata.monitor_stop(code)
            self.trader.delete_list_db(code)

        if code == self.chartdrawer.code:
            self.chartdrawer.set_code(None)

        for index in range(self.firstListBox.count()):
            item = self.firstListBox.item(index)
            if item and item.text() == code:
                self.firstListBox.takeItem(index)
                break
    
    @pyqtSlot(str)
    def on_stock_bought(self, code):
        if code not in [self.secondListBox.item(i).text() for i in range(self.secondListBox.count())]:
            self.secondListBox.addItem(code)

    @pyqtSlot(str)
    def on_stock_sold(self, code):
        for index in range(self.secondListBox.count()):
            item = self.secondListBox.item(index)
            if item and item.text() == code:
                self.secondListBox.takeItem(index)
                break

    @pyqtSlot(int)
    def update_counter_label(self, counter):
        self.counterlabel.setText(f"타이머: {counter}")

    def save_last_stg(self):
        if not self.login_handler.config.has_section('SETTINGS'):
            self.login_handler.config.add_section('SETTINGS')
        self.login_handler.config.set('SETTINGS', 'last_strategy', self.comboStg.currentText())
        with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
            self.login_handler.config.write(configfile)

    def save_buystrategy(self):
        try:
            investment_strategy = self.comboStg.currentText()
            buy_strategy = self.comboBuyStg.currentText()
            buy_key = self.comboBuyStg.currentData()
            strategy_content = self.buystgInputWidget.toPlainText()

            # 'STRATEGIES' 섹션이 없으면 추가
            if not self.login_handler.config.has_section('STRATEGIES'):
                self.login_handler.config.add_section('STRATEGIES')

            # 투자 전략 섹션이 없으면 추가
            if not self.login_handler.config.has_section(investment_strategy):
                self.login_handler.config.add_section(investment_strategy)

            # 매수 전략 내용을 JSON 형식으로 저장
            strategy_data = {
                'name': buy_strategy,
                'content': strategy_content
            }
            strategy_json = json.dumps(strategy_data, ensure_ascii=False)
            self.login_handler.config.set(investment_strategy, str(buy_key), strategy_json)

            with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                self.login_handler.config.write(configfile)

            existing_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == buy_strategy), None)
            if existing_strategy:
                existing_strategy['content'] = strategy_content
                logging.info(f"매수전략 '{buy_strategy}'이(가) 업데이트되었습니다.")

            QMessageBox.information(self, "수정 완료", f"매수전략 '{buy_strategy}'이 수정되었습니다.")

        except Exception as ex:
            logging.error(f"save_strategy -> {ex}")
            QMessageBox.critical(self, "수정 실패", f"전략 수정 중 오류가 발생했습니다:\n{str(ex)}")

    def save_sellstrategy(self):
        try:
            investment_strategy = self.comboStg.currentText()
            sell_strategy = self.comboSellStg.currentText()
            sell_key = self.comboSellStg.currentData()
            strategy_content = self.sellstgInputWidget.toPlainText()

            # 'STRATEGIES' 섹션이 없으면 추가
            if not self.login_handler.config.has_section('STRATEGIES'):
                self.login_handler.config.add_section('STRATEGIES')

            # 투자 전략 섹션이 없으면 추가
            if not self.login_handler.config.has_section(investment_strategy):
                self.login_handler.config.add_section(investment_strategy)

            # 매도 전략 내용을 JSON 형식으로 저장
            strategy_data = {
                'name': sell_strategy,
                'content': strategy_content
            }
            strategy_json = json.dumps(strategy_data, ensure_ascii=False)
            self.login_handler.config.set(investment_strategy, str(sell_key), strategy_json)

            with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                self.login_handler.config.write(configfile)

            existing_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == sell_strategy), None)
            if existing_strategy:
                existing_strategy['content'] = strategy_content
                logging.info(f"매도전략 '{sell_strategy}'이(가) 업데이트되었습니다.")

            QMessageBox.information(self, "수정 완료", f"매도전략 '{sell_strategy}'이 수정되었습니다.")

        except Exception as ex:
            logging.error(f"save_strategy -> {ex}")
            QMessageBox.critical(self, "수정 실패", f"전략 수정 중 오류가 발생했습니다:\n{str(ex)}")

    def load_strategy(self):
        try:
            self.dataStg = []
            self.data8537 = {}
            self.strategies = {}

            self.comboStg.clear()
            self.comboBuyStg.clear()
            self.buystgInputWidget.clear()

            # 투자 전략 섹션 불러오기
            if self.login_handler.config.has_section('STRATEGIES'):
                existing_stgnames = set(self.login_handler.config['STRATEGIES'].values())

            self.data8537 = self.objstg.requestList()
            for stgname, v in self.data8537.items():
                if stgname not in existing_stgnames:
                    existing_keys = self.login_handler.config['STRATEGIES'].keys()
                    existing_numbers = []
                    for k in existing_keys:
                        match = re.match(r'stg_?(\d+)', k)
                        if match:
                            existing_numbers.append(int(match.group(1)))
                    next_number = max(existing_numbers, default=0) + 1
                    new_key = f'stg_{next_number}'

                    self.login_handler.config.set('STRATEGIES', new_key, stgname)
                    existing_stgnames.add(stgname)
                    with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                        self.login_handler.config.write(configfile)

            # 모든 전략의 매수 매도 조건 로드
            for investment_strategy in existing_stgnames:
                if self.login_handler.config.has_section(investment_strategy):
                    self.strategies[investment_strategy] = []
                    for buy_key in sorted([k for k in self.login_handler.config[investment_strategy] if k.startswith('buy_stg_')], key=lambda x: int(x.split('_')[-1])):
                        buy_strategy = json.loads(self.login_handler.config.get(investment_strategy, buy_key))
                        buy_strategy['key'] = buy_key
                        self.strategies[investment_strategy].append(buy_strategy)

                    for sell_key in sorted([k for k in self.login_handler.config[investment_strategy] if k.startswith('sell_stg_')], key=lambda x: int(x.split('_')[-1])):
                        sell_strategy = json.loads(self.login_handler.config.get(investment_strategy, sell_key))
                        sell_strategy['key'] = sell_key
                        self.strategies[investment_strategy].append(sell_strategy)

            self.comboStg.blockSignals(True)
            for stgname in existing_stgnames:
                self.comboStg.addItem(stgname)
            
            last_strategy = self.login_handler.config.get('SETTINGS', 'last_strategy', fallback='VI 발동')
            index = self.comboStg.findText(last_strategy)
            if index != -1:
                self.comboStg.setCurrentIndex(index)
            self.comboStg.blockSignals(False)

            self.is_loading_strategy = True
            self.stgChanged()
            self.is_loading_strategy = False

        except Exception as ex:
            logging.error(f"load_strategy -> {ex}")

    def stgChanged(self, *args):
        stgName = self.comboStg.currentText()
        self.save_last_stg()

        if not self.is_loading_strategy:
            self.sell_all_item()
            self.trader.clear_list_db('mylist.db')
            
        if stgName == 'VI 발동':
            self.objstg.Clear()

            logging.info(f"전략 초기화")
            
            self.trader.init_stock_balance()
            self.trader.load_from_list_db('mylist.db')
            for code in list(self.trader.database_set):
                if code not in self.trader.monistock_set:
                    if self.trader.daydata.select_code(code) and self.trader.tickdata.monitor_code(code) and self.trader.mindata.monitor_code(code):
                        if code not in self.trader.starting_time:
                            self.trader.starting_time[code] = datetime.now().strftime('%m/%d 09:00:00')
                        self.trader.monistock_set.add(code)
                        self.firstListBox.addItem(code)
                        self.trader.save_vi_data(code)
                    else:
                        self.trader.daydata.monitor_stop(code)
                        self.trader.mindata.monitor_stop(code)
                        self.trader.tickdata.monitor_stop(code)
            
            if not hasattr(self, 'pb9619'):
                self.pb9619 = CpPB9619()  # 시장조치사항 실시간 구독 객체 생성
                self.pb9619.Subscribe("", self.trader)  # 구독 시작
            self.trader_thread.start_cnn_process()

        elif stgName == 'VI 발동 D1':
            if hasattr(self, 'pb9619'):
                self.pb9619.Unsubscribe()
            self.objstg.Clear()
            
            logging.info(f"전략 초기화")

            self.trader.init_stock_balance()
            self.trader.download_vi()
            self.trader_thread.start_cnn_process()

        else:
            if hasattr(self, 'pb9619'):
                self.pb9619.Unsubscribe()

            logging.info(f"전략 초기화")
            self.trader.init_stock_balance()
            self.trader.load_from_list_db('mylist.db')
            for code in list(self.trader.database_set):
                if code not in self.trader.monistock_set:
                    if self.trader.daydata.select_code(code) and self.trader.tickdata.monitor_code(code) and self.trader.mindata.monitor_code(code):
                        if code not in self.trader.starting_time:
                            self.trader.starting_time[code] = datetime.now().strftime('%m/%d 09:00:00')
                        self.trader.monistock_set.add(code)
                        self.firstListBox.addItem(code)
                        self.trader.save_vi_data(code)
                    else:
                        self.trader.daydata.monitor_stop(code)
                        self.trader.mindata.monitor_stop(code)
                        self.trader.tickdata.monitor_stop(code)
            
            item = self.data8537[stgName]
            id = item['ID']
            name = item['전략명']
            if name != '급등주':
                ret, self.dataStg = self.objstg.requestStgID(id) # 종목검색 조회
                if ret == False :
                    return
                for s in self.dataStg:
                    if self.trader.daydata.select_code(s['code']) and self.trader.tickdata.monitor_code(s['code']) and self.trader.mindata.monitor_code(s['code']):
                        if s['code'] not in self.trader.starting_time:
                            self.trader.starting_time[s['code']] = datetime.now().strftime('%m/%d 09:00:00')
                        self.trader.monistock_set.add(s['code'])
                        self.firstListBox.addItem(s['code'])
                    else:
                        self.trader.daydata.monitor_stop(s['code'])
                        self.trader.mindata.monitor_stop(s['code'])
                        self.trader.tickdata.monitor_stop(s['code'])

            ret, monid = self.objstg.requestMonitorID(id) # 전략의 감시 일련번호 요청
            if ret == False:
                return
            ret, status = self.objstg.requestStgControl(id, monid, True, stgName)  # 전략 감시 시작 요청
            if ret == False:
                return
            self.trader_thread.start_cnn_process()
            
        logging.info(f"{stgName} 전략 감시 시작")

        # 현재 투자 전략에 속한 매수 전략 로드
        self.comboBuyStg.clear()
        self.comboSellStg.clear()
        if stgName in self.strategies:
            for strategy in self.strategies[stgName]:
                strategy_name = strategy.get('name')
                strategy_key = strategy.get('key')
                if strategy_key.startswith('buy'):
                    self.comboBuyStg.addItem(strategy_name, strategy_key)
                elif strategy_key.startswith('sell'):
                    self.comboSellStg.addItem(strategy_name, strategy_key)

            if self.comboBuyStg.count() > 0:
                self.comboBuyStg.setCurrentIndex(0)
                self.buyStgChanged()

            if self.comboSellStg.count() > 0:
                self.comboSellStg.setCurrentIndex(0)
                self.sellStgChanged()

    def buyStgChanged(self):
        try:
            investment_strategy = self.comboStg.currentText()
            buy_strategy = self.comboBuyStg.currentText()

            if not investment_strategy or not buy_strategy:
                self.buystgInputWidget.clear()
                return

            if investment_strategy in self.strategies:
                selected_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == buy_strategy), None)
                if selected_strategy:
                    strategy_content = selected_strategy.get('content', '')
                    self.buystgInputWidget.setPlainText(strategy_content)
                else:
                    self.buystgInputWidget.clear()
            else:
                self.buystgInputWidget.clear()
        except Exception as ex:
            logging.error(f"buyStgChanged -> {ex}")

    def sellStgChanged(self):
        try:
            investment_strategy = self.comboStg.currentText()
            sell_strategy = self.comboSellStg.currentText()

            if not investment_strategy or not sell_strategy:
                self.sellstgInputWidget.clear()
                return

            if investment_strategy in self.strategies:
                selected_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == sell_strategy), None)
                if selected_strategy:
                    strategy_content = selected_strategy.get('content', '')
                    self.sellstgInputWidget.setPlainText(strategy_content)
                else:
                    self.sellstgInputWidget.clear()
            else:
                self.sellstgInputWidget.clear()
        except Exception as ex:
            logging.error(f"sellStgChanged -> {ex}")

    def close_external_popup(self):
        try:
            windows = gw.getWindowsWithTitle('공지사항')
            for window in windows:
                window.close()
        except Exception as ex:
            logging.error(f"close_external_popup -> {ex}")

    def listBoxChanged(self, current):
        if current:
            self.chartdrawer.set_code(current.text())
        else:
            self.chartdrawer.set_code(None)

    def delete_select_item(self):
        selected_items = self.firstListBox.selectedItems()
        if selected_items:
            for item in selected_items:
                self.firstListBox.takeItem(self.firstListBox.row(item))
                code = item.text()
                if code in self.trader.monistock_set:
                    self.trader.monistock_set.remove(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                self.trader.daydata.monitor_stop(code)
                self.trader.delete_list_db(code)

                if code == self.chartdrawer.code:
                    self.chartdrawer.set_code(None)
                logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 종목 삭제")

    def buy_item(self):
        selected_items = self.firstListBox.selectedItems()
        if selected_items:
            for item in selected_items:
                self.trader.buy_stock(item.text(), '기타', '0', '03')

    def sell_item(self):
        selected_items = self.secondListBox.selectedItems()
        if selected_items:
            for item in selected_items:
                self.trader.sell_stock(item.text(), '직접 매도')

    def sell_all_item(self):
        if self.trader.sell_all():
            for code in list(self.trader.monistock_set):
                self.on_stock_removed(code)

    def print_chart(self):
        # QPrinter 설정
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPrinter.A4)  # 용지 크기를 A4로 설정
        printer.setOrientation(QPrinter.Portrait)  # 세로 방향으로 설정

        # QPrintDialog를 통해 출력 설정
        printDialog = QPrintDialog(printer, self)
        if printDialog.exec_() == QPrintDialog.Accepted:
            painter = QPainter(printer)
            pageRect = printer.pageRect()
            canvasSize = self.canvas.size()
            xScale = pageRect.width() / canvasSize.width()
            yScale = pageRect.height() / canvasSize.height()
            scale = min(xScale, yScale)
            painter.scale(scale, scale)
            self.canvas.render(painter)
            painter.end()

    def output_current_data(self):
        try:
            # 현재 모니터링 중인 종목들의 데이터를 수집합니다.
            tick_data_all = self.trader.tickdata.stockdata
            min_data_all = self.trader.mindata.stockdata
            day_data_all = self.trader.daydata.stockdata

            if not tick_data_all and not min_data_all and not day_data_all:
                QMessageBox.warning(self, "데이터 없음", "현재 모니터링 중인 종목의 데이터가 없습니다.")
                return

            wb = Workbook()
            
            # Tick Data 시트 생성
            ws_tick = wb.active
            ws_tick.title = "Tick Data"
            for code, tick_data in tick_data_all.items():
                ws_tick.append([f"종목코드: {code}"])
                tick_keys = list(tick_data.keys())
                ws_tick.append(tick_keys)

                max_len_tick = max(len(v) for v in tick_data.values())

                for i in range(max_len_tick):
                    row = []
                    for key in tick_keys:
                        if i < len(tick_data[key]):
                            row.append(tick_data[key][i])
                        else:
                            row.append(None)  # 데이터 부족 시 None으로 패딩
                    ws_tick.append(row)
                ws_tick.append([])  # 종목 간 구분을 위해 빈 행 추가

            # Minute Data 시트 생성
            ws_min = wb.create_sheet(title="Minute Data")
            for code, min_data in min_data_all.items():
                ws_min.append([f"종목코드: {code}"])
                min_keys = list(min_data.keys())
                ws_min.append(min_keys)

                max_len_min = max(len(v) for v in min_data.values())

                for i in range(max_len_min):
                    row = []
                    for key in min_keys:
                        if i < len(min_data[key]):
                            row.append(min_data[key][i])
                        else:
                            row.append(None)
                    ws_min.append(row)
                ws_min.append([])

            # Day Data 시트 생성
            ws_day = wb.create_sheet(title="Day Data")
            for code, day_data in day_data_all.items():
                ws_day.append([f"종목코드: {code}"])
                day_keys = list(day_data.keys())
                ws_day.append(day_keys)

                max_len_day = max(len(v) for v in day_data.values())

                for i in range(max_len_day):
                    row = []
                    for key in day_keys:
                        if i < len(day_data[key]):
                            row.append(day_data[key][i])
                        else:
                            row.append(None)
                    ws_day.append(row)
                ws_day.append([])

            # 엑셀 파일 저장 경로 선택
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            filename, _ = QFileDialog.getSaveFileName(self, "Save Excel File",
                                                    f"stock_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                                    "Excel Files (*.xlsx);;All Files (*)", options=options)
            if not filename:
                QMessageBox.warning(self, "저장 취소", "파일 저장이 취소되었습니다.")
                return

            # 엑셀 파일 저장
            wb.save(filename)
            QMessageBox.information(self, "성공", f"데이터가 '{filename}'에 저장되었습니다.")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"데이터 저장 중 오류가 발생했습니다:\n{str(e)}")

    def print_terminal(self):
        printer = QPrinter()
        printDialog = QPrintDialog(printer, self)
        if printDialog.exec_() == QPrintDialog.Accepted:
            self.terminalOutput.print_(printer)

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Message', "Are you sure you want to quit?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.save_last_stg()

            # 데이터베이스 작업 쓰레드 종료 처리
            if hasattr(self.trader, 'db_worker'):
                self.trader.db_worker.stop()
                self.trader.db_thread.quit()
                self.trader.db_thread.wait()

            # 동작 중인 타이머 정지 (예상치 못한 콜백 방지)
            if getattr(self.trader, 'save_data_timer', None):
                self.trader.save_data_timer.stop()
            # if getattr(self.trader, 'monitor_vi_timer', None):
            #     self.trader.monitor_vi_timer.stop()
            if getattr(self, 'cnn_train_timer', None):
                self.cnn_train_timer.stop()
            if getattr(self, 'update_chart_status_timer', None):
                self.update_chart_status_timer.stop()
            if getattr(self.trader, 'tickdata', None):
                self.trader.tickdata.update_data_timer.stop()
            if getattr(self.trader, 'mindata', None):
                self.trader.mindata.update_data_timer.stop()
            if getattr(self.trader, 'daydata', None):
                self.trader.daydata.update_data_timer.stop()

            if self.chartdrawer.chart_thread:
                self.chartdrawer.chart_thread.stop()

            if self.trader_thread:
                self.trader_thread.stop()

            for code in list(self.trader.monistock_set):
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                self.trader.daydata.monitor_stop(code)
            self.trader.monistock_set.clear()

            self.objstg.Clear()

            self.clean_up_processes()

            # 핸들러 정리
            logger = logging.getLogger()
            for handler in logger.handlers[:]:
                handler.close()
                logger.removeHandler(handler)

            cpStatus.PlusDisconnect()

            event.accept()
        else:
            event.ignore()

    def clean_up_processes(self):
        process_names = ['coStarter', 'CpStart', 'DibServer']
        for process_name in process_names:
            for proc in psutil.process_iter(attrs=['pid', 'name']):
                if process_name.lower() in proc.info['name'].lower():
                    proc.kill()
                    logging.warning(f"Killed  {proc.info['name']} (PID: {proc.info['pid']})")

    def init_ui(self):
        self.setWindowTitle("초단타 매매 프로그램 v2.3")
        self.setGeometry(0, 0, 1900, 980)

        loginLayout = QVBoxLayout()

        loginLabelLayout = QHBoxLayout()
        loginLabel = QLabel("아이디:")
        loginLabelLayout.addWidget(loginLabel)
        self.loginEdit = QLineEdit()
        loginLabelLayout.addWidget(self.loginEdit)

        passwordLabelLayout = QHBoxLayout()
        passwordLabel = QLabel("비밀번호:")
        passwordLabelLayout.addWidget(passwordLabel)
        self.passwordEdit = QLineEdit()
        self.passwordEdit.setEchoMode(QLineEdit.Password)
        passwordLabelLayout.addWidget(self.passwordEdit)

        certpasswordLabelLayout = QHBoxLayout()
        certpasswordLabel = QLabel("인증번호:")
        certpasswordLabelLayout.addWidget(certpasswordLabel)
        self.certpasswordEdit = QLineEdit()
        self.certpasswordEdit.setEchoMode(QLineEdit.Password)
        certpasswordLabelLayout.addWidget(self.certpasswordEdit)

        label_width = 70  # 원하는 고정 길이 (픽셀 단위로 설정)
        loginLabel.setFixedWidth(label_width)
        passwordLabel.setFixedWidth(label_width)
        certpasswordLabel.setFixedWidth(label_width)

        loginButtonLayout = QHBoxLayout()
        self.autoLoginCheckBox = QCheckBox("자동 로그인")
        loginButtonLayout.addWidget(self.autoLoginCheckBox)
        self.loginButton = QPushButton("로그인")
        loginButtonLayout.addWidget(self.loginButton)

        loginLayout.addLayout(loginLabelLayout)
        loginLayout.addLayout(passwordLabelLayout)
        loginLayout.addLayout(certpasswordLabelLayout)
        loginLayout.addLayout(loginButtonLayout)

        buycountLayout = QHBoxLayout()
        buycountLabel = QLabel("최대투자 종목수 :")
        buycountLayout.addWidget(buycountLabel)
        self.buycountEdit = QLineEdit()
        buycountLayout.addWidget(self.buycountEdit)
        self.buycountButton = QPushButton("설정")
        self.buycountButton.setFixedWidth(70)
        buycountLayout.addWidget(self.buycountButton)

        firstListBoxLayout = QVBoxLayout()
        listBoxLabel = QLabel("투자 대상 종목 :")
        firstListBoxLayout.addWidget(listBoxLabel)
        self.firstListBox = QListWidget()
        firstListBoxLayout.addWidget(self.firstListBox, 1)
        firstButtonLayout = QHBoxLayout()
        self.buyButton = QPushButton("매입")
        firstButtonLayout.addWidget(self.buyButton)
        self.deleteFirstButton = QPushButton("삭제")        
        firstButtonLayout.addWidget(self.deleteFirstButton)        
        firstListBoxLayout.addLayout(firstButtonLayout)

        secondListBoxLayout = QVBoxLayout()
        secondListBoxLabel = QLabel("투자 종목 :")
        secondListBoxLayout.addWidget(secondListBoxLabel)
        self.secondListBox = QListWidget()        
        secondListBoxLayout.addWidget(self.secondListBox, 1)
        secondButtonLayout = QHBoxLayout()
        self.sellButton = QPushButton("매도")
        secondButtonLayout.addWidget(self.sellButton)
        self.sellAllButton = QPushButton("전부 매도")
        secondButtonLayout.addWidget(self.sellAllButton)     
        secondListBoxLayout.addLayout(secondButtonLayout)

        printLayout = QHBoxLayout()
        self.printChartButton = QPushButton("차트 출력")
        printLayout.addWidget(self.printChartButton)
        self.dataOutputButton2 = QPushButton("차트데이터 저장")
        printLayout.addWidget(self.dataOutputButton2)

        listBoxesLayout = QVBoxLayout()
        listBoxesLayout.addLayout(loginLayout)
        listBoxesLayout.addLayout(buycountLayout)
        listBoxesLayout.addLayout(firstListBoxLayout, 6)
        listBoxesLayout.addLayout(secondListBoxLayout, 4)
        listBoxesLayout.addLayout(printLayout)

        chartLayout = QVBoxLayout()
        self.fig = Figure(figsize=(12, 8))
        self.canvas = FigureCanvas(self.fig)
        chartLayout.addWidget(self.canvas)

        chartAndListLayout = QHBoxLayout()
        chartAndListLayout.addLayout(listBoxesLayout, 1)
        chartAndListLayout.addLayout(chartLayout, 4)

        strategyAndTradeLayout = QVBoxLayout()

        strategyLayout = QHBoxLayout()
        strategyLabel = QLabel("투자전략:")
        strategyLabel.setFixedWidth(label_width)
        strategyLayout.addWidget(strategyLabel, Qt.AlignLeft)
        self.comboStg = QComboBox()
        self.comboStg.setFixedWidth(200)
        strategyLayout.addWidget(self.comboStg, Qt.AlignLeft)
        strategyLayout.addStretch()
        self.counterlabel = QLabel('타이머: 0')
        self.counterlabel.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        strategyLayout.addWidget(self.counterlabel)
        self.chart_status_label = QLabel("Chart: None")
        self.chart_status_label.setStyleSheet("color: red")
        self.chart_status_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        strategyLayout.addWidget(self.chart_status_label)

        buyStrategyLayout = QHBoxLayout()
        buyStgLabel = QLabel("매수전략:")
        buyStgLabel.setFixedWidth(label_width)
        buyStrategyLayout.addWidget(buyStgLabel, alignment=Qt.AlignLeft)
        self.comboBuyStg = QComboBox()
        self.comboBuyStg.setFixedWidth(200)
        buyStrategyLayout.addWidget(self.comboBuyStg, alignment=Qt.AlignLeft)
        buyStrategyLayout.addStretch()
        self.saveBuyStgButton = QPushButton("수정")
        self.saveBuyStgButton.setFixedWidth(100)
        buyStrategyLayout.addWidget(self.saveBuyStgButton, alignment=Qt.AlignRight)
        self.buystgInputWidget = QTextEdit()
        self.buystgInputWidget.setPlaceholderText("매수전략의 내용을 입력하세요...")
        self.buystgInputWidget.setFixedHeight(80)

        sellStrategyLayout = QHBoxLayout()
        sellStgLabel = QLabel("매도전략:")
        sellStgLabel.setFixedWidth(label_width)
        sellStrategyLayout.addWidget(sellStgLabel, alignment=Qt.AlignLeft)
        self.comboSellStg = QComboBox()
        self.comboSellStg.setFixedWidth(200)
        sellStrategyLayout.addWidget(self.comboSellStg, alignment=Qt.AlignLeft)
        sellStrategyLayout.addStretch()
        self.saveSellStgButton = QPushButton("수정")
        self.saveSellStgButton.setFixedWidth(100)
        sellStrategyLayout.addWidget(self.saveSellStgButton, alignment=Qt.AlignRight)
        self.sellstgInputWidget = QTextEdit()
        self.sellstgInputWidget.setPlaceholderText("매도전략의 내용을 입력하세요...")
        self.sellstgInputWidget.setFixedHeight(63)

        self.stock_table = QTableWidget()
        self.stock_table.setRowCount(0)
        self.stock_table.setColumnCount(6)
        self.stock_table.setHorizontalHeaderLabels(["종목코드", "현재가", "상승확률(%)", "매수가", "평가손익", "수익률(%)"])
        self.stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)  # 편집 비활성화
        self.stock_table.setFixedHeight(220)
        self.stock_table.verticalHeader().setDefaultSectionSize(20)

        strategyAndTradeLayout.addLayout(strategyLayout)
        strategyAndTradeLayout.addLayout(buyStrategyLayout)
        strategyAndTradeLayout.addWidget(self.buystgInputWidget)
        strategyAndTradeLayout.addLayout(sellStrategyLayout)
        strategyAndTradeLayout.addWidget(self.sellstgInputWidget)
        strategyAndTradeLayout.addWidget(self.stock_table)

        self.terminalOutput = QTextEdit()
        self.terminalOutput.setReadOnly(True)

        counterAndterminalLayout = QVBoxLayout()
        
        counterAndterminalLayout.addLayout(strategyAndTradeLayout)
        counterAndterminalLayout.addWidget(self.terminalOutput)

        mainLayout = QHBoxLayout()
        mainLayout.addLayout(chartAndListLayout, 70)
        mainLayout.addLayout(counterAndterminalLayout, 30)
        self.setLayout(mainLayout)

        self.loginButton.clicked.connect(self.login_handler.handle_login)
        self.buycountButton.clicked.connect(self.login_handler.buycount_setting)

        self.buyButton.clicked.connect(self.buy_item)
        self.deleteFirstButton.clicked.connect(self.delete_select_item)
        self.sellButton.clicked.connect(self.sell_item)
        self.sellAllButton.clicked.connect(self.sell_all_item)

        self.firstListBox.currentItemChanged.connect(self.listBoxChanged)
        self.firstListBox.itemClicked.connect(self.listBoxChanged)
        self.secondListBox.currentItemChanged.connect(self.listBoxChanged)
        self.secondListBox.itemClicked.connect(self.listBoxChanged)

        self.printChartButton.clicked.connect(self.print_chart)
        self.dataOutputButton2.clicked.connect(self.output_current_data)

        self.comboStg.currentIndexChanged.connect(self.stgChanged)
        self.comboBuyStg.currentIndexChanged.connect(self.buyStgChanged)
        self.comboSellStg.currentIndexChanged.connect(self.sellStgChanged)
        self.saveBuyStgButton.clicked.connect(self.save_buystrategy)
        self.saveSellStgButton.clicked.connect(self.save_sellstrategy)

class QTextEditLogger(QObject, logging.Handler):
    log_signal = pyqtSignal(str)

    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit
        self.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
        self.log_signal.connect(self.append_log)

    @pyqtSlot(str)
    def append_log(self, msg):
        self.text_edit.append(msg)

    def emit(self, record):
        msg = self.format(record)
        # 특정 키워드에 대해 색상 변경
        if '매매이익' in msg:
            msg = f"<span style='color:green;'>{msg}</span>"
        elif '매매손실' in msg:
            msg = f"<span style='color:red;'>{msg}</span>"
        elif '매매실현손익' in msg:
            msg = f"<span style='font-weight:bold;'>{msg}</span>"
        else:
            msg = f"<span>{msg}</span>"

        self.log_signal.emit(msg)

if __name__ == "__main__":
    setup_logging()

    app = QApplication(sys.argv)
    app.setFont(QFont("Malgun Gothic", 9))
    myWindow = MyWindow()
    myWindow.setWindowIcon(QIcon('stock_trader.ico'))
    myWindow.showMaximized()
    app.exec_()