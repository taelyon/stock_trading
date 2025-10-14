
import sys
import ctypes
import matplotlib
matplotlib.use('Qt5Agg')
from PyQt5.QtWidgets import *
from PyQt5.QtCore import (
    QTimer, pyqtSignal, QProcess, QObject, QThread, Qt, 
    pyqtSlot, QRunnable, QThreadPool, QEventLoop
)
from PyQt5.QtGui import QIcon, QPainter, QFont, QColor
from PyQt5.QtPrintSupport import QPrintDialog, QPrinter
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from datetime import datetime, timedelta
import pandas as pd
# Windows 전용 의존성은 사용할 수 없는 환경이 있을 수 있으므로 지연 확인
import win32com.client
import time
import numpy as np
import re
import mplfinance as mpf
import sqlite3
import os
import psutil
import configparser
import pygetwindow as gw
import requests
import warnings
import logging
from openpyxl import Workbook
import json
from slacker import Slacker
import threading
import queue
import copy
import talib
import traceback
import pyautogui
from collections import deque

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

IS_WINDOWS = sys.platform.startswith('win')


def _prevent_system_sleep():
    """Windows 환경에서만 동작하는 절전 모드 해제 처리"""
    if not IS_WINDOWS or not hasattr(ctypes, "windll"):
        return

    try:
        ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001)
    except Exception as ex:
        logging.warning(f"시스템 절전 방지 설정 실패: {ex}")


_prevent_system_sleep()

plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

# PLUS 공통 OBJECT (전역 변수 선언만, 실제 초기화는 init_plus_objects()에서)
cpCodeMgr = None
cpStatus = None
cpCybos = None  # API 제한 상태 확인용
cpTrade = None
cpBalance = None
cpCash = None
cpOrder = None
cpStock = None

def init_plus_objects():
    """크레온 PLUS COM 객체 초기화 (지연 초기화)"""
    global cpCodeMgr, cpStatus, cpCybos, cpTrade, cpBalance, cpCash, cpOrder, cpStock
    
    try:
        cpCodeMgr = win32com.client.Dispatch('CpUtil.CpCodeMgr')
        cpStatus = win32com.client.Dispatch('CpUtil.CpCybos')
        cpCybos = cpStatus  # API 제한 상태 확인용 (동일한 객체)
        cpTrade = win32com.client.Dispatch('CpTrade.CpTdUtil')
        cpBalance = win32com.client.Dispatch('CpTrade.CpTd6033')
        cpCash = win32com.client.Dispatch('CpTrade.CpTdNew5331A')
        cpOrder = win32com.client.Dispatch('CpTrade.CpTd0311')
        cpStock = win32com.client.Dispatch('DsCbo1.StockMst')
        return True
    except Exception as ex:
        logging.error(f"크레온 PLUS COM 객체 초기화 실패: {ex}")
        return False

def init_plus_check():
    """크레온 PLUS 연결 및 권한 확인"""
    if not IS_WINDOWS or not hasattr(ctypes, "windll"):
        logging.error("크레온 PLUS 기능은 Windows 환경에서만 사용할 수 있습니다.")
        return False

    # 관리자 권한 체크
    if not ctypes.windll.shell32.IsUserAnAdmin():
        logging.error(f"오류: 일반권한으로 실행됨. 관리자 권한으로 실행해 주세요")
        return False
    
    # COM 객체 초기화
    if not init_plus_objects():
        return False
    
    # 연결 체크
    if (cpStatus.IsConnect == 0):
        logging.error(f"PLUS가 정상적으로 연결되지 않음")
        return False
    
    # 거래 초기화
    if (cpTrade.TradeInit(0) != 0):
        logging.error(f"주문 초기화 실패")
        return False
    
    return True

def setup_logging():
    """로그 설정 (PyInstaller 대응)"""
    try:
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)
        logging.getLogger('matplotlib').setLevel(logging.WARNING)

        # ✅ 실행 파일 경로 확인
        if getattr(sys, 'frozen', False):
            # PyInstaller로 빌드된 경우
            application_path = os.path.dirname(sys.executable)
        else:
            # 일반 Python 실행
            application_path = os.path.dirname(os.path.abspath(__file__))

        # ✅ 로그 디렉토리 생성 (안전하게)
        log_dir = os.path.join(application_path, 'log')
        try:
            if not os.path.exists(log_dir):
                os.makedirs(log_dir)
        except Exception as e:
            # 로그 폴더 생성 실패 시 임시 폴더 사용
            log_dir = os.path.join(os.environ.get('TEMP', 'C:\\Temp'), 'stock_trader_log')
            os.makedirs(log_dir, exist_ok=True)

        # 로그 파일 경로
        log_path = os.path.join(log_dir, f"trading_{datetime.now().strftime('%Y%m%d')}.log")
        
        # 파일 핸들러 (버퍼링 비활성화)
        file_handler = logging.FileHandler(log_path, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(file_formatter)
        # 버퍼링 비활성화로 즉시 플러시
        file_handler.stream.flush = lambda: None
        logger.addHandler(file_handler)

        # 콘솔 핸들러 (INFO 레벨로 변경하여 더 많은 로그 표시)
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)  # WARNING → INFO로 변경
        console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_handler.setFormatter(console_formatter)
        # 버퍼링 비활성화
        console_handler.stream.flush = lambda: None
        logger.addHandler(console_handler)
        
        logging.info(f"로그 초기화 완료: {log_path}")
        
        # 로깅 즉시 플러시 함수 등록
        def flush_logs():
            """모든 로그 핸들러 즉시 플러시"""
            for handler in logger.handlers:
                if hasattr(handler, 'stream'):
                    handler.stream.flush()
        
        # 전역 함수로 등록
        import builtins
        builtins.flush_logs = flush_logs
        
    except Exception as ex:
        # 로그 설정 실패 시에도 프로그램은 계속 실행
        print(f"로그 설정 오류: {ex}")
        traceback.print_exc()

def send_slack_message(login_handler, channel, message):
    if login_handler is None or login_handler.slack is None:
        logging.warning("Slack 설정이 되어 있지 않습니다.")
        return
    sender = SlackMessageSender(login_handler, channel, message)
    QThreadPool.globalInstance().start(sender)

class SlackMessageSender(QRunnable):
    def __init__(self, login_handler, channel, message):
        super().__init__()
        self.login_handler = login_handler
        self.channel = channel
        self.message = message

    def run(self):
        try:
            if self.login_handler.slack:
                self.login_handler.slack.chat.post_message(self.channel, self.message)
        except Exception as ex:
            logging.error(f"Slack 메시지 전송 실패: {ex}")

# ==================== 데이터 캐시 ====================
class DataCache:
    """종목 정보 캐싱"""
    
    def __init__(self, expire_seconds=300):
        self.cache = {}
        self.expire_seconds = expire_seconds
        self.lock = threading.Lock()
    
    def get(self, key):
        with self.lock:
            if key in self.cache:
                data, timestamp = self.cache[key]
                if time.time() - timestamp < self.expire_seconds:
                    return data
                else:
                    del self.cache[key]
            return None
    
    def set(self, key, value):
        with self.lock:
            self.cache[key] = (value, time.time())
    
    def clear(self):
        with self.lock:
            self.cache.clear()

# 전역 캐시
stock_info_cache = DataCache(expire_seconds=300)  # 5분

# ==================== 영업일 찾기 유틸리티 ==================== ✅ 여기에 추가
# 전역 영업일 캐시
_global_trading_date = None
_global_trading_date_lock = threading.Lock()

def get_last_trading_date(target_date=None, max_attempts=10):
    """가장 최근 영업일 찾기 (삼성전자 기준) - 전역 캐시 사용
    
    Args:
        target_date: 검색 시작일 (datetime 또는 YYYYMMDD 정수)
        max_attempts: 최대 시도 횟수 (기본 10일)
    
    Returns:
        (success, trading_date): (성공 여부, 영업일 YYYYMMDD 정수)
    """
    global _global_trading_date
    
    # ✅ 전역 캐시 확인 (API 호출 방지)
    with _global_trading_date_lock:
        if _global_trading_date is not None:
            logging.info(f"✅ 캐시된 영업일 사용: {_global_trading_date}")
            return (True, _global_trading_date)
    
    try:
        if target_date is None:
            check_date = datetime.now()
        elif isinstance(target_date, int):
            date_str = str(target_date)
            check_date = datetime.strptime(date_str, '%Y%m%d')
        else:
            check_date = target_date
        
        objRq = win32com.client.Dispatch("CpSysDib.StockChart")
        attempts = 0
        
        for i in range(max_attempts * 2):  # 주말 포함
            if attempts >= max_attempts:
                break
            
            test_date = check_date - timedelta(days=i)
            
            # 주말 스킵
            if test_date.weekday() >= 5:
                continue
            
            attempts += 1
            test_date_int = test_date.year * 10000 + test_date.month * 100 + test_date.day
                      
            # 삼성전자로 영업일 확인
            objRq.SetInputValue(0, 'A005930')
            objRq.SetInputValue(1, ord('1'))
            objRq.SetInputValue(2, test_date_int)
            objRq.SetInputValue(3, test_date_int)
            objRq.SetInputValue(5, [0])  # 날짜만
            objRq.SetInputValue(6, ord('D'))
            objRq.SetInputValue(9, ord('1'))
            objRq.BlockRequest2(1)
            
            rqStatus = objRq.GetDibStatus()
            if rqStatus != 0:
                continue
            
            len_data = objRq.GetHeaderValue(3)
            
            if len_data > 0:
                actual_date = objRq.GetDataValue(0, 0)
                
                # ✅ 전역 캐시에 저장
                with _global_trading_date_lock:
                    _global_trading_date = actual_date
                
                return (True, actual_date)
        
        logging.warning(f"최근 {max_attempts}일 내 영업일을 찾지 못했습니다")
        return (False, None)
        
    except Exception as ex:
        logging.error(f"get_last_trading_date: {ex}")
        return (False, None)
    
# ==================== 급등주 스캐너 (검증용으로만 사용) ====================
class MomentumScanner(QObject):
    """급등주 검증 - 조건검색 편입 종목 재확인용"""
    
    stock_found = pyqtSignal(dict)
    
    def __init__(self, trader):
        super().__init__()
        self.trader = trader
        self.cpStock = win32com.client.Dispatch('DsCbo1.StockMst')
    
    def verify_momentum_conditions(self, code):
        """급등주 조건 재확인 (메모리 데이터 기반)
        
        조건검색으로 들어온 종목이 실제로 급등주 조건을 만족하는지 검증
        
        Returns:
            (is_valid, score, message): (검증 통과 여부, 점수, 메시지)
        """
        try:
            stock_name = cpCodeMgr.CodeToName(code)
            logging.debug(f"🔍 [MomentumScanner] {stock_name}({code}): 검증 시작")
            
            # 캐시 확인
            cached_score = stock_info_cache.get(f"score_{code}")
            if cached_score is not None:
                logging.debug(f"🔍 [MomentumScanner] {stock_name}({code}): 캐시에서 조회 (점수: {cached_score})")
                return (cached_score >= 70, cached_score, "캐시에서 조회")
            
            # ===== ✅ 메모리 데이터 기반 검증 (stockdata 직접 조회) =====
            with self.trader.tickdata.stockdata_lock:
                tick_data = self.trader.tickdata.stockdata.get(code, {})
            
            with self.trader.mindata.stockdata_lock:
                min_data = self.trader.mindata.stockdata.get(code, {})
            
            with self.trader.daydata.stockdata_lock:
                day_data = self.trader.daydata.stockdata.get(code, {})
            
            logging.debug(f"🔍 [MomentumScanner] {stock_name}({code}): 데이터 확인 - 틱:{len(tick_data.get('C', []))}개, 분:{len(min_data.get('C', []))}개, 일:{len(day_data.get('C', []))}개")
            
            # ✅ 최소 데이터 확인 (일봉 데이터만 있어도 진행)
            if not day_data or len(day_data.get('C', [])) < 2:
                logging.warning(f"🔍 [MomentumScanner] {stock_name}({code}): 일봉 데이터 부족 - 일봉:{len(day_data.get('C', [])) if day_data else 0}개")
                return (False, 0, "일봉 데이터 부족")
            
            # 틱/분 데이터가 없어도 경고만 출력하고 계속 진행
            if not tick_data or len(tick_data.get('C', [])) == 0:
                logging.warning(f"⚠️ [MomentumScanner] {stock_name}({code}): 틱 데이터 없음, 일봉 데이터로만 검증")
                # 틱 데이터가 없을 때 기본값 설정
                tick_data = {'C': [0], 'O': [0], 'H': [0], 'L': [0], 'V': [0]}
            if not min_data or len(min_data.get('C', [])) == 0:
                logging.warning(f"⚠️ [MomentumScanner] {stock_name}({code}): 분봉 데이터 없음, 일봉 데이터로만 검증")
                # 분봉 데이터가 없을 때 기본값 설정
                min_data = {'C': [0], 'O': [0], 'H': [0], 'L': [0], 'V': [0]}
            
            # ===== 데이터 추출 (메모리에서) =====
            try:
                # ✅ 현재가 추출 (틱 데이터 우선, 없으면 일봉 데이터 사용)
                if tick_data.get('C') and len(tick_data.get('C', [])) > 0:
                    current_price = tick_data.get('C', [0])[-1]
                elif day_data.get('C') and len(day_data.get('C', [])) > 0:
                    current_price = day_data.get('C', [0])[-1]  # 일봉 데이터에서 현재가 사용
                    logging.info(f"🔍 {stock_name}({code}): 틱 데이터 없음, 일봉 데이터에서 현재가 사용: {current_price}")
                else:
                    current_price = 0
                
                # 일봉 데이터에서 당일 정보
                open_price = day_data.get('O', [0])[-1] if day_data.get('O') else 0
                high_price = day_data.get('H', [0])[-1] if day_data.get('H') else 0
                low_price = day_data.get('L', [0])[-1] if day_data.get('L') else 0
                volume = day_data.get('V', [0])[-1] if day_data.get('V') else 0
                
                # 전일 데이터
                prev_close = day_data.get('C', [0])[-2] if len(day_data.get('C', [])) >= 2 else 0
                prev_volume = day_data.get('V', [0])[-2] if len(day_data.get('V', [])) >= 2 else 0
                
                market_cap = 0  # 일봉 데이터에는 시가총액 정보가 없음
                
                if current_price == 0 or open_price == 0:
                    return (False, 0, "가격 데이터 없음")
                
                # ✅ 거래량 데이터 유효성 확인
                if volume <= 0:
                    logging.warning(f"🔍 {code}: 당일 거래량이 0 또는 음수 ({volume})")
                    return (False, 0, f"당일 거래량 오류 ({volume})")
                
                if prev_volume < 0:
                    logging.warning(f"🔍 {code}: 전일 거래량이 음수 ({prev_volume})")
                    prev_volume = 0  # 음수는 0으로 처리
                    
            except Exception as ex:
                logging.error(f"{code}: 메모리 데이터 추출 실패: {ex}")
                return (False, 0, "메모리 데이터 추출 실패")
            
            # ===== 1차 필터링 =====
            if current_price < 2000 or current_price > 50000:
                return (False, 0, f"가격대 미달 ({current_price}원)")
            
            # 시가총액 체크 (일봉 데이터에는 시가총액 정보가 없으므로 0일 때는 스킵)
            # if market_cap > 0 and (market_cap < 50000 or market_cap > 500000):
            #     return (False, 0, f"시가총액 미달 ({market_cap/10000:.0f}억)")
            
            score = 0
            
            # ✅ 현재 시간 변수 정의 (다른 지표에서 사용)
            now = datetime.now()
            
            # ===== 1. 시가 대비 상승률 (장초반 가중치 증가) =====
            if open_price > 0:
                price_change_pct = (current_price - open_price) / open_price * 100
                
                # 장초반(9-10시)에는 상승률에 더 높은 가중치 부여
                max_price_score = 30 if 9 <= now.hour < 10 else 30
                
                if 2.0 <= price_change_pct < 3.5:
                    score += max_price_score
                elif 3.5 <= price_change_pct < 5.0:
                    score += max_price_score * 0.7  # 21점
                elif 5.0 <= price_change_pct < 7.0:
                    score += max_price_score * 0.4  # 12점
                elif price_change_pct < 0:
                    return (False, 0, "시가 대비 하락")
            
            # ===== 2. 거래량 비율 (0-25점) =====
            # ✅ 개선: 장초반에는 거래량 검증 스킵, 다른 지표로 대체
            volume_score = 0
            
            # 장초반(9-10시)에는 거래량 검증 스킵하고 기본 점수 부여
            if 9 <= now.hour < 10:
                logging.info(f"🔍 {code}: 장초반 거래량 검증 스킵 (시간: {now.hour:02d}:{now.minute:02d})")
                volume_score = 15  # 장초반 기본 점수
            else:
                # 10시 이후부터는 정상적인 거래량 검증
                if prev_volume > 0:
                    volume_ratio = volume / prev_volume
                    
                    # 시간대별 보정 (10시 이후)
                    time_factor = 1.0
                    if 10 <= now.hour < 11:  # 10-11시
                        time_factor = 0.6  # 60%만 거래되어도 정상
                    elif 11 <= now.hour < 12:  # 11-12시
                        time_factor = 0.8  # 80%만 거래되어도 정상
                    
                    # 보정된 거래량 비율
                    adjusted_ratio = volume_ratio / time_factor
                    
                    if adjusted_ratio >= 5.0:
                        volume_score = 25
                    elif adjusted_ratio >= 3.0:
                        volume_score = 20
                    elif adjusted_ratio >= 2.0:
                        volume_score = 15
                    elif adjusted_ratio >= 1.0:
                        volume_score = 10
                    elif adjusted_ratio >= 0.7:  # 70% 이상이면 통과
                        volume_score = 5
                    else:
                        # 거래량 부족이지만 탈락하지는 않음 (점수만 낮게)
                        logging.warning(f"🔍 {code}: 거래량 부족 ({adjusted_ratio:.1f}배, 원래:{volume_ratio:.1f}배)")
                        volume_score = 0
                else:
                    # 전일 거래량이 없는 경우
                    if volume > 0:
                        logging.info(f"🔍 {code}: 전일 거래량 없음, 당일 거래량으로 판단 ({volume:,}주)")
                        volume_score = 15  # 기본 점수
                    else:
                        logging.warning(f"🔍 {code}: 거래량 데이터 없음")
                        volume_score = 0
            
            score += volume_score
            
            # ===== 3. 당일 고가 근처 유지 (0-20점) =====
            if high_price > 0 and low_price > 0:
                if (high_price - low_price) > 0:
                    position = (current_price - low_price) / (high_price - low_price)
                else:
                    position = 0
                
                if position >= 0.8:
                    score += 20
                elif position >= 0.6:
                    score += 15
                elif position >= 0.4:
                    score += 10
            
            # ===== 4. 시가 상승 유지 (0-15점) =====
            if current_price > open_price * 1.015:
                score += 15
            elif current_price > open_price:
                score += 10
            
            # ===== 5. 시간대 가중치 (장초반 가중치 증가) =====
            if 9 <= now.hour < 10:
                score += 15  # 장초반 가중치 증가 (10→15점)
            elif 10 <= now.hour < 12:
                score += 10  # 오전 가중치 증가 (7→10점)
            elif 13 <= now.hour < 14:
                score += 7   # 오후 가중치 증가 (5→7점)
            
            # ===== 6. 장초반 급등주 보너스 (0-10점) =====
            if 9 <= now.hour < 10 and price_change_pct >= 2.0:
                # 장초반에 2% 이상 상승한 종목에 보너스 점수
                if price_change_pct >= 5.0:
                    score += 10  # 5% 이상 급등
                elif price_change_pct >= 3.5:
                    score += 7   # 3.5% 이상 급등
                elif price_change_pct >= 2.0:
                    score += 5   # 2% 이상 상승
            
            # 캐시 저장
            stock_info_cache.set(f"score_{code}", score)
            
            is_valid = score >= 70
            message = f"급등주 점수: {score}/100"
            
            # ✅ 장초반 급등주 검증 로깅 개선
            time_info = f"시간: {now.hour:02d}:{now.minute:02d}"
            volume_info = f"거래량: {volume:,}주"
            if prev_volume > 0 and 9 < now.hour:  # 10시 이후만 거래량 비율 표시
                volume_ratio = volume / prev_volume
                volume_info += f" (전일:{prev_volume:,}주, 비율:{volume_ratio:.1f}배)"
            else:
                volume_info += " (거래량 검증 스킵)" if 9 <= now.hour < 10 else " (전일거래량 없음)"
            
            logging.info(f"🚀 [급등주검증] {stock_name}({code}): {time_info}, 상승률:{price_change_pct:.1f}%, 점수:{score}/100, 유효:{is_valid}, {volume_info}")
            return (is_valid, score, message)
            
        except Exception as ex:
            logging.error(f"verify_momentum_conditions({code}): {ex}\n{traceback.format_exc()}")
            return (False, 0, f"검증 오류: {ex}")
        
# ==================== 갭 상승 스캐너 (검증 + 매수조건) ====================
class GapUpScanner:
    """갭 상승 스캐너 - 검증 + 매수 조건 체크"""
    
    def __init__(self, trader):
        self.trader = trader
        self.cpStock = win32com.client.Dispatch('DsCbo1.StockMst')

        # ===== ✅ 캐시 및 안전장치 추가 =====
        self.verification_cache = {}
        self.last_verification_time = {}
    
    def verify_gap_conditions(self, code):
        """갭상승 조건 재확인 (메모리 데이터 기반)
        
        Returns:
            (is_valid, gap_pct, message): (검증 통과 여부, 갭 비율, 메시지)
        """
        try:
            stock_name = cpCodeMgr.CodeToName(code)
            
            # ===== ✅ 캐시 확인 (5초 이내 재검증 방지) =====
            now = time.time()
            if code in self.last_verification_time:
                if now - self.last_verification_time[code] < 5.0:
                    cached = self.verification_cache.get(code)
                    if cached:
                        logging.debug(f"{code}: 캐시에서 조회 (5초 이내)")
                        return cached
            
            # ===== ✅ 메모리 데이터 기반 검증 (stockdata 직접 조회) =====
            with self.trader.tickdata.stockdata_lock:
                tick_data = self.trader.tickdata.stockdata.get(code, {})
            
            with self.trader.daydata.stockdata_lock:
                day_data = self.trader.daydata.stockdata.get(code, {})
            
            # 최소 데이터 확인
            if not tick_data or not day_data:
                return (False, 0, "메모리 데이터 없음")
            
            if len(tick_data.get('C', [])) == 0 or len(day_data.get('C', [])) < 2:
                return (False, 0, f"데이터 부족 (틱:{len(tick_data.get('C', []))}, 일:{len(day_data.get('C', []))})")
            
            # ===== ✅ 데이터 안전하게 추출 (메모리에서) =====
            try:
                current_price = tick_data.get('C', [0])[-1] if tick_data.get('C') else 0  # 현재가 (최신 값)
                open_price = day_data.get('O', [0])[-1] if day_data.get('O') else 0  # 시가
                prev_close = day_data.get('C', [0])[-2] if len(day_data.get('C', [])) >= 2 else 0  # 전일종가
                
                # 데이터 유효성 검증
                if not all([current_price, open_price, prev_close]):
                    logging.warning(f"{stock_name}({code}): 가격 데이터 없음")
                    return (False, 0, "가격 데이터 없음")
                
                if prev_close <= 0:
                    logging.warning(f"{stock_name}({code}): 전일종가 0 이하")
                    return (False, 0, "전일종가 오류")
                
            except Exception as ex:
                logging.error(f"{stock_name}({code}): 데이터 추출 실패: {ex}")
                return (False, 0, "데이터 추출 실패")
            
            # ===== ✅ 갭 비율 계산 =====
            gap_pct = (open_price - prev_close) / prev_close * 100
            
            # ===== ✅ 갭상승 조건 검증 =====
            is_valid = False
            message = ""
            
            # 조건 1: 시가가 전일종가 대비 2% 이상 상승
            if gap_pct >= 2.0:
                # 조건 2: 현재가가 시가 대비 하락하지 않음
                if current_price >= open_price * 0.98:  # 2% 이내 허용
                    is_valid = True
                    message = f"갭상승 {gap_pct:.2f}%, 시가 유지"
                else:
                    message = f"갭상승 {gap_pct:.2f}%이나 시가 대비 하락"
            else:
                message = f"갭 비율 부족 ({gap_pct:.2f}%)"
            
            # ===== ✅ 결과 캐시 저장 =====
            result = (is_valid, gap_pct, message)
            self.verification_cache[code] = result
            self.last_verification_time[code] = now
            
            return result
            
        except Exception as ex:
            logging.error(f"verify_gap_conditions({code}): {ex}\n{traceback.format_exc()}")
            return (False, 0, f"검증 오류: {ex}")
        
    def check_gap_hold(self, code):
        """갭 유지 확인 (매수 조건)
        
        매수 시점에 갭이 여전히 유지되고 있는지 확인
        시가 대비 -0.3% 이내면 갭 유지로 판단
        """
        try:
            day_data = self.trader.daydata.stockdata.get(code, {})
            
            if len(day_data.get('O', [])) == 0:
                return False
            
            today_open = day_data.get('O', [0])[-1] if day_data.get('O') else 0
            
            # 현재가
            tick_latest = self.trader.tickdata.get_latest_data(code)
            current_price = tick_latest.get('C', 0)
            
            if current_price == 0:
                return False
            
            # 시가 대비 -0.3% 이내 (갭 유지)
            if current_price >= today_open * 0.997:
                return True
            
            return False
            
        except Exception as ex:
            logging.error(f"check_gap_hold({code}): {ex}")
            return False
        
# ==================== 변동성 돌파 전략 ====================
class VolatilityBreakout:
    """변동성 돌파 전략"""
    
    def __init__(self, trader):
        self.trader = trader
        self.K_value = 0.5
        self.target_prices = {}
        self.breakout_checked = set()
    
    def calculate_target_price(self, code):
        """목표가 계산"""
        
        try:
            # 일봉 데이터
            day_data = self.trader.daydata.stockdata.get(code, {})
            
            if len(day_data.get('H', [])) < 2:
                return None
            
            # 전일 고가/저가
            prev_high = day_data.get('H', [0])[-2] if len(day_data.get('H', [])) >= 2 else 0
            prev_low = day_data.get('L', [0])[-2] if len(day_data.get('L', [])) >= 2 else 0
            
            # 당일 시가
            today_open = day_data.get('O', [0])[-1] if day_data.get('O') else 0
            
            # 변동폭
            range_value = prev_high - prev_low
            
            # 목표가
            target = today_open + (range_value * self.K_value)
            
            self.target_prices[code] = target
            
            return target
            
        except Exception as ex:
            logging.error(f"calculate_target_price({code}): {ex}")
            return None
    
    def check_breakout(self, code):
        """돌파 확인"""
        
        try:
            # 이미 체크했으면 스킵
            if code in self.breakout_checked:
                return False
            
            # 목표가 계산
            if code not in self.target_prices:
                target = self.calculate_target_price(code)
                if not target:
                    return False
            else:
                target = self.target_prices[code]
            
            # 현재가
            tick_latest = self.trader.tickdata.get_latest_data(code)
            current_price = tick_latest.get('C', 0)
            
            if current_price == 0:
                return False
            
            # 돌파 확인
            if current_price >= target:
                # 거래량 확인
                volume_ratio = self._get_volume_ratio(code)
                
                if volume_ratio >= 1.5:
                    self.breakout_checked.add(code)
                    
                    logging.info(
                        f"{cpCodeMgr.CodeToName(code)}({code}): "
                        f"변동성 돌파 (목표: {target:.0f}, "
                        f"현재: {current_price:.0f}, "
                        f"거래량비: {volume_ratio:.1f}배)"
                    )
                    
                    return True
            
            return False
            
        except Exception as ex:
            logging.error(f"check_breakout({code}): {ex}")
            return False
    
    def _get_volume_ratio(self, code):
        """거래량 비율"""
        try:
            day_data = self.trader.daydata.stockdata.get(code, {})
            
            if len(day_data.get('V', [])) < 2:
                return 0
            
            today_vol = day_data.get('V', [0])[-1] if day_data.get('V') else 0
            prev_vol = day_data.get('V', [0])[-2] if len(day_data.get('V', [])) >= 2 else 0
            
            if prev_vol > 0:
                return today_vol / prev_vol
            return 0
        except:
            return 0

# ==================== 기존 CpEvent 클래스 (유지) ====================
class CpEvent:
    def __init__(self):
        self.last_update_time = None

    def set_params(self, client, name, caller):
        self.client = client
        self.name = name
        self.caller = caller
        self.dic = {ord('1'): "종목별 VI", ord('2'): "배분정보", ord('3'): "기준가결정", ord('4'): "임의종료", ord('5'): "종목정보공개", ord('6'): "종목조치", ord('7'): "시장조치"}

    def OnReceived(self):
        if self.name == '9619s':
            time_num = self.client.GetHeaderValue(0)
            flag = self.client.GetHeaderValue(1)
            time_str = datetime.strptime(f"{time_num:06d}", '%H%M%S')
            combined_datetime = datetime.now().replace(hour=time_str.hour, minute=time_str.minute, second=time_str.second)
            time = combined_datetime.strftime('%m/%d %H:%M:%S')

            if self.dic.get(flag) == "종목별 VI":
                code = self.client.GetHeaderValue(3)
                event = self.client.GetHeaderValue(5)
                event2 = self.client.GetHeaderValue(6)
                match1 = re.search(r'^A\d{6}$', code)
                match2 = re.search(r"괴리율:(-?\d+\.\d+)%", event2)

                if (cpCodeMgr.GetStockControlKind(code) == 0 and
                    cpCodeMgr.GetStockSectionKind(code) == 1 and
                    match1 and match2 and "정적" in event):
                        gap_rate = float(match2.group(1))
                        if gap_rate > 0 and combined_datetime < combined_datetime.replace(hour=15, minute=15, second=0):
                            logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> VI 발동")
                            self.caller.monitor_vi(time, code, event2)
            return
        
        if self.name == 'stockcur':
            code = self.client.GetHeaderValue(0)
            timess = self.client.GetHeaderValue(18)
            exFlag = self.client.GetHeaderValue(19)
            cprice = self.client.GetHeaderValue(13)
            cVol = self.client.GetHeaderValue(17)

            if (exFlag == ord('2')):
                item = {'code': code, 'time': timess, 'cur': cprice, 'vol': cVol}
                self.caller.updateCurData(item)
            return
        
        if self.name == 'cssalert':
            stgid = self.client.GetHeaderValue(0)
            stgmonid = self.client.GetHeaderValue(1)
            code = self.client.GetHeaderValue(2)
            inoutflag = self.client.GetHeaderValue(3)
            stgtime = self.client.GetHeaderValue(4)
            stgprice = self.client.GetHeaderValue(5)
            time_obj = datetime.strptime(stgtime.zfill(6), '%H%M%S')
            combined_datetime = datetime.now().replace(hour=time_obj.hour, minute=time_obj.minute, second=time_obj.second)
            time_str = combined_datetime.strftime('%m/%d %H:%M:%S')
            
            # 진입 처리
            if inoutflag == ord('1') and combined_datetime < combined_datetime.replace(hour=15, minute=15, second=0):
                self.caller.checkRealtimeStg(stgid, stgmonid, code, stgprice, time_str)
            # 퇴출 처리 (로그만 기록, 중복 필터링)
            elif inoutflag == ord('2'):
                # ===== ✅ 퇴출 이벤트도 중복 필터링 =====
                import time as time_module  # 모듈 이름 충돌 방지
                event_key = f"exit_{stgid}_{code}"
                current_time = time_module.time()
                
                if hasattr(self.caller, 'last_event_time'):
                    if event_key in self.caller.last_event_time:
                        elapsed = current_time - self.caller.last_event_time[event_key]
                        if elapsed < self.caller.event_dedup_seconds:
                            return  # 3초 이내 중복 퇴출 이벤트 무시
                    
                    self.caller.last_event_time[event_key] = current_time
                
                stock_name = cpCodeMgr.CodeToName(code)
                stgname = self.caller.stgname.get(stgid, '알 수 없음') if hasattr(self.caller, 'stgname') else '알 수 없음'
                logging.info(f"📤 [{stgname}] 조건검색 퇴출: {stock_name}({code}) @{stgprice:,}원")
            return
        
        if self.name == 'conclusion':
            conflag = self.client.GetHeaderValue(14)
            ordernum = self.client.GetHeaderValue(5)
            qty = self.client.GetHeaderValue(3)
            price = self.client.GetHeaderValue(4)
            code = self.client.GetHeaderValue(9)
            bs = self.client.GetHeaderValue(12)
            buyprice = self.client.GetHeaderValue(21)
            balance = self.client.GetHeaderValue(23)
            conflags = {"1": "체결", "2": "확인", "3": "거부", "4": "접수"}.get(conflag, "")
            self.caller.monitorOrderStatus(code, ordernum, conflags, price, qty, bs, balance, buyprice)

# ==================== 기존 Publish 클래스들 (유지) ====================
class CpPublish:
    def __init__(self, name, service_id):
        self.name = name
        self.obj = win32com.client.Dispatch(service_id)
        self.bIsSB = False

    def Subscribe(self, var, caller):
        if self.bIsSB:
            self.Unsubscribe()

        if len(var) > 0:
            self.obj.SetInputValue(0, var)
        handler = win32com.client.WithEvents(self.obj, CpEvent)
        handler.set_params(self.obj, self.name, caller)
        self.obj.Subscribe()
        self.bIsSB = True

    def Unsubscribe(self):
        if self.bIsSB:
            self.obj.Unsubscribe()
            self.bIsSB = False

class CpPBConclusion(CpPublish):
    def __init__(self):
        super().__init__('conclusion', 'DsCbo1.CpConclusion')

class CpPBStockCur(CpPublish):
    def __init__(self):
        super().__init__('stockcur', 'DsCbo1.StockCur')

class CpPB9619(CpPublish):
    def __init__(self):
        super().__init__('9619s', 'CpSysDib.CpSvr9619s')

class CpPBCssAlert(CpPublish):
    def __init__(self):
        super().__init__('cssalert', 'CpSysDib.CssAlert')

class CpRequest:
    def __init__(self):
        self.result = None
        self.loop = QEventLoop()
        self.is_finished = False
        self.client = None
        self.name = None
        self.caller = None
        self.params = {}
        self.is_requesting = False

    def set_params(self, client, name, caller, params=None):
        self.client = client
        self.name = name
        self.caller = caller
        self.params = params if params else {}

    def wait(self):
        if not self.is_finished:
            self.loop.exec_()
        return self.result

    def OnReceived(self):
        try:
            if self.name == 'order':
                rqStatus = self.client.GetDibStatus()
                if rqStatus != 0:
                    rqRet = self.client.GetDibMsg1()
                    stock_name = self.params.get('stock_name', 'Unknown')
                    code = self.params.get('code', 'Unknown')
                    
                    # 주문 수량 관련 오류인 경우 더 자세한 정보 로깅
                    if '수량' in rqRet or '0' in rqRet:
                        logging.warning(f"{stock_name}({code}) 주문 요청 오류, 주문 수량이 없습니다.[{code}][0]")
                        logging.debug(f"상세 오류 메시지: {rqRet}")
                    else:
                        logging.warning(f"{stock_name}({code}) 주문 요청 오류, {rqRet}")
                    
                    self.result = False
                    return
                self.result = True

        except Exception as ex:
            logging.error(f"OnReceived -> {ex}")
            self.result = False

        finally:
            self.is_finished = True
            if self.loop.isRunning():
                self.loop.quit()
            if self.caller and hasattr(self.caller, 'cp_request'):
                self.caller.cp_request.is_requesting = False

# ==================== CpStrategy (조건검색 편입 처리 - 큐 기반) ====================
class CpStrategy(QThread):
    # 시그널 정의
    stock_processed = pyqtSignal(str, bool)  # (종목코드, 성공여부)
    processing_error = pyqtSignal(str, str)  # (종목코드, 오류메시지)
    
    def __init__(self, trader):
        super().__init__()
        self.monList = {}
        self.trader = trader
        self.stgname = {}
        self.objpb = CpPBCssAlert()
        
        # ===== ✅ Scanner 즉시 초기화 (타이밍 문제 해결) =====
        self.momentum_scanner = MomentumScanner(trader)
        self.gap_scanner = GapUpScanner(trader)
        
        # ===== ✅ 큐 처리 강화 (QThread 사용) =====
        self.stock_queue = queue.Queue()
        self.is_processing = False
        self.processing_lock = threading.Lock()
        self.is_thread_started = False
        
        # ===== ✅ 처리 제한 추가 =====
        self.max_concurrent_stocks = 1  # 동시 처리 종목 수 제한
        self.processing_stocks = set()  # 현재 처리 중인 종목
        self.failed_stocks = {}  # 실패한 종목 기록
        
        # ===== ✅ 중복 이벤트 필터링 =====
        self.last_event_time = {}  # 종목별 마지막 이벤트 시간
        self.event_dedup_seconds = 3.0  # 3초 이내 중복 이벤트 무시

    def start_processing_queue(self):
        """큐 처리 시작 (QThread 시작)"""
        if self.is_thread_started:
            logging.debug("큐 처리는 이미 시작됨")
            return
        
        self.is_thread_started = True  # ✅ 플래그 설정
        
        # ✅ QThread 시작
        self.start()
        logging.info("✅ 종목 처리 큐 QThread 시작")
    
    def stop_processing_queue(self):
        """큐 처리 중지 (QThread 종료)"""
        if self.is_thread_started:
            self.requestInterruption()
            self.wait(5000)  # 5초 대기
            self.is_thread_started = False
            logging.info("✅ 종목 처리 큐 QThread 종료")

    def run(self):
        """QThread 실행 메서드 - 큐에서 종목 순차 처리 (안전성 강화)"""
        while not self.isInterruptionRequested():
            try:
                # 큐에서 종목 가져오기
                try:
                    stock_data = self.stock_queue.get(timeout=1.0)
                except queue.Empty:
                    continue
                
                # QThread에서는 None 체크 대신 isInterruptionRequested 사용
                
                code = stock_data.get('code')
                
                # ===== ✅ 중복 처리 방지 =====
                if code in self.processing_stocks:
                    logging.debug(f"{code}: 이미 처리 중, 스킵")
                    self.stock_queue.task_done()
                    continue
                
                # ===== ✅ 실패 이력 확인 (3회 실패 시 스킵) =====
                if code in self.failed_stocks and self.failed_stocks[code] >= 3:
                    logging.warning(f"{code}: 3회 실패, 더 이상 시도 안 함")
                    self.stock_queue.task_done()
                    continue
                
                # 처리 중 표시
                self.processing_stocks.add(code)
                
                try:
                    # ===== ✅ 타임아웃 적용 처리 =====
                    success = self._process_single_stock_with_timeout(stock_data, timeout=60.0)
                    
                    # 처리 결과 시그널 전송
                    self.stock_processed.emit(code, success)
                    
                except Exception as ex:
                    error_msg = f"{code} 처리 중 예외: {ex}"
                    logging.error(error_msg)
                    
                    # 실패 카운트 증가
                    self.failed_stocks[code] = self.failed_stocks.get(code, 0) + 1
                    
                    # 오류 시그널 전송
                    self.processing_error.emit(code, error_msg)
                    
                finally:
                    # 처리 완료 표시
                    if code in self.processing_stocks:
                        self.processing_stocks.remove(code)
                    
                    self.stock_queue.task_done()
                
                # ===== ✅ 다음 종목 전 더 긴 대기 (API 제한 고려) =====
                time.sleep(1.5)  # 1.0초 → 1.5초로 증가
                
            except Exception as ex:
                logging.error(f"_process_stock_queue: {ex}\n{traceback.format_exc()}")
                continue

    def _check_api_limit_and_wait(self, operation_name="API 요청", rqtype=0):
        """API 제한 확인 및 대기 (요청 건수 기반)"""
        try:
            # 시세 조회 제한 상태 확인 (문서 기준 올바른 함수명 사용)
            remain_count = cpCybos.GetLimitRemainCount(rqtype)   # 시세 조회 남은 횟수 (0: 시세 조회)

            logging.debug(f"🔍 {operation_name} 전 API 상태 - 남은 요청: {remain_count}건")
            
            if remain_count > 0:
                # ✅ 요청 가능: 남은 건수가 있음
                if remain_count <= 5:  # 남은 요청이 5건 이하일 때 경고
                    logging.warning(f"⚠️ API 요청 한계 근접: 남은 요청 {remain_count}건")
                return True
            else:
                # ❌ 요청 불가: 남은 건수가 0건
                logging.warning(f"🚫 {operation_name} 거부: 요청 제한 (남은 요청: {remain_count}건)")
                return False
            
        except Exception as ex:
            logging.error(f"❌ API 제한 체크 실패: {ex}")
            return False

    def _log_api_status(self, operation_name="API 요청", success=True):
        """API 상태 로깅"""
        try:
            remain_count = cpCybos.GetLimitRemainCount(0)
            
            if success:
                logging.debug(f"✅ {operation_name} 완료 - API 상태: 남은 요청 {remain_count}건")
            else:
                logging.warning(f"❌ {operation_name} 실패 - API 상태: 남은 요청 {remain_count}건")
                
            # API 제한 상태 경고
            if remain_count <= 10:
                logging.warning(f"⚠️ API 요청 한계 근접: 남은 요청 {remain_count}건")
                
        except Exception as ex:
            logging.debug(f"API 상태 로깅 실패: {ex}")

    def _process_single_stock_with_timeout(self, stock_data, timeout=60.0):
        """종목 처리 (메인 스레드에서 직접 실행)"""
        try:
            # 메인 스레드에서 직접 처리
            return self._process_single_stock(stock_data)
        except Exception as ex:
            code = stock_data.get('code')
            logging.error(f"{code}: 처리 실패 - {ex}")
            raise ex

    def _process_single_stock(self, stock_data):
        """단일 종목 처리"""
        code = None
        try:
            stgid = stock_data['stgid']
            stgmonid = stock_data['stgmonid']
            code = stock_data['code']
            stgprice = stock_data['stgprice']
            time_str = stock_data['time']
            stgname = stock_data.get('stgname', '')
            
            stock_name = cpCodeMgr.CodeToName(code)
            
            logging.info(f"{'='*40}")
            if stgprice > 0:
                logging.info(f"🔍 [{stgname}] {stock_name}({code}) 검증 시작 - 가격: {stgprice:,}원")
            else:
                logging.info(f"🔍 [{stgname}] {stock_name}({code}) 검증 시작 (기존 편입 종목)")
            logging.info(f"{'='*40}")
            
            # 중복 확인
            if code in self.trader.monistock_set:
                logging.info(f"⚠️ [{stgname}] {stock_name}({code}): 이미 모니터링 중, 스킵")
                return True
            
            if code in self.trader.bought_set:
                logging.info(f"⚠️ [{stgname}] {stock_name}({code}): 이미 보유 중, 스킵")
                return True
            
            # ===== ✅ 대신증권 API 제한만 확인 =====
            if not self._check_api_limit_and_wait("기타 종목 처리", 0):
                logging.warning(f"❌ [{stgname}] {stock_name}({code}): API 제한으로 처리 거부")
                return False
            
            # 장 시작 후에만 처리
            now = datetime.now()
            market_open = now.replace(hour=9, minute=3, second=0, microsecond=0)
            if now < market_open:
                logging.debug(f"{code}: 장 시작 전, 스킵")
                return False
            
            # 전략별 처리
            if stgname == '급등주':
                result = self._process_momentum_stock(code, stgprice, time_str)
                if result:
                    logging.info(f"✅ [{stgname}] {stock_name}({code}): 검증 완료 → 투자대상 추가")
                else:
                    logging.info(f"❌ [{stgname}] {stock_name}({code}): 검증 실패 → 제외")
                return result
            elif stgname == '갭상승':
                result = self._process_gap_stock(code, stgprice, time_str)
                if result:
                    logging.info(f"✅ [{stgname}] {stock_name}({code}): 검증 완료 → 투자대상 추가")
                else:
                    logging.info(f"❌ [{stgname}] {stock_name}({code}): 검증 실패 → 제외")
                return result
            else:
                result = self._process_other_stock(code, stgprice, time_str)
                if result:
                    logging.info(f"✅ [{stgname}] {stock_name}({code}): 검증 완료 → 투자대상 추가")
                else:
                    logging.info(f"❌ [{stgname}] {stock_name}({code}): 검증 실패 → 제외")
                return result
                
        except Exception as ex:
            logging.error(f"_process_single_stock({code}): {ex}\n{traceback.format_exc()}")
            return False
        
    def _process_momentum_stock(self, code, stgprice, time_str):
        """급등주 처리 (안전)"""
        try:
            stock_name = cpCodeMgr.CodeToName(code)
            
            if not self.momentum_scanner:
                logging.warning(f"❌ [{stock_name}] MomentumScanner가 초기화되지 않음")
                return False
            
            # ===== ✅ 1단계: 일봉 데이터 먼저 로드 =====            
            # API 제한 확인
            if not self._check_api_limit_and_wait("일봉 데이터 로드"):
                logging.warning(f"❌ [급등주] {stock_name}({code}): API 제한으로 일봉 로드 거부")
                return False
            
            if not self.trader.daydata.select_code(code):
                logging.warning(f"❌ [급등주] {stock_name}({code}): 일봉 로드 실패")
                self._log_api_status("일봉 데이터 로드", False)
                return False
            
            # 일봉 로드 성공 후 API 상태 로깅
            self._log_api_status("일봉 데이터 로드", True)
            
            # 일봉 로드 후 더 긴 대기 (API 제한 고려)
            time.sleep(0.5)
            
            # ===== ✅ 2단계: 틱/분 데이터 로드 (순차 로드로 변경) =====
            try:
                # ✅ 틱 데이터 먼저 로드 (API 제한 고려)
                logging.debug(f"🔍 [급등주] {stock_name}({code}): 틱 데이터 로드 시작...")
                
                # API 제한 확인
                if not self._check_api_limit_and_wait("틱 데이터 로드"):
                    logging.warning(f"❌ [급등주] {stock_name}({code}): API 제한으로 틱 로드 거부")
                    tick_ok = False
                else:
                    tick_ok = self._load_with_timeout(
                        self.trader.tickdata.monitor_code,
                        code,
                        timeout=45.0  # 60초 → 45초로 단축
                    )
                logging.debug(f"🔍 [급등주] {stock_name}({code}): 틱 데이터 로드 결과: {tick_ok}")
                
                # 틱 로드 후 대기
                time.sleep(0.3)
                
                # ✅ 분봉 데이터 로드
                logging.debug(f"🔍 [급등주] {stock_name}({code}): 분봉 데이터 로드 시작...")
                
                # API 제한 확인
                if not self._check_api_limit_and_wait("분봉 데이터 로드"):
                    logging.warning(f"❌ [급등주] {stock_name}({code}): API 제한으로 분봉 로드 거부")
                    min_ok = False
                else:
                    min_ok = self._load_with_timeout(
                        self.trader.mindata.monitor_code,
                        code,
                        timeout=45.0  # 60초 → 45초로 단축
                    )
                logging.debug(f"🔍 [급등주] {stock_name}({code}): 분봉 데이터 로드 결과: {min_ok}")
                
                # ✅ 부분 실패 시에도 계속 진행 (일봉 데이터만으로도 기본 검증 가능)
                if not tick_ok:
                    logging.warning(f"⚠️ [급등주] {stock_name}({code}): 틱 데이터 로드 실패, 일봉 데이터로만 검증 진행")
                    self._log_api_status("틱 데이터 로드", False)
                else:
                    self._log_api_status("틱 데이터 로드", True)
                    
                if not min_ok:
                    logging.warning(f"⚠️ [급등주] {stock_name}({code}): 분봉 데이터 로드 실패, 일봉 데이터로만 검증 진행")
                    self._log_api_status("분봉 데이터 로드", False)
                else:
                    self._log_api_status("분봉 데이터 로드", True)
                
                # ✅ 틱/분 모두 실패 시 종목 데이터 부재 가능성 체크
                if not tick_ok and not min_ok:
                    logging.warning(f"🚫 [급등주] {stock_name}({code}): 틱/분 모두 로드 실패 - 종목 데이터 부재 또는 거래정지 가능성")
                    # API 제한이 아닌 경우 데이터 부재로 판단하고 스킵
                    remain_count = cpCybos.GetLimitRemainCount(0)
                    if remain_count > 10:
                        logging.info(f"📋 [급등주] {stock_name}({code}): API 제한 아님(요청 {remain_count}건) - 데이터 부재로 판단하여 스킵")
                        self.trader.daydata.monitor_stop(code)
                        self.trader.tickdata.monitor_stop(code)
                        self.trader.mindata.monitor_stop(code)
                        return False
            except Exception as ex:
                logging.error(f"❌ [급등주] {stock_name}({code}): 데이터 로드 중 오류: {ex}")
                self.trader.daydata.monitor_stop(code)
                return False
            
            # ✅ 부분 실패 시에도 일봉 데이터로 기본 검증 진행
            if not tick_ok and not min_ok:
                logging.warning(f"❌ [급등주] {stock_name}({code}): 틱/분 모두 로드 실패 (틱:{tick_ok}, 분:{min_ok})")
                self.trader.daydata.monitor_stop(code)
                return False
            elif not (tick_ok and min_ok):
                logging.warning(f"⚠️ [급등주] {stock_name}({code}): 일부 데이터 로드 실패 (틱:{tick_ok}, 분:{min_ok}), 일봉 데이터로 검증 진행")
            
            # ===== ✅ 데이터 준비 대기 (최대 15초, 조건 완화) =====
            data_ready = False
            for attempt in range(12):  # 최대 12회 시도 (12초로 단축)
                time.sleep(1.0)
                
                # 데이터 확인 (조건 완화: 최소 데이터만 있으면 OK)
                tick_data = self.trader.tickdata.stockdata.get(code, {})
                min_data = self.trader.mindata.stockdata.get(code, {})
                day_data = self.trader.daydata.stockdata.get(code, {})
                
                # ✅ 조건 완화: 일봉 데이터만 있어도 진행 (틱/분 데이터는 선택사항)
                tick_has_data = tick_data and len(tick_data.get('C', [])) > 0 if tick_ok else True  # 틱 로드 실패 시 무시
                min_has_data = min_data and len(min_data.get('C', [])) > 0 if min_ok else True   # 분 로드 실패 시 무시
                day_has_data = day_data and len(day_data.get('C', [])) >= 2  # 일봉은 필수
                
                # 일봉 데이터만 있어도 진행
                if day_has_data and (tick_has_data or not tick_ok) and (min_has_data or not min_ok):
                    data_ready = True
                    break
                else:
                    logging.info(f"⏳ [급등주] {stock_name}({code}): 데이터 대기 중... ({attempt+1}/15초) - 틱:{len(tick_data.get('C', []))}개, 분:{len(min_data.get('C', []))}개, 일:{len(day_data.get('C', []))}개)")
            
            if not data_ready:
                logging.warning(f"❌ [급등주] {stock_name}({code}): 데이터 준비 시간 초과 (15초)")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            # ===== ✅ 3단계: 급등주 조건 재확인 (메모리 데이터 기반) =====
            logging.info(f"🔍 [급등주] {stock_name}({code}): 조건 검증 시작...")
            try:
                logging.debug(f"🔍 [급등주] {stock_name}({code}): MomentumScanner 검증 호출...")
                is_valid, score, message = self._verify_with_timeout(
                    self.momentum_scanner.verify_momentum_conditions,
                    code,
                    timeout=10.0
                )
                logging.debug(f"🔍 [급등주] {stock_name}({code}): 검증 결과 - 유효:{is_valid}, 점수:{score}, 메시지:{message}")
            except Exception as ex:
                logging.error(f"❌ [급등주] {stock_name}({code}): 검증 중 오류: {ex}")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            if not is_valid:
                logging.info(f"❌ [급등주] {stock_name}({code}): 재검증 실패 - {message}")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            logging.info(
                f"✅ [급등주] {stock_name}({code}) → 재검증 통과 "
                f"({message})"
            )
            
            # 체결강도 확인
            logging.info(f"🔍 [급등주] {stock_name}({code}): 체결강도 확인 중...")
            try:
                strength = self.trader.tickdata.get_strength(code)
                logging.debug(f"🔍 [급등주] {stock_name}({code}): 체결강도 조회 결과: {strength}")
            except Exception as ex:
                logging.error(f"❌ [급등주] {stock_name}({code}): 체결강도 조회 실패: {ex}")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            # ✅ 장초반에는 체결강도 기준 완화 (9-10시: 100, 10시 이후: 120)
            now = datetime.now()
            min_strength = 100 if 9 <= now.hour < 10 else 120
            
            if strength >= min_strength:
                # 투자대상 추가
                logging.info(f"✅ [급등주] {stock_name}({code}): 투자대상 추가 시작... (체결강도: {strength:.0f}, 기준: {min_strength})")
                try:
                    self._add_to_monitoring(code, stgprice, time_str, f"급등주 (점수: {score}, 체결강도: {strength:.0f})")
                    logging.info(f"✅ [급등주] {stock_name}({code}): 투자대상 추가 완료!")
                    return True
                except Exception as ex:
                    logging.error(f"❌ [급등주] {stock_name}({code}): 투자대상 추가 실패: {ex}")
                    self.trader.daydata.monitor_stop(code)
                    self.trader.tickdata.monitor_stop(code)
                    self.trader.mindata.monitor_stop(code)
                    return False
            else:
                logging.info(f"❌ [급등주] {stock_name}({code}): 체결강도 부족 (현재: {strength:.0f}, 최소: {min_strength})")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
                
        except Exception as ex:
            logging.error(f"❌ [급등주] _process_momentum_stock({code}): {ex}\n{traceback.format_exc()}")
            return False

    def _process_gap_stock(self, code, stgprice, time_str):
        """갭상승 처리 (안전성 대폭 강화)"""
        stock_name = cpCodeMgr.CodeToName(code)
        
        try:
            logging.info(f"🔍 [갭상승] {stock_name}({code}): 검증 시작")
            
            # ===== ✅ GapUpScanner 확인 =====
            if not self.gap_scanner:
                logging.error(f"❌ [갭상승] {stock_name}({code}): GapUpScanner 미초기화")
                return False
            
            # ===== ✅ 1단계: 일봉 데이터 먼저 로드 =====
            logging.info(f"📊 [갭상승] {stock_name}({code}): 일봉 로드 중...")
            
            try:
                # API 제한 확인
                if not self._check_api_limit_and_wait("일봉 데이터 로드"):
                    logging.warning(f"❌ [갭상승] {stock_name}({code}): API 제한으로 일봉 로드 거부")
                    return False
                
                if not self.trader.daydata.select_code(code):
                    logging.warning(f"❌ [갭상승] {stock_name}({code}): 일봉 로드 실패")
                    self._log_api_status("일봉 데이터 로드", False)
                    return False
                
                # 일봉 로드 성공 후 API 상태 로깅
                self._log_api_status("일봉 데이터 로드", True)
                
                # 일봉 로드 후 더 긴 대기 (API 제한 고려)
                time.sleep(0.5)
            except Exception as ex:
                logging.error(f"❌ [갭상승] {stock_name}({code}): 일봉 로드 중 오류: {ex}")
                return False
            
            # ===== ✅ 2단계: 틱/분 데이터 로드 (순차 로드로 변경) =====            
            try:
                # ✅ 틱 데이터 먼저 로드 (API 제한 고려)
                # API 제한 확인
                if not self._check_api_limit_and_wait("틱 데이터 로드"):
                    logging.warning(f"❌ [갭상승] {stock_name}({code}): API 제한으로 틱 로드 거부")
                    self.trader.daydata.monitor_stop(code)
                    return False
                
                tick_ok = self._load_with_timeout(
                    self.trader.tickdata.monitor_code,
                    code,
                    timeout=35.0  # 40초 → 35초로 단축
                )
                
                if not tick_ok:
                    logging.warning(f"❌ [갭상승] {stock_name}({code}): 틱 데이터 로드 실패")
                    self._log_api_status("틱 데이터 로드", False)
                    self.trader.daydata.monitor_stop(code)
                    return False
                else:
                    self._log_api_status("틱 데이터 로드", True)
                
                # 틱 로드 후 대기
                time.sleep(0.3)
                
                # ✅ 분봉 데이터 로드
                # API 제한 확인
                if not self._check_api_limit_and_wait("분봉 데이터 로드"):
                    logging.warning(f"❌ [갭상승] {stock_name}({code}): API 제한으로 분봉 로드 거부")
                    self.trader.daydata.monitor_stop(code)
                    self.trader.tickdata.monitor_stop(code)
                    return False
                
                min_ok = self._load_with_timeout(
                    self.trader.mindata.monitor_code,
                    code,
                    timeout=35.0  # 40초 → 35초로 단축
                )
                
                if not min_ok:
                    logging.warning(f"❌ [갭상승] {stock_name}({code}): 분 데이터 로드 실패")
                    self._log_api_status("분봉 데이터 로드", False)
                    self.trader.daydata.monitor_stop(code)
                    self.trader.tickdata.monitor_stop(code)
                    return False
                else:
                    self._log_api_status("분봉 데이터 로드", True)
                
            except TimeoutError:
                logging.error(f"❌ [갭상승] {stock_name}({code}): 데이터 로드 타임아웃")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                return False
            except Exception as ex:
                logging.error(f"❌ [갭상승] {stock_name}({code}): 데이터 로드 중 오류: {ex}")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                return False
            
            # ===== ✅ 데이터 준비 대기 (최대 15초, 조건 완화) =====
            data_ready = False
            for attempt in range(12):  # 최대 12회 시도 (12초로 단축)
                time.sleep(1.0)
                
                # 데이터 확인 (조건 완화: 최소 데이터만 있으면 OK)
                tick_data = self.trader.tickdata.stockdata.get(code, {})
                min_data = self.trader.mindata.stockdata.get(code, {})
                day_data = self.trader.daydata.stockdata.get(code, {})
                
                # 최소 조건: 각 데이터에 종가(C)가 있고, 일봉에 최소 2개 이상의 데이터 (전일종가 필요)
                tick_has_data = tick_data and len(tick_data.get('C', [])) > 0
                min_has_data = min_data and len(min_data.get('C', [])) > 0
                day_has_data = day_data and len(day_data.get('C', [])) >= 2
                
                if tick_has_data and min_has_data and day_has_data:
                    data_ready = True
                    break
                else:
                    logging.info(f"⏳ [갭상승] {stock_name}({code}): 데이터 대기 중... ({attempt+1}/15초) - 틱:{len(tick_data.get('C', []))}개, 분:{len(min_data.get('C', []))}개, 일:{len(day_data.get('C', []))}개)")
            
            if not data_ready:
                logging.warning(f"❌ [갭상승] {stock_name}({code}): 데이터 준비 시간 초과 (15초)")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            # ===== ✅ 3단계: 갭상승 조건 재확인 (타임아웃 15초) =====
            logging.info(f"🔍 [갭상승] {stock_name}({code}): 갭상승 조건 검증 시작...")
            
            try:
                is_valid, gap_pct, message = self._verify_with_timeout(
                    self.gap_scanner.verify_gap_conditions,
                    code,
                    timeout=15.0
                )
            except TimeoutError:
                logging.error(f"❌ [갭상승] {stock_name}({code}): 검증 타임아웃")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            except Exception as ex:
                logging.error(f"❌ [갭상승] {stock_name}({code}): 검증 중 오류: {ex}")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            if not is_valid:
                logging.info(f"❌ [갭상승] {stock_name}({code}): 재검증 실패 - {message}")
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
            logging.info(f"✅ [갭상승] {stock_name}({code}): 재검증 통과 - {message}")
            
            # ===== ✅ 투자대상 추가 =====
            try:
                self._add_to_monitoring(code, stgprice, time_str, f"갭상승 ({gap_pct:.2f}%)")
                logging.info(f"✅ [갭상승] {stock_name}({code}): 투자대상 추가 완료")
                return True
            except Exception as ex:
                logging.error(f"{stock_name}({code}): 투자대상 추가 실패: {ex}")
                # 실패 시 정리
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return False
            
        except Exception as ex:
            logging.error(f"_process_gap_stock({code}): {ex}\n{traceback.format_exc()}")
            # 예외 발생 시 리소스 정리
            try:
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
            except:
                pass
            return False
        
    def _process_other_stock(self, code, stgprice, time_str):
        """기타 전략 처리"""
        try:
            stock_name = cpCodeMgr.CodeToName(code)
            
            # 데이터 로드
            if not self.trader.daydata.select_code(code):
                logging.warning(f"❌ [기타] {stock_name}({code}): 일봉 로드 실패")
                return False
            
            if not (self.trader.tickdata.monitor_code(code) and self.trader.mindata.monitor_code(code)):
                logging.warning(f"❌ [기타] {stock_name}({code}): 틱/분 로드 실패")
                self.trader.daydata.monitor_stop(code)
                return False
            
            # 투자대상 추가
            self._add_to_monitoring(code, stgprice, time_str, "기타 전략")
            return True
            
        except Exception as ex:
            logging.error(f"❌ [기타] _process_other_stock({code}): {ex}")
            return False

    def _verify_with_timeout(self, func, code, timeout=10.0):
        """검증 (타임아웃 포함)"""
        import threading
        import time
        
        result = [None]
        exception = [None]
        
        def target():
            try:
                result[0] = func(code)
            except Exception as ex:
                exception[0] = ex
        
        thread = threading.Thread(target=target)
        thread.daemon = True
        thread.start()
        thread.join(timeout)
        
        if thread.is_alive():
            logging.warning(f"{code}: 검증 타임아웃 ({timeout}초)")
            return (False, 0, f"타임아웃 ({timeout}초)")
        
        if exception[0]:
            logging.warning(f"{code}: 검증 실패 - {exception[0]}")
            return (False, 0, str(exception[0]))
        
        return result[0]

    def _load_with_timeout(self, func, code, timeout=30.0):
        """로드 (타임아웃 포함)"""
        import threading
        import time
        
        result = [None]
        exception = [None]
        
        def target():
            try:
                result[0] = func(code)
            except Exception as ex:
                exception[0] = ex
        
        thread = threading.Thread(target=target)
        thread.daemon = True
        thread.start()
        thread.join(timeout)
        
        if thread.is_alive():
            # 타임아웃 시 API 제한 상태 확인
            try:
                remain_count = cpCybos.GetLimitRemainCount(0)
                logging.warning(f"⏰ {code}: 로드 타임아웃 ({timeout}초) - API 상태: 남은 요청 {remain_count}건")
            except:
                logging.warning(f"⏰ {code}: 로드 타임아웃 ({timeout}초)")
            return False
        
        if exception[0]:
            # 오류 시 API 제한 상태 확인
            try:
                remain_count = cpCybos.GetLimitRemainCount(0)
                logging.warning(f"❌ {code}: 로드 실패 - {exception[0]} | API 상태: 남은 요청 {remain_count}건")
            except:
                logging.warning(f"❌ {code}: 로드 실패 - {exception[0]}")
            return False
        
        return result[0]

    def _add_to_monitoring(self, code, price, time_str, reason):
        """투자대상 종목 추가"""
        try:
            stock_name = cpCodeMgr.CodeToName(code)
            
            # ===== ✅ 20개 종목 제한 체크 =====
            # 보유 종목은 제외하고 순수 모니터링 종목만 카운트
            monitoring_only = self.trader.monistock_set - self.trader.bought_set
            MAX_MONITORING_STOCKS = 20
            
            if len(monitoring_only) >= MAX_MONITORING_STOCKS:
                logging.warning(
                    f"⚠️ {stock_name}({code}) 추가 거부: "
                    f"모니터링 종목이 이미 {MAX_MONITORING_STOCKS}개 (보유 제외)"
                )
                # 리소스 정리
                self.trader.daydata.monitor_stop(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                return
            
            self.trader.starting_time[code] = time_str
            self.trader.starting_price[code] = price
            self.trader.monistock_set.add(code)
            
            logging.info(f"📋 {stock_name}({code}) 모니터링 세트에 추가 완료")
            logging.info(f"📋 현재 모니터링 종목 수: {len(self.trader.monistock_set)} (순수 모니터링: {len(monitoring_only)+1}/{MAX_MONITORING_STOCKS})")
            
            self.trader.stock_added_to_monitor.emit(code)
            logging.info(f"📋 {stock_name}({code}) UI 업데이트 시그널 발송 완료")
            
            self.trader.save_list_db(code, time_str, price, 1)
            
            logging.info(f"{stock_name}({code}) -> 투자 대상 추가: {reason}")
            
        except Exception as ex:
            logging.error(f"_add_to_monitoring({code}): {ex}")

    def requestList(self):
        retStgList = {}
        objRq = win32com.client.Dispatch("CpSysDib.CssStgList")
        objRq.SetInputValue(0, ord('1'))
        objRq.BlockRequest2(1)

        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"나의전략 조회실패, {rqStatus}, {rqRet}")
            return (False, retStgList)

        cnt = objRq.GetHeaderValue(0)
        flag = objRq.GetHeaderValue(1)

        for i in range(cnt):
            item = {}
            item['전략명'] = objRq.GetDataValue(0, i)
            item['ID'] = objRq.GetDataValue(1, i)
            item['평균수익률'] = objRq.GetDataValue(6, i)
            retStgList[item['전략명']] = item
        return retStgList

    def requestStgID(self, id):
        retStgstockList = []
        objRq = None
        objRq = win32com.client.Dispatch("CpSysDib.CssStgFind")
        objRq.SetInputValue(0, id)
        objRq.BlockRequest2(1)
        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"전략ID 조회실패, {rqStatus}, {rqRet}")
            return (False, retStgstockList)

        cnt = objRq.GetHeaderValue(0)
        totcnt = objRq.GetHeaderValue(1)
        stime = objRq.GetHeaderValue(2)

        for i in range(cnt):
            item = {}
            item['code'] = objRq.GetDataValue(0, i)
            item['종목명'] = cpCodeMgr.CodeToName(item['code'])
            retStgstockList.append(item)

        return (True, retStgstockList)

    def requestMonitorID(self, id):
        objRq = win32com.client.Dispatch("CpSysDib.CssWatchStgSubscribe")
        objRq.SetInputValue(0, id)
        objRq.BlockRequest2(1)

        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            logging.warning(f"감시번호 조회실패, {rqStatus}, {rqRet}")
            return (False, 0)

        monID = objRq.GetHeaderValue(0)
        if monID == 0:
            logging.warning(f"감시 일련번호 구하기 실패")
            return (False, 0)

        return (True, monID)

    def requestStgControl(self, id, monID, bStart, stgname):
        objRq = win32com.client.Dispatch("CpSysDib.CssWatchStgControl")
        objRq.SetInputValue(0, id)
        objRq.SetInputValue(1, monID)

        if bStart == True:
            objRq.SetInputValue(2, ord('1'))
            self.stgname[id] = stgname
        else:
            objRq.SetInputValue(2, ord('3'))
            if id in self.stgname:
                del self.stgname[id]
        objRq.BlockRequest2(1)

        rqStatus = objRq.GetDibStatus()
        if rqStatus != 0:
            rqRet = objRq.GetDibMsg1()
            
            if bStart:
                logging.warning(f"감시시작 실패 - 전략: {stgname}, 감시번호: {monID}, 상태: {rqStatus}, 메시지: {rqRet}")
            else:
                logging.info(f"감시취소 완료 - 전략: {stgname}, 감시번호: {monID}, 상태: {rqStatus}, 메시지: {rqRet}")
            return (False, '')

        status = objRq.GetHeaderValue(0)

        # ===== ✅ Subscribe는 첫 조건검색 시작 시 한 번만 호출 =====
        if bStart == True:
            # 아직 구독하지 않은 경우에만 Subscribe
            if not self.objpb.bIsSB:
                self.objpb.Subscribe('', self)
                logging.debug(f"조건검색 이벤트 구독 시작 (CssAlert)")
            
            self.monList[id] = monID
            logging.info(f"✅ 조건검색 감시 시작 - 전략: {stgname}, ID: {id}, 감시번호: {monID}")
        else:
            if id in self.monList:
                del self.monList[id]
            
            # ===== ✅ 모든 조건검색이 종료되면 Unsubscribe =====
            if len(self.monList) == 0 and self.objpb.bIsSB:
                self.objpb.Unsubscribe()
                logging.debug(f"모든 조건검색 종료 - 이벤트 구독 해제")
            
            logging.info(f"✅ 조건검색 감시 취소 - 전략: {stgname}, ID: {id}, 감시번호: {monID}")

        return (True, status)

    def checkRealtimeStg(self, stgid, stgmonid, code, stgprice, time_str):
        """조건검색 편입 시 호출 - 큐에 추가 (중복 필터링)"""
        
        if stgid not in self.monList:
            return
        if stgmonid != self.monList[stgid]:
            return
        
        # ===== ✅ 중복 이벤트 필터링 (3초 이내 중복 무시) =====
        import time as time_module  # 모듈 이름 충돌 방지
        event_key = f"{stgid}_{code}"
        current_time = time_module.time()
        
        if event_key in self.last_event_time:
            elapsed = current_time - self.last_event_time[event_key]
            if elapsed < self.event_dedup_seconds:
                # 3초 이내 중복 이벤트 무시
                return
        
        # 마지막 이벤트 시간 기록
        self.last_event_time[event_key] = current_time
        
        stgname = self.stgname.get(stgid, '')
        stock_name = cpCodeMgr.CodeToName(code)
        
        # ===== ✅ 조건검색 편입 로그 (로그창 표시) =====
        logging.info(f"📢 [{stgname}] 조건검색 편입: {stock_name}({code}) @{stgprice:,}원")
        
        # ===== ✅ 큐에 추가 (즉시 반환) =====
        stock_data = {
            'stgid': stgid,
            'stgmonid': stgmonid,
            'code': code,
            'stgprice': stgprice,
            'time': time_str,
            'stgname': stgname
        }
        
        self.stock_queue.put(stock_data)
        logging.info(f"🔄 {stock_name}({code}): 처리 큐에 추가됨 (대기: {self.stock_queue.qsize()}개)")

    def Clear(self):
        """정리 (큐 종료 포함)"""
        delitem = []
        for id, monId in self.monList.items():
            delitem.append((id, monId))

        for item in delitem:
            self.requestStgControl(item[0], item[1], False, "Unknown")

        # ===== ✅ Subscribe 확실하게 정리 =====
        if self.objpb.bIsSB:
            self.objpb.Unsubscribe()
            logging.debug("조건검색 이벤트 구독 해제")
        
        # 중복 필터링 변수 초기화
        self.last_event_time.clear()
        
        # 큐 종료 (QThread 방식)
        if self.is_thread_started:
            self.requestInterruption()
            self.wait(3000)  # 3초 대기
            self.is_thread_started = False
            logging.info("✅ 종목 처리 큐 QThread 종료 (Clear)")
        
        # ===== ✅ 큐 비우기 (남은 데이터 제거) =====
        while not self.stock_queue.empty():
            try:
                self.stock_queue.get_nowait()
            except:
                break
        
        # ===== ✅ 플래그 초기화 (다음 전략에서 QThread 재시작 가능) =====
        self.is_thread_started = False
        
        # ===== ✅ 처리 중 종목 및 실패 기록 초기화 =====
        self.processing_stocks.clear()
        self.failed_stocks.clear()

# ==================== CpIndicators (유지) ====================
class CpIndicators:
    def __init__(self, chart_type):
        self.chart_type = chart_type
        self.params = self._get_default_params()
    
    def _get_default_params(self):
        if self.chart_type == 'T':
            return {
                'MA_PERIODS': [5, 20, 60, 120],
                'RSI_PERIOD': 12,
                'RSI_SIGNAL_PERIOD': 12,
                'MACD': (12, 26, 9),
                'STOCH': (7, 3, 3),
                'ATR_PERIOD': 10,
                'CCI_PERIOD': 12,
                'BB_PERIOD': 20,
                'BB_STD': 2,
                'WILLIAMS_R_PERIOD': 14,  # 추가
                'ROC_PERIOD': 10,
            }
        elif self.chart_type == 'm':
            return {
                'MA_PERIODS': [5, 10, 20],
                'RSI_PERIOD': 9,
                'RSI_SIGNAL_PERIOD': 9,
                'MACD': (12, 26, 9),
                'STOCH': (7, 3, 3),
                'ATR_PERIOD': 10,
                'CCI_PERIOD': 12,
                'BB_PERIOD': 20,
                'BB_STD': 2,
                'WILLIAMS_R_PERIOD': 14,  # 추가
                'ROC_PERIOD': 10,
            }
        elif self.chart_type == 'D':
            return {
                'MA_PERIODS': [5, 10],
                'MACD': (12, 26, 9)
            }
        return {}
    
    def _validate_data(self, chart_data, min_length):
        required_keys = ['C', 'H', 'L', 'V', 'D', 'T']
        for key in required_keys:
            if key not in chart_data:
                return False
            if len(chart_data[key]) < min_length:
                return False
        return True
    
    def _fill_nan(self, data, method='smart'):
        if method == 'smart':
            series = pd.Series(data)
            filled = series.fillna(method='ffill').fillna(method='bfill').fillna(0)
            return filled.tolist()
        else:
            return np.nan_to_num(data, nan=0.0).tolist()

    def calculate_williams_r(self, highs, lows, closes, period=14):
        """Williams %R 계산
        
        과매수/과매도 지표
        -20 이상: 과매수 (매도 신호)
        -80 이하: 과매도 (매수 신호)
        
        Args:
            highs: 고가 리스트
            lows: 저가 리스트
            closes: 종가 리스트
            period: 계산 기간 (기본 14)
        
        Returns:
            Williams %R 리스트 (-100 ~ 0)
        """
        williams_r = []
        
        for i in range(len(closes)):
            if i < period - 1:
                williams_r.append(-50)  # 기본값
                continue
            
            # 최근 N일의 최고가/최저가
            high_max = max(highs[i-period+1:i+1])
            low_min = min(lows[i-period+1:i+1])
            
            if high_max - low_min == 0:
                williams_r.append(-50)
            else:
                # Williams %R = (최고가 - 현재가) / (최고가 - 최저가) * -100
                wr = ((high_max - closes[i]) / (high_max - low_min)) * -100
                williams_r.append(wr)
        
        return williams_r
    
    def calculate_roc(self, closes, period=10):
        """Price Rate of Change (ROC) 계산
        
        가격 변화율 - 모멘텀 지표
        양수: 상승 추세
        음수: 하락 추세
        
        Args:
            closes: 종가 리스트
            period: 계산 기간 (기본 10)
        
        Returns:
            ROC 리스트 (%)
        """
        roc = []
        
        for i in range(len(closes)):
            if i < period:
                roc.append(0)
            else:
                if closes[i-period] != 0:
                    # ROC = (현재가 - N일전 가격) / N일전 가격 * 100
                    roc_value = ((closes[i] - closes[i-period]) / closes[i-period]) * 100
                    roc.append(roc_value)
                else:
                    roc.append(0)
        
        return roc
    
    def calculate_obv(self, closes, volumes):
        """On-Balance Volume (OBV) 계산
        
        거래량 기반 추세 확인
        OBV 상승 + 가격 상승: 강한 상승 추세
        OBV 하락 + 가격 상승: 약한 상승 추세 (다이버전스)
        
        Args:
            closes: 종가 리스트
            volumes: 거래량 리스트
        
        Returns:
            OBV 리스트
        """
        obv = [0]
        
        for i in range(1, len(closes)):
            if closes[i] > closes[i-1]:
                # 상승 시 거래량 더함
                obv.append(obv[-1] + volumes[i])
            elif closes[i] < closes[i-1]:
                # 하락 시 거래량 뺌
                obv.append(obv[-1] - volumes[i])
            else:
                # 보합 시 유지
                obv.append(obv[-1])
        
        return obv
    
    def calculate_obv_ma(self, obv, period=20):
        """OBV 이동평균 계산
        
        Args:
            obv: OBV 리스트
            period: 계산 기간 (기본 20)
        
        Returns:
            OBV MA 리스트
        """
        obv_ma = []
        
        for i in range(len(obv)):
            if i < period - 1:
                obv_ma.append(obv[i])
            else:
                ma = sum(obv[i-period+1:i+1]) / period
                obv_ma.append(ma)
        
        return obv_ma
    
    def calculate_volume_profile(self, closes, volumes, bins=20):
        """Volume Profile 계산
        
        가격대별 거래량 분포 분석
        
        Args:
            closes: 종가 리스트
            volumes: 거래량 리스트
            bins: 가격대 구간 수
        
        Returns:
            (max_volume_price, current_vs_poc): 최대 거래량 가격, 현재가 위치
        """
        if len(closes) == 0 or len(volumes) == 0:
            return 0, 0
        
        # 가격 범위
        price_min = min(closes)
        price_max = max(closes)
        
        if price_max == price_min:
            return closes[-1], 0
        
        # 가격대별 거래량 집계
        bin_size = (price_max - price_min) / bins
        volume_profile = {}
        
        for price, volume in zip(closes, volumes):
            bin_index = int((price - price_min) / bin_size) if bin_size > 0 else 0
            bin_index = min(bin_index, bins - 1)  # 상한선
            
            volume_profile[bin_index] = volume_profile.get(bin_index, 0) + volume
        
        # 최대 거래량 가격대 (POC: Point of Control)
        if volume_profile:
            max_volume_bin = max(volume_profile, key=volume_profile.get)
            max_volume_price = price_min + (max_volume_bin + 0.5) * bin_size
        else:
            max_volume_price = closes[-1]
        
        # 현재가 vs POC
        current_price = closes[-1]
        current_vs_poc = (current_price - max_volume_price) / max_volume_price if max_volume_price > 0 else 0
        
        return max_volume_price, current_vs_poc
    
    def _get_default_result(self, indicator_type, length):
        """기본 결과값 반환 (데이터 부족 시)"""
        default_value = [0] * length
        
        if indicator_type == 'MA':
            if self.chart_type == 'T':
                return {
                    'MAT5': default_value, 'MAT20': default_value,
                    'MAT60': default_value, 'MAT120': default_value,
                    'MAT5_MAT20_DIFF': default_value, 'MAT20_MAT60_DIFF': default_value,
                    'MAT60_MAT120_DIFF': default_value, 'C_MAT5_DIFF': default_value,
                    'MAT5_CHANGE': default_value, 'MAT20_CHANGE': default_value,
                    'MAT60_CHANGE': default_value, 'MAT120_CHANGE': default_value
                }
            elif self.chart_type == 'm':
                return {
                    'MAM5': default_value, 'MAM10': default_value, 'MAM20': default_value,
                    'MAM5_MAM10_DIFF': default_value, 'MAM10_MAM20_DIFF': default_value,
                    'C_MAM5_DIFF': default_value, 'C_ABOVE_MAM5': default_value
                }
            elif self.chart_type == 'D':
                return {'MAD5': default_value, 'MAD10': default_value}
        
        elif indicator_type == 'MACD':
            if self.chart_type == 'T':
                return {
                    'MACDT': default_value, 'MACDT_SIGNAL': default_value,
                    'OSCT': default_value
                }
            else:
                return {
                    'MACD': default_value, 'MACD_SIGNAL': default_value,
                    'OSC': default_value
                }
        
        elif indicator_type == 'RSI':
            if self.chart_type == 'T':
                return {'RSIT': default_value, 'RSIT_SIGNAL': default_value}
            else:
                return {'RSI': default_value, 'RSI_SIGNAL': default_value}
        
        elif indicator_type == 'STOCH':
            return {'STOCHK': default_value, 'STOCHD': default_value}
        
        elif indicator_type == 'ATR':
            return {'ATR': default_value}
        
        elif indicator_type == 'CCI':
            return {'CCI': default_value}
        
        elif indicator_type == 'BBANDS':
            return {
                'BB_UPPER': default_value, 'BB_MIDDLE': default_value,
                'BB_LOWER': default_value, 'BB_POSITION': default_value,
                'BB_BANDWIDTH': default_value
            }
        
        elif indicator_type == 'VWAP':
            return {'VWAP': default_value}
        
        # === 새로운 지표들 ===
        elif indicator_type == 'WILLIAMS_R':
            return {'WILLIAMS_R': default_value}
        
        elif indicator_type == 'ROC':
            return {'ROC': default_value}
        
        elif indicator_type == 'OBV':
            return {'OBV': default_value, 'OBV_MA20': default_value}
        
        elif indicator_type == 'VOLUME_PROFILE':
            return {'VP_POC': 0, 'VP_POSITION': 0}
        
        return {}

    def make_indicator(self, indicator_type, code, chart_data):
        try:
            result = {}
            closes = np.array(chart_data['C'], dtype=np.float64)
            highs = np.array(chart_data['H'], dtype=np.float64)
            lows = np.array(chart_data['L'], dtype=np.float64)
            volumes = np.array(chart_data['V'], dtype=np.float64)
            dates = np.array(chart_data['D'], dtype=np.int64)
            
            desired_length = len(closes)
            
            min_lengths = {
                'MA': max(self.params.get('MA_PERIODS', [5])),
                'MACD': 35,
                'RSI': self.params.get('RSI_PERIOD', 14),
                'STOCH': 14,
                'ATR': self.params.get('ATR_PERIOD', 14),
                'CCI': self.params.get('CCI_PERIOD', 14),
                'BBANDS': self.params.get('BB_PERIOD', 20),
                'VWAP': 1,
                'WILLIAMS_R': self.params.get('WILLIAMS_R_PERIOD', 14),
                'ROC': self.params.get('ROC_PERIOD', 10),
                'OBV': 2,
                'VOLUME_PROFILE': 20
            }
            
            min_required = min_lengths.get(indicator_type, 20)
            
            if not self._validate_data(chart_data, min_required):
                logging.debug(f"{code}: {indicator_type} 데이터 부족 ({len(closes)} < {min_required})")
                return self._get_default_result(indicator_type, desired_length)
            
            if indicator_type == 'MA':
                if self.chart_type == 'T':
                    terms = self.params['MA_PERIODS']
                    index_names = ["MAT5", "MAT20", "MAT60", "MAT120"]
                elif self.chart_type == 'm':
                    terms = self.params['MA_PERIODS']
                    index_names = ["MAM5", "MAM10", "MAM20"]
                elif self.chart_type == 'D':
                    terms = self.params['MA_PERIODS']
                    index_names = ["MAD5", "MAD10"]
                
                for term, name in zip(terms, index_names):
                    sma = talib.SMA(closes, timeperiod=term)
                    result[name] = self._fill_nan(sma)
                
                if self.chart_type == 'T':
                    result['MAT5_MAT20_DIFF'] = [
                        (result['MAT5'][i] - result['MAT20'][i]) / result['MAT20'][i] 
                        if result['MAT20'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['MAT20_MAT60_DIFF'] = [
                        (result['MAT20'][i] - result['MAT60'][i]) / result['MAT60'][i] 
                        if result['MAT60'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['MAT60_MAT120_DIFF'] = [
                        (result['MAT60'][i] - result['MAT120'][i]) / result['MAT120'][i] 
                        if result['MAT120'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['C_MAT5_DIFF'] = [
                        (closes[i] - result['MAT5'][i]) / result['MAT5'][i] 
                        if result['MAT5'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    
                    for name in ['MAT5', 'MAT20', 'MAT60', 'MAT120']:
                        ma_values = result[name]
                        changes = [0] + [
                            (ma_values[i] - ma_values[i-1]) / ma_values[i-1]
                            if ma_values[i-1] != 0 else 0
                            for i in range(1, len(ma_values))
                        ]
                        result[f'{name}_CHANGE'] = changes
                
                elif self.chart_type == 'm':
                    result['MAM5_MAM10_DIFF'] = [
                        (result['MAM5'][i] - result['MAM10'][i]) / result['MAM10'][i] 
                        if result['MAM10'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['MAM10_MAM20_DIFF'] = [
                        (result['MAM10'][i] - result['MAM20'][i]) / result['MAM20'][i] 
                        if result['MAM20'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['C_MAM5_DIFF'] = [
                        (closes[i] - result['MAM5'][i]) / result['MAM5'][i] 
                        if result['MAM5'][i] != 0 else 0 
                        for i in range(desired_length)
                    ]
                    result['C_ABOVE_MAM5'] = [
                        1 if closes[i] > result['MAM5'][i] else 0 
                        for i in range(desired_length)
                    ]

            elif indicator_type == 'MACD':
                fast, slow, signal = self.params['MACD']
                macd_line, signal_line, macd_hist = talib.MACD(
                    closes, fastperiod=fast, slowperiod=slow, signalperiod=signal
                )
                
                if self.chart_type == 'T':
                    result['MACDT'] = self._fill_nan(macd_line)
                    result['MACDT_SIGNAL'] = self._fill_nan(signal_line)
                    result['OSCT'] = self._fill_nan(macd_hist)
                else:
                    result['MACD'] = self._fill_nan(macd_line)
                    result['MACD_SIGNAL'] = self._fill_nan(signal_line)
                    result['OSC'] = self._fill_nan(macd_hist)

            elif indicator_type == 'RSI':
                rsi_period = self.params['RSI_PERIOD']
                rsi_signal_period = self.params['RSI_SIGNAL_PERIOD']
                
                rsi_values = talib.RSI(closes, timeperiod=rsi_period)
                rsi_signal = talib.SMA(rsi_values, timeperiod=rsi_signal_period)
                
                if self.chart_type == 'T':
                    result['RSIT'] = self._fill_nan(rsi_values)
                    result['RSIT_SIGNAL'] = self._fill_nan(rsi_signal)
                else:
                    result['RSI'] = self._fill_nan(rsi_values)
                    result['RSI_SIGNAL'] = self._fill_nan(rsi_signal)

            elif indicator_type == 'STOCH':
                fastk_period, slowk_period, slowd_period = self.params['STOCH']
                stoch_k, stoch_d = talib.STOCH(
                    highs, lows, closes,
                    fastk_period=fastk_period,
                    slowk_period=slowk_period,
                    slowd_period=slowd_period
                )
                result['STOCHK'] = self._fill_nan(stoch_k)
                result['STOCHD'] = self._fill_nan(stoch_d)

            elif indicator_type == 'ATR':
                atr_period = self.params['ATR_PERIOD']
                atr = talib.ATR(highs, lows, closes, timeperiod=atr_period)
                result['ATR'] = self._fill_nan(atr)

            elif indicator_type == 'CCI':
                cci_period = self.params['CCI_PERIOD']
                cci = talib.CCI(highs, lows, closes, timeperiod=cci_period)
                result['CCI'] = self._fill_nan(cci)

            elif indicator_type == 'BBANDS':
                bb_period = self.params['BB_PERIOD']
                bb_std = self.params['BB_STD']
                
                upper, middle, lower = talib.BBANDS(
                    closes, timeperiod=bb_period,
                    nbdevup=bb_std, nbdevdn=bb_std
                )
                
                result['BB_UPPER'] = self._fill_nan(upper)
                result['BB_MIDDLE'] = self._fill_nan(middle)
                result['BB_LOWER'] = self._fill_nan(lower)
                
                bandwidth = upper - lower
                # ✅ 경고 억제하여 안전하게 나누기
                with np.errstate(divide='ignore', invalid='ignore'):
                    bb_position = np.where(
                        bandwidth > 1e-6,
                        (closes - middle) / bandwidth,
                        0.5
                    )
                bb_position = np.clip(bb_position, -2, 2)
                result['BB_POSITION'] = bb_position.tolist()
                
                # ✅ 경고 억제하여 안전하게 나누기
                with np.errstate(divide='ignore', invalid='ignore'):
                    bb_bandwidth = np.where(
                        middle > 1e-6,
                        bandwidth / middle,
                        0
                    )
                result['BB_BANDWIDTH'] = bb_bandwidth.tolist()

            elif indicator_type == 'VWAP':
                vwap = np.zeros_like(closes)
                
                if len(dates) == len(closes):
                    unique_dates = np.unique(dates)
                    
                    for d in unique_dates:
                        mask = dates == d
                        day_closes = closes[mask]
                        day_volumes = volumes[mask]
                        
                        cumsum_pv = np.cumsum(day_closes * day_volumes)
                        cumsum_v = np.cumsum(day_volumes)
                        
                        day_vwap = np.divide(
                            cumsum_pv, cumsum_v,
                            out=np.zeros_like(cumsum_pv),
                            where=cumsum_v != 0
                        )
                        vwap[mask] = day_vwap
                
                result['VWAP'] = vwap.tolist()

            elif indicator_type == 'WILLIAMS_R':
                """Williams %R 계산"""
                period = self.params.get('WILLIAMS_R_PERIOD', 14)
                
                williams_r = self.calculate_williams_r(
                    highs.tolist(), 
                    lows.tolist(), 
                    closes.tolist(), 
                    period
                )
                
                result['WILLIAMS_R'] = williams_r
            
            elif indicator_type == 'ROC':
                """ROC 계산"""
                period = self.params.get('ROC_PERIOD', 10)
                
                roc = self.calculate_roc(closes.tolist(), period)
                
                result['ROC'] = roc
            
            elif indicator_type == 'OBV':
                """OBV 및 OBV MA 계산"""
                obv = self.calculate_obv(closes.tolist(), volumes.tolist())
                obv_ma20 = self.calculate_obv_ma(obv, period=20)
                
                result['OBV'] = obv
                result['OBV_MA20'] = obv_ma20
            
            elif indicator_type == 'VOLUME_PROFILE':
                """Volume Profile 계산"""
                max_volume_price, current_vs_poc = self.calculate_volume_profile(
                    closes.tolist(), 
                    volumes.tolist()
                )
                
                # 단일 값이므로 리스트 대신 스칼라
                result['VP_POC'] = max_volume_price
                result['VP_POSITION'] = current_vs_poc

            else:
                logging.error(f"알 수 없는 지표 유형: {indicator_type}")
                return self._get_default_result(indicator_type, desired_length)

            return result

        except Exception as ex:
            logging.error(f"make_indicator -> {code}, {indicator_type}{self.chart_type} {ex}\n{traceback.format_exc()}")
            return self._get_default_result(indicator_type, len(chart_data.get('C', [])))

# ==================== CpData (체결강도 추가) ====================
class CpData(QObject):
    new_bar_completed = pyqtSignal(str)
    data_updated = pyqtSignal(str)  # 실시간 데이터 업데이트 시그널

    def __init__(self, interval, chart_type, number, trader):
        super().__init__()
        self.interval = interval
        self.number = number
        self.chart_type = chart_type
        self.objCur = {}
        self.stockdata = {}
        self.trader = trader  # trader 객체 참조 추가
        self.is_updating = {}  # 업데이트 상태 추적
        self.buy_volumes = {}  # 매수 거래량 추적
        self.sell_volumes = {}  # 매도 거래량 추적
        
        # 누락된 속성들 초기화
        self.objIndicators = {}
        self.code = ''
        self.LASTTIME = 1530
        self.is_initial_loaded = {}
        self.stockdata_lock = threading.Lock()
        self.last_update_time = {}
        self.last_indicator_update = {}
        self.latest_snapshot = {}
        self.strength_cache = {}
        self.indicator_update_interval = 1.0
        
        # 타이머 초기화
        self.update_data_timer = QTimer()
        self.update_data_timer.timeout.connect(self.periodic_update_data)
        self.update_data_timer.start(20000)  # 20초 - API 제한 고려

    def _check_api_limit_and_wait(self, operation_name="API 요청", rqtype=0):
        """API 제한 확인 및 대기 (요청 건수 기반)"""
        try:
            # 시세 조회 제한 상태 확인 (문서 기준 올바른 함수명 사용)
            remain_count = cpCybos.GetLimitRemainCount(rqtype)   # 시세 조회 남은 횟수 (0: 시세 조회)

            logging.debug(f"🔍 {operation_name} 전 API 상태 - 남은 요청: {remain_count}건")
            
            if remain_count > 0:
                # ✅ 요청 가능: 남은 건수가 있음
                if remain_count <= 5:  # 남은 요청이 5건 이하일 때 경고
                    logging.warning(f"⚠️ API 요청 한계 근접: 남은 요청 {remain_count}건")
                return True
            else:
                # ❌ 요청 불가: 남은 건수가 0건
                logging.warning(f"🚫 {operation_name} 거부: 요청 제한 (남은 요청: {remain_count}건)")
                return False
            
        except Exception as ex:
            logging.error(f"❌ API 제한 체크 실패: {ex}")
            return False


        # ===== ✅ 영업일은 start_timers()에서 설정 =====
        now = time.localtime()
        self.todayDate = now.tm_year * 10000 + now.tm_mon * 100 + now.tm_mday

    def get_strength(self, code):
        """체결강도 반환 (매수세 / 매도세 * 100)"""
        
        # 캐시 확인 (1초)
        if code in self.strength_cache:
            cached_strength, cached_time = self.strength_cache[code]
            if time.time() - cached_time < 1.0:
                return cached_strength
        
        with self.stockdata_lock:
            if code not in self.buy_volumes or len(self.buy_volumes[code]) < 3:
                return 100
            
            total_buy = sum(self.buy_volumes[code])
            total_sell = sum(self.sell_volumes[code])
            
            if total_sell > 0:
                strength = (total_buy / total_sell) * 100
            else:
                strength = 200
            
            # 캐시 저장
            self.strength_cache[code] = (strength, time.time())
            
            return strength

    def periodic_update_data(self):
        """주기적 데이터 업데이트 (수정 버전)"""
        try:
            current_time = time.time()
            with self.stockdata_lock:
                codes = list(self.stockdata.keys())
            
            for code in codes:
                if (code in self.trader.vistock_set and 
                    code not in self.trader.monistock_set and 
                    code not in self.trader.bought_set):
                    continue
                
                # ✅ 20개 종목 최적화: 빠른 업데이트 (보유: 15초, 모니터링: 30초)
                interval = 15 if code in self.trader.bought_set else 30

                last_time = self.last_update_time.get(code, 0)
                if current_time - last_time < interval:
                    continue

                with self.stockdata_lock:
                    if code not in self.stockdata:
                        logging.debug(f"{code}: stockdata에서 제거됨, 스킵")
                        continue
                    if self.is_updating.get(code, False):
                        logging.debug(f"{code}: 데이터 업데이트 진행 중, 스킵")
                        continue
                
                self.update_chart_data(code, self.interval, self.number)
                self.last_update_time[code] = current_time

                with self.stockdata_lock:
                    if code not in self.stockdata:
                        logging.debug(f"{code}: 업데이트 후 stockdata에서 제거됨, 스킵")
                        continue
                    if code not in self.objIndicators:
                        self.objIndicators[code] = CpIndicators(self.chart_type)
                        
                        # === 모든 지표 재계산 (새 지표 포함) ===
                        indicator_types = [
                            "MA", "MACD", "RSI", "STOCH", 
                            "ATR", "CCI", "BBANDS", "VWAP",
                            "WILLIAMS_R", "ROC", "OBV", "VOLUME_PROFILE"  # 추가
                        ]
                        
                        # ✅ 안전하게 지표 계산
                        for ind in indicator_types:
                            try:
                                result = self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                                if result:
                                    self.stockdata[code].update(result)
                            except Exception as ind_ex:
                                logging.debug(f"{code}: {ind} 계산 실패: {ind_ex}")
                        
                        self._update_snapshot(code)
                            
        except Exception as ex:
            logging.error(f"periodic_update_data -> {ex}")

    def select_code(self, code):
        try:
            if code in self.stockdata:
                return True
            
            self.stockdata[code] = {
                'D': [], 'T': [], 'O': [], 'H': [], 'L': [], 'C': [], 'V': [], 'TV': [], 
                'MAD5': [], 'MAD10': [], 'VWAP': []
            }
            
            # 상태 및 거래량 추적 초기화
            self.is_updating[code] = False
            self.buy_volumes[code] = deque(maxlen=10)
            self.sell_volumes[code] = deque(maxlen=10)
            self.update_chart_data(code, self.interval, self.number)
            self.is_initial_loaded[code] = False
            
            if code not in self.objIndicators:
                self.objIndicators[code] = CpIndicators(self.chart_type)
                result = self.objIndicators[code].make_indicator("MA", code, self.stockdata[code])

                if result:
                    self.stockdata[code].update(result)
                    self._update_snapshot(code)
                    return True
                else:
                    return False
            return True
        except Exception as ex:
            logging.error(f"select_code -> {ex}")
            return False

    def monitor_code(self, code):
        """종목 모니터링 시작"""
        try:
            if code in self.stockdata:
                return True
            
            # 데이터 구조 초기화
            self.stockdata[code] = {
                'D': [], 'T': [], 'O': [], 'H': [], 'L': [], 'C': [], 'V': [], 'TV': [],
                'MAT5': [], 'MAT20': [], 'MAT60': [], 'MAT120': [],
                'MAM5': [], 'MAM10': [], 'MAM20': [],
                'MACDT': [], 'MACDT_SIGNAL': [], 'OSCT': [],
                'MACD': [], 'MACD_SIGNAL': [], 'OSC': [],
                'RSIT': [], 'RSIT_SIGNAL': [],
                'RSI': [], 'RSI_SIGNAL': [],
                'STOCHK': [], 'STOCHD': [],
                'ATR': [], 'CCI': [],
                'BB_UPPER': [], 'BB_MIDDLE': [], 'BB_LOWER': [],
                'BB_POSITION': [], 'BB_BANDWIDTH': [],
                'VWAP': [],
                'WILLIAMS_R': [], 'ROC': [], 'OBV': [], 'OBV_MA20': [],
                'VP_POC': 0, 'VP_POSITION': 0,
                'TICKS': [],
                'MAT5_MAT20_DIFF': [], 'MAT20_MAT60_DIFF': [],
                'MAT60_MAT120_DIFF': [], 'C_MAT5_DIFF': [],
                'MAM5_MAM10_DIFF': [], 'MAM10_MAM20_DIFF': [],
                'C_MAM5_DIFF': [], 'C_ABOVE_MAM5': [],
                'MAT5_CHANGE': [], 'MAT20_CHANGE': [],
                'MAT60_CHANGE': [], 'MAT120_CHANGE': []
            }
            
            # 상태 및 거래량 추적 초기화
            self.is_updating[code] = False
            self.buy_volumes[code] = deque(maxlen=10)
            self.sell_volumes[code] = deque(maxlen=10)

            # ===== ✅ 모든 차트 타입: 충분한 과거 데이터 로드 =====
            success = self.update_chart_data_from_market_open(code)
            if not success:
                logging.warning(f"{code}: 과거 데이터 로드 실패, 일반 개수 기준 폴백")
                self.update_chart_data(code, self.interval, self.number)
                self.is_initial_loaded[code] = False
            else:
                self.is_initial_loaded[code] = True

            # 지표 계산 (데이터가 부족해도 실시간 구독은 시작)
            with self.stockdata_lock:
                if code not in self.objIndicators:
                    self.objIndicators[code] = CpIndicators(self.chart_type)
                    
                    indicator_types = [
                        "MA", "MACD", "RSI", "STOCH",
                        "ATR", "CCI", "BBANDS", "VWAP",
                        "WILLIAMS_R", "ROC", "OBV", "VOLUME_PROFILE"
                    ]
                    
                    # ✅ 데이터 길이 확인
                    data_length = len(self.stockdata[code].get('C', []))
                    
                    if data_length < 20:
                        # 데이터 부족 - 경고 1회만 출력
                        logging.info(f"⚠️ {code}: 초기 데이터 부족 ({data_length}개), 실시간 구독 시작 후 지표 계산 예정")
                        
                        # ✅ 실시간 구독만 시작 (지표는 나중에)
                        if code not in self.objCur:
                            self.objCur[code] = CpPBStockCur()
                            self.objCur[code].Subscribe(code, self)
                        
                        return True
                    
                    # 데이터 충분 - 지표 계산 시도
                    results = []
                    for ind in indicator_types:
                        try:
                            result = self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                            results.append(result)
                        except Exception as ind_ex:
                            logging.debug(f"{code}: {ind} 계산 실패: {ind_ex}")
                            results.append({})
                    
                    # ✅ 성공한 지표만 업데이트
                    for result in results:
                        if result:
                            self.stockdata[code].update(result)
                    
                    self._update_snapshot(code)
                    
                    # ✅ 실시간 구독 시작 (지표 계산 성공 여부와 무관)
                    if code not in self.objCur:
                        self.objCur[code] = CpPBStockCur()
                        self.objCur[code].Subscribe(code, self)
                    
                    return True
            
            return True
            
        except Exception as ex:
            logging.error(f"monitor_code({code}): {ex}")
            return False
        
    def monitor_stop(self, code):
        try:
            if self.is_updating.get(code, False):
                logging.debug(f"{code}: 데이터 업데이트 진행 중, 1초 후 재시도")
                QTimer.singleShot(1000, lambda: self.monitor_stop(code))
                return
            
            with self.stockdata_lock:
                # ===== ✅ COM 객체 Unsubscribe (스레드 에러 무시) =====
                if code in self.objCur:
                    try:
                        self.objCur[code].Unsubscribe()
                    except Exception as com_ex:
                        # COM 스레드 에러는 무시 (다른 스레드에서 호출 시 발생)
                        logging.debug(f"{code}: Unsubscribe COM 에러 무시 (스레드 충돌)")
                    
                    # Unsubscribe 실패해도 객체는 삭제
                    try:
                        del self.objCur[code]
                    except:
                        pass
                
                # ===== ✅ 나머지 데이터 정리 (안전) =====
                if code in self.stockdata:
                    del self.stockdata[code]
                if code in self.objIndicators:
                    del self.objIndicators[code]
                if code in self.is_updating:
                    del self.is_updating[code]
                if code in self.is_initial_loaded:
                    del self.is_initial_loaded[code]
                if code in self.last_indicator_update:
                    del self.last_indicator_update[code]
                if code in self.latest_snapshot:
                    del self.latest_snapshot[code]
                if code in self.buy_volumes:
                    del self.buy_volumes[code]
                if code in self.sell_volumes:
                    del self.sell_volumes[code]
                if code in self.strength_cache:
                    del self.strength_cache[code]
                    
        except Exception as ex:
            logging.debug(f"{code}: monitor_stop 정리 중 에러 (무시): {ex}")
            return False

    def _request_chart_data(self, code, request_type='count', count=None, start_date=None, end_date=None):
        """공통 차트 데이터 요청 로직"""
        try:
            # ===== ✅ 대신증권 API 제한만 확인 =====
            if not self._check_api_limit_and_wait("차트 데이터 요청", 0):
                logging.warning(f"❌ {code}: API 제한으로 차트 데이터 요청 거부")
                return False
            
            objRq = win32com.client.Dispatch("CpSysDib.StockChart")
            objRq.SetInputValue(0, code)
            
            if request_type == 'count':
                objRq.SetInputValue(1, ord('2'))
                objRq.SetInputValue(4, count)
            elif request_type == 'period':
                objRq.SetInputValue(1, ord('1'))
                objRq.SetInputValue(2, end_date)
                objRq.SetInputValue(3, start_date)
            else:
                logging.error(f"잘못된 request_type: {request_type}")
                return None
            
            objRq.SetInputValue(5, [0, 1, 2, 3, 4, 5, 8, 9, 13])
            objRq.SetInputValue(6, ord(self.chart_type))
            objRq.SetInputValue(7, self.interval)
            objRq.SetInputValue(9, ord('1'))
            objRq.BlockRequest2(1)
            
            rqStatus = objRq.GetDibStatus()
            if rqStatus != 0:
                rqRet = objRq.GetDibMsg1()
                logging.warning(f"{code} 데이터 조회 실패, {rqStatus}, {rqRet}")
                return None
            
            len_data = objRq.GetHeaderValue(3)
            lastCount = objRq.GetHeaderValue(4)
            
            new_data = {
                'D': [], 'T': [], 'O': [], 'H': [], 'L': [], 'C': [], 'V': [], 'TV': [], 'TICKS': []
            }
            
            for i in range(len_data):
                new_data['D'].append(objRq.GetDataValue(0, i))
                new_data['T'].append(objRq.GetDataValue(1, i))
                new_data['O'].append(objRq.GetDataValue(2, i))
                new_data['H'].append(objRq.GetDataValue(3, i))
                new_data['L'].append(objRq.GetDataValue(4, i))
                new_data['C'].append(objRq.GetDataValue(5, i))
                new_data['V'].append(objRq.GetDataValue(6, i))
                new_data['TV'].append(objRq.GetDataValue(7, i))
                
                if self.chart_type == 'T':
                    if i == (len_data - 1):
                        new_data['TICKS'].append(lastCount)
                    else:
                        new_data['TICKS'].append(self.interval)
            
            for key in ['D', 'T', 'O', 'H', 'L', 'C', 'V', 'TV']:
                new_data[key].reverse()
            if self.chart_type == 'T':
                new_data['TICKS'].reverse()
            
            return new_data
            
        except Exception as ex:
            logging.error(f"_request_chart_data -> {code}, {ex}")
            return None
        
    def update_chart_data(self, code, interval, number):
        """실시간 증분 업데이트 (주기적 호출)"""
        try:
            self.is_updating[code] = True
            
            new_data = self._request_chart_data(code, request_type='count', count=number)
            
            if new_data is None:
                self.is_updating[code] = False
                return False
            
            with self.stockdata_lock:
                if code not in self.stockdata:
                    logging.debug(f"{code}: stockdata에 없음, 업데이트 중단")
                    self.is_updating[code] = False
                    return False
                
                if self.is_initial_loaded.get(code, False):
                    if self.stockdata[code].get('T') and len(self.stockdata[code]['T']) > 0:
                        last_time = self.stockdata[code]['T'][-1]
                        last_date = self.stockdata[code]['D'][-1]
                        
                        # ===== ✅ 안전한 인덱스 생성 (배열 길이 확인) =====
                        max_len = min(len(new_data.get('T', [])), len(new_data.get('D', [])))
                        if max_len == 0:
                            logging.debug(f"{code}: 업데이트할 데이터 없음")
                            self.is_updating[code] = False
                            return False
                        
                        new_indices = [
                            i for i in range(max_len)
                            if (new_data['D'][i] > last_date) or 
                               (new_data['D'][i] == last_date and new_data['T'][i] > last_time)
                        ]
                        
                        if new_indices:
                            for key in new_data:
                                # ===== ✅ 인덱스 범위 확인 =====
                                arr = new_data[key]
                                if not isinstance(arr, list) or len(arr) == 0:
                                    continue
                                
                                # 유효한 인덱스만 필터링
                                valid_indices = [i for i in new_indices if i < len(arr)]
                                if valid_indices:
                                    filtered_data = [arr[i] for i in valid_indices]
                                    self.stockdata[code][key].extend(filtered_data)
                            
                            if self.chart_type == 'T':
                                max_length = 300
                            elif self.chart_type == 'm':
                                max_length = 150
                            elif self.chart_type == 'D':
                                max_length = 80
                            else:
                                max_length = 50
                            
                            for key in self.stockdata[code]:
                                if isinstance(self.stockdata[code][key], list):
                                    self.stockdata[code][key] = self.stockdata[code][key][-max_length:]
                    else:
                        for key in new_data:
                            self.stockdata[code][key] = new_data[key]
                else:
                    for key in new_data:
                        self.stockdata[code][key] = new_data[key]
            
            self.is_updating[code] = False
            return True
            
        except Exception as ex:
            logging.error(f"update_chart_data -> {ex}")
            self.is_updating[code] = False
            return False

    def update_chart_data_from_market_open(self, code):
        """충분한 과거 데이터 로드 (현시점 기준으로 지정된 개수만큼)
        
        ✅ 개선: 모든 차트 타입(틱봉/분봉/일봉)에서 과거 데이터 포함하여 충분한 데이터 확보
        """
        try:
            self.is_updating[code] = True
            
            # ===== ✅ 개수 기준으로 충분한 데이터 로드 =====
            chart_type_name = {'T': '틱봉', 'm': '분봉', 'D': '일봉'}.get(self.chart_type, self.chart_type)
            logging.debug(f"{code}: {self.number}개 {chart_type_name} 데이터 로드 시도")
            
            # 개수 기준으로 데이터 조회 (과거 데이터 포함)
            new_data = self._request_chart_data(
                code,
                request_type='count',
                count=self.number
            )
            
            if new_data is None:
                logging.warning(f"{code}: API 조회 실패")
                self.is_updating[code] = False
                return False
            
            # 데이터 확인
            if len(new_data.get('D', [])) == 0:
                logging.warning(f"{code}: 데이터 없음")
                self.is_updating[code] = False
                return False
            
            # 데이터 적용
            with self.stockdata_lock:
                if code not in self.stockdata:
                    logging.debug(f"{code}: stockdata에 없음, 중단")
                    self.is_updating[code] = False
                    return False
                
                for key in new_data:
                    self.stockdata[code][key] = new_data[key]
            
            data_count = len(new_data['D'])
            chart_type_name = {'T': '틱봉', 'm': '분봉', 'D': '일봉'}.get(self.chart_type, self.chart_type)
            logging.debug(
                f"✅ {code}: {data_count}개 {chart_type_name} 데이터 로드 완료 "
                f"(요청: {self.number}개)"
            )
            
            # ===== ✅ 데이터 충분성 확인 =====
            if data_count < self.number * 0.6:  # 60% 미만이면 경고 (기준 완화)
                logging.warning(
                    f"⚠️ {code}: {chart_type_name} 데이터 부족 "
                    f"(로드: {data_count}개, 요청: {self.number}개, 부족률: {(1-data_count/self.number)*100:.1f}%)"
                )
            else:
                logging.info(
                    f"✅ {code}: {chart_type_name} 데이터 충분 "
                    f"(로드: {data_count}개, 요청: {self.number}개)"
                )
            
            self.is_updating[code] = False
            return True
            
        except Exception as ex:
            logging.error(f"update_chart_data_from_market_open({code}): {ex}\n{traceback.format_exc()}")
            self.is_updating[code] = False
            return False

    def update_chart_data_from_today_only(self, code):
        """일봉용: 당일 영업일 데이터만 로드"""
        try:
            self.is_updating[code] = True
            
            # ===== ✅ 일봉은 당일 영업일만 로드 =====
            date_str = str(self.todayDate)
            formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            
            logging.debug(f"{code}: {formatted_date} 일봉 데이터 로드 시도")
            
            # 데이터 조회
            new_data = self._request_chart_data(
                code,
                request_type='period',
                start_date=date_str,
                end_date=date_str
            )
            
            if new_data is None:
                logging.warning(f"{code}: 일봉 API 조회 실패")
                self.is_updating[code] = False
                return False
            
            # 데이터 확인
            if len(new_data.get('D', [])) == 0:
                logging.warning(f"{code}: {formatted_date} 일봉 데이터 없음")
                self.is_updating[code] = False
                return False
            
            # 데이터 적용
            with self.stockdata_lock:
                if code not in self.stockdata:
                    logging.debug(f"{code}: stockdata에 없음, 중단")
                    self.is_updating[code] = False
                    return False
                
                for key in new_data:
                    self.stockdata[code][key] = new_data[key]
            
            logging.debug(
                f"✅ {code}: {formatted_date} 일봉 데이터 로드 완료 "
                f"({len(new_data['D'])}개)"
            )
            
            self.is_updating[code] = False
            return True
            
        except Exception as ex:
            logging.error(f"update_chart_data_from_today_only({code}): {ex}\n{traceback.format_exc()}")
            self.is_updating[code] = False
            return False
        
    def verify_data_coverage(self, code):
        """데이터가 장 시작부터 커버하는지 확인"""
        try:
            with self.stockdata_lock:
                if code not in self.stockdata:
                    return False
                
                stock_data = self.stockdata[code]
                if not stock_data.get('T') or len(stock_data['T']) == 0:
                    return False
                
                first_time = stock_data['T'][0]
                market_open = 900
                
                if first_time > market_open + 30:
                    logging.warning(
                        f"{code}: {self.chart_type} 데이터가 {first_time}부터 시작 "
                        f"(장 시작 데이터 부족, 총 {len(stock_data['T'])}개)"
                    )
                    return False
                
                logging.info(
                    f"{code}: {self.chart_type} 데이터 커버리지 양호 "
                    f"({first_time}부터 시작, 총 {len(stock_data['T'])}개)"
                )
                return True
                
        except Exception as ex:
            logging.error(f"verify_data_coverage -> {ex}")
            return False

    def _safe_get_last(self, data, key, default=0):
        """안전하게 마지막 값 추출 (리스트/스칼라 자동 처리)"""
        try:
            value = data.get(key)
            if value is None:
                return default
            
            # 리스트인 경우
            if isinstance(value, (list, tuple)):
                if len(value) == 0:
                    return default
                return value[-1]
            
            # 스칼라인 경우 (int, float, bool 등)
            if isinstance(value, (int, float, bool, str)):
                return value
            
            # 기타 (dict 등)
            return default
            
        except Exception as ex:
            logging.debug(f"_safe_get_last({key}): {ex}")
            return default
    
    def _safe_get_recent(self, data, key, count=3, default_list=None):
        """안전하게 최근 N개 값 추출 (리스트만 처리)"""
        try:
            value = data.get(key)
            if value is None:
                return default_list or [0] * count
            
            # 리스트인 경우
            if isinstance(value, (list, tuple)):
                if len(value) == 0:
                    return default_list or [0] * count
                elif len(value) >= count:
                    return list(value[-count:])
                else:
                    # 부족한 경우 앞을 0으로 채움
                    padding = [0] * (count - len(value))
                    return padding + list(value)
            
            # 스칼라인 경우 복제
            if isinstance(value, (int, float)):
                return [value] * count
            
            # 기타 (dict, str 등)
            return default_list or [0] * count
            
        except Exception as ex:
            logging.debug(f"_safe_get_recent({key}): {ex}")
            return default_list or [0] * count

    def _update_snapshot(self, code):
        """읽기 전용 스냅샷 업데이트 (안전 버전)"""
        try:
            if code not in self.stockdata:
                return
            
            data = self.stockdata[code]
            
            # ✅ 기본 데이터가 없으면 스킵
            if not data.get('C') or len(data.get('C', [])) == 0:
                return
            
            if self.chart_type == 'T':
                self.latest_snapshot[code] = {
                    # 기본 가격
                    'C': self._safe_get_last(data, 'C', 0),
                    'O': self._safe_get_last(data, 'O', 0),
                    'H': self._safe_get_last(data, 'H', 0),
                    'L': self._safe_get_last(data, 'L', 0),
                    'V': self._safe_get_last(data, 'V', 0),
                    
                    # 이동평균
                    'MAT5': self._safe_get_last(data, 'MAT5', 0),
                    'MAT20': self._safe_get_last(data, 'MAT20', 0),
                    'MAT60': self._safe_get_last(data, 'MAT60', 0),
                    'MAT120': self._safe_get_last(data, 'MAT120', 0),
                    
                    # RSI
                    'RSIT': self._safe_get_last(data, 'RSIT', 0),
                    'RSIT_SIGNAL': self._safe_get_last(data, 'RSIT_SIGNAL', 0),
                    
                    # MACD
                    'MACDT': self._safe_get_last(data, 'MACDT', 0),
                    'MACDT_SIGNAL': self._safe_get_last(data, 'MACDT_SIGNAL', 0),
                    'OSCT': self._safe_get_last(data, 'OSCT', 0),
                    
                    # Stochastic
                    'STOCHK': self._safe_get_last(data, 'STOCHK', 0),
                    'STOCHD': self._safe_get_last(data, 'STOCHD', 0),
                    
                    # 기타
                    'ATR': self._safe_get_last(data, 'ATR', 0),
                    'CCI': self._safe_get_last(data, 'CCI', 0),
                    'BB_UPPER': self._safe_get_last(data, 'BB_UPPER', 0),
                    'BB_MIDDLE': self._safe_get_last(data, 'BB_MIDDLE', 0),
                    'BB_LOWER': self._safe_get_last(data, 'BB_LOWER', 0),
                    'BB_POSITION': self._safe_get_last(data, 'BB_POSITION', 0),
                    'BB_BANDWIDTH': self._safe_get_last(data, 'BB_BANDWIDTH', 0),
                    'VWAP': self._safe_get_last(data, 'VWAP', 0),
                    
                    # === 새로운 지표들 ===
                    'WILLIAMS_R': self._safe_get_last(data, 'WILLIAMS_R', -50),
                    'ROC': self._safe_get_last(data, 'ROC', 0),
                    'OBV': self._safe_get_last(data, 'OBV', 0),
                    'OBV_MA20': self._safe_get_last(data, 'OBV_MA20', 0),
                    'VP_POC': self._safe_get_last(data, 'VP_POC', 0),
                    'VP_POSITION': self._safe_get_last(data, 'VP_POSITION', 0),
                    
                    # 최근 추이
                    'C_recent': self._safe_get_recent(data, 'C', 3, [0, 0, 0]),
                    'H_recent': self._safe_get_recent(data, 'H', 3, [0, 0, 0]),
                    'L_recent': self._safe_get_recent(data, 'L', 3, [0, 0, 0]),
                }
            
            elif self.chart_type == 'm':
                self.latest_snapshot[code] = {
                    # 기본 가격
                    'C': self._safe_get_last(data, 'C', 0),
                    'O': self._safe_get_last(data, 'O', 0),
                    'H': self._safe_get_last(data, 'H', 0),
                    'L': self._safe_get_last(data, 'L', 0),
                    'V': self._safe_get_last(data, 'V', 0),
                    
                    # 이동평균
                    'MAM5': self._safe_get_last(data, 'MAM5', 0),
                    'MAM10': self._safe_get_last(data, 'MAM10', 0),
                    'MAM20': self._safe_get_last(data, 'MAM20', 0),
                    
                    # RSI
                    'RSI': self._safe_get_last(data, 'RSI', 0),
                    'RSI_SIGNAL': self._safe_get_last(data, 'RSI_SIGNAL', 0),
                    
                    # MACD
                    'MACD': self._safe_get_last(data, 'MACD', 0),
                    'MACD_SIGNAL': self._safe_get_last(data, 'MACD_SIGNAL', 0),
                    'OSC': self._safe_get_last(data, 'OSC', 0),
                    
                    # Stochastic
                    'STOCHK': self._safe_get_last(data, 'STOCHK', 0),
                    'STOCHD': self._safe_get_last(data, 'STOCHD', 0),
                    
                    # 기타
                    'CCI': self._safe_get_last(data, 'CCI', 0),
                    'VWAP': self._safe_get_last(data, 'VWAP', 0),
                    
                    # === 새로운 지표들 ===
                    'WILLIAMS_R': self._safe_get_last(data, 'WILLIAMS_R', -50),
                    'ROC': self._safe_get_last(data, 'ROC', 0),
                    'OBV': self._safe_get_last(data, 'OBV', 0),
                    'OBV_MA20': self._safe_get_last(data, 'OBV_MA20', 0),
                    'VP_POC': self._safe_get_last(data, 'VP_POC', 0),
                    'VP_POSITION': self._safe_get_last(data, 'VP_POSITION', 0),
                    
                    # 최근 추이
                    'C_recent': self._safe_get_recent(data, 'C', 2, [0, 0]),
                    'O_recent': self._safe_get_recent(data, 'O', 2, [0, 0]),
                    'H_recent': self._safe_get_recent(data, 'H', 2, [0, 0]),
                    'L_recent': self._safe_get_recent(data, 'L', 2, [0, 0]),
                }
            
            elif self.chart_type == 'D':
                self.latest_snapshot[code] = {
                    'C': self._safe_get_last(data, 'C', 0),
                    'V': self._safe_get_last(data, 'V', 0),
                    'MAD5': self._safe_get_last(data, 'MAD5', 0),
                    'MAD10': self._safe_get_last(data, 'MAD10', 0),
                    'VWAP': self._safe_get_last(data, 'VWAP', 0),
                }
                
        except Exception as ex:
            logging.error(f"_update_snapshot -> {code}, {ex}\n{traceback.format_exc()}")

    def get_latest_data(self, code):
        """빠른 읽기 - 스냅샷 반환 (락 불필요)"""
        return self.latest_snapshot.get(code, {})
    
    def get_full_data(self, code):
        """전체 데이터 읽기 (락 필요, 차트 그리기용)"""
        with self.stockdata_lock:
            return copy.deepcopy(self.stockdata.get(code, {}))
    
    def get_recent_data(self, code, count=10):
        """최근 N개 데이터 읽기 (락 필요)"""
        with self.stockdata_lock:
            if code not in self.stockdata:
                return {}
            
            data = self.stockdata[code]
            result = {}
            for key, values in data.items():
                if isinstance(values, list) and len(values) > 0:
                    result[key] = values[-count:] if len(values) >= count else values
                else:
                    result[key] = values
            return result

    def updateCurData(self, item):
        """실시간 체결 데이터 업데이트"""
        try:
            code = item.get('code')
            if not code:
                return
            
            if self.is_updating.get(code, False):
                return
            
            # ✅ stockdata에 종목이 없으면 스킵
            if code not in self.stockdata:
                return
            
            time_val = item.get('time', 0)
            cur = item.get('cur', 0)
            vol = item.get('vol', 0)
            
            # ✅ 유효한 데이터인지 확인
            if cur <= 0 or time_val <= 0:
                return
            
            current_time = time.time()
            
            # 체결강도 업데이트
            with self.stockdata_lock:
                if code in self.buy_volumes:
                    # ✅ 안전하게 이전 가격 가져오기
                    c_data = self.stockdata.get(code, {}).get('C')
                    if c_data and isinstance(c_data, list) and len(c_data) > 0:
                        prev_price = c_data[-1]
                        if cur > prev_price:
                            self.buy_volumes[code].append(vol)
                            self.sell_volumes[code].append(0)
                        elif cur < prev_price:
                            self.buy_volumes[code].append(0)
                            self.sell_volumes[code].append(vol)
                        else:
                            self.buy_volumes[code].append(vol / 2)
                            self.sell_volumes[code].append(vol / 2)

            with self.stockdata_lock:
                bar_completed = False
                
                if self.chart_type == 'T':
                    hh, mm = divmod(time_val, 10000)
                    mm, tt = divmod(mm, 100)
                    if mm == 60:
                        hh += 1
                        mm = 0
                    lCurTime = hh * 100 + mm

                    bFind = False
                    if lCurTime > self.LASTTIME:
                        lCurTime = self.LASTTIME

                    if code in self.stockdata:
                        if len(self.stockdata[code]['T']) > 0:
                            lastCount = self.stockdata[code]['TICKS'][-1]
                            if 1 <= lastCount < self.interval:
                                bFind = True
                                self.stockdata[code]['T'][-1] = lCurTime
                                self.stockdata[code]['C'][-1] = cur
                                if self.stockdata[code]['H'][-1] < cur:
                                    self.stockdata[code]['H'][-1] = cur
                                if self.stockdata[code]['L'][-1] > cur:
                                    self.stockdata[code]['L'][-1] = cur
                                self.stockdata[code]['V'][-1] += vol
                                self.stockdata[code]['TICKS'][-1] += 1

                        if not bFind:
                            bar_completed = True
                            
                            self.stockdata[code]['D'].append(self.todayDate)
                            self.stockdata[code]['T'].append(lCurTime)
                            self.stockdata[code]['O'].append(cur)
                            self.stockdata[code]['H'].append(cur)
                            self.stockdata[code]['L'].append(cur)
                            self.stockdata[code]['C'].append(cur)
                            self.stockdata[code]['V'].append(vol)
                            self.stockdata[code]['TICKS'].append(1)

                        desired_length = 300
                        for key in self.stockdata[code]:
                            # ✅ 리스트인 경우에만 슬라이싱
                            if isinstance(self.stockdata[code][key], list):
                                self.stockdata[code][key] = self.stockdata[code][key][-desired_length:]

                        self._update_snapshot(code)
                        
                        # ✅ 실시간 데이터 업데이트 시그널 발생
                        self.data_updated.emit(code)

                        last_update = self.last_indicator_update.get(code, 0)
                        if current_time - last_update >= self.indicator_update_interval:
                            if code in self.objIndicators:
                                # ✅ 안전하게 지표 업데이트
                                indicator_types = ["MA", "RSI", "MACD", "STOCH", "ATR", "CCI", "BBANDS", "VWAP",
                                                 "WILLIAMS_R", "ROC", "OBV", "VOLUME_PROFILE"]
                                for ind in indicator_types:
                                    try:
                                        result = self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                                        if result:
                                            self.stockdata[code].update(result)
                                    except Exception as ind_ex:
                                        logging.debug(f"{code}: {ind} 업데이트 실패: {ind_ex}")
                                
                                self._update_snapshot(code)
                            
                            self.last_indicator_update[code] = current_time
                    
                elif self.chart_type == 'm':
                    hh, mm = divmod(time_val, 10000)
                    mm, tt = divmod(mm, 100)
                    convertedMintime = hh * 60 + mm

                    bFind = False
                    a, b = divmod(convertedMintime, self.interval)
                    intervaltime = a * self.interval
                    lChartTime = intervaltime + self.interval
                    hour, minute = divmod(lChartTime, 60)
                    lCurTime = hour * 100 + minute

                    if lCurTime > self.LASTTIME:
                        lCurTime = self.LASTTIME

                    if code in self.stockdata:
                        if len(self.stockdata[code]['T']) > 0:
                            lLastTime = self.stockdata[code]['T'][-1]
                            if lLastTime == lCurTime:
                                bFind = True
                                self.stockdata[code]['C'][-1] = cur
                                if self.stockdata[code]['H'][-1] < cur:
                                    self.stockdata[code]['H'][-1] = cur
                                if self.stockdata[code]['L'][-1] > cur:
                                    self.stockdata[code]['L'][-1] = cur
                                self.stockdata[code]['V'][-1] += vol
        
                        if not bFind:
                            # ✅ 새 봉 생성 = 완성 이벤트
                            bar_completed = True
                            
                            self.stockdata[code]['D'].append(self.todayDate)
                            self.stockdata[code]['T'].append(lCurTime)
                            self.stockdata[code]['O'].append(cur)
                            self.stockdata[code]['H'].append(cur)
                            self.stockdata[code]['L'].append(cur)
                            self.stockdata[code]['C'].append(cur)
                            self.stockdata[code]['V'].append(vol)

                        desired_length = 150
                        for key in self.stockdata[code]:
                            # ✅ 리스트인 경우에만 슬라이싱
                            if isinstance(self.stockdata[code][key], list):
                                self.stockdata[code][key] = self.stockdata[code][key][-desired_length:]

                        self._update_snapshot(code)
                        
                        # ✅ 실시간 데이터 업데이트 시그널 발생
                        self.data_updated.emit(code)

                        last_update = self.last_indicator_update.get(code, 0)
                        if current_time - last_update >= self.indicator_update_interval:
                            if code in self.objIndicators:
                                # ✅ 안전하게 지표 업데이트
                                indicator_types = ["MA", "MACD", "RSI", "STOCH", "ATR", "CCI", "BBANDS", "VWAP"]
                                for ind in indicator_types:
                                    try:
                                        result = self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                                        if result:
                                            self.stockdata[code].update(result)
                                    except Exception as ind_ex:
                                        logging.debug(f"{code}: {ind} 업데이트 실패: {ind_ex}")
                                
                                self._update_snapshot(code)
                            
                            self.last_indicator_update[code] = current_time
                
                elif self.chart_type == 'D':
                    # 일봉은 당일 데이터만 업데이트 (실시간 OHLC 업데이트)
                    if code in self.stockdata and len(self.stockdata[code]['T']) > 0:
                        # 현재 봉의 OHLC 업데이트
                        self.stockdata[code]['C'][-1] = cur
                        if self.stockdata[code]['H'][-1] < cur:
                            self.stockdata[code]['H'][-1] = cur
                        if self.stockdata[code]['L'][-1] > cur:
                            self.stockdata[code]['L'][-1] = cur
                        self.stockdata[code]['V'][-1] += vol
                        
                        # 스냅샷 업데이트
                        self._update_snapshot(code)
                        
                        # ✅ 실시간 데이터 업데이트 시그널 발생
                        self.data_updated.emit(code)
                        
                        # 지표 업데이트 (1초 간격)
                        last_update = self.last_indicator_update.get(code, 0)
                        if current_time - last_update >= self.indicator_update_interval:
                            if code in self.objIndicators:
                                indicator_types = ["MA", "RSI", "MACD", "STOCH", "ATR", "CCI", "BBANDS", "VWAP",
                                                 "WILLIAMS_R", "ROC", "OBV", "VOLUME_PROFILE"]
                                for ind in indicator_types:
                                    try:
                                        result = self.objIndicators[code].make_indicator(ind, code, self.stockdata[code])
                                        if result:
                                            self.stockdata[code].update(result)
                                    except Exception as ind_ex:
                                        logging.debug(f"{code}: {ind} 업데이트 실패: {ind_ex}")
                                
                                self._update_snapshot(code)
                            
                            self.last_indicator_update[code] = current_time
                
                # ✅ 새 봉 완성 시 signal 발생
                if bar_completed:
                    self.new_bar_completed.emit(code)
        
        except Exception as ex:
            logging.error(f"updateCurData -> {ex}\n{traceback.format_exc()}")

# ==================== CTrader (계속) ====================
class CTrader(QObject):
    """트레이더 클래스 (DatabaseWorker 제거, combined_tick_data 단일 사용)"""
    
    stock_added_to_monitor = pyqtSignal(str)
    stock_bought = pyqtSignal(str)
    stock_sold = pyqtSignal(str)

    def __init__(self, obj_cp_trade, cp_balance, cpCodeMgr, cp_cash, cp_order, cp_stock, buycount, window):
        super().__init__()
        self.cpTrade = obj_cp_trade
        self.cpBalance = cp_balance
        self.cpCodeMgr = cpCodeMgr
        self.cpCash = cp_cash
        self.cpOrder = cp_order
        self.cpStock = cp_stock
        self.target_buy_count = buycount

        self.bought_set = set()
        self.database_set = set()
        self.monistock_set = set()
        self.vistock_set = set()
        self.sell_half_set = set()
        self.buyorder_set = set()
        self.buyordering_set = set()
        self.sellorder_set = set()
        
        self.starting_price = {}
        self.starting_time = {}
        self.highest_price = {}
        self.buy_price = {}
        self.buy_qty = {}
        self.buyorder_qty = {}
        self.sell_half_qty = {}
        self.balance_qty = {}
        self.sell_amount = {}
        self.buy_amount = 0
        self.conclusion = CpPBConclusion()
        self.conclusion.Subscribe('', self)
        self.window = window

        self.cp_request = CpRequest()

        self.daydata = CpData(1, 'D', 80, self)
        self.mindata = CpData(3, 'm', 150, self)
        self.tickdata = CpData(30, 'T', 300, self)

        self.db_name = 'vi_stock_data.db'

        # ===== 설정 파일 읽기 (간소화) =====
        config = configparser.RawConfigParser()
        if os.path.exists('settings.ini'):
            config.read('settings.ini', encoding='utf-8')
        
        # combined_tick_data 저장 주기만 사용
        self.combined_save_interval = config.getint('DATA_SAVING', 'interval_seconds', fallback=5)
        
        logging.debug(f"데이터 저장 설정: combined_tick_data 간격={self.combined_save_interval}초")

    def init_database(self):
        """데이터베이스 초기화 (combined_tick_data 단일 테이블)"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            # ===== combined_tick_data (백테스팅용 메인 테이블) =====
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS combined_tick_data (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT NOT NULL,
                    timestamp DATETIME NOT NULL,
                    date TEXT NOT NULL,
                    time TEXT NOT NULL,
                    
                    -- 틱 데이터 (현재 시점)
                    tick_C REAL, tick_O REAL, tick_H REAL, tick_L REAL, tick_V INTEGER,
                    tick_MAT5 REAL, tick_MAT20 REAL, tick_MAT60 REAL, tick_MAT120 REAL,
                    tick_RSIT REAL, tick_RSIT_SIGNAL REAL,
                    tick_MACDT REAL, tick_MACDT_SIGNAL REAL, tick_OSCT REAL,
                    tick_STOCHK REAL, tick_STOCHD REAL,
                    tick_ATR REAL, tick_CCI REAL,
                    tick_BB_UPPER REAL, tick_BB_MIDDLE REAL, tick_BB_LOWER REAL,
                    tick_BB_POSITION REAL, tick_BB_BANDWIDTH REAL,
                    tick_VWAP REAL,
                    
                    -- === 새 지표: 틱 ===
                    tick_WILLIAMS_R REAL,
                    tick_ROC REAL,
                    tick_OBV REAL,
                    tick_OBV_MA20 REAL,
                    tick_VP_POC REAL,
                    tick_VP_POSITION REAL,
                    
                    -- 분봉 데이터 (가장 최근 완성된 분봉)
                    min_C REAL, min_O REAL, min_H REAL, min_L REAL, min_V INTEGER,
                    min_MAM5 REAL, min_MAM10 REAL, min_MAM20 REAL,
                    min_RSI REAL, min_RSI_SIGNAL REAL,
                    min_MACD REAL, min_MACD_SIGNAL REAL, min_OSC REAL,
                    min_STOCHK REAL, min_STOCHD REAL,
                    min_CCI REAL, min_VWAP REAL,
                    
                    -- === 새 지표: 분봉 ===
                    min_WILLIAMS_R REAL,
                    min_ROC REAL,
                    min_OBV REAL,
                    min_OBV_MA20 REAL,
                    
                    -- 추가 정보
                    strength REAL,
                    buy_price REAL,
                    position_type TEXT,
                    save_reason TEXT,
                    
                    UNIQUE(code, timestamp)
                )
            ''')
            
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_combined_code ON combined_tick_data(code)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_combined_date ON combined_tick_data(date)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_combined_timestamp ON combined_tick_data(timestamp)')
            
            # ===== 오래된 데이터 자동 정리 (30일 이상 된 데이터 삭제) =====
            try:
                from datetime import datetime, timedelta
                cutoff_date = (datetime.now() - timedelta(days=30)).strftime('%Y%m%d')
                
                # combined_tick_data 정리
                cursor.execute('DELETE FROM combined_tick_data WHERE date < ?', (cutoff_date,))
                deleted_rows = cursor.rowcount
                if deleted_rows > 0:
                    logging.info(f"🗑️ 오래된 데이터 정리: combined_tick_data {deleted_rows}개 레코드 삭제 (30일 이전)")
                
                # trades 정리 (선택적 - 거래 기록은 보관할 수도 있음)
                # cursor.execute('DELETE FROM trades WHERE date < ?', (cutoff_date,))
                
                # ===== 트랜잭션 커밋 후 VACUUM 실행 =====
                conn.commit()
                
                # VACUUM은 트랜잭션 외부에서 실행해야 함
                cursor.execute('VACUUM')
                logging.info(f"✅ DB 최적화 완료 (VACUUM)")
                
            except Exception as ex:
                logging.warning(f"오래된 데이터 정리 중 오류 (무시): {ex}")
            
            # ===== trades 테이블 (실거래 기록) =====
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS trades (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    code TEXT,
                    stock_name TEXT,
                    date TEXT,
                    time TEXT,
                    action TEXT,
                    price REAL,
                    quantity INTEGER,
                    amount REAL,
                    strategy TEXT,
                    buy_reason TEXT,
                    sell_reason TEXT,
                    buy_price REAL,
                    profit REAL,
                    profit_pct REAL,
                    hold_minutes REAL
                )
            ''')
            
            # ===== daily_summary 테이블 (일별 요약) =====
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS daily_summary (
                    date TEXT PRIMARY KEY,
                    strategy TEXT,
                    total_trades INTEGER DEFAULT 0,
                    win_trades INTEGER DEFAULT 0,
                    lose_trades INTEGER DEFAULT 0,
                    win_rate REAL DEFAULT 0,
                    total_profit REAL DEFAULT 0,
                    total_return_pct REAL DEFAULT 0,
                    avg_profit_pct REAL DEFAULT 0,
                    max_profit_pct REAL DEFAULT 0,
                    max_loss_pct REAL DEFAULT 0,
                    total_buy_amount REAL DEFAULT 0,
                    final_cash REAL DEFAULT 0,
                    portfolio_value REAL DEFAULT 0,
                    cash REAL DEFAULT 0,
                    holdings_value REAL DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # ===== backtest_results 테이블 (백테스팅 결과) =====
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS backtest_results (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    strategy TEXT NOT NULL,
                    start_date TEXT NOT NULL,
                    end_date TEXT NOT NULL,
                    initial_cash REAL NOT NULL,
                    final_cash REAL NOT NULL,
                    total_profit REAL NOT NULL,
                    total_return_pct REAL NOT NULL,
                    total_trades INTEGER NOT NULL,
                    win_trades INTEGER NOT NULL,
                    lose_trades INTEGER NOT NULL,
                    win_rate REAL NOT NULL,
                    avg_profit_pct REAL NOT NULL,
                    max_profit_pct REAL NOT NULL,
                    max_loss_pct REAL NOT NULL,
                    mdd REAL NOT NULL,
                    sharpe_ratio REAL,
                    avg_hold_minutes REAL,
                    parameters TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_backtest_strategy ON backtest_results(strategy)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_backtest_date ON backtest_results(start_date, end_date)')
            
            conn.commit()
            conn.close()
            
            logging.info("데이터베이스 초기화 완료 (새 지표 포함)")
            
        except Exception as ex:
            logging.error(f"init_database -> {ex}\n{traceback.format_exc()}")
            raise

    def save_trade_record(self, code, action, price, quantity, **kwargs):
        """실거래 기록 저장"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            now = datetime.now()
            date = now.strftime('%Y%m%d')
            time_str = now.strftime('%H%M%S')
            
            stock_name = cpCodeMgr.CodeToName(code)
            amount = price * quantity
            
            # 전략 이름
            strategy = kwargs.get('strategy', '')
            if not strategy and hasattr(self, 'window'):
                strategy = self.window.comboStg.currentText()
            
            cursor.execute('''
                INSERT INTO trades (
                    code, stock_name, date, time, action, price, quantity, amount,
                    strategy, buy_reason, sell_reason, buy_price, profit, profit_pct, hold_minutes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                code, 
                stock_name, 
                date, 
                time_str, 
                action, 
                price, 
                quantity, 
                amount,
                strategy,
                kwargs.get('buy_reason', ''),
                kwargs.get('sell_reason', ''),
                kwargs.get('buy_price', None),
                kwargs.get('profit', None),
                kwargs.get('profit_pct', None),
                kwargs.get('hold_minutes', None)
            ))
            
            conn.commit()
            conn.close()
            
            logging.debug(f"거래 기록 저장: {stock_name}({code}) {action} {quantity}주 @{price:,}원")
            
        except Exception as ex:
            logging.error(f"save_trade_record -> {ex}\n{traceback.format_exc()}")

    def update_daily_summary(self):
        """일별 요약 업데이트"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            today = datetime.now().strftime('%Y%m%d')
            strategy = self.window.comboStg.currentText() if hasattr(self, 'window') else ''
            
            # 오늘의 매도 거래만 집계
            cursor.execute('''
                SELECT 
                    COUNT(*) as total_trades,
                    SUM(CASE WHEN profit > 0 THEN 1 ELSE 0 END) as win_trades,
                    SUM(CASE WHEN profit <= 0 THEN 1 ELSE 0 END) as lose_trades,
                    AVG(profit_pct) as avg_profit_pct,
                    MAX(profit_pct) as max_profit_pct,
                    MIN(profit_pct) as max_loss_pct,
                    SUM(profit) as total_profit
                FROM trades
                WHERE date = ? AND action = 'SELL'
            ''', (today,))
            
            row = cursor.fetchone()
            
            if row and row[0] > 0:
                total_trades, win_trades, lose_trades, avg_profit_pct, max_profit_pct, max_loss_pct, total_profit = row
                win_rate = (win_trades / total_trades * 100) if total_trades > 0 else 0
                
                # 매수 금액 합계
                cursor.execute('''
                    SELECT SUM(amount) FROM trades
                    WHERE date = ? AND action = 'BUY'
                ''', (today,))
                
                total_buy_amount = cursor.fetchone()[0] or 0
                
                cursor.execute('''
                    INSERT OR REPLACE INTO daily_summary (
                        date, strategy, total_trades, win_trades, lose_trades, win_rate,
                        total_profit, avg_profit_pct, max_profit_pct, max_loss_pct, total_buy_amount
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    today, strategy, total_trades, win_trades, lose_trades, win_rate,
                    total_profit or 0, avg_profit_pct or 0, max_profit_pct or 0, max_loss_pct or 0,
                    total_buy_amount
                ))
                
                conn.commit()
                
                logging.debug(f"일별 요약 업데이트: {today} - 거래 {total_trades}회, 승률 {win_rate:.1f}%")
            
            conn.close()
            
        except Exception as ex:
            logging.error(f"update_daily_summary -> {ex}\n{traceback.format_exc()}")

    def get_stock_balance(self, code, func):
        """계좌 잔고 조회"""
        try:
            stock_name = self.cpCodeMgr.CodeToName(code)
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)
            self.cpBalance.SetInputValue(0, acc)
            self.cpBalance.SetInputValue(1, accFlag[0])
            self.cpBalance.SetInputValue(2, 50)
            ret = self.cpBalance.BlockRequest2(1)
            if ret != 0:
                logging.warning(f"{stock_name}({code}) 잔고 조회 실패, {ret}")
                return False
            
            if code == 'START':
                logging.info(f"계좌명 : {str(self.cpBalance.GetHeaderValue(0))}")
                logging.info(f"결제잔고수량 : {str(self.cpBalance.GetHeaderValue(1))}")
                logging.info(f"평가금액 : {str(self.cpBalance.GetHeaderValue(3))}")
                logging.info(f"평가손익 : {str(self.cpBalance.GetHeaderValue(4))}")
                logging.info(f"종목수 : {str(self.cpBalance.GetHeaderValue(7))}")
                return

            stocks = []
            for i in range(self.cpBalance.GetHeaderValue(7)):
                stock_code = self.cpBalance.GetDataValue(12, i)
                stock_name = self.cpBalance.GetDataValue(0, i)
                stock_qty = self.cpBalance.GetDataValue(15, i)
                buy_price = self.cpBalance.GetDataValue(17, i)
                stocks.append({'code': stock_code, 'name': stock_name, 'qty': stock_qty, 'buy_price': buy_price})

            if code == 'ALL':
                logging.debug("잔고 전부 조회 성공")
                return stocks

            for s in stocks:
                if s['code'] == code:
                    logging.debug(f"{s['name']}({s['code']}) 잔고 조회 성공")
                    return s['name'], s['qty'], s['buy_price']             

        except Exception as ex:
            logging.error(f"get_stock_balance({func}) -> {ex}")
            return False

    def init_stock_balance(self):
        """시작 시 계좌 잔고 초기화 (중복 방지)"""
        try:
            # 중복 호출 방지 플래그 확인
            if hasattr(self, '_stock_balance_initialized') and self._stock_balance_initialized:
                logging.info("계좌 잔고 이미 초기화됨, 스킵")
                return
                
            stocks = self.get_stock_balance('ALL', 'init_stock_balance')

            for s in stocks:
                if self.daydata.select_code(s['code']) and self.tickdata.monitor_code(s['code']) and self.mindata.monitor_code(s['code']):
                    if s['code'] not in self.starting_time:
                        self.starting_time[s['code']] = datetime.now().strftime('%m/%d 09:00:00')
                    self.monistock_set.add(s['code'])
                    self.bought_set.add(s['code'])
                    self.buy_price[s['code']] = s['buy_price']
                    self.buy_qty[s['code']] = s['qty']
                    self.balance_qty[s['code']] = s['qty']
                    
                    logging.info(f"📋 잔고 종목 {s['code']} 처리 완료")

            remaining_count = self.target_buy_count - len(stocks)
            self.buy_percent = 1/remaining_count if remaining_count > 0 else 0
            self.total_cash = self.get_current_cash() * 0.9
            self.buy_amount = int(self.total_cash * self.buy_percent)
            
            logging.info(f"주문 가능 금액 : {self.total_cash}")
            logging.info(f"종목별 주문 비율 : {self.buy_percent}")
            logging.info(f"종목별 주문 금액 : {self.buy_amount}")
            
            # 초기화 완료 플래그 설정
            self._stock_balance_initialized = True
            
            logging.info(f"📋 잔고 초기화 완료: {len(self.monistock_set)}개 종목")

        except Exception as ex:
            logging.error(f"init_stock_balance -> {ex}")

    def get_current_cash(self):
        """현재 주문 가능 금액 조회"""
        try:
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)
            self.cpCash.SetInputValue(0, acc)
            self.cpCash.SetInputValue(1, accFlag[0])
            self.cpCash.SetInputValue(5, "Y")
            self.cpCash.BlockRequest2(1)

            rqStatus = self.cpCash.GetDibStatus()
            if rqStatus != 0:
                rqRet = self.cpCash.GetDibMsg1()
                logging.warning(f"주문가능금액 조회실패, {rqStatus}, {rqRet}")
                return (False, '')
            current_cash = self.cpCash.GetHeaderValue(9)
            return current_cash
        except Exception as ex:
            logging.error(f"get_current_cash -> {ex}")
            return False

    def get_current_price(self, code):
        """현재가 조회"""
        try:
            self.cpStock.SetInputValue(0, code)
            ret = self.cpStock.BlockRequest2(1)
            if ret != 0:
                logging.warning(f"{code} 현재가 조회 실패, {ret}")
                return False

            rqStatus = self.cpStock.GetDibStatus()
            if rqStatus != 0:
                rqRet = self.cpStock.GetDibMsg1()
                logging.warning(f"현재가 조회 오류, {rqStatus}, {rqRet}")
                return False
            
            item = {'cur_price': self.cpStock.GetHeaderValue(11), 'ask': self.cpStock.GetHeaderValue(16), 'upper': self.cpStock.GetHeaderValue(8)}
            return item['cur_price'], item['ask'], item['upper']
        
        except Exception as ex:
            logging.error(f"get_current_price -> {ex}")
            return False
    
    def get_trade_profit(self):
        """매매실현손익 조회"""
        try:
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            self.objRq = win32com.client.Dispatch("CpTrade.CpTd6032")
            self.objRq.SetInputValue(0, acc)
            self.objRq.SetInputValue(1, accFlag[0])
            ret = self.objRq.BlockRequest2(1)
            if ret != 0:
                logging.warning(f"매매실현손익 조회 실패, {ret}")
                return
            
            rqStatus = self.objRq.GetDibStatus()
            if rqStatus != 0:
                rqRet = self.objRq.GetDibMsg1()
                logging.warning(f"매매실현손익 조회 오류, {rqStatus}, {rqRet}")
                return False
            
            profit = self.objRq.GetHeaderValue(2)
            logging.info(f"매매실현손익 : {self.objRq.GetHeaderValue(2)}원")
            send_slack_message(self.window.login_handler, "#stock", f"매매실현손익 : {profit}원")

        except Exception as ex:
            logging.error(f"get_trade_profit -> {ex}")
    
    def update_highest_price(self, code, current_price):
        """최고가 업데이트"""
        if code not in self.highest_price:
            self.highest_price[code] = current_price
        elif current_price > self.highest_price[code]:
            self.highest_price[code] = current_price    

    def monitor_vi(self, time, code, event2):
        """VI 발동 모니터링"""
        try:
            # ===== ✅ 20개 종목 제한 체크 (보유 종목 제외) =====
            monitoring_only = self.monistock_set - self.bought_set
            MAX_MONITORING_STOCKS = 20
            
            if code in self.monistock_set or len(monitoring_only) >= MAX_MONITORING_STOCKS or code in self.bought_set:
                if len(monitoring_only) >= MAX_MONITORING_STOCKS and code not in self.monistock_set and code not in self.bought_set:
                    logging.debug(f"{code}: 모니터링 종목 제한({MAX_MONITORING_STOCKS}개) 도달, VI 추가 거부")
                return

            if not self.daydata.select_code(code):
                logging.debug(f"{code}: 일봉 데이터 로드 실패")
                return
            
            day_data = self.daydata.stockdata[code]
            if not (day_data['MAD5'][-1] > day_data['MAD10'][-1]):
                logging.debug(f"{code}: MAD5 < MAD10, 추세 미달")
                self.daydata.monitor_stop(code)
                return
            
            recent_ma5 = day_data['MAD5'][-3:]
            if len(recent_ma5) >= 3 and not all(recent_ma5[i] < recent_ma5[i+1] for i in range(2)):
                logging.debug(f"{code}: MAD5 추세 약화")
                self.daydata.monitor_stop(code)
                return
            
            if len(day_data['V']) < 30:
                logging.debug(f"{code}: 데이터 부족 ({len(day_data['V'])}일)")
                self.daydata.monitor_stop(code)
                return
            
            Trading_amount = sum(day_data['V'][-30:]) / 30
            Trading_Value = sum(day_data['TV'][-30:]) / 30
            
            MIN_VOLUME = 100000
            MIN_VALUE = 3000000000
            
            if Trading_amount < MIN_VOLUME or Trading_Value < MIN_VALUE:
                logging.debug(
                    f"{code}: 유동성 부족 - "
                    f"거래량 {Trading_amount:.0f}/{MIN_VOLUME}, "
                    f"거래금액 {Trading_Value/100000000:.1f}/{MIN_VALUE/100000000}억"
                )
                self.daydata.monitor_stop(code)
                return
            
            now = datetime.now()
            elapsed_minutes = (now - now.replace(hour=9, minute=0, second=0)).total_seconds() / 60
            
            if elapsed_minutes <= 120:
                volume_threshold = 0.5
            elif elapsed_minutes <= 240:
                volume_threshold = 0.7
            else:
                volume_threshold = 1.0
            
            current_volume = day_data['V'][-1]
            prev_volume = day_data['V'][-2]
            
            if current_volume < prev_volume * volume_threshold:
                logging.debug(
                    f"{code}: 거래량 부족 - "
                    f"{current_volume}/{prev_volume * volume_threshold:.0f} ({volume_threshold*100}%)"
                )
                self.daydata.monitor_stop(code)
                return
            
            current_price = day_data['C'][-1]
            if current_price < 1000 or current_price > 500000:
                logging.debug(f"{code}: 가격 {current_price}원 범위 초과")
                self.daydata.monitor_stop(code)
                return
            
            match_gap = re.search(r"괴리율:(-?\d+\.\d+)%", event2)
            if match_gap:
                gap_rate = float(match_gap.group(1))
                if gap_rate < 3.0:
                    logging.debug(f"{code}: 괴리율 {gap_rate}% (3% 미만)")
                    self.daydata.monitor_stop(code)
                    return
            
            sector = cpCodeMgr.GetStockSectionKind(code)
            sector_count = sum(1 for c in self.monistock_set 
                            if cpCodeMgr.GetStockSectionKind(c) == sector)
            if sector_count >= 2:
                logging.debug(f"{code}: 동일 섹터 종목 {sector_count}개 초과")
                self.daydata.monitor_stop(code)
                return
            
            if not (self.tickdata.monitor_code(code) and self.mindata.monitor_code(code)):
                logging.error(f"{code}: 틱/분 데이터 모니터링 시작 실패")
                self.daydata.monitor_stop(code)
                return
            
            logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 장 시작부터 데이터 로드 시작")
            
            tick_loaded = self.tickdata.update_chart_data_from_market_open(code)
            min_loaded = self.mindata.update_chart_data_from_market_open(code)
            
            if not tick_loaded or not min_loaded:
                logging.warning(f"{code}: 장 시작 데이터 로드 실패, 개수 기준으로 폴백")
                self.tickdata.update_chart_data(code, 60, 400)
                self.mindata.update_chart_data(code, 3, 150)
            
            self.starting_time[code] = time
            
            match_price = re.search(r"발동가격:\s*(\d+)", event2)
            if match_price:
                self.starting_price[code] = int(match_price.group(1))
            else:
                self.starting_price[code] = current_price
            
            self.monistock_set.add(code)
            self.stock_added_to_monitor.emit(code)
            
            logging.info(
                f"{cpCodeMgr.CodeToName(code)}({code}) -> "
                f"투자 대상 추가 (VI: {time}, 발동가: {self.starting_price[code]:.0f}원, "
                f"괴리율: {gap_rate if match_gap else 'N/A'}%)"
            )
            
            self.save_list_db(code, self.starting_time[code], self.starting_price[code], 1)

        except Exception as ex:
            logging.error(f"monitor_vi -> {code}, {ex}\n{traceback.format_exc()}")

    def save_list_db(self, code, starting_time, starting_price, is_moni=0, db_file='mylist.db'):
        """종목 리스트 DB 저장"""
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS items (
                codes TEXT PRIMARY KEY,
                starting_time DATETIME,
                starting_price REAL,
                is_moni INTEGER
            )
        ''')
        c.execute('INSERT OR REPLACE INTO items (codes, starting_time, starting_price, is_moni) VALUES (?, ?, ?, ?)', 
                  (code, starting_time, starting_price, is_moni))
        conn.commit()
        conn.close()

    def delete_list_db(self, code, db_file='mylist.db'):
        """종목 리스트 DB 삭제"""
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS items (
                codes TEXT PRIMARY KEY,
                starting_time DATETIME,
                starting_price REAL,
                is_moni INTEGER
            )
        ''')
        c.execute("DELETE FROM items WHERE codes = ?", (code,))
        conn.commit()
        conn.close()

    def clear_list_db(self, db_file='mylist.db'):
        """종목 리스트 DB 전체 삭제"""
        try:
            if os.path.exists(db_file):
                os.remove(db_file)
        except Exception as ex:
            logging.error(f"clear_list_db -> {ex}")
        self.monistock_set = set()
        self.vistock_set = set()
        self.database_set = set()
        self.starting_time = {}
        self.starting_price = {}

    def save_combined_tick_data(self, code, save_data):
        """결합 틱 데이터 DB 저장"""
        try:
            timestamp = save_data.get('timestamp', datetime.now())
            tick_data = save_data.get('tick_data', {})
            min_data = save_data.get('min_data', {})
            reason = save_data.get('reason', '')
            
            # 체결강도 조회
            strength = 0
            try:
                strength = self.tickdata.get_strength(code)
            except:
                pass
            
            # 포지션 타입 및 매수가 결정
            if code in self.bought_set:
                position_type = 'HOLD'
                buy_price = self.buy_price.get(code, 0)
            elif code in self.buyorder_set:
                position_type = 'BUY_ORDER'
                buy_price = self.starting_price.get(code, 0)
            else:
                position_type = 'NONE'
                buy_price = None
            
            conn = sqlite3.connect('vi_stock_data.db')
            cursor = conn.cursor()
            
            # combined_tick_data 테이블에 저장 (컬럼 순서를 테이블 구조와 정확히 맞춤)
            cursor.execute('''
                INSERT OR REPLACE INTO combined_tick_data (
                    code, timestamp, date, time,
                    tick_C, tick_O, tick_H, tick_L, tick_V,
                    tick_MAT5, tick_MAT20, tick_MAT60, tick_MAT120,
                    tick_RSIT, tick_RSIT_SIGNAL,
                    tick_MACDT, tick_MACDT_SIGNAL, tick_OSCT,
                    tick_STOCHK, tick_STOCHD,
                    tick_ATR, tick_CCI,
                    tick_BB_UPPER, tick_BB_MIDDLE, tick_BB_LOWER,
                    tick_BB_POSITION, tick_BB_BANDWIDTH,
                    tick_VWAP,
                    tick_WILLIAMS_R, tick_ROC, tick_OBV, tick_OBV_MA20,
                    tick_VP_POC, tick_VP_POSITION,
                    min_C, min_O, min_H, min_L, min_V,
                    min_MAM5, min_MAM10, min_MAM20,
                    min_RSI, min_RSI_SIGNAL,
                    min_MACD, min_MACD_SIGNAL, min_OSC,
                    min_STOCHK, min_STOCHD,
                    min_CCI, min_VWAP,
                    min_WILLIAMS_R, min_ROC, min_OBV, min_OBV_MA20,
                    strength, buy_price, position_type
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                code, timestamp, timestamp.strftime('%Y%m%d'), timestamp.strftime('%H%M%S'),  # 4개
                # 틱 데이터 (5개)
                tick_data.get('C', 0), tick_data.get('O', 0), 
                tick_data.get('H', 0), tick_data.get('L', 0), tick_data.get('V', 0),
                # 틱 이동평균 (4개)
                tick_data.get('MAT5', 0), tick_data.get('MAT20', 0), 
                tick_data.get('MAT60', 0), tick_data.get('MAT120', 0),
                # 틱 RSI (2개)
                tick_data.get('RSIT', 0), tick_data.get('RSIT_SIGNAL', 0),
                # 틱 MACD (3개)
                tick_data.get('MACDT', 0), tick_data.get('MACDT_SIGNAL', 0), 
                tick_data.get('OSCT', 0),
                # 틱 스토캐스틱 (2개)
                tick_data.get('STOCHK', 0), tick_data.get('STOCHD', 0),
                # 틱 기타 (2개)
                tick_data.get('ATR', 0), tick_data.get('CCI', 0),
                # 틱 볼린저밴드 (3개)
                tick_data.get('BB_UPPER', 0), tick_data.get('BB_MIDDLE', 0), 
                tick_data.get('BB_LOWER', 0),
                # 틱 볼린저밴드 위치/폭 (2개)
                tick_data.get('BB_POSITION', 0), tick_data.get('BB_BANDWIDTH', 0),
                # 틱 VWAP (1개)
                tick_data.get('VWAP', 0),
                # 새 지표 - 틱 (6개)
                tick_data.get('WILLIAMS_R', -50), tick_data.get('ROC', 0),
                tick_data.get('OBV', 0), tick_data.get('OBV_MA20', 0),
                tick_data.get('VP_POC', 0), tick_data.get('VP_POSITION', 0),
                # 분봉 데이터 (5개)
                min_data.get('C', 0), min_data.get('O', 0), 
                min_data.get('H', 0), min_data.get('L', 0), min_data.get('V', 0),
                # 분봉 이동평균 (3개)
                min_data.get('MAM5', 0), min_data.get('MAM10', 0), 
                min_data.get('MAM20', 0),
                # 분봉 RSI (2개)
                min_data.get('RSI', 0), min_data.get('RSI_SIGNAL', 0),
                # 분봉 MACD (3개)
                min_data.get('MACD', 0), min_data.get('MACD_SIGNAL', 0), 
                min_data.get('OSC', 0),
                # 분봉 스토캐스틱 (2개)
                min_data.get('STOCHK', 0), min_data.get('STOCHD', 0),
                # 분봉 기타 (2개)
                min_data.get('CCI', 0), min_data.get('VWAP', 0),
                # 새 지표 - 분봉 (4개)
                min_data.get('WILLIAMS_R', -50), min_data.get('ROC', 0),
                min_data.get('OBV', 0), min_data.get('OBV_MA20', 0),
                # 추가 정보 (3개)
                strength, buy_price, position_type
            ))
            
            conn.commit()
            conn.close()
            
        except Exception as ex:
            logging.error(f"save_combined_tick_data 오류 ({code}): {ex}")

    def load_from_list_db(self, db_file='mylist.db'):
        """종목 리스트 DB 로드"""
        if not os.path.exists(db_file):
            self.monistock_set = set()
            self.vistock_set = set()
            self.database_set = set()
            self.stgname = {}
            self.starting_time = {}
            self.starting_price = {}
            return
        
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS items (
                codes TEXT PRIMARY KEY,
                starting_time DATETIME,
                starting_price REAL,
                is_moni INTEGER
            )
        ''')
        c.execute('SELECT codes, starting_time, starting_price, is_moni FROM items')
        rows = c.fetchall()
        conn.close()
        if rows:
            self.vistock_set = {row[0] for row in rows if not row[3]}
            self.database_set = {row[0] for row in rows if row[3]}
            self.starting_time = {row[0]: row[1] for row in rows}
            self.starting_price = {row[0]: row[2] for row in rows}

    @pyqtSlot(str, str, str, str)
    def buy_stock(self, code, buy_message, order_condition, order_style):
        """매수 주문"""
        try:
            if code in self.bought_set or code in self.buyorder_set or code in self.buyordering_set:
                return
            self.buyordering_set.add(code)
            stock_name = self.cpCodeMgr.CodeToName(code)
            cur_price, ask_price, upper_price = self.get_current_price(code)

            if not ask_price:
                if ask_price == 0:
                    self.buyorder_set.add(code)
                    logging.info(f"{stock_name}({code}) 상한가 주문 불가")
                else:
                    logging.error(f"{stock_name}({code}) 현재가 조회 실패")
                return
            if (len(self.buyorder_set) + len(self.bought_set)) >= self.target_buy_count:
                return
            buy_qty = self.buy_amount // ask_price
            total_amount = self.total_cash - self.buy_amount * (len(self.buyorder_set) + len(self.bought_set))
            max_buy_qty = total_amount // upper_price
            if max_buy_qty <= 0 or buy_qty <= 0: 
                logging.warning(f"{stock_name}({code}) 주문 수량 계산 결과 0 또는 음수: buy_qty={buy_qty}, max_buy_qty={max_buy_qty}")
                return
            self.buyorder_qty[code] = int(min(buy_qty, max_buy_qty))
            
            # 추가 안전 검증
            if self.buyorder_qty[code] <= 0:
                logging.warning(f"{stock_name}({code}) 최종 주문 수량이 0 이하: {self.buyorder_qty[code]}")
                return

            if self.buyorder_qty[code] > 0:
                acc = self.cpTrade.AccountNumber[0]
                accFlag = self.cpTrade.GoodsList(acc, 1)
                if self.cp_request.is_requesting:
                    logging.debug("요청 진행 중, 매수 주문 스킵")
                    return None

                self.cp_request.is_requesting = True
                self.cpOrder.SetInputValue(0, "2")
                self.cpOrder.SetInputValue(1, acc)
                self.cpOrder.SetInputValue(2, accFlag[0])
                self.cpOrder.SetInputValue(3, code)
                self.cpOrder.SetInputValue(4, self.buyorder_qty[code])
                if buy_message == '발동가':
                    self.cpOrder.SetInputValue(5, self.starting_price[code])
                else:
                    self.cpOrder.SetInputValue(5, ask_price)
                self.cpOrder.SetInputValue(7, order_condition)
                self.cpOrder.SetInputValue(8, order_style)
                
                remain_count0 = cpCybos.GetLimitRemainCount(0)
                if remain_count0 == 0:
                    logging.error(f"🚫 매수 주문 거부: 거래 요청 제한")
                    return
                elif remain_count0 <= 3:
                    logging.warning(f"⚠️ 거래 요청 한계 근접: 남은 요청 {remain_count0}건")
                
                # 주문 전 최종 수량 검증
                if self.buyorder_qty[code] <= 0:
                    logging.warning(f"{stock_name}({code}) 주문 전 수량 재검증 실패: {self.buyorder_qty[code]}")
                    return
                
                logging.info(f"{stock_name}({code}), {buy_message} -> 매수 요청({self.buyorder_qty[code]}주)")
                
                handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
                handler.set_params(self.cpOrder, 'order', self, {'code': code, 'stock_name': stock_name})
                self.cpOrder.Request()          
                self.buyorder_set.add(code)
                if code in self.buyordering_set:
                    self.buyordering_set.remove(code)
                result = handler.wait()
                if not result:
                    logging.warning(f"{stock_name}({code}) 매수주문 실패")
                    if code in self.buyorder_set:
                        self.buyorder_set.remove(code)
                    return            

        except Exception as ex:
            logging.error(f"buy_stock -> {code}, {ex}")

    @pyqtSlot(str, str)
    def sell_stock(self, code, message):
        """매도 주문"""
        try:
            if code in self.sellorder_set:
                return
            
            stock_name = self.cpCodeMgr.CodeToName(code)
            if code in self.buy_qty:
                stock_qty = self.buy_qty[code]
            else:
                return
            
            # ===== ✅ 실제 잔고 수량 재조회 =====
            try:
                _, actual_qty, _ = self.get_stock_balance(code, '')
                if actual_qty is None or actual_qty <= 0:
                    logging.warning(f"{stock_name}({code}) 실제 잔고 수량이 0 이하: {actual_qty}")
                    return
                # balance_qty 업데이트
                self.balance_qty[code] = actual_qty
                logging.debug(f"{stock_name}({code}) 잔고 수량 업데이트: {actual_qty}주")
            except Exception as ex:
                logging.warning(f"{stock_name}({code}) 잔고 재조회 실패: {ex}")
                # 재조회 실패 시 기존 값 사용
                if code not in self.balance_qty:
                    logging.warning(f"{stock_name}({code}) 잔고 정보 없음")
                    return
            
            sell_order_qty = min(stock_qty, self.balance_qty[code])
            
            # 매도 수량 검증
            if sell_order_qty <= 0:
                logging.warning(f"{stock_name}({code}) 매도 주문 수량이 0 이하: {sell_order_qty}")
                return
            
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            if self.cp_request.is_requesting:
                logging.debug("요청 진행 중, 매도 주문 스킵")
                return None

            self.cp_request.is_requesting = True
            self.cpOrder.SetInputValue(0, "1")
            self.cpOrder.SetInputValue(1, acc)
            self.cpOrder.SetInputValue(2, accFlag[0])
            self.cpOrder.SetInputValue(3, code)
            self.cpOrder.SetInputValue(4, sell_order_qty)
            self.cpOrder.SetInputValue(7, "0")
            self.cpOrder.SetInputValue(8, "03")
        
            remain_count0 = cpCybos.GetLimitRemainCount(0)
            if remain_count0 == 0:
                logging.error(f"거래 요청 제한")
                return
            
            # 매도 주문 전 최종 수량 검증
            if sell_order_qty <= 0:
                logging.warning(f"{stock_name}({code}) 매도 주문 전 수량 재검증 실패: {sell_order_qty}")
                return
            
            logging.info(f"{stock_name}({code}), {message} -> 매도 요청({sell_order_qty}주)")
            handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
            handler.set_params(self.cpOrder, 'order', self, {'code': code, 'stock_name': stock_name})
            self.cpOrder.Request()
            self.sellorder_set.add(code) 

            result = handler.wait()
            if not result:
                logging.warning(f"{stock_name}({code}) 매도주문 실패")
                if code in self.sellorder_set:
                    self.sellorder_set.remove(code)
                return
                       
        except Exception as ex:
            logging.error(f"sell_stock -> {code}, {ex}")

    @pyqtSlot(str, str)
    def sell_half_stock(self, code, message):
        """분할 매도 주문"""
        try:
            if code in self.sellorder_set:
                return
            
            stock_name = self.cpCodeMgr.CodeToName(code)
            if code in self.buy_qty:
                stock_qty = self.buy_qty[code]
            else:
                return
            # ===== ✅ 실제 잔고 수량 재조회 =====
            try:
                _, actual_qty, _ = self.get_stock_balance(code, '')
                if actual_qty is None or actual_qty <= 0:
                    logging.warning(f"{stock_name}({code}) 분할매도 - 실제 잔고 수량이 0 이하: {actual_qty}")
                    return
                # balance_qty 업데이트
                self.balance_qty[code] = actual_qty
                logging.debug(f"{stock_name}({code}) 분할매도 - 잔고 수량 업데이트: {actual_qty}주")
            except Exception as ex:
                logging.warning(f"{stock_name}({code}) 분할매도 - 잔고 재조회 실패: {ex}")
                # 재조회 실패 시 기존 값 사용
                if code not in self.balance_qty:
                    logging.warning(f"{stock_name}({code}) 분할매도 - 잔고 정보 없음")
                    return
            
            self.sell_half_qty[code] = stock_qty - ((stock_qty + 1) // 2)
            sell_half_order_qty = min(((stock_qty + 1) // 2), self.balance_qty.get(code, 0))

            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            if self.cp_request.is_requesting:
                logging.debug("요청 진행 중, 분할매도 주문 스킵")
                return None

            self.cp_request.is_requesting = True
            self.cpOrder.SetInputValue(0, "1")
            self.cpOrder.SetInputValue(1, acc)
            self.cpOrder.SetInputValue(2, accFlag[0])
            self.cpOrder.SetInputValue(3, code)
            self.cpOrder.SetInputValue(4, sell_half_order_qty)
            self.cpOrder.SetInputValue(7, "0")
            self.cpOrder.SetInputValue(8, "03")

            remain_count0 = cpCybos.GetLimitRemainCount(0)
            if remain_count0 == 0:
                logging.error(f"거래 요청 제한")
                return
            
            logging.info(f"{stock_name}({code}), {message} -> 분할 매도 요청({sell_half_order_qty}주)")
            handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
            handler.set_params(self.cpOrder, 'order', self, {'code': code, 'stock_name': stock_name})
            self.cpOrder.Request()
            self.sellorder_set.add(code) 

            result = handler.wait()
            if not result:
                logging.warning(f"{stock_name}({code}) 분할매도주문 실패")
                if code in self.sellorder_set:
                    self.sellorder_set.remove(code)
                return
                       
        except Exception as ex:
            logging.error(f"sell_half_stock -> {code}, {ex}")

    @pyqtSlot()
    def sell_all(self):
        """전량 매도"""
        try:
            stocks = self.get_stock_balance('ALL', 'sell_all')
            acc = self.cpTrade.AccountNumber[0]
            accFlag = self.cpTrade.GoodsList(acc, 1)

            for s in stocks:
                if s['qty'] > 0:
                    if self.cp_request.is_requesting:
                        logging.debug("요청 진행 중, 전부매도 주문 스킵")
                        return None

                    self.cp_request.is_requesting = True
                    self.cpOrder.SetInputValue(0, "1")
                    self.cpOrder.SetInputValue(1, acc)
                    self.cpOrder.SetInputValue(2, accFlag[0])
                    self.cpOrder.SetInputValue(3, s['code'])
                    self.cpOrder.SetInputValue(4, s['qty'])
                    self.cpOrder.SetInputValue(7, "0")
                    self.cpOrder.SetInputValue(8, "03")

                    logging.info(f"{s['name']}({s['code']}) -> 매도 요청({s['qty']}주)")
                    handler = win32com.client.WithEvents(self.cpOrder, CpRequest)
                    handler.set_params(self.cpOrder, 'order', self, {'code': s['code'], 'stock_name': s['name']})
                    self.cpOrder.Request()
                    self.sellorder_set.add(s['code'])

                    result = handler.wait()
                    if not result:
                        logging.warning(f"{s['name']}({s['code']}) 전부매도주문 실패")
                        if s['code'] in self.sellorder_set:
                            self.sellorder_set.remove(s['code'])
            return True

        except Exception as ex:
            logging.error(f"sell_all -> {ex}")
            return False

    def monitorOrderStatus(self, code, ordernum, conflags, price, qty, bs, balance, buyprice):
        """주문 체결 모니터링"""
        try:
            stock_name = self.cpCodeMgr.CodeToName(code)

            # ===== 매도 체결 =====
            if bs == '1' and conflags == "체결":
                logging.debug(f"{stock_name}({code}), {price}원, {qty}주 매도, 잔고: {balance}주")
                self.balance_qty[code] = balance

                if code not in self.sell_amount:
                    self.sell_amount[code] = 0
                self.sell_amount[code] += price * qty

                # 분할 매도 완료
                if code in self.sell_half_qty and balance == self.sell_half_qty[code]:
                    logging.info(f"{stock_name}({code}), 분할 매도 완료")                  
                    
                    stock_profit = self.sell_amount[code] * 0.99835 - self.buy_price[code] * (self.buy_qty[code] - balance) * 1.00015
                    stock_rate = (stock_profit / (self.buy_price[code] * (self.buy_qty[code] - balance))) * 100
                    
                    if stock_profit > 0:
                        logging.info(f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                    else:
                        logging.info(f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                    
                    self.get_trade_profit()
                    
                    if code in self.sellorder_set:
                        self.sellorder_set.remove(code)
                    if code not in self.sell_half_set:
                        self.sell_half_set.add(code)
                    
                    self.buy_qty[code] = balance
                    self.sell_amount[code] = 0
                    
                    if code in self.sell_half_qty:
                        del self.sell_half_qty[code]

                # 전량 매도 완료
                if balance == 0:
                    logging.info(f"{stock_name}({code}), 매도 완료")

                    stock_profit = self.sell_amount[code] * 0.99835 - self.buy_price[code] * self.buy_qty[code] * 1.00015
                    stock_rate = (stock_profit / (self.buy_price[code] * self.buy_qty[code])) * 100
                    
                    # 보유 시간 계산
                    hold_minutes = 0
                    if code in self.starting_time:
                        try:
                            buy_time = datetime.strptime(
                                f"{datetime.now().year}/{self.starting_time[code]}", 
                                '%Y/%m/%d %H:%M:%S'
                            )
                            hold_minutes = (datetime.now() - buy_time).total_seconds() / 60
                        except:
                            pass
                    
                    # 매도 거래 기록 저장
                    self.save_trade_record(
                        code=code,
                        action='SELL',
                        price=price,
                        quantity=self.buy_qty[code],
                        buy_price=self.buy_price[code],
                        profit=stock_profit,
                        profit_pct=stock_rate,
                        hold_minutes=hold_minutes,
                        sell_reason='매도 완료'
                    )
                    
                    # 일별 요약 업데이트
                    self.update_daily_summary()
                    
                    if stock_profit > 0:
                        logging.info(f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매이익({int(stock_profit)}원, {stock_rate:.2f}%)")
                    else:
                        logging.info(f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                        send_slack_message(self.window.login_handler, "#stock", f"{stock_name}({code}), 매매손실({int(stock_profit)}원, {stock_rate:.2f}%)")
                    
                    self.get_trade_profit()                    
                                        
                    self.stock_sold.emit(code)                    
                    
                    if code in self.bought_set:
                        self.bought_set.remove(code) 
                    if code in self.sell_half_set:
                        self.sell_half_set.remove(code)
                    if code in self.sellorder_set:
                        self.sellorder_set.remove(code)                 
                    if code in self.buy_price:
                        del self.buy_price[code]
                    if code in self.buy_qty:
                        del self.buy_qty[code]
                    if code in self.sell_amount:
                        del self.sell_amount[code]

            # ===== 매수 체결 =====
            elif bs == '2' and conflags == "체결":
                logging.debug(f"{stock_name}({code}), {qty}주 매수, 잔고: {balance}주")
                self.balance_qty[code] = balance
                
                if code in self.buyorder_qty and balance >= self.buyorder_qty[code]:
                    self.buy_qty[code] = balance
                    self.buy_price[code] = buyprice
                    logging.info(f"{stock_name}({code}), {self.buy_qty[code]}주, 매수 완료({int(self.buy_price[code])}원)")
                    
                    # 매수 거래 기록 저장
                    self.save_trade_record(
                        code=code,
                        action='BUY',
                        price=buyprice,
                        quantity=self.buy_qty[code],
                        buy_reason='매수 완료'
                    )
                    
                    if code not in self.bought_set:
                        self.bought_set.add(code)
                        self.stock_bought.emit(code)
                    if code in self.buyorder_set:
                        self.buyorder_set.remove(code)
                    if code in self.buyorder_qty:
                        del self.buyorder_qty[code]                
                    return

        except Exception as ex:
            logging.error(f"monitorOrderStatus -> {ex}\n{traceback.format_exc()}")

# ==================== CalculationWorker (계산 및 DB 저장 전용) ====================
class CalculationWorker(QThread):
    """계산 및 DB 저장 전용 워커 스레드"""
    
    calculation_completed = pyqtSignal(str, dict)  # (종목코드, 계산결과)
    db_save_completed = pyqtSignal(str, bool)  # (종목코드, 성공여부)
    
    def __init__(self, trader):
        super().__init__()
        self.trader = trader
        self.running = True
        self.calculation_queue = queue.Queue()
        self.db_save_queue = queue.Queue()
        
    def run(self):
        """워커 스레드 메인 루프"""
        while self.running:
            try:
                # 계산 작업 처리
                if not self.calculation_queue.empty():
                    try:
                        code, tick_data, min_data = self.calculation_queue.get_nowait()
                        result = self._perform_calculation(code, tick_data, min_data)
                        self.calculation_completed.emit(code, result)
                    except queue.Empty:
                        pass
                
                # DB 저장 작업 처리
                if not self.db_save_queue.empty():
                    try:
                        code, save_data = self.db_save_queue.get_nowait()
                        success = self._perform_db_save(code, save_data)
                        self.db_save_completed.emit(code, success)
                    except queue.Empty:
                        pass
                
                # CPU 사용률 조절
                self.msleep(10)  # 10ms 대기
                
            except Exception as ex:
                logging.error(f"CalculationWorker 오류: {ex}")
                self.msleep(100)
    
    def _perform_calculation(self, code, tick_data, min_data):
        """기술적 지표 계산 수행"""
        try:
            # 기술적 지표 계산 로직
            result = {
                'timestamp': datetime.now(),
                'tick_data': tick_data,
                'min_data': min_data,
                'calculated_indicators': {
                    'rsi': tick_data.get('RSIT', 0),
                    'macd': tick_data.get('MACDT', 0),
                    'bb_position': tick_data.get('BB_POSITION', 0),
                    'vwap': tick_data.get('VWAP', 0),
                    'atr': tick_data.get('ATR', 0),
                    'cci': tick_data.get('CCI', 0),
                    'williams_r': tick_data.get('WILLIAMS_R', 0),
                    'roc': tick_data.get('ROC', 0),
                    'obv': tick_data.get('OBV', 0),
                    'vp_position': tick_data.get('VP_POSITION', 0)
                }
            }
            return result
        except Exception as ex:
            logging.error(f"계산 오류 ({code}): {ex}")
            return {}
    
    def _perform_db_save(self, code, save_data):
        """DB 저장 수행"""
        try:
            # DB 저장 로직
            self.trader.save_combined_tick_data(code, save_data)
            return True
        except Exception as ex:
            logging.error(f"DB 저장 오류 ({code}): {ex}")
            return False
    
    def add_calculation_task(self, code, tick_data, min_data):
        """계산 작업 추가"""
        self.calculation_queue.put((code, tick_data, min_data))
    
    def add_db_save_task(self, code, save_data):
        """DB 저장 작업 추가"""
        self.db_save_queue.put((code, save_data))
    
    def stop(self):
        """워커 스레드 중지"""
        self.running = False

# ==================== AutoTraderThread (통합 전략 적용) ====================
class AutoTraderThread(QThread):
    """자동매매 스레드 - 통합 전략 (DatabaseWorker 제거 반영)"""
    
    buy_signal = pyqtSignal(str, str, str, str)
    sell_signal = pyqtSignal(str, str)
    sell_half_signal = pyqtSignal(str, str)
    sell_all_signal = pyqtSignal()
    stock_removed_from_monitor = pyqtSignal(str)
    counter_updated = pyqtSignal(int)
    stock_data_updated = pyqtSignal(list)
    
    # 계산 작업 시그널
    calculation_requested = pyqtSignal(str, dict, dict)  # (종목코드, 틱데이터, 분봉데이터)
    db_save_requested = pyqtSignal(str, dict)  # (종목코드, 저장데이터)

    def __init__(self, trader, window):
        super().__init__()
        
        self.trader = trader
        self.window = window
        
        self.running = True
        self.sell_all_emitted = False
        
        self.counter = 0
        
        self.volatility_strategy = None
        
        self.load_trading_settings()
        
        self.last_evaluation_time = {}
        self.evaluation_lock = threading.Lock()
        
        # DB 저장용
        self.last_save_time = {}
        self.save_lock = threading.Lock()
        
        # 계산 워커 스레드 초기화
        self.calculation_worker = CalculationWorker(trader)
        self.calculation_worker.calculation_completed.connect(self._on_calculation_completed)
        self.calculation_worker.db_save_completed.connect(self._on_db_save_completed)
        self.calculation_worker.start()

    def _on_calculation_completed(self, code, result):
        """계산 완료 처리"""
        try:
            if result:
                # 계산 결과를 이용한 매매 판단 로직
                self._process_calculation_result(code, result)
        except Exception as ex:
            logging.error(f"계산 완료 처리 오류 ({code}): {ex}")

    def _on_db_save_completed(self, code, success):
        """DB 저장 완료 처리"""
        try:
            if success:
                logging.debug(f"DB 저장 완료: {code}")
            else:
                logging.warning(f"DB 저장 실패: {code}")
        except Exception as ex:
            logging.error(f"DB 저장 완료 처리 오류 ({code}): {ex}")

    def _process_calculation_result(self, code, result):
        """계산 결과 처리 및 매매 판단"""
        try:
            if code not in self.trader.monistock_set:
                return
            
            t_now = datetime.now()
            if not self._is_trading_hours(t_now):
                return
            
            # 계산된 지표 데이터 사용
            calculated_indicators = result.get('calculated_indicators', {})
            tick_data = result.get('tick_data', {})
            min_data = result.get('min_data', {})
            
            # 매매 조건 평가
            current_strategy = self.window.comboStg.currentText()
            buy_strategies = [
                stg for stg in self.window.strategies.get(current_strategy, []) 
                if stg['key'].startswith('buy')
            ]
            sell_strategies = [
                stg for stg in self.window.strategies.get(current_strategy, []) 
                if stg['key'].startswith('sell')
            ]
            
            # 매수 조건 평가
            if buy_strategies and code not in self.trader.bought_set and code not in self.trader.buyorder_set:
                for strategy in buy_strategies:
                    if self._evaluate_strategy_conditions(code, [strategy], tick_data, min_data):
                        self.buy_signal.emit(code, strategy['name'], strategy.get('order_condition', '0'), strategy.get('order_style', '03'))
                        break
            
            # 매도 조건 평가
            if sell_strategies and code in self.trader.bought_set and code not in self.trader.sellorder_set:
                for strategy in sell_strategies:
                    if self._evaluate_sell_condition(code, t_now, strategy, sell_strategies):
                        if strategy['key'] == 'sell_half':
                            self.sell_half_signal.emit(code, strategy['name'])
                        else:
                            self.sell_signal.emit(code, strategy['name'])
                        break
                        
        except Exception as ex:
            logging.error(f"계산 결과 처리 오류 ({code}): {ex}")

    def load_trading_settings(self):
        """매매 평가 설정 로드"""
        config = configparser.RawConfigParser()
        if os.path.exists('settings.ini'):
            config.read('settings.ini', encoding='utf-8')
        
        # 기본값 설정
        self.evaluation_interval = config.getint('TRADING', 'evaluation_interval', fallback=3)  # 5초 → 3초
        self.event_based_evaluation = config.getboolean('TRADING', 'event_based_evaluation', fallback=True)
        self.min_evaluation_gap = config.getfloat('TRADING', 'min_evaluation_gap', fallback=3.0)
        
        logging.debug(
            f"매매 평가 설정: 주기={self.evaluation_interval}초, "
            f"이벤트기반={self.event_based_evaluation}, "
            f"최소간격={self.min_evaluation_gap}초"
        )

    def set_volatility_strategy(self, strategy):
        """변동성 돌파 전략 설정"""
        self.volatility_strategy = strategy

    def connect_bar_signals(self):
        """봉 완성 signal 연결"""
        if self.event_based_evaluation:
            # 틱봉 완성 시
            self.trader.tickdata.new_bar_completed.connect(self.on_tick_bar_completed)
            # 분봉 완성 시
            self.trader.mindata.new_bar_completed.connect(self.on_min_bar_completed)
            logging.info("이벤트 기반 매매 평가 활성화")

    @pyqtSlot(str)
    def on_tick_bar_completed(self, code):
        """틱봉 완성 시 즉시 평가"""
        self._evaluate_code_if_ready(code, "틱봉 완성")

    @pyqtSlot(str)
    def on_min_bar_completed(self, code):
        """분봉 완성 시 즉시 평가"""
        self._evaluate_code_if_ready(code, "분봉 완성")

    def _evaluate_code_if_ready(self, code, reason):
        """종목 평가 (최소 간격 체크) + DB 저장"""
        
        if code not in self.trader.monistock_set:
            return
        
        with self.evaluation_lock:
            now = time.time()
            last_time = self.last_evaluation_time.get(code, 0)
            
            if now - last_time < self.min_evaluation_gap:
                return
            
            self.last_evaluation_time[code] = now
        
        t_now = datetime.now()
        
        if not self._is_trading_hours(t_now):
            return
        
        try:
            # 현재 데이터 조회 (API 호출은 메인 스레드에서)
            tick_latest = self.trader.tickdata.get_latest_data(code)
            min_latest = self.trader.mindata.get_latest_data(code)
            
            if not tick_latest or not min_latest:
                return
            
            # 계산 작업을 워커 스레드로 위임
            self.calculation_worker.add_calculation_task(code, tick_latest, min_latest)
            
            # DB 저장 작업을 워커 스레드로 위임
            save_data = {
                'timestamp': t_now,
                'tick_data': tick_latest,
                'min_data': min_latest,
                'reason': reason
            }
            self.calculation_worker.add_db_save_task(code, save_data)
            
            # 매매 조건 평가
            current_strategy = self.window.comboStg.currentText()
            buy_strategies = [
                stg for stg in self.window.strategies.get(current_strategy, []) 
                if stg['key'].startswith('buy')
            ]
            sell_strategies = [
                stg for stg in self.window.strategies.get(current_strategy, []) 
                if stg['key'].startswith('sell')
            ]
            
            if code not in self.trader.buyorder_set and code not in self.trader.bought_set:
                logging.debug(f"{code}: {reason} - 매수 조건 평가")
                self._evaluate_buy_condition(code, t_now, current_strategy, buy_strategies)
            
            elif (code in self.trader.bought_set and 
                  code not in self.trader.buyorder_set and 
                  code not in self.trader.sellorder_set):
                logging.debug(f"{code}: {reason} - 매도 조건 평가")
                self._evaluate_sell_condition(code, t_now, current_strategy, sell_strategies)
                
        except Exception as ex:
            logging.error(f"{code} 이벤트 기반 평가 오류: {ex}")

    def save_to_db_if_needed(self, code, timestamp, tick_data, min_data, trigger_reason):
        """조건부 DB 저장 (새 지표 포함)"""
        
        should_save = False
        save_reason = ""
        last_save_backup = 0
        
        # === 저장 필요 여부 판단 ===
        with self.save_lock:
            now = time.time()
            last_save = self.last_save_time.get(code, 0)
            last_save_backup = last_save
            
            if now - last_save >= 5.0:
                should_save = True
                save_reason = "주기적 저장"
                self.last_save_time[code] = now
            elif "완성" in trigger_reason and now - last_save >= 1.0:
                should_save = True
                save_reason = trigger_reason
                self.last_save_time[code] = now
            elif (code in self.trader.buyorder_set or code in self.trader.sellorder_set):
                should_save = True
                save_reason = "매매 발생"
                self.last_save_time[code] = now
        
        if not should_save:
            return
        
        # === 실제 DB 저장 ===
        try:
            conn = sqlite3.connect(self.trader.db_name, timeout=5)
            cursor = conn.cursor()
            
            date_str = timestamp.strftime('%Y%m%d')
            time_str = timestamp.strftime('%H%M%S')
            
            # 체결강도
            strength = self.trader.tickdata.get_strength(code)
            
            # 포지션 정보
            if code in self.trader.bought_set:
                position_type = 'BOUGHT'
                buy_price = self.trader.buy_price.get(code, None)
            elif code in self.trader.buyorder_set:
                position_type = 'BUYORDER'
                buy_price = None
            else:
                position_type = 'NONE'
                buy_price = None
            
            # INSERT OR REPLACE
            cursor.execute('''
                INSERT OR REPLACE INTO combined_tick_data (
                    code, timestamp, date, time,
                    tick_C, tick_O, tick_H, tick_L, tick_V,
                    tick_MAT5, tick_MAT20, tick_MAT60, tick_MAT120,
                    tick_RSIT, tick_RSIT_SIGNAL,
                    tick_MACDT, tick_MACDT_SIGNAL, tick_OSCT,
                    tick_STOCHK, tick_STOCHD,
                    tick_ATR, tick_CCI,
                    tick_BB_UPPER, tick_BB_MIDDLE, tick_BB_LOWER,
                    tick_BB_POSITION, tick_BB_BANDWIDTH,
                    tick_VWAP,
                    tick_WILLIAMS_R, tick_ROC, tick_OBV, tick_OBV_MA20,
                    tick_VP_POC, tick_VP_POSITION,
                    min_C, min_O, min_H, min_L, min_V,
                    min_MAM5, min_MAM10, min_MAM20,
                    min_RSI, min_RSI_SIGNAL,
                    min_MACD, min_MACD_SIGNAL, min_OSC,
                    min_STOCHK, min_STOCHD,
                    min_CCI, min_VWAP,
                    min_WILLIAMS_R, min_ROC, min_OBV, min_OBV_MA20,
                    strength, buy_price, position_type,
                    save_reason
                ) VALUES (
                    ?, ?, ?, ?,
                    ?, ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?,
                    ?, ?, ?, ?,
                    ?, ?,
                    ?, ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?, ?, ?,
                    ?, ?,
                    ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?,
                    ?
                )
            ''', (
                code, timestamp, date_str, time_str,
                # 틱 데이터
                tick_data.get('C', 0), tick_data.get('O', 0), 
                tick_data.get('H', 0), tick_data.get('L', 0), tick_data.get('V', 0),
                tick_data.get('MAT5', 0), tick_data.get('MAT20', 0), 
                tick_data.get('MAT60', 0), tick_data.get('MAT120', 0),
                tick_data.get('RSIT', 0), tick_data.get('RSIT_SIGNAL', 0),
                tick_data.get('MACDT', 0), tick_data.get('MACDT_SIGNAL', 0), 
                tick_data.get('OSCT', 0),
                tick_data.get('STOCHK', 0), tick_data.get('STOCHD', 0),
                tick_data.get('ATR', 0), tick_data.get('CCI', 0),
                tick_data.get('BB_UPPER', 0), tick_data.get('BB_MIDDLE', 0), 
                tick_data.get('BB_LOWER', 0),
                tick_data.get('BB_POSITION', 0), tick_data.get('BB_BANDWIDTH', 0),
                tick_data.get('VWAP', 0),
                # 새 지표 - 틱
                tick_data.get('WILLIAMS_R', -50), tick_data.get('ROC', 0),
                tick_data.get('OBV', 0), tick_data.get('OBV_MA20', 0),
                tick_data.get('VP_POC', 0), tick_data.get('VP_POSITION', 0),
                # 분봉 데이터
                min_data.get('C', 0), min_data.get('O', 0), 
                min_data.get('H', 0), min_data.get('L', 0), min_data.get('V', 0),
                min_data.get('MAM5', 0), min_data.get('MAM10', 0), 
                min_data.get('MAM20', 0),
                min_data.get('RSI', 0), min_data.get('RSI_SIGNAL', 0),
                min_data.get('MACD', 0), min_data.get('MACD_SIGNAL', 0), 
                min_data.get('OSC', 0),
                min_data.get('STOCHK', 0), min_data.get('STOCHD', 0),
                min_data.get('CCI', 0), min_data.get('VWAP', 0),
                # 새 지표 - 분봉
                min_data.get('WILLIAMS_R', -50), min_data.get('ROC', 0),
                min_data.get('OBV', 0), min_data.get('OBV_MA20', 0),
                # 추가 정보
                strength, buy_price, position_type,
                save_reason
            ))
            
            conn.commit()
            conn.close()
            
            logging.debug(f"{code}: DB 저장 완료 ({save_reason})")
            
        except Exception as ex:
            logging.error(f"{code}: DB 저장 실패 - {ex}")
            
            with self.save_lock:
                self.last_save_time[code] = last_save_backup

    def run(self):
        """메인 루프"""
        while self.running:
            self.autotrade()
            # 5초 주기
            self.msleep(self.evaluation_interval * 1000)

    def stop(self):
        """스레드 정지"""
        logging.info("AutoTraderThread 정지 시작...")
        self.running = False
        
        # 계산 워커 스레드 정지
        if hasattr(self, 'calculation_worker'):
            self.calculation_worker.stop()
            self.calculation_worker.quit()
            self.calculation_worker.wait()
        
        self.quit()
        self.wait()
        logging.info("AutoTraderThread 정지 완료")

    def autotrade(self):
        """자동매매 메인 루프 (주기적 평가)"""
        try:
            t_now = datetime.now()
            
            self.counter += 1
            self.counter_updated.emit(self.counter)
            
            self._update_stock_data_table()
            
            if self._is_trading_hours(t_now):
                # 주기적 평가 + 저장
                self._execute_trading_logic(t_now)
            elif self._is_market_close_time(t_now):
                self._handle_market_close()
                
        except Exception as ex:
            logging.error(f"autotrade 오류: {ex}\n{traceback.format_exc()}")

    def _is_trading_hours(self, t_now):
        """거래 시간인지 확인"""
        t_0903 = t_now.replace(hour=9, minute=3, second=0, microsecond=0)
        t_1515 = t_now.replace(hour=15, minute=15, second=0, microsecond=0)
        return t_0903 < t_now <= t_1515

    def _is_market_close_time(self, t_now):
        """장 종료 시간인지 확인"""
        t_1515 = t_now.replace(hour=15, minute=15, second=0, microsecond=0)
        t_exit = t_now.replace(hour=15, minute=20, second=0, microsecond=0)
        return t_1515 < t_now < t_exit and not self.sell_all_emitted

    def _update_stock_data_table(self):
        """주식 데이터 테이블 업데이트"""
        stock_data_list = []
        
        monitored_codes = list(self.trader.monistock_set.copy())
        
        for code in monitored_codes:
            if code not in self.trader.monistock_set:
                continue
            
            tick_latest = self.trader.tickdata.get_latest_data(code)
            current_price = tick_latest.get('C', 0.0) if tick_latest else 0.0
            buy_price = self.trader.buy_price.get(code, 0.0)
            quantity = self.trader.buy_qty.get(code, 0)

            stock_data_list.append({
                'code': code,
                'current_price': float(current_price),
                'upward_probability': 0.0,  # CNN 제거로 0
                'buy_price': float(buy_price),
                'quantity': quantity
            })
        
        self.stock_data_updated.emit(stock_data_list)

    def _execute_trading_logic(self, t_now):
        """거래 로직 실행 (5초 주기)"""
        
        monitored_codes = list(self.trader.monistock_set.copy())
        
        for code in monitored_codes:
            if code not in self.trader.monistock_set:
                continue
            
            try:
                # 5초마다 평가 + 저장
                self._evaluate_code_if_ready(code, "주기적 평가")
                    
            except Exception as ex:
                logging.error(f"{code} 거래 로직 오류: {ex}")

    def _handle_market_close(self):
        """장 종료 처리 (간소화)"""
        
        if self.trader.buyorder_set or self.trader.sellorder_set:
            return
        
        # 보유 주식 전부 매도
        if self.trader.bought_set:
            logging.info("보유 주식 전부 매도")
            self.sell_all_signal.emit()
        
        self.sell_all_emitted = True

    # ===== 헬퍼 함수들 (변경 없음) =====
    
    def get_threshold_by_hour(self):
        """시간대별 임계값 반환"""
        now = datetime.now()
        hour = now.hour
        
        if hour == 9:
            return 65
        elif hour >= 14:
            return 85
        else:
            return 75
    
    def is_after_time(self, hour, minute):
        """특정 시각 이후인지 확인"""
        now = datetime.now()
        return now >= now.replace(hour=hour, minute=minute, second=0)

    # ===== 매수 조건 평가 =====

    def _evaluate_buy_condition(self, code, t_now, strategy, buy_strategies):
        """매수 조건 평가"""
        tick_latest = self.trader.tickdata.get_latest_data(code)
        min_latest = self.trader.mindata.get_latest_data(code)
        
        if not tick_latest or not min_latest:
            return
        
        if tick_latest.get('MAT5', 0) == 0:
            logging.debug(f"{code}: 지표 준비 미완료")
            return
        
        if self._should_remove_from_monitor(code, tick_latest, min_latest, t_now):
            return
        
        # ===== 통합 전략: 텍스트 기반 평가 =====
        if strategy == "통합 전략" and buy_strategies:
            if self._evaluate_integrated_buy(code, buy_strategies, tick_latest, min_latest):
                return  # 매수 신호 발생
        
        # ===== 기타 전략 =====
        elif self._evaluate_strategy_conditions(code, buy_strategies, tick_latest, min_latest):
            self.buy_signal.emit(code, "사용자 전략", "0", "03")

    def _evaluate_integrated_buy(self, code, buy_strategies, tick_latest, min_latest):
        """매수 평가 - 공통 함수 사용"""
        from strategy_utils import (
            STRATEGY_SAFE_GLOBALS,
            evaluate_strategies,
            build_realtime_buy_locals
        )
        
        # ===== 공통 함수로 변수 구성 =====
        safe_locals = build_realtime_buy_locals(
            code=code,
            tick_latest=tick_latest,
            min_latest=min_latest,
            trader=self.trader,
            window=self.window
        )
        
        # === 전략 평가 (공통 함수 사용) ===
        matched, strategy = evaluate_strategies(
            buy_strategies,
            safe_locals,
            code=code,
            strategy_type="매수"
        )
        
        if matched:
            buy_reason = strategy.get('name', '통합 전략')
            
            # 로그용 변수 추출
            strength = safe_locals.get('strength', 0)
            momentum_score = safe_locals.get('momentum_score', 0)
            WILLIAMS_R = safe_locals.get('WILLIAMS_R', -50)
            ROC = safe_locals.get('ROC', 0)
            
            logging.info(
                f"{cpCodeMgr.CodeToName(code)}({code}): {buy_reason} 매수 "
                f"(체결강도: {strength:.0f}, 점수: {momentum_score}, "
                f"Williams %R: {WILLIAMS_R:.1f}, ROC: {ROC:.2f}%)"
            )
            self.buy_signal.emit(code, buy_reason, "0", "03")
            return True
        
        return False

    def _should_remove_from_monitor(self, code, tick_latest, min_latest, t_now):
        """투자 대상에서 제거해야 하는지 확인"""
        min_close = min_latest.get('C', 0)
        MAM5 = min_latest.get('MAM5', 0)
        MAM10 = min_latest.get('MAM10', 0)
        
        # 급격한 하락
        if code in self.trader.starting_price:
            if min_close < self.trader.starting_price[code] * 0.97:
                if MAM5 < MAM10:
                    try:
                        start_time_str = self.trader.starting_time.get(code, '')
                        if start_time_str:
                            start_time = datetime.strptime(
                                f"{datetime.now().year}/{start_time_str}", 
                                '%Y/%m/%d %H:%M:%S'
                            )
                            if t_now - start_time > timedelta(hours=1):
                                logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 제거 (하락)")
                                self.stock_removed_from_monitor.emit(code)
                                return True
                    except Exception as ex:
                        logging.error(f"{code} 시작 시각 파싱 오류: {ex}")
        
        # 상한가 체크
        min_high_recent = min_latest.get('H_recent', [0, 0])
        min_low_recent = min_latest.get('L_recent', [0, 0])
        
        if len(min_high_recent) >= 2 and len(min_low_recent) >= 2:
            if all(h == l for h, l in zip(min_high_recent[-2:], min_low_recent[-2:])):
                logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 제거 (상한가)")
                self.stock_removed_from_monitor.emit(code)
                return True
        
        return False

    def _evaluate_strategy_conditions(self, code, strategies, tick_latest, min_latest):
        """전략별 조건 평가 (기존 전략용)"""
        if not strategies:
            return False
        
        tick_data_full = self.trader.tickdata.get_recent_data(code, 10)
        min_data_full = self.trader.mindata.get_recent_data(code, 10)
        
        # ===== 매도 조건용 변수들 계산 =====
        tick_close = tick_latest.get('C', 0)
        buy_price = self.trader.buy_price.get(code, 0)
        current_profit_pct = (tick_close / buy_price - 1) * 100 if buy_price > 0 else 0
        
        highest_price = self.trader.highest_price.get(code, buy_price)
        from_peak_pct = (tick_close / highest_price - 1) * 100 if highest_price > 0 else 0
        
        # 보유 시간 계산
        buy_time_str = self.trader.starting_time.get(code)
        if buy_time_str:
            try:
                from datetime import datetime
                buy_time = datetime.strptime(
                    f"{datetime.now().year}/{buy_time_str}", 
                    '%Y/%m/%d %H:%M:%S'
                )
                hold_minutes = (datetime.now() - buy_time).total_seconds() / 60
            except:
                hold_minutes = 0
        else:
            hold_minutes = 0
        
        tick_close_price = tick_data_full.get('C', [0])
        MAT5 = tick_data_full.get('MAT5', [0])
        MAT20 = tick_data_full.get('MAT20', [0])
        MAT60 = tick_data_full.get('MAT60', [0])
        RSIT = tick_data_full.get('RSIT', [0])
        OSCT = tick_data_full.get('OSCT', [0])
        bb_upper = tick_data_full.get('BB_UPPER', [0])
        
        min_close_price = min_data_full.get('C', [0])
        MAM5 = min_data_full.get('MAM5', [0])
        MAM10 = min_data_full.get('MAM10', [0])
        RSI = min_data_full.get('RSI', [0])
        OSC = min_data_full.get('OSC', [0])
        VWAP = min_data_full.get('VWAP', [0])
        
        min_close_recent = min_data_full.get('C', [0, 0])[-2:]
        min_open_recent = min_data_full.get('O', [0, 0])[-2:]
        positive_candle = all(
            min_close_recent[i] > min_open_recent[i] 
            for i in range(min(2, len(min_close_recent), len(min_open_recent)))
        )
        
        # ===== safe_locals 변수 정의 =====
        safe_locals = {
            # 최신 데이터 변수들 (전략에서 사용) - 단일 값만 사용
            'C': tick_latest.get('C', 0),  # 현재 틱 종가
            'MAT5': tick_latest.get('MAT5', 0),  # 현재 MAT5
            'MAT20': tick_latest.get('MAT20', 0),  # 현재 MAT20
            'MAT60': tick_latest.get('MAT60', 0),  # 현재 MAT60
            'MAM5': min_latest.get('MAM5', 0),  # 현재 MAM5
            'MAM10': min_latest.get('MAM10', 0),  # 현재 MAM10
            'min_MAM5': min_latest.get('MAM5', 0),  # 분봉 MAM5 (호환성)
            'min_MAM10': min_latest.get('MAM10', 0),  # 분봉 MAM10 (호환성)
            'RSIT': tick_latest.get('RSIT', 50),  # 현재 RSIT
            'RSI': min_latest.get('RSI', 50),  # 현재 RSI
            'min_RSI': min_latest.get('RSI', 50),  # 분봉 RSI (호환성)
            'OSCT': tick_latest.get('OSCT', 0),  # 현재 OSCT
            'OSC': min_latest.get('OSC', 0),  # 현재 OSC
            'min_OSC': min_latest.get('OSC', 0),  # 분봉 OSC (호환성)
            'VWAP': min_latest.get('VWAP', 0),  # 현재 VWAP
            'min_VWAP': min_latest.get('VWAP', 0),  # 분봉 VWAP (호환성)
            'min_CCI': min_latest.get('CCI', 0),  # 분봉 CCI (호환성)
            'min_MACD': min_latest.get('MACD', 0),  # 분봉 MACD (호환성)
            'min_MACD_SIGNAL': min_latest.get('MACD_SIGNAL', 0),  # 분봉 MACD_SIGNAL (호환성)
            'strength': tick_latest.get('strength', 0),  # 체결강도
            
            # 추가 변수들
            'positive_candle': positive_candle,
            'tick_VWAP': tick_latest.get('VWAP', 0),  # 틱 VWAP
            'STOCHK': tick_latest.get('STOCHK', 50),  # 스토캐스틱 K
            'STOCHD': tick_latest.get('STOCHD', 50),  # 스토캐스틱 D
            'tick_STOCHK': tick_latest.get('STOCHK', 50),  # 틱 스토캐스틱 K (호환성)
            'tick_STOCHD': tick_latest.get('STOCHD', 50),  # 틱 스토캐스틱 D (호환성)
            'tick_RSI': tick_latest.get('RSIT', 50),  # 틱 RSI (호환성)
            'tick_CCI': tick_latest.get('CCI', 0),  # 틱 CCI (호환성)
            'tick_MACD': tick_latest.get('MACDT', 0),  # 틱 MACD (호환성)
            'MACDT': tick_latest.get('MACDT', 0),  # 틱 MACDT (호환성)
            'MACD': tick_latest.get('MACDT', 0),  # 틱 MACD (호환성)
            'MAT5': tick_latest.get('MAT5', 0),  # 틱 MAT5 (호환성)
            'MAT20': tick_latest.get('MAT20', 0),  # 틱 MAT20 (호환성)
            'MAT60': tick_latest.get('MAT60', 0),  # 틱 MAT60 (호환성)
            'ATR': tick_latest.get('ATR', 0),  # 틱 ATR (호환성)
            'CCI': tick_latest.get('CCI', 0),  # 틱 CCI (호환성)
            'min_STOCHK': min_latest.get('STOCHK', 50),  # 분봉 스토캐스틱 K (호환성)
            'min_STOCHD': min_latest.get('STOCHD', 50),  # 분봉 스토캐스틱 D (호환성)
            'MAM20': min_latest.get('MAM20', 0),  # 분봉 MAM20 (호환성)
            'MAM60': min_latest.get('MAM60', 0),  # 분봉 MAM60 (호환성)
            'MAM120': min_latest.get('MAM120', 0),  # 분봉 MAM120 (호환성)
            
            # 갭상승 전략 변수들
            'gap_hold': self._check_gap_hold(code, tick_latest),  # 갭 유지 확인
            'BB_MIDDLE': tick_latest.get('BB_MIDDLE', 0),  # 볼린저 밴드 중간선
            'BB_POSITION': tick_latest.get('BB_POSITION', 0),  # 볼린저 밴드 위치
            
            # 추가 지표들
            'min_close': min_latest.get('C', 0),  # 분봉 종가
            'tick_close_price': tick_latest.get('C', 0),  # 틱 종가 (배열이 아닌 현재값)
            
            # 신호 변수들 (단일 값으로 수정)
            'RSIT_SIGNAL': tick_latest.get('RSIT_SIGNAL', 0),  # RSIT 신호 (단일 값)
            'tick_C_recent': tick_latest.get('C', 0),  # 틱 종가 (단일 값)
            
            # 누락된 변수들 추가 (단일 값으로 수정)
            'bb_upper': tick_latest.get('BB_UPPER', 0),  # 볼린저 상단선 (단일 값)
            'MACDT_SIGNAL': tick_latest.get('MACDT_SIGNAL', 0),  # MACDT 신호 (단일 값)
            
            # 통합 전략용 추가 변수들
            'WILLIAMS_R': tick_latest.get('WILLIAMS_R', -50),  # Williams %R
            'min_WILLIAMS_R': min_latest.get('WILLIAMS_R', -50),  # 분봉 Williams %R
            'ROC': tick_latest.get('ROC', 0),  # Rate of Change
            'min_ROC': min_latest.get('ROC', 0),  # 분봉 ROC
            'OBV': tick_latest.get('OBV', 0),  # On Balance Volume
            'OBV_MA20': tick_latest.get('OBV_MA20', 0),  # OBV 20일 이동평균
            'min_OBV': min_latest.get('OBV', 0),  # 분봉 OBV
            'min_OBV_MA20': min_latest.get('OBV_MA20', 0),  # 분봉 OBV MA20
            'VP_POC': tick_latest.get('VP_POC', 0),  # Volume Profile POC
            'VP_POSITION': tick_latest.get('VP_POSITION', 0),  # Volume Profile Position
            'volume_profile_breakout': self._check_volume_profile_breakout(code, tick_latest),  # Volume Profile 돌파
            'volatility_breakout': self._check_volatility_breakout(code, tick_latest),  # 변동성 돌파
            'BB_BANDWIDTH': tick_latest.get('BB_BANDWIDTH', 0),  # 볼린저 밴드폭
            'ATR': tick_latest.get('ATR', 0),  # Average True Range
            
            # 매도 조건용 변수들
            'current_profit_pct': current_profit_pct,  # 현재 수익률 (%)
            'from_peak_pct': from_peak_pct,  # 고점 대비 수익률 (%)
            'hold_minutes': hold_minutes,  # 보유 시간 (분)
            
            # 특수 변수들 (매도 조건용)
            'self': self,  # self 객체 접근용
            'code': code,  # 종목 코드
            'after_market_close': self._is_after_market_close(),  # 장 마감 후 여부
        }
        
        # ===== safe_globals 정의 =====
        safe_globals = {
            '__builtins__': {
                'min': min, 'max': max, 'abs': abs, 'round': round,
                'int': int, 'float': float, 'bool': bool, 'str': str,
                'sum': sum, 'all': all, 'any': any,
                'True': True, 'False': False, 'None': None,
                # len 함수를 안전하게 래핑
                'len': lambda x: len(x) if hasattr(x, '__len__') else 1
            }
        }
        
        for strategy in strategies:
            try:
                condition = strategy.get('content', '')
                if eval(condition, safe_globals, safe_locals):
                    logging.debug(f"{code}: {strategy.get('name')} 조건 만족")
                    return True
            except Exception as ex:
                logging.error(f"{code} 전략 평가 오류: {ex}")
        
        return False

    def _check_gap_hold(self, code, tick_latest):
        """갭 유지 확인 (매수 조건)"""
        try:
            # 전일 종가 가져오기
            day_data = self.trader.daydata.stockdata.get(code, {})
            if not day_data or 'C' not in day_data:
                return False
            
            # 전일 종가 (마지막 값)
            prev_close = day_data['C'][-1] if day_data['C'] else 0
            
            if prev_close == 0:
                return False
            
            # 현재가
            current_price = tick_latest.get('C', 0)
            
            if current_price == 0:
                return False
            
            # 갭 유지 확인 (시가 대비 -0.3% 이내면 갭 유지로 판단)
            gap_ratio = (current_price - prev_close) / prev_close
            gap_hold = gap_ratio >= -0.003  # -0.3% 이상이면 갭 유지
            
            return gap_hold
            
        except Exception as ex:
            logging.error(f"_check_gap_hold({code}): {ex}")
            return False

    def _is_after_market_close(self):
        """장 마감 후 여부 확인 (14:45 이후)"""
        try:
            from datetime import datetime
            now = datetime.now()
            market_close_time = now.replace(hour=14, minute=45, second=0, microsecond=0)
            return now >= market_close_time
        except Exception as ex:
            logging.error(f"_is_after_market_close: {ex}")
            return False

    def _check_volume_profile_breakout(self, code, tick_latest):
        """Volume Profile 돌파 확인"""
        try:
            vp_position = tick_latest.get('VP_POSITION', 0)
            return vp_position > 0  # 현재가가 POC 위에 있으면 돌파
        except Exception as ex:
            logging.error(f"_check_volume_profile_breakout({code}): {ex}")
            return False

    def _check_volatility_breakout(self, code, tick_latest):
        """변동성 돌파 확인"""
        try:
            atr = tick_latest.get('ATR', 0)
            current_price = tick_latest.get('C', 0)
            if current_price == 0:
                return False
            
            # ATR이 현재가의 1% 이상 5% 이하면 변동성 돌파로 판단
            atr_ratio = atr / current_price
            return 0.01 <= atr_ratio <= 0.05
        except Exception as ex:
            logging.error(f"_check_volatility_breakout({code}): {ex}")
            return False

    # ===== 매도 조건 평가 =====

    def _evaluate_sell_condition(self, code, t_now, strategy, sell_strategies):
        """매도 조건 평가"""
        tick_latest = self.trader.tickdata.get_latest_data(code)
        min_latest = self.trader.mindata.get_latest_data(code)
        
        if not tick_latest or not min_latest:
            return
        
        tick_close = tick_latest.get('C', 0)
        
        self.trader.update_highest_price(code, tick_close)
        
        buy_price = self.trader.buy_price.get(code, 0)
        if buy_price == 0:
            return
        
        # ===== 수익률 계산 =====
        current_profit_pct = (tick_close / buy_price - 1) * 100
        highest_price = self.trader.highest_price.get(code, buy_price)
        from_peak_pct = (tick_close / highest_price - 1) * 100
        
        # ===== 보유 시간 계산 =====
        buy_time_str = self.trader.starting_time.get(code)
        if buy_time_str:
            try:
                buy_time = datetime.strptime(
                    f"{datetime.now().year}/{buy_time_str}", 
                    '%Y/%m/%d %H:%M:%S'
                )
                hold_minutes = (t_now - buy_time).total_seconds() / 60
            except:
                hold_minutes = 0
        else:
            hold_minutes = 0
        
        # ===== 통합 전략: 텍스트 기반 평가 =====
        if strategy == "통합 전략" and sell_strategies:
            if self._evaluate_integrated_sell(code, sell_strategies, tick_latest, min_latest,
                                             current_profit_pct, from_peak_pct, hold_minutes):
                return
        
        # ===== 기타 전략 =====
        elif self._evaluate_strategy_conditions(code, sell_strategies, tick_latest, min_latest):
            self.sell_signal.emit(code, "전략 매도")

    def _evaluate_integrated_sell(self, code, sell_strategies, tick_latest, min_latest,
                              current_profit_pct, from_peak_pct, hold_minutes):
        """매도 평가 - 공통 함수 사용"""
        from strategy_utils import (
            STRATEGY_SAFE_GLOBALS,
            evaluate_strategies,
            build_realtime_sell_locals
        )
        
        # 매수 시간 가져오기
        buy_time_str = self.trader.starting_time.get(code, '')
        buy_price = self.trader.buy_price.get(code, 0)
        highest_price = self.trader.highest_price.get(code, buy_price)
        
        # ===== 공통 함수로 변수 구성 =====
        safe_locals = build_realtime_sell_locals(
            code=code,
            tick_latest=tick_latest,
            min_latest=min_latest,
            trader=self.trader,
            buy_price=buy_price,
            highest_price=highest_price,
            buy_time_str=buy_time_str,
            window=self.window
        )
        
        # === 전략 평가 (공통 함수 사용) ===
        matched, strategy = evaluate_strategies(
            sell_strategies,
            safe_locals,
            code=code,
            strategy_type="매도"
        )
        
        if matched:
            sell_reason = strategy.get('name', '통합 전략')
            
            # 로그용 변수 추출
            WILLIAMS_R = safe_locals.get('WILLIAMS_R', -50)
            ROC = safe_locals.get('ROC', 0)
            current_profit_pct = safe_locals.get('current_profit_pct', 0)
            hold_minutes = safe_locals.get('hold_minutes', 0)
            
            logging.info(
                f"{cpCodeMgr.CodeToName(code)}({code}): {sell_reason} "
                f"({current_profit_pct:+.2f}%, {hold_minutes:.0f}분 보유, "
                f"Williams %R: {WILLIAMS_R:.1f}, ROC: {ROC:.2f}%)"
            )
            
            if '분할' in sell_reason:
                self.sell_half_signal.emit(code, sell_reason)
            else:
                self.sell_signal.emit(code, sell_reason)
            
            return True
        
        return False

# ==================== ChartDrawer 관련 클래스 (유지) ====================
class ChartDrawerThread(QThread):
    data_ready = pyqtSignal(dict)

    def __init__(self, trader, code):
        super().__init__()
        self.trader = trader
        self.code = code
        self.is_running = True

    def run(self):
        while self.is_running:
            if self.code:
                tick_data = self.trader.tickdata.get_full_data(self.code)
                min_data = self.trader.mindata.get_full_data(self.code)

                # ===== ✅ 데이터 유효성 체크 =====
                tick_valid = (tick_data and 
                             len(tick_data.get('D', [])) > 0 and 
                             len(tick_data.get('C', [])) > 0)
                
                min_valid = (min_data and 
                            len(min_data.get('D', [])) > 0 and 
                            len(min_data.get('C', [])) > 0)
                
                # 둘 다 유효한 경우에만 emit
                if tick_valid and min_valid:
                    data = {'tick_data': tick_data, 'min_data': min_data, 'code': self.code}
                    self.data_ready.emit(data)
                    logging.debug(f"📊 {self.code}: 차트 데이터 업데이트 완료")
                else:
                    # 디버그 로그 (10초마다만 출력)
                    current_time = time.time()
                    if not hasattr(self, '_last_debug_time'):
                        self._last_debug_time = 0
                    
                    if current_time - self._last_debug_time >= 10:
                        if not tick_valid:
                            logging.debug(f"📊 {self.code}: 틱 데이터 대기 중 (D:{len(tick_data.get('D', [])) if tick_data else 0}, C:{len(tick_data.get('C', [])) if tick_data else 0})")
                        if not min_valid:
                            logging.debug(f"📊 {self.code}: 분봉 데이터 대기 중 (D:{len(min_data.get('D', [])) if min_data else 0}, C:{len(min_data.get('C', [])) if min_data else 0})")
                        self._last_debug_time = current_time
                
            self.msleep(2000)  # 1초 → 2초로 조정 (UI 부하 감소)

    def stop(self):
        self.is_running = False
        self.quit()
        self.wait()

class ChartDrawer(QObject):
    def __init__(self, fig, canvas, trader, trader_thread, window):
        super().__init__()
        self.fig = fig
        self.canvas = canvas
        self.trader = trader
        self.trader_thread = trader_thread
        self.window = window
        self.code = None
        self.create_initial_chart()
        self.chart_thread = None
        self.stop_loss = {}
        self.last_chart_update_time = None

    def set_code(self, code):
        self.code = code
        if self.chart_thread:
            self.chart_thread.stop()
            self.chart_thread = None

        if code:
            self.chart_thread = ChartDrawerThread(self.trader, code)
            self.chart_thread.data_ready.connect(self.update_chart)
            self.chart_thread.start()
        else:
            self.create_initial_chart()

    def create_initial_chart(self):
        self.fig.clear()
        self.tick_axes = [self.fig.add_subplot(18, 1, (1, 4), title='Tick Chart'), self.fig.add_subplot(18, 1, (5, 6)), self.fig.add_subplot(18, 1, (7, 8))]
        self.min_axes = [self.fig.add_subplot(18, 1, (10, 13), title='3 Min Chart'), self.fig.add_subplot(18, 1, 14), self.fig.add_subplot(18, 1, (15, 16)), self.fig.add_subplot(18, 1, (17, 18))]

        for ax in self.tick_axes + self.min_axes:
            ax.clear()
            ax.set_xticklabels([])
            ax.set_yticklabels([])
            ax.grid(True)

        self.fig.subplots_adjust(hspace=0)
        plt.tight_layout(rect=[0, 0.03, 1, 0.95])
        self.canvas.draw()
        self.canvas.flush_events()

    @pyqtSlot(dict)
    def update_chart(self, data):
        self.draw_chart(data)
        self.last_chart_update_time = data['tick_data'].get('T', [None])[-1]
        
        # 차트 업데이트 시 상태 라벨도 함께 업데이트
        if hasattr(self, 'window') and hasattr(self.window, 'chart_status_label'):
            self.window._update_chart_status_label()

    def draw_chart(self, data):
        try:
            if data['code']:
                code_name = cpCodeMgr.CodeToName(data['code'])
                for ax in self.tick_axes:
                    ax.clear()
                for ax in self.min_axes:
                    ax.clear()

                self.draw_chart_data(data['tick_data'], self.tick_axes, data['code'], data_type='tick')
                self.draw_chart_data(data['min_data'], self.min_axes, data['code'], data_type='min')
                self.fig.suptitle(f"Chart for {data['code']}({code_name})", fontsize=15)
                self.fig.subplots_adjust(hspace=0)
                self.canvas.draw()
            else:
                self.create_initial_chart()

        except Exception as ex:
            logging.error(f"draw_chart -> {ex}")

    def draw_chart_data(self, chart_data, axes, code, data_type):
        """차트 데이터 그리기"""
        try:
            current_strategy = self.window.comboStg.currentText()
            all_strategies = self.window.strategies.get(current_strategy, [])
            
            if not chart_data:
                logging.debug(f"{code}: 차트 데이터 없음 ({data_type})")
                return
            
            if data_type == 'tick':
                keys_to_keep = ['O', 'H', 'L', 'C', 'V', 'D', 'T', 'MAT5', 'MAT20', 'MAT60', 'MAT120', 'RSIT', 'RSIT_SIGNAL', 'MACDT', 'MACDT_SIGNAL', 'OSCT']
            elif data_type == 'min':
                keys_to_keep = ['O', 'H', 'L', 'C', 'V', 'D', 'T', 'MAM5', 'MAM10', 'MAM20', 'RSI', 'RSI_SIGNAL', 'MACD', 'MACD_SIGNAL', 'OSC']
            else:
                return

            if data_type == 'tick':
                filtered_data = {
                    key: [x for x in (chart_data[key][-90:] if len(chart_data[key]) >= 90 
                                    else chart_data[key][:])] 
                    for key in keys_to_keep
                }

                # ===== ✅ 데이터 유효성 검사 (경고 → 디버그) =====
                if len(filtered_data['D']) == 0 or len(filtered_data['C']) == 0:
                    logging.debug(f"{code}: 틱 차트 데이터 부족")
                    return

                # ===== ✅ 배열 길이 확인 및 조정 =====
                base_length = len(filtered_data['C'])
                
                # 모든 배열을 base_length로 맞춤
                def ensure_length(data, key, default_value=0):
                    arr = data.get(key, [])
                    if len(arr) == base_length:
                        return arr
                    elif len(arr) > base_length:
                        return arr[-base_length:]
                    else:
                        # 부족한 경우 앞을 default_value로 채움
                        return [default_value] * (base_length - len(arr)) + list(arr)
                
                df = pd.DataFrame({
                    'Open': ensure_length(filtered_data, 'O'), 
                    'High': ensure_length(filtered_data, 'H'), 
                    'Low': ensure_length(filtered_data, 'L'), 
                    'Close': ensure_length(filtered_data, 'C'), 
                    'Volume': ensure_length(filtered_data, 'V'), 
                    'Date': ensure_length(filtered_data, 'D'), 
                    'Time': ensure_length(filtered_data, 'T'), 
                    'MAT5': ensure_length(filtered_data, 'MAT5'), 
                    'MAT20': ensure_length(filtered_data, 'MAT20'), 
                    'MAT60': ensure_length(filtered_data, 'MAT60'), 
                    'MAT120': ensure_length(filtered_data, 'MAT120'), 
                    'RSIT': ensure_length(filtered_data, 'RSIT', 50), 
                    'RSIT_SIGNAL': ensure_length(filtered_data, 'RSIT_SIGNAL', 50), 
                    'MACDT': ensure_length(filtered_data, 'MACDT'), 
                    'MACDT_SIGNAL': ensure_length(filtered_data, 'MACDT_SIGNAL'), 
                    'OSCT': ensure_length(filtered_data, 'OSCT')
                })
                df.index = pd.to_datetime(df['Date'].astype(str) + df['Time'].astype(str), format='%Y%m%d%H%M', errors='coerce')

                addplots = [
                    mpf.make_addplot(df['MAT5'].values, color='magenta', label='MAT5', ax=axes[0], width=0.9),
                    mpf.make_addplot(df['MAT20'].values, color='blue', label='MAT20', ax=axes[0], width=0.9),
                    mpf.make_addplot(df['MAT60'].values, color='darkorange', label='MAT60', ax=axes[0], width=1.0),                        
                    mpf.make_addplot(df['MAT120'].values, color='black', label='MAT120', ax=axes[0], width=0.9),
                    mpf.make_addplot(df['RSIT'].values, color='red', label='RSIT', ax=axes[1], width=0.9),
                    mpf.make_addplot(df['RSIT_SIGNAL'].values, color='blue', label='RSIT_SIGNAL', ax=axes[1], width=0.9),
                    mpf.make_addplot(df['MACDT'].values, color='red', label='MACDT', ax=axes[2], width=0.9),
                    mpf.make_addplot(df['MACDT_SIGNAL'].values, color='blue', label='MACDT_SIGNAL', ax=axes[2], width=0.9),
                    mpf.make_addplot(df['OSCT'].values, color='purple', type='bar', label='OSCT', ax=axes[2]),
                ]
                mpf.plot(df, ax=axes[0], type='candle', style='yahoo', addplot=addplots)

                # ===== ✅ Y축 범위를 현재 주가 수준에 맞게 조정 =====
                price_high = df['High'].max()
                price_low = df['Low'].min()
                price_range = price_high - price_low
                
                # 여유 공간 추가 (위아래 5%)
                margin = price_range * 0.05 if price_range > 0 else price_low * 0.05
                y_min = max(0, price_low - margin)  # 0 이하로 내려가지 않도록
                y_max = price_high + margin
                
                axes[0].set_ylim(y_min, y_max)
                
                current_price = df['Close'].iloc[-1]
                axes[0].text(
                    1.01, current_price, f'{current_price}',
                    transform=axes[0].get_yaxis_transform(), color='red',
                    verticalalignment='center', bbox=dict(facecolor='white', edgecolor='none')
                )
                axes[0].legend(loc='upper left')

                x_raw = list(range(len(df['Date'])))
                axes[1].fill_between(x_raw, df['RSIT'], 20, where=(df['RSIT'] <= 20), color='blue', alpha=0.5)
                axes[1].fill_between(x_raw, df['RSIT'], 80, where=(df['RSIT'] >= 80), color='red', alpha=0.5)
                axes[1].set_yticks([20, 50, 80])
                axes[1].legend(loc='upper left')
                axes[2].legend(loc='upper left')

                x_labels = self.create_labels(filtered_data)
                for index, ax in enumerate(axes):
                    ax.set_xlim(-2, len(filtered_data['D']) + 1)
                    ax.grid(True, axis='y')
                    ax.set_xticks(x_raw)
                    if index == 0 or index == 1:
                        ax.set_xticklabels([])
                    if index == 2:
                        ax.set_xticklabels(x_labels)

            elif data_type == 'min':
                filtered_data = {key: [x for x in (chart_data[key][-50:] if len(chart_data[key]) >= 50 else chart_data[key][:])] for key in keys_to_keep}

                # ===== ✅ 데이터 유효성 검사 (경고 → 디버그) =====
                if len(filtered_data['D']) == 0 or len(filtered_data['C']) == 0:
                    logging.debug(f"{code}: 분봉 차트 데이터 부족")
                    return

                # ===== ✅ 배열 길이 확인 및 조정 =====
                base_length = len(filtered_data['C'])
                
                # 모든 배열을 base_length로 맞춤
                def ensure_length_min(data, key, default_value=0):
                    arr = data.get(key, [])
                    if len(arr) == base_length:
                        return arr
                    elif len(arr) > base_length:
                        return arr[-base_length:]
                    else:
                        # 부족한 경우 앞을 default_value로 채움
                        return [default_value] * (base_length - len(arr)) + list(arr)
                
                df = pd.DataFrame({
                    'Open': ensure_length_min(filtered_data, 'O'), 
                    'High': ensure_length_min(filtered_data, 'H'), 
                    'Low': ensure_length_min(filtered_data, 'L'), 
                    'Close': ensure_length_min(filtered_data, 'C'), 
                    'Volume': ensure_length_min(filtered_data, 'V'),
                    'Date': ensure_length_min(filtered_data, 'D'), 
                    'Time': ensure_length_min(filtered_data, 'T'), 
                    'MAM5': ensure_length_min(filtered_data, 'MAM5'), 
                    'MAM10': ensure_length_min(filtered_data, 'MAM10'), 
                    'MAM20': ensure_length_min(filtered_data, 'MAM20'),
                    'RSI': ensure_length_min(filtered_data, 'RSI', 50), 
                    'RSI_SIGNAL': ensure_length_min(filtered_data, 'RSI_SIGNAL', 50), 
                    'MACD': ensure_length_min(filtered_data, 'MACD'), 
                    'MACD_SIGNAL': ensure_length_min(filtered_data, 'MACD_SIGNAL'), 
                    'OSC': ensure_length_min(filtered_data, 'OSC')
                })
                df.index = pd.to_datetime(df['Date'].astype(str) + df['Time'].astype(str), format='%Y%m%d%H%M')

                addplots = [
                    mpf.make_addplot(df['MAM5'].values, color='magenta', label='MAM5', ax=axes[0], width=0.9),
                    mpf.make_addplot(df['MAM10'].values, color='blue', label='MAM10', ax=axes[0], width=0.9),
                    mpf.make_addplot(df['MAM20'].values, color='darkorange', label='MAM20', ax=axes[0], width=1.0),
                    mpf.make_addplot(df['RSI'].values, color='red', label='RSI', ax=axes[2], width=0.9),
                    mpf.make_addplot(df['RSI_SIGNAL'].values, color='blue', label='RSI_SIGNAL', ax=axes[2], width=0.9),
                    mpf.make_addplot(df['MACD'].values, color='red', label='MACD', ax=axes[3], width=0.9),
                    mpf.make_addplot(df['MACD_SIGNAL'].values, color='blue', label='MACD_SIGNAL', ax=axes[3], width=0.9),
                    mpf.make_addplot(df['OSC'].values, color='purple', type='bar', label='OSC', ax=axes[3])
                ]
                mpf.plot(df, ax=axes[0], type='candle', style='yahoo', volume=axes[1], addplot=addplots)

                # ===== ✅ Y축 범위를 현재 주가 수준에 맞게 조정 =====
                price_high = df['High'].max()
                price_low = df['Low'].min()
                price_range = price_high - price_low
                
                # 여유 공간 추가 (위아래 5%)
                margin = price_range * 0.05 if price_range > 0 else price_low * 0.05
                y_min = max(0, price_low - margin)  # 0 이하로 내려가지 않도록
                y_max = price_high + margin
                
                axes[0].set_ylim(y_min, y_max)
                
                current_price = df['Close'].iloc[-1]
                axes[0].text(
                    1.01, current_price, f'<{current_price}>',
                    transform=axes[0].get_yaxis_transform(), color='red',
                    verticalalignment='center', bbox=dict(facecolor='white', edgecolor='none')
                )
                axes[0].legend(loc='upper left')
                
                x_raw = list(range(len(df['Date'])))
                axes[2].fill_between(x_raw, df['RSI'], 30, where=(df['RSI'] <= 30), color='blue', alpha=0.5)
                axes[2].fill_between(x_raw, df['RSI'], 70, where=(df['RSI'] >= 70), color='red', alpha=0.5)
                axes[2].set_yticks([30, 50, 70])
                axes[2].legend(loc='upper left')
                axes[3].legend(loc='upper left')

                x_labels = self.create_labels(filtered_data)
                for index, ax in enumerate(axes):
                    ax.set_xlim(-2, len(filtered_data['D']) + 1)
                    ax.grid(True, axis='y')
                    ax.set_xticks(x_raw)
                    if index == 0 or index == 1 or index == 2:
                        ax.set_xticklabels([])
                    if index == 3:
                        ax.set_xticklabels(x_labels)
                axes[1].yaxis.set_label_position("right")
                axes[1].yaxis.tick_right()
                axes[1].grid(False)

            if code in self.trader.starting_price:
                starting_price_line = self.trader.starting_price[code]
                
                # ===== ✅ Starting Price가 0이 아닌 경우만 처리 =====
                if starting_price_line > 0:
                    y_min, y_max = axes[0].get_ylim()
                    
                    # Starting Price가 범위 밖이면 범위 확장
                    if starting_price_line < y_min or starting_price_line > y_max:
                        axes[0].set_ylim(min(y_min, starting_price_line * 0.99), max(y_max, starting_price_line * 1.01))
                    
                    axes[0].axhline(y=starting_price_line, color='orangered', linestyle='--', linewidth=1, label='Starting Price')

            if code in self.trader.buy_price:
                buy_price_line = self.trader.buy_price[code]
                
                # ===== ✅ Buy Price가 0이 아닌 경우만 처리 =====
                if buy_price_line > 0:
                    y_min, y_max = axes[0].get_ylim()
                    
                    # Buy Price가 범위 밖이면 범위 확장
                    if buy_price_line < y_min or buy_price_line > y_max:
                        axes[0].set_ylim(min(y_min, buy_price_line * 0.99), max(y_max, buy_price_line * 1.01))
                    
                    axes[0].axhline(y=buy_price_line, color='red', linestyle='-', linewidth=1, label='Buy Price')

        except Exception as ex:
            logging.error(f"draw_chart_data -> {code}, {ex}\n{traceback.format_exc()}")

    def create_labels(self, filtered_data):
        labels = []
        current_hour = None
        mmm_0_added = False
        mmm_15_added = False
        mmm_30_added = False
        mmm_45_added = False

        for i in range(len(filtered_data['D'])):
            date = filtered_data['D'][i]
            time = filtered_data['T'][i]

            yy, mmdd = divmod(date, 10000)
            mm, dd = divmod(mmdd, 100)
            formatted_date = f"{mm:02}/{dd:02} "

            hhh, mmm = divmod(time, 100)
            formatted_time = f"{hhh:02}:{mmm:02}"

            if hhh != current_hour:
                current_hour = hhh
                mmm_0_added = False
                mmm_15_added = False
                mmm_30_added = False
                mmm_45_added = False

            label = ''

            if i == 0:
                label = f"{formatted_date}{formatted_time}"
            elif i == len(filtered_data['D']) - 1:
                label = f"{formatted_date}{formatted_time}"
            else:
                if mmm == 0:
                    if not mmm_0_added:
                        label = formatted_time
                        mmm_0_added = True
                elif mmm == 15:
                    if not mmm_15_added:
                        label = formatted_time
                        mmm_15_added = True
                elif mmm == 30:
                    if not mmm_30_added:
                        label = formatted_time
                        mmm_30_added = True
                elif mmm == 45:
                    if not mmm_45_added:
                        label = formatted_time
                        mmm_45_added = True

            labels.append(label if label else '')

        return labels

# ==================== LoginHandler ====================
class LoginHandler:
    def __init__(self, parent_window):
        self.parent = parent_window
        # ✅ RawConfigParser 사용 (% 문자 이슈 완전 해결)
        self.config = configparser.RawConfigParser()
        self.config_file = 'settings.ini'
        self.process = None
        self.slack = None
        self.slack_channel = '#stock'
        
        # ===== ✅ 자동 클릭 타이머 추가 =====
        self.auto_click_timer = QTimer()
        self.auto_click_timer.timeout.connect(self.check_and_click_popup)
        self.auto_click_attempts = 0
        self.max_auto_click_attempts = 50  # ✅ 30회 → 50회 (25초, 여유 확보)
        self.window_found = False  # 창을 찾았는지 플래그

        self.init_default_strategy_types()

    def init_default_strategy_types(self):
        """전략 타입 기본값 초기화 (settings.ini에 없으면 생성)"""
        try:
            # 설정 파일 읽기
            if os.path.exists(self.config_file):
                self.config.read(self.config_file, encoding='utf-8')
            
            # STRATEGY_TYPES 섹션이 없으면 생성
            if not self.config.has_section('STRATEGY_TYPES'):
                self.config.add_section('STRATEGY_TYPES')
                
                # 기본값 설정
                self.config.set('STRATEGY_TYPES', 'static', 
                    '전일급등,전일거래량급증,52주신고가,이격도상승,신고가돌파')
                self.config.set('STRATEGY_TYPES', 'dynamic', 
                    '급등주,갭상승,거래량급증,돌파,급등돌파')
                self.config.set('STRATEGY_TYPES', 'max_static_load', '10')
                
                # 파일 저장
                with open(self.config_file, 'w', encoding='utf-8') as configfile:
                    self.config.write(configfile)
                
                logging.info("✅ 전략 타입 기본 설정 생성 완료")
            
        except Exception as ex:
            logging.error(f"init_default_strategy_types: {ex}")

    def load_settings(self):
        """설정 로드"""
        if os.path.exists(self.config_file):
            self.config.read(self.config_file, encoding='utf-8')
            self.parent.loginEdit.setText(self.config.get('LOGIN', 'username', fallback=''))
            self.parent.passwordEdit.setText(self.config.get('LOGIN', 'password', fallback=''))
            self.parent.certpasswordEdit.setText(self.config.get('LOGIN', 'certpassword', fallback=''))
            self.parent.autoLoginCheckBox.setChecked(self.config.getboolean('LOGIN', 'autologin', fallback=False))

            self.parent.buycountEdit.setText(self.config.get('BUYCOUNT', 'target_buy_count', fallback='3'))

            if self.config.has_section('SLACK'):
                token = self.config.get('SLACK', 'token', fallback='')
                self.slack_channel = self.config.get('SLACK', 'channel', fallback='#stock')
                if token:
                    self.slack = Slacker(token)

    def save_settings(self):
        if not self.config.has_section('LOGIN'):
            self.config.add_section('LOGIN')
        self.config.set('LOGIN', 'username', self.parent.loginEdit.text())
        self.config.set('LOGIN', 'password', self.parent.passwordEdit.text())
        self.config.set('LOGIN', 'certpassword', self.parent.certpasswordEdit.text())
        self.config.set('LOGIN', 'autologin', str(self.parent.autoLoginCheckBox.isChecked()))
        with open(self.config_file, 'w', encoding='utf-8') as configfile:
            self.config.write(configfile)

    def attempt_auto_login(self):
        """자동 로그인 시도"""
        try:
            if (self.parent.autoLoginCheckBox.isChecked() and 
                self.parent.loginEdit.text() and 
                self.parent.passwordEdit.text() and 
                self.parent.certpasswordEdit.text()):
                self.handle_login()
        except Exception as ex:
            logging.error(f"attempt_auto_login: {ex}")

    def handle_login(self):
        username = self.parent.loginEdit.text()
        password = self.parent.passwordEdit.text()
        certpassword = self.parent.certpasswordEdit.text()

        self.save_settings()
        self.parent.clean_up_processes()

        self.process = QProcess(self.parent)
        creon_path = 'C:\\CREON\\STARTER\\coStarter.exe'
        args = ["/prj:cp", f"/id:{username}", f"/pwd:{password}", f"/pwdcert:{certpassword}"]
        if self.parent.autoLoginCheckBox.isChecked():
            args.append('/autostart')
        self.process.start(creon_path, args)
        self.process.finished.connect(self.init_plus_check_and_continue)
        
        # ===== ✅ 창이 뜨는 시간(10초) 고려하여 지연 시작 =====
        logging.info("로그인 프로세스 시작... 모의투자 창 대기 중...")
        
        # 5초 후에 타이머 시작 (창이 뜨기 시작할 때쯤)
        QTimer.singleShot(5000, self.start_auto_click_timer)

    def start_auto_click_timer(self):
        """자동 클릭 타이머 시작 (지연 시작)"""
        self.auto_click_attempts = 0
        self.auto_click_timer.start(500)  # 0.5초마다 체크
        logging.info("⏰ 모의투자 창 자동 클릭 감지 시작...")

    def check_and_click_popup(self):
        """모의투자 선택 창 감지 및 자동 클릭 (10초 지연 고려)"""
        try:
            self.auto_click_attempts += 1
            
            # 최대 시도 횟수 초과
            if self.auto_click_attempts > self.max_auto_click_attempts:
                self.auto_click_timer.stop()
                logging.warning("⚠️ 모의투자 창 자동 클릭 타임아웃 (수동 클릭 필요)")
                return
            
            # ===== ✅ 진행 상황 로그 (더 자세히) =====
            elapsed_time = self.auto_click_attempts * 0.5
            if self.auto_click_attempts == 1:
                logging.info(f"🔍 모의투자 창 감지 시작... (최대 {self.max_auto_click_attempts * 0.5:.0f}초 대기)")
            elif self.auto_click_attempts % 10 == 0:  # 5초마다
                logging.info(f"🔍 감지 중... {elapsed_time:.0f}초 경과 ({self.auto_click_attempts}/{self.max_auto_click_attempts})")
            
            # ===== ✅ 창 찾기 =====
            possible_titles = [
                '모의투자 선택',
                '모의투자',
                'CREON',
                'Creon',
                'creon'
            ]
            
            target_window = None
            found_title = None
            
            for title in possible_titles:
                try:
                    windows = gw.getWindowsWithTitle(title)
                    if windows and len(windows) > 0:
                        # 창이 실제로 보이는지 확인
                        for window in windows:
                            if window.visible and window.width > 0 and window.height > 0:
                                target_window = window
                                found_title = title
                                break
                        if target_window:
                            break
                except Exception as ex:
                    logging.debug(f"창 검색 중 오류 ({title}): {ex}")
                    continue
            
            # 창을 못 찾으면 계속 시도
            if not target_window:
                if self.auto_click_attempts % 10 == 0:
                    logging.debug(f"창을 아직 찾지 못함 ({elapsed_time:.0f}초 경과)")
                return  # 타이머 계속 실행
            
            # ===== ✅ 창을 처음 찾았을 때 =====
            if not self.window_found:
                self.window_found = True
                logging.info(f"✨ '{found_title}' 창 발견! ({elapsed_time:.0f}초 경과)")
                # 창이 완전히 로드될 때까지 잠깐 대기
                time.sleep(0.5)
            
            # ===== ✅ 클릭 시도 =====
            logging.info(f"🎯 '{found_title}' 버튼 클릭 시도 중...")
            
            # 창 활성화
            try:
                target_window.activate()
                time.sleep(0.3)
            except Exception as ex:
                logging.debug(f"창 활성화 실패: {ex}")
            
            # 여러 방법으로 클릭 시도
            success = False
            
            # 방법 1: 화면 중앙 클릭
            if not success:
                success = self._click_screen_center(target_window)
            
            # ===== ✅ 성공 확인 =====
            if success:
                self.auto_click_timer.stop()
                logging.info(f"✅ 모의투자 접속 버튼 자동 클릭 성공! (총 {elapsed_time:.0f}초 소요)")
                self.window_found = False  # 플래그 리셋
            else:
                # 실패 시 다음 타이머에서 재시도
                if self.auto_click_attempts % 5 == 0:
                    logging.debug(f"클릭 실패, 재시도 중... ({self.auto_click_attempts}회)")
            
        except Exception as ex:
            logging.error(f"check_and_click_popup: {ex}\n{traceback.format_exc()}")

    def _click_screen_center(self, window):
        """방법 1: 화면 중앙 클릭"""
        try:
            # 창 정보
            left = window.left
            top = window.top
            width = window.width
            height = window.height
            
            # ===== ✅ 창 크기 검증 =====
            if width < 100 or height < 100:
                logging.debug(f"창 크기가 너무 작음 ({width}x{height})")
                return False
            
            # 창의 정중앙
            center_x = left + width // 2
            center_y = top + height // 2
            
            logging.info(f"🎯 화면 중앙 클릭: ({center_x}, {center_y}) [창 크기: {width}x{height}]")
            
            # 마우스 이동
            pyautogui.moveTo(center_x, center_y, duration=0.3)
            time.sleep(0.2)
            
            # 클릭
            pyautogui.click()
            time.sleep(0.5)  # ✅ 클릭 후 대기 시간 증가
            
            # ===== ✅ 클릭 성공 확인 (창이 사라졌는지) =====
            try:
                windows = gw.getWindowsWithTitle(window.title)
                if not windows:
                    logging.info("✅ 창이 사라짐 - 클릭 성공")
                    return True
                
                # 창이 여전히 있는지 확인
                for w in windows:
                    if w.visible and w.width > 0:
                        logging.debug("창이 여전히 존재 - 클릭 실패 또는 미반응")
                        return False
                
                logging.info("✅ 창이 사라짐 - 클릭 성공")
                return True
                
            except Exception as ex:
                logging.debug(f"창 확인 중 오류: {ex}")
                # 확인 실패 시 성공으로 간주
                return True
            
        except Exception as ex:
            logging.debug(f"_click_screen_center 실패: {ex}")
            return False
       
    def buycount_setting(self):
        if not self.config.has_section('BUYCOUNT'):
            self.config.add_section('BUYCOUNT')
        self.config.set('BUYCOUNT', 'target_buy_count', self.parent.buycountEdit.text())

        with open(self.config_file, 'w', encoding='utf-8') as configfile:
            self.config.write(configfile)

        logging.info(f"최대투자 종목수가 업데이트되었습니다.")

    def init_plus_check_and_continue(self):
        """PLUS 체크 후 초기화"""
        if not init_plus_check():
            logging.error("PLUS 연결 실패")
            exit()
    
        # ===== ✅ 연결 안정화 대기 추가 =====
        logging.info("🔗 PLUS 연결 안정화 대기 중...")
        time.sleep(2.0)  # 2초 대기
    
        self.parent.post_login_setup()

    def auto_select_creon_popup(self):
        """기존 메서드 (호환성 유지)"""
        try:
            button_x, button_y = 960, 500
            pyautogui.moveTo(button_x, button_y, duration=0.5)
            pyautogui.click()
            
            logging.info("모의투자 접속 버튼 클릭 완료")
        except Exception as e:
            logging.error(f"모의투자 접속 버튼 클릭 실패: {e}")

class StockLoaderThread(QThread):
    """종목 로딩 백그라운드 스레드"""
    
    # 시그널 정의
    progress_updated = pyqtSignal(int, int, str)  # (현재, 전체, 메시지)
    stock_loaded = pyqtSignal(str, bool)  # (종목코드, 성공여부)
    loading_completed = pyqtSignal(int, int)  # (성공, 실패)
    error_occurred = pyqtSignal(str)  # 에러 메시지
    
    def __init__(self, parent, stock_list, loader_func):
        """
        Args:
            parent: 부모 위젯
            stock_list: 로드할 종목 리스트
            loader_func: 단일 종목 로드 함수
        """
        super().__init__()
        self.parent = parent
        self.stock_list = stock_list
        self.loader_func = loader_func
        self.is_cancelled = False
        self.loaded_count = 0
        self.failed_count = 0
    
    def run(self):
        """스레드 실행"""
        try:
            total = len(self.stock_list)
            
            for idx, stock_item in enumerate(self.stock_list, 1):
                # 취소 확인
                if self.is_cancelled:
                    logging.info("종목 로딩 취소됨")
                    break
                
                try:
                    code = stock_item.get('code')
                    if not code:
                        continue
                    
                    stock_name = cpCodeMgr.CodeToName(code)
                    
                    # 진행 상황 업데이트
                    self.progress_updated.emit(idx, total, f"{stock_name}({code}) 로딩 중...")
                    
                    # 종목 로드
                    success = self.loader_func(code)
                    
                    if success:
                        self.loaded_count += 1
                        self.stock_loaded.emit(code, True)
                    else:
                        self.failed_count += 1
                        self.stock_loaded.emit(code, False)
                    
                    # 3개마다 더 긴 대기
                    if idx % 3 == 0:
                        time.sleep(2.0)
                    else:
                        time.sleep(1.5)  # API 제한 고려하여 증가
                    
                except Exception as ex:
                    logging.error(f"{code} 로드 실패: {ex}")
                    self.failed_count += 1
                    self.stock_loaded.emit(code, False)
            
            # 완료 시그널
            self.loading_completed.emit(self.loaded_count, self.failed_count)
            
        except Exception as ex:
            logging.error(f"StockLoaderThread: {ex}\n{traceback.format_exc()}")
            self.error_occurred.emit(str(ex))
    
    def cancel(self):
        """로딩 취소"""
        self.is_cancelled = True

class StockLoadingProgressDialog(QDialog):
    """종목 로딩 진행 상황 다이얼로그"""
    
    def __init__(self, parent, total_count):
        super().__init__(parent)
        self.setWindowTitle("종목 로딩 중")
        self.setModal(True)
        self.resize(500, 200)
        self.setWindowFlags(Qt.Dialog | Qt.WindowTitleHint)
        
        layout = QVBoxLayout()
        
        # 상태 레이블
        self.status_label = QLabel("종목 로딩을 시작합니다...")
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)
        
        # 프로그레스 바
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximum(total_count)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # 상세 정보
        info_layout = QHBoxLayout()
        self.success_label = QLabel("성공: 0")
        self.failed_label = QLabel("실패: 0")
        self.remaining_label = QLabel(f"남은 종목: {total_count}")
        info_layout.addWidget(self.success_label)
        info_layout.addWidget(self.failed_label)
        info_layout.addWidget(self.remaining_label)
        info_layout.addStretch()
        layout.addLayout(info_layout)
        
        # 취소 버튼
        button_layout = QHBoxLayout()
        self.cancel_button = QPushButton("취소")
        self.close_button = QPushButton("닫기")
        self.close_button.setEnabled(False)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.close_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # 변수
        self.total_count = total_count
        self.success_count = 0
        self.failed_count = 0
    
    def update_progress(self, current, total, message):
        """진행 상황 업데이트"""
        self.progress_bar.setValue(current)
        self.status_label.setText(message)
        self.remaining_label.setText(f"남은 종목: {total - current}")
        QApplication.processEvents()
    
    def on_stock_loaded(self, code, success):
        """종목 로드 완료"""
        if success:
            self.success_count += 1
            self.success_label.setText(f"✅ 성공: {self.success_count}")
        else:
            self.failed_count += 1
            self.failed_label.setText(f"❌ 실패: {self.failed_count}")
        
        QApplication.processEvents()
    
    def on_loading_completed(self, success, failed):
        """로딩 완료"""
        self.status_label.setText(
            f"✅ 종목 로딩 완료!\n"
            f"성공: {success}개, 실패: {failed}개"
        )
        self.cancel_button.setEnabled(False)
        self.close_button.setEnabled(True)
        
        # 자동 닫기 (3초 후)
        QTimer.singleShot(3000, self.accept)

# ==================== MyWindow ====================
class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        
        # ===== ✅ 기본 변수만 초기화 =====
        self.is_loading_strategy = False
        self.market_close_emitted = False
        self.pending_updates = set()  # 업데이트 대기 중인 종목들
        self.ui_update_timer = None  # 통합 UI 업데이트 타이머
        
        # 전략 객체들 (나중에 초기화)
        self.momentum_scanner = None
        self.gap_scanner = None
        self.volatility_strategy = None
        
        # ===== ✅ LoginHandler 먼저 생성 (UI보다 먼저!) =====
        self.login_handler = LoginHandler(self)
        
        # ===== ✅ UI 생성 (이제 login_handler 사용 가능) =====
        self.init_ui()
        
        # ===== ✅ 설정 로드 (login_handler의 메서드 호출) =====
        self.login_handler.load_settings()
        
        # ===== ✅ 자동 로그인 (비동기) =====
        QTimer.singleShot(100, self.login_handler.attempt_auto_login)

    def __del__(self):
        if hasattr(self, 'objstg'):
            self.objstg.Clear()

    def _attempt_auto_login(self):
        """자동 로그인 시도 (비동기)"""
        try:
            if (self.login_handler.config.getboolean('LOGIN', 'autologin', fallback=False) and 
                self.loginEdit.text() and 
                self.passwordEdit.text() and 
                self.certpasswordEdit.text()):
                
                # 약간의 지연 후 로그인
                QTimer.singleShot(500, self.login_handler.handle_login)
        except Exception as ex:
            logging.error(f"_attempt_auto_login: {ex}")

    def post_login_setup(self):
        """로그인 후 설정"""
        
        # 1. 모의투자 서버 연결 확인 (PLUS 연결 체크에서 이미 검증됨)
        logging.info("📡 모의투자 서버 연결 확인 중...")
        time.sleep(0.5)  # 최소한의 안정화 대기
        
        # 2. 로거 초기화 (계좌 조회 이전에 먼저 실행)
        logger = logging.getLogger()
        if not any(isinstance(handler, QTextEditLogger) for handler in logger.handlers):
            text_edit_logger = QTextEditLogger(self.terminalOutput)
            text_edit_logger.setLevel(logging.INFO)
            logger.addHandler(text_edit_logger)
        
        # 3. 팝업 닫기 (로그인 직후 바로 실행)
        self.close_external_popup()
        
        # 4. 계좌 정보 조회 (트레이더 객체 생성 전에 먼저 실행)
        try:
            acc = cpTrade.AccountNumber[0]
            accFlag = cpTrade.GoodsList(acc, 1)
            cpBalance.SetInputValue(0, acc)
            cpBalance.SetInputValue(1, accFlag[0])
            cpBalance.SetInputValue(2, 50)
            ret = cpBalance.BlockRequest2(1)
            if ret == 0:
                logging.info(f"계좌명 : {str(cpBalance.GetHeaderValue(0))}")
                logging.info(f"결제잔고수량 : {str(cpBalance.GetHeaderValue(1))}")
                logging.info(f"평가금액 : {str(cpBalance.GetHeaderValue(3))}")
                logging.info(f"평가손익 : {str(cpBalance.GetHeaderValue(4))}")
                logging.info(f"종목수 : {str(cpBalance.GetHeaderValue(7))}")
            else:
                logging.warning(f"계좌 잔고 조회 실패, {ret}")
        except Exception as ex:
            logging.error(f"계좌 정보 조회 오류: {ex}")
        
        logging.info(f"시작 시간: {datetime.now().strftime('%m/%d %H:%M:%S')}")
        
        # 5. 트레이더 객체 생성
        buycount = int(self.buycountEdit.text())
        self.trader = CTrader(cpTrade, cpBalance, cpCodeMgr, cpCash, cpOrder, cpStock, buycount, self)
        self.objstg = CpStrategy(self.trader)
        self.trader_thread = AutoTraderThread(self.trader, self)
        
        # ✅ CpStrategy 시그널 연결
        self.objstg.stock_processed.connect(self.on_stock_processed)
        self.objstg.processing_error.connect(self.on_processing_error)
        
        # ✅ 실시간 데이터 업데이트 시그널 연결
        self.trader.tickdata.data_updated.connect(self.on_realtime_data_updated)
        self.trader.mindata.data_updated.connect(self.on_realtime_data_updated)
        self.trader.daydata.data_updated.connect(self.on_realtime_data_updated)

        self.chartdrawer = ChartDrawer(self.fig, self.canvas, self.trader, self.trader_thread, self)

        self.code = ''
        self.stocks = []
        self.counter = 0

        # 6. 전략 로드 (메인 스레드)
        self.load_strategy()

        # ===== ✅ 7. 큐 스레드 시작 (백그라운드 로드와 동시에 시작) =====
        self.objstg.start_processing_queue()

        # 8. 타이머 시작
        self.start_timers()
        
        # 9. 시그널 연결
        self.trader.stock_added_to_monitor.connect(self.on_stock_added)
        self.trader.stock_bought.connect(self.on_stock_bought)
        self.trader.stock_sold.connect(self.on_stock_sold)
        
        self.trader_thread.buy_signal.connect(self.trader.buy_stock)
        self.trader_thread.sell_signal.connect(self.trader.sell_stock)
        self.trader_thread.sell_half_signal.connect(self.trader.sell_half_stock)
        self.trader_thread.sell_all_signal.connect(self.trader.sell_all)
        self.trader_thread.stock_removed_from_monitor.connect(self.on_stock_removed)
        self.trader_thread.counter_updated.connect(self.update_counter_label)
        self.trader_thread.stock_data_updated.connect(self.update_stock_table)
        
        self.trader_thread.connect_bar_signals()
        
        self.trader_thread.start()
        
    def get_strategy_type(self, strategy_name):
        """전략 타입 확인
        
        Args:
            strategy_name: 전략명 (예: '전일급등', '급등주')
        
        Returns:
            'static' or 'dynamic'
        """
        try:
            # 설정 파일에서 읽기
            if self.login_handler.config.has_section('STRATEGY_TYPES'):
                static_str = self.login_handler.config.get('STRATEGY_TYPES', 'static', fallback='')
                dynamic_str = self.login_handler.config.get('STRATEGY_TYPES', 'dynamic', fallback='')
                
                # 리스트로 변환 (공백 제거)
                static_list = [s.strip() for s in static_str.split(',') if s.strip()]
                dynamic_list = [s.strip() for s in dynamic_str.split(',') if s.strip()]
                
                # 전략명 확인
                if strategy_name in static_list:
                    return 'static'
                elif strategy_name in dynamic_list:
                    return 'dynamic'
            
            # ===== ✅ 설정에 없으면 이름으로 자동 판단 =====
            static_keywords = ['전일', '전날', '52주', '이격도', '신고가']
            if any(keyword in strategy_name for keyword in static_keywords):
                logging.debug(f"'{strategy_name}' → 정적 전략으로 자동 분류")
                return 'static'
            else:
                logging.debug(f"'{strategy_name}' → 동적 전략으로 자동 분류")
                return 'dynamic'
                
        except Exception as ex:
            logging.error(f"get_strategy_type({strategy_name}): {ex}")
            return 'static'  # 안전하게 정적으로 처리

    def get_max_static_load(self):
        """정적 전략 최대 로드 개수"""
        try:
            if self.login_handler.config.has_section('STRATEGY_TYPES'):
                return self.login_handler.config.getint('STRATEGY_TYPES', 'max_static_load', fallback=10)
            return 10
        except Exception as ex:
            logging.error(f"get_max_static_load: {ex}")
            return 10

    def _background_initialization(self):
        """백그라운드 초기화 (무거운 작업들)"""
        try:
            # 60% - 계좌 정보
            QTimer.singleShot(0, lambda: self.splash.update_progress(60, "계좌 정보 조회 중..."))
            self.trader.get_stock_balance('START', 'post_login_setup')
            
            # 70% - 외부 팝업 닫기
            QTimer.singleShot(0, lambda: self.splash.update_progress(70, "팝업 정리 중..."))
            self.close_external_popup()
            
            # 80% - 전략 로드
            QTimer.singleShot(0, lambda: self.splash.update_progress(80, "전략 로드 중..."))
            self.load_strategy()
            
            # 90% - 시그널 연결
            QTimer.singleShot(0, lambda: self.splash.update_progress(90, "시그널 연결 중..."))
            self._connect_signals()
            
            # 95% - 타이머 시작
            QTimer.singleShot(0, lambda: self.splash.update_progress(95, "타이머 시작 중..."))
            self.start_timers()
            
            # 100% - 완료
            QTimer.singleShot(0, lambda: self.splash.update_progress(100, "완료!"))
            time.sleep(0.5)
            
            # 스플래시 닫기 (메인 스레드에서)
            QTimer.singleShot(0, self._finish_initialization)
            
        except Exception as ex:
            logging.error(f"_background_initialization: {ex}\n{traceback.format_exc()}")
            QTimer.singleShot(0, lambda: QMessageBox.critical(
                self, "초기화 오류", f"초기화 중 오류 발생:\n{str(ex)}"
            ))

    def _connect_signals(self):
        """시그널 연결"""
        self.trader.stock_added_to_monitor.connect(self.on_stock_added)
        self.trader.stock_bought.connect(self.on_stock_bought)
        self.trader.stock_sold.connect(self.on_stock_sold)
        
        self.trader_thread.buy_signal.connect(self.trader.buy_stock)
        self.trader_thread.sell_signal.connect(self.trader.sell_stock)
        self.trader_thread.sell_half_signal.connect(self.trader.sell_half_stock)
        self.trader_thread.sell_all_signal.connect(self.trader.sell_all)
        self.trader_thread.stock_removed_from_monitor.connect(self.on_stock_removed)
        self.trader_thread.counter_updated.connect(self.update_counter_label)
        self.trader_thread.stock_data_updated.connect(self.update_stock_table)
        
        # ✅ 실시간 데이터 업데이트 시그널 연결
        self.trader.tickdata.data_updated.connect(self.on_realtime_data_updated)
        self.trader.mindata.data_updated.connect(self.on_realtime_data_updated)
        self.trader.daydata.data_updated.connect(self.on_realtime_data_updated)
        
        self.trader_thread.connect_bar_signals()
        
        logging.info(f"시작 시간 : {datetime.now().strftime('%m/%d %H:%M:%S')}")

    def _finish_initialization(self):
        """초기화 완료"""
        try:
            if hasattr(self, 'splash'):
                self.splash.close()
                del self.splash
            
            # 트레이더 스레드 시작
            self.trader_thread.start()
            
            logging.info("=== 초기화 완료 ===")
            
        except Exception as ex:
            logging.error(f"_finish_initialization: {ex}")

    def start_timers(self):
        """타이머 시작 (휴일 대응)"""
        now = datetime.now()
        today = datetime.today().weekday()

        # ===== ✅ 주말 체크 =====
        if today == 5 or today == 6:
            logging.info(f"Today is {'Saturday.' if today == 5 else 'Sunday.'}")
            logging.info(f"오늘은 장이 쉽니다. 최근 영업일 데이터를 표시합니다.")
            
            # 데이터베이스만 초기화
            self.trader.init_database()
            return
        
        # ===== 평일 - 영업일 확인 (API 호출) =====
        today_int = now.year * 10000 + now.month * 100 + now.day
        
        # 영업일 확인 API 호출
        success, trading_date = get_last_trading_date(today_int, max_attempts=10)
        
        if success:
            date_str = str(trading_date)
            formatted = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            logging.info(f"📅 최근 영업일: {formatted}")
            
            # CpData 객체들의 todayDate 설정
            if hasattr(self.trader, 'daydata') and self.trader.daydata:
                self.trader.daydata.todayDate = trading_date
            if hasattr(self.trader, 'mindata') and self.trader.mindata:
                self.trader.mindata.todayDate = trading_date
            if hasattr(self.trader, 'tickdata') and self.trader.tickdata:
                self.trader.tickdata.todayDate = trading_date
            
            # 오늘이 영업일이 아니면 (공휴일)
            if today_int > trading_date:
                logging.info(f"오늘은 공휴일입니다. 최근 영업일({formatted}) 데이터를 표시합니다.")
                self.trader.init_database()
                return
        else:
            logging.warning(f"⚠️ 영업일 찾기 실패, 오늘 날짜 사용: {today_int}")
        
        # ===== 영업일 - 장 시간 확인 =====
        start_time = now.replace(hour=9, minute=0, second=0, microsecond=0)
        end_time = now.replace(hour=15, minute=20, second=0, microsecond=0)
        
        if now < start_time:
            logging.info(f"자동 매매 시작 대기")
            self.trader.init_database()
            QTimer.singleShot(int((start_time - now).total_seconds() * 1000) + 1000, self.start_timers)

        elif start_time <= now < end_time:
            logging.info(f"자동 매매 시작")
            send_slack_message(self.login_handler, "#stock", f"자동 매매 시작")

            # ✅ 통합 UI 업데이트 타이머 시작 (2초마다)
            self.ui_update_timer = QTimer(self)
            self.ui_update_timer.timeout.connect(self.process_pending_updates)
            self.ui_update_timer.start(2000)  # 2초마다 통합 UI 업데이트
            
            QTimer.singleShot(int((end_time - now).total_seconds() * 1000) + 1000, self.start_timers)
            
        elif end_time <= now and not self.market_close_emitted:
            logging.info("=== 장 종료 처리 시작 ===")
                    
            # 차트 업데이트 타이머 정지
            if self.trader.tickdata is not None:
                self.trader.tickdata.update_data_timer.stop()
            if self.trader.mindata is not None:
                self.trader.mindata.update_data_timer.stop()
            if self.trader.daydata is not None:
                self.trader.daydata.update_data_timer.stop()
            if self.ui_update_timer is not None:
                self.ui_update_timer.stop()
            
            # 미보유 종목 정리
            for code in list(self.trader.monistock_set):
                if code not in self.trader.bought_set:
                    self.on_stock_removed(code)
                    self.trader.delete_list_db(code)
            
            for code in list(self.trader.vistock_set):
                if code not in self.trader.monistock_set and code not in self.trader.bought_set:
                    self.trader.delete_list_db(code)

            self.market_close_emitted = True
            logging.info(f"자동 매매 종료")
            send_slack_message(self.login_handler, "#stock", f"자동 매매 종료")

    @pyqtSlot(str)
    def on_realtime_data_updated(self, code):
        """실시간 데이터 업데이트 시 즉시 차트 업데이트"""
        try:
            # 현재 선택된 종목이면 즉시 차트 업데이트
            if hasattr(self, 'chartdrawer') and self.chartdrawer.code == code:
                # 차트 스레드가 실행 중이면 강제 업데이트
                if hasattr(self.chartdrawer, 'chart_thread') and self.chartdrawer.chart_thread and self.chartdrawer.chart_thread.isRunning():
                    # 차트 스레드에서 데이터를 다시 가져와서 업데이트
                    QTimer.singleShot(100, lambda: self._force_chart_update(code))
                else:
                    # 차트 스레드가 없으면 재시작
                    self.chartdrawer.set_code(code)
            
            # ✅ 투자현황 표도 즉시 업데이트
            if code in self.trader.monistock_set:
                QTimer.singleShot(200, lambda: self._update_investment_table(code))
            
            # 업데이트가 필요한 종목들을 기록 (백업용)
            if not hasattr(self, 'pending_updates'):
                self.pending_updates = set()
            self.pending_updates.add(code)
                
        except Exception as ex:
            logging.debug(f"실시간 데이터 업데이트 처리 오류: {ex}")
    
    @pyqtSlot(str, bool)
    def on_stock_processed(self, code, success):
        """종목 처리 완료 시그널 핸들러"""
        try:
            if success:
                logging.debug(f"✅ {code}: 종목 검증 완료")
            else:
                logging.debug(f"❌ {code}: 종목 검증 실패")
        except Exception as ex:
            logging.error(f"종목 처리 완료 핸들러 오류: {ex}")
    
    @pyqtSlot(str, str)
    def on_processing_error(self, code, error_msg):
        """종목 처리 오류 시그널 핸들러"""
        try:
            logging.error(f"❌ {code}: {error_msg}")
        except Exception as ex:
            logging.error(f"종목 처리 오류 핸들러 오류: {ex}")
    
    def _force_chart_update(self, code):
        """차트 강제 업데이트 (메인 스레드에서 안전하게 실행)"""
        try:
            # 메인 스레드에서 실행되도록 QTimer.singleShot 사용
            if hasattr(self.chartdrawer, 'chart_thread') and self.chartdrawer.chart_thread and self.chartdrawer.chart_thread.isRunning():
                tick_data = self.trader.tickdata.get_full_data(code)
                min_data = self.trader.mindata.get_full_data(code)
                
                if tick_data and min_data and len(tick_data.get('C', [])) > 0 and len(min_data.get('C', [])) > 0:
                    data = {'tick_data': tick_data, 'min_data': min_data, 'code': code}
                    
                    # 메인 스레드에서 차트 업데이트 실행
                    QTimer.singleShot(0, lambda: self._safe_chart_update(data))
                    logging.debug(f"📊 {code}: 차트 업데이트 요청")
        except Exception as ex:
            logging.debug(f"강제 차트 업데이트 오류: {ex}")
    
    def _safe_chart_update(self, data):
        """메인 스레드에서 안전하게 차트 업데이트"""
        try:
            self.chartdrawer.update_chart(data)
            logging.debug(f"📊 {data.get('code')}: 차트 업데이트 완료")
        except Exception as ex:
            logging.error(f"차트 업데이트 오류: {ex}")
    
    def _update_investment_table(self, code):
        """특정 종목의 투자현황 표 업데이트 (메인 스레드에서 안전하게 실행)"""
        try:
            if code not in self.trader.monistock_set:
                return
            
            # 해당 종목의 데이터만 업데이트
            tick_latest = self.trader.tickdata.get_latest_data(code)
            current_price = tick_latest.get('C', 0.0) if tick_latest else 0.0
            buy_price = self.trader.buy_price.get(code, 0.0)
            quantity = self.trader.buy_qty.get(code, 0)
            
            if current_price == 0 or buy_price == 0:
                return
            
            # 메인 스레드에서 테이블 업데이트 실행
            QTimer.singleShot(0, lambda: self._safe_table_update(code, current_price, buy_price, quantity))
            logging.debug(f"📊 {code}: 투자현황 표 업데이트 요청")
                    
        except Exception as ex:
            logging.debug(f"투자현황 표 업데이트 오류 ({code}): {ex}")
    
    def _safe_table_update(self, code, current_price, buy_price, quantity):
        """메인 스레드에서 안전하게 테이블 업데이트"""
        try:
            # 테이블에서 해당 종목 찾기
            for row in range(self.stock_table.rowCount()):
                item = self.stock_table.item(row, 0)
                if item and item.text() == code:
                    # 현재가 업데이트
                    current_price_item = QTableWidgetItem(f"{current_price:,.0f}")
                    current_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.stock_table.setItem(row, 1, current_price_item)
                    
                    # 수익/손실 업데이트
                    profit_loss = (current_price - buy_price) * quantity
                    profit_loss_item = QTableWidgetItem(f"{profit_loss:,.0f}")
                    profit_loss_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.stock_table.setItem(row, 4, profit_loss_item)
                    if profit_loss > 0:
                        profit_loss_item.setForeground(Qt.green)
                    elif profit_loss < 0:
                        profit_loss_item.setForeground(Qt.red)
                    else:
                        profit_loss_item.setForeground(Qt.black)
                    
                    # 수익률 업데이트
                    return_pct = ((current_price - buy_price) / buy_price * 100) if buy_price != 0 else 0.0
                    return_item = QTableWidgetItem(f"{return_pct:.2f}")
                    return_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.stock_table.setItem(row, 5, return_item)
                    if return_pct > 0:
                        return_item.setForeground(Qt.green)
                    elif return_pct < 0:
                        return_item.setForeground(Qt.red)
                    else:
                        return_item.setForeground(Qt.black)
                    
                    logging.debug(f"📊 {code}: 투자현황 표 업데이트 완료 (현재가: {current_price:,.0f}, 수익률: {return_pct:.2f}%)")
                    break
                    
        except Exception as ex:
            logging.error(f"테이블 업데이트 오류 ({code}): {ex}")
    
    def _update_all_investment_table(self):
        """전체 투자현황 표 업데이트 (3초마다) - 메인 스레드에서 안전하게 실행"""
        try:
            if not self.trader.monistock_set:
                return
            
            # 메인 스레드에서 전체 테이블 업데이트 실행
            QTimer.singleShot(0, lambda: self._safe_full_table_update())
            logging.debug(f"📊 투자현황 표 전체 업데이트 요청 ({len(self.trader.monistock_set)}개 종목)")
            
        except Exception as ex:
            logging.debug(f"전체 투자현황 표 업데이트 오류: {ex}")
    
    def _safe_full_table_update(self):
        """메인 스레드에서 안전하게 전체 테이블 업데이트"""
        try:
            # 전체 테이블 업데이트
            self.trader_thread._update_stock_data_table()
            logging.debug(f"📊 투자현황 표 전체 업데이트 완료 ({len(self.trader.monistock_set)}개 종목)")
            
        except Exception as ex:
            logging.error(f"전체 테이블 업데이트 오류: {ex}")

    def process_pending_updates(self):
        """통합 UI 업데이트 처리 (2초마다 실행)"""
        try:
            # 1. 투자현황 표 업데이트
            self._update_all_investment_table()
            
            # 2. 대기 중인 업데이트 처리
            if not hasattr(self, 'pending_updates') or not self.pending_updates:
                return
            
            # 현재 선택된 종목이 업데이트 대기 중이면 차트 업데이트
            if hasattr(self, 'chartdrawer') and self.chartdrawer.code in self.pending_updates:
                code = self.chartdrawer.code
                
                # ✅ 차트 스레드가 실행 중인지 확인
                if hasattr(self.chartdrawer, 'chart_thread') and self.chartdrawer.chart_thread and self.chartdrawer.chart_thread.isRunning():
                    logging.debug(f"📊 {code}: 차트 스레드가 이미 실행 중")
                else:
                    # 차트 스레드 재시작
                    logging.debug(f"📊 {code}: 차트 스레드 재시작")
                    self.chartdrawer.set_code(code)
                
                # 대기 중인 업데이트 제거
                self.pending_updates.discard(code)
            
            # 모니터링 중인 종목이 업데이트 대기 중이면 투자현황표 업데이트
            monitoring_updated = any(code in self.trader.monistock_set for code in self.pending_updates)
            if monitoring_updated:
                self.trader_thread._update_stock_data_table()
            
            # 처리 완료 후 대기 목록 클리어
            self.pending_updates.clear()
                
        except Exception as ex:
            logging.debug(f"대기 중인 업데이트 처리 오류: {ex}")

    def _update_chart_status_label(self):
        """차트 상태 라벨 업데이트"""
        if hasattr(self, 'chartdrawer') and self.chartdrawer.last_chart_update_time:
            # 현재 시간과 차트 업데이트 시간의 차이 계산 (분 단위)
            current_time = datetime.now()
            chart_time = self.chartdrawer.last_chart_update_time
            
            # 시간 차이 계산 (분 단위)
            if isinstance(chart_time, int):
                # HHMM 형식인 경우
                chart_hour = chart_time // 100
                chart_minute = chart_time % 100
                chart_datetime = current_time.replace(hour=chart_hour, minute=chart_minute, second=0, microsecond=0)
            else:
                chart_datetime = chart_time
            
            time_diff = (current_time - chart_datetime).total_seconds() / 60  # 분 단위
            
            if time_diff < 2:
                chart_color = "green"
                status_text = f"Chart: {time_diff:.1f}m ago"
            else:
                chart_color = "red"
                status_text = f"Chart: {time_diff:.1f}m ago"
            
            self.chart_status_label.setText(status_text)
            self.chart_status_label.setStyleSheet(f"color: {chart_color}")
        else:
            self.chart_status_label.setText("Chart: None")
            self.chart_status_label.setStyleSheet("color: red")

    def update_stock_table(self, stock_data_list):
        """monistock_set 데이터를 기반으로 표 업데이트"""
        self.stock_table.setRowCount(len(stock_data_list))
        for row, stock_data in enumerate(stock_data_list):
            code = stock_data.get('code', '')
            current_price = stock_data.get('current_price', 0.0)
            upward_prob = stock_data.get('upward_probability', 0.0)
            buy_price = stock_data.get('buy_price', 0.0)
            quantity = stock_data.get('quantity', 0)
            profit_loss = (current_price - buy_price) * quantity
            return_pct = ((current_price - buy_price) / buy_price * 100) if buy_price != 0 else 0.0

            code_item = QTableWidgetItem(code)
            code_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.stock_table.setItem(row, 0, code_item)

            current_price_item = QTableWidgetItem(f"{current_price:,.0f}")
            current_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 1, current_price_item)

            upward_prob_item = QTableWidgetItem(f"{upward_prob:.2f}")
            upward_prob_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 2, upward_prob_item)

            buy_price_item = QTableWidgetItem(f"{buy_price:,.0f}")
            buy_price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 3, buy_price_item)

            profit_loss_item = QTableWidgetItem(f"{profit_loss:,.0f}")
            profit_loss_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 4, profit_loss_item)
            profit_loss_item = self.stock_table.item(row, 4)
            if profit_loss > 0:
                profit_loss_item.setForeground(Qt.green)
            elif profit_loss < 0:
                profit_loss_item.setForeground(Qt.red)
            else:
                profit_loss_item.setForeground(Qt.black)

            return_item = QTableWidgetItem(f"{return_pct:.2f}")
            return_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.stock_table.setItem(row, 5, return_item)
            return_item = self.stock_table.item(row, 5)
            if return_pct > 0:
                return_item.setForeground(Qt.green)
            elif return_pct < 0:
                return_item.setForeground(Qt.red)
            else:
                return_item.setForeground(Qt.black)

    @pyqtSlot(str)
    def on_stock_added(self, code):
        logging.info(f"📋 on_stock_added 슬롯 호출됨: {code}")
        try:
            existing_items = [self.firstListBox.item(i).text() for i in range(self.firstListBox.count())]
            if code not in existing_items:
                self.firstListBox.addItem(code)
                logging.info(f"📋 {code} 투자대상종목 리스트박스에 추가됨")
            else:
                logging.debug(f"📋 {code} 이미 리스트박스에 존재함")
        except Exception as ex:
            logging.error(f"📋 on_stock_added 슬롯 오류: {ex}")

    @pyqtSlot(str)
    def on_stock_removed(self, code):
        if code in self.trader.vistock_set:
            self.trader.vistock_set.remove(code)
        if code in self.trader.monistock_set:
            self.trader.monistock_set.remove(code)
            self.trader.tickdata.monitor_stop(code)
            self.trader.mindata.monitor_stop(code)
            self.trader.daydata.monitor_stop(code)
            self.trader.delete_list_db(code)

        if code == self.chartdrawer.code:
            self.chartdrawer.set_code(None)

        for index in range(self.firstListBox.count()):
            item = self.firstListBox.item(index)
            if item and item.text() == code:
                self.firstListBox.takeItem(index)
                break
    
    @pyqtSlot(str)
    def on_stock_bought(self, code):
        if code not in [self.secondListBox.item(i).text() for i in range(self.secondListBox.count())]:
            self.secondListBox.addItem(code)

    @pyqtSlot(str)
    def on_stock_sold(self, code):
        for index in range(self.secondListBox.count()):
            item = self.secondListBox.item(index)
            if item and item.text() == code:
                self.secondListBox.takeItem(index)
                break

    @pyqtSlot(int)
    def update_counter_label(self, counter):
        self.counterlabel.setText(f"타이머: {counter}")

    def save_last_stg(self):
        if not self.login_handler.config.has_section('SETTINGS'):
            self.login_handler.config.add_section('SETTINGS')
        self.login_handler.config.set('SETTINGS', 'last_strategy', self.comboStg.currentText())
        with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
            self.login_handler.config.write(configfile)

    def validate_strategy_condition(self, condition, strategy_type='buy'):
        """전략 조건식 검증
        
        Args:
            condition: 검증할 조건식 문자열
            strategy_type: 'buy' 또는 'sell'
        
        Returns:
            (is_valid, message): (True/False, 메시지)
        """
        try:
            import ast
            
            # 빈 문자열 체크
            if not condition or not condition.strip():
                return False, "조건식이 비어있습니다"
            
            # 문법 검증
            try:
                tree = ast.parse(condition, mode='eval')
            except SyntaxError as e:
                return False, f"문법 오류: {e.msg} (라인 {e.lineno})"
            
            # 사용 가능한 변수 정의
            if strategy_type == 'buy':
                available_vars = {
                    # 틱 데이터
                    'MAT5', 'MAT20', 'MAT60', 'MAT120', 'C', 'VWAP', 
                    'RSIT', 'RSIT_SIGNAL', 'MACDT', 'MACDT_SIGNAL', 'OSCT',
                    'STOCHK', 'STOCHD', 'ATR', 'CCI',
                    'BB_UPPER', 'BB_MIDDLE', 'BB_LOWER', 'BB_POSITION', 'BB_BANDWIDTH',
                    
                    # 분봉 데이터
                    'MAM5', 'MAM10', 'MAM20', 'min_close',
                    'RSI', 'RSI_SIGNAL', 'MACD', 'MACD_SIGNAL', 'OSC',
                    
                    # 계산 변수
                    'strength', 'momentum_score', 'threshold',
                    'volatility_breakout', 'gap_hold',
                    
                    # 기타
                    'positive_candle', 'tick_close_price', 'min_close_price'
                }
            else:  # sell
                available_vars = {
                    # 기본 변수
                    'min_close', 'MAM5', 'MAM10',
                    'current_profit_pct', 'from_peak_pct', 'hold_minutes',
                    'code', 'osct_negative', 'after_market_close',
                    
                    # trader 객체 접근 (self.trader)
                    'self'
                }
            
            # 사용된 변수명 추출
            used_vars = set()
            for node in ast.walk(tree):
                if isinstance(node, ast.Name):
                    used_vars.add(node.id)
                elif isinstance(node, ast.Attribute):
                    # self.trader.sell_half_set 같은 경우
                    if isinstance(node.value, ast.Name):
                        used_vars.add(node.value.id)
            
            # 정의되지 않은 변수 체크
            undefined = used_vars - available_vars - {
                # Python 내장 함수 허용
                'True', 'False', 'None', 
                'min', 'max', 'abs', 'round', 'int', 'float', 'len', 'sum', 'all', 'any'
            }
            
            if undefined:
                return False, f"정의되지 않은 변수: {', '.join(sorted(undefined))}"
            
            # 위험한 함수 호출 체크
            dangerous_calls = {'eval', 'exec', 'compile', '__import__', 'open', 'file'}
            for node in ast.walk(tree):
                if isinstance(node, ast.Call):
                    if isinstance(node.func, ast.Name):
                        if node.func.id in dangerous_calls:
                            return False, f"위험한 함수 사용 금지: {node.func.id}"
            
            return True, "검증 성공"
            
        except Exception as ex:
            return False, f"검증 중 오류: {str(ex)}"

    def save_buystrategy(self):
        """매수 전략 저장 (검증 추가)"""
        try:
            investment_strategy = self.comboStg.currentText()
            buy_strategy = self.comboBuyStg.currentText()
            buy_key = self.comboBuyStg.currentData()
            strategy_content = self.buystgInputWidget.toPlainText()

            # ===== 여기에서 검증 =====
            is_valid, message = self.validate_strategy_condition(strategy_content, 'buy')
            if not is_valid:
                QMessageBox.warning(
                    self, 
                    "전략 검증 실패", 
                    f"매수 전략 '{buy_strategy}'의 조건식이 올바르지 않습니다.\n\n{message}"
                )
                return

            if not self.login_handler.config.has_section('STRATEGIES'):
                self.login_handler.config.add_section('STRATEGIES')

            if not self.login_handler.config.has_section(investment_strategy):
                self.login_handler.config.add_section(investment_strategy)

            strategy_data = {
                'name': buy_strategy,
                'content': strategy_content
            }
            strategy_json = json.dumps(strategy_data, ensure_ascii=False)
            self.login_handler.config.set(investment_strategy, str(buy_key), strategy_json)

            with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                self.login_handler.config.write(configfile)

            existing_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == buy_strategy), None)
            if existing_strategy:
                existing_strategy['content'] = strategy_content
                logging.info(f"매수전략 '{buy_strategy}'이(가) 업데이트되었습니다.")

            QMessageBox.information(self, "수정 완료", f"매수전략 '{buy_strategy}'이 수정되었습니다.")

        except Exception as ex:
            logging.error(f"save_buystrategy -> {ex}")
            QMessageBox.critical(self, "수정 실패", f"전략 수정 중 오류가 발생했습니다:\n{str(ex)}")

    def save_sellstrategy(self):
        try:
            investment_strategy = self.comboStg.currentText()
            sell_strategy = self.comboSellStg.currentText()
            sell_key = self.comboSellStg.currentData()
            strategy_content = self.sellstgInputWidget.toPlainText()

            # ===== 여기에서 검증 =====
            is_valid, message = self.validate_strategy_condition(strategy_content, 'sell')
            if not is_valid:
                QMessageBox.warning(
                    self, 
                    "전략 검증 실패", 
                    f"매도 전략 '{sell_strategy}'의 조건식이 올바르지 않습니다.\n\n{message}"
                )
                return

            if not self.login_handler.config.has_section('STRATEGIES'):
                self.login_handler.config.add_section('STRATEGIES')

            if not self.login_handler.config.has_section(investment_strategy):
                self.login_handler.config.add_section(investment_strategy)

            strategy_data = {
                'name': sell_strategy,
                'content': strategy_content
            }
            strategy_json = json.dumps(strategy_data, ensure_ascii=False)
            self.login_handler.config.set(investment_strategy, str(sell_key), strategy_json)

            with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                self.login_handler.config.write(configfile)

            existing_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == sell_strategy), None)
            if existing_strategy:
                existing_strategy['content'] = strategy_content
                logging.info(f"매도전략 '{sell_strategy}'이(가) 업데이트되었습니다.")

            QMessageBox.information(self, "수정 완료", f"매도전략 '{sell_strategy}'이 수정되었습니다.")

        except Exception as ex:
            logging.error(f"save_strategy -> {ex}")
            QMessageBox.critical(self, "수정 실패", f"전략 수정 중 오류가 발생했습니다:\n{str(ex)}")

    # ===== load_strategy() 통합 (메인 스레드에서 실행) =====

    def load_strategy(self):
        """전략 로드 (메인 스레드, 순서 보장, API 호출 최소화)"""
        try:
            # ===== 1. 초기화 =====
            self.dataStg = []
            self.data8537 = {}  # 일단 빈 딕셔너리로 초기화
            self.strategies = {}

            self.comboStg.clear()
            self.comboBuyStg.clear()
            self.buystgInputWidget.clear()

            # ===== 2. 설정 파일에서 전략 목록 읽기 (API 호출 없음) =====
            if self.login_handler.config.has_section('STRATEGIES'):
                existing_stgnames = set(self.login_handler.config['STRATEGIES'].values())
            else:
                existing_stgnames = set()
                
            logging.debug(f"설정 파일에서 {len(existing_stgnames)}개 전략 로드")

            # ===== 3. 설정 파일에서 전략별 매수/매도 조건 읽기 =====
            for investment_strategy in existing_stgnames:
                if self.login_handler.config.has_section(investment_strategy):
                    self.strategies[investment_strategy] = []
                    
                    # 매수 전략 로드
                    buy_keys = sorted(
                        [k for k in self.login_handler.config[investment_strategy] if k.startswith('buy_stg_')],
                        key=lambda x: int(x.split('_')[-1])
                    )
                    for buy_key in buy_keys:
                        try:
                            buy_strategy = json.loads(self.login_handler.config.get(investment_strategy, buy_key))
                            buy_strategy['key'] = buy_key
                            self.strategies[investment_strategy].append(buy_strategy)
                        except json.JSONDecodeError as ex:
                            logging.warning(f"{investment_strategy} - {buy_key} 파싱 실패: {ex}")

                    # 매도 전략 로드
                    sell_keys = sorted(
                        [k for k in self.login_handler.config[investment_strategy] if k.startswith('sell_stg_')],
                        key=lambda x: int(x.split('_')[-1])
                    )
                    for sell_key in sell_keys:
                        try:
                            sell_strategy = json.loads(self.login_handler.config.get(investment_strategy, sell_key))
                            sell_strategy['key'] = sell_key
                            self.strategies[investment_strategy].append(sell_strategy)
                        except json.JSONDecodeError as ex:
                            logging.warning(f"{investment_strategy} - {sell_key} 파싱 실패: {ex}")

            # ===== 4. "통합 전략"이 없으면 추가 (API 호출 없음) =====
            if "통합 전략" not in existing_stgnames:
                if not self.login_handler.config.has_section('STRATEGIES'):
                    self.login_handler.config.add_section('STRATEGIES')
                
                self.login_handler.config.set('STRATEGIES', 'stg_integrated', "통합 전략")
                existing_stgnames.add("통합 전략")
                
                with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                    self.login_handler.config.write(configfile)
                
                logging.debug("'통합 전략' 설정 파일에 추가")

            # ===== 5. 전략 콤보박스 채우기 =====
            self.comboStg.blockSignals(True)
            for stgname in existing_stgnames:
                self.comboStg.addItem(stgname)
            
            # ===== 5-1. 백테스팅 탭 전략 콤보박스에도 추가 =====
            if hasattr(self, 'bt_strategy_combo'):
                self.bt_strategy_combo.clear()
                for stgname in existing_stgnames:
                    self.bt_strategy_combo.addItem(stgname)
                
                # 기본값: 통합 전략
                index = self.bt_strategy_combo.findText("통합 전략")
                if index != -1:
                    self.bt_strategy_combo.setCurrentIndex(index)
                logging.info(f"✅ 백테스팅 전략 콤보박스 초기화 완료 ({len(existing_stgnames)}개)")
            
            # ===== 6. 마지막 선택 전략 복원 =====
            last_strategy = self.login_handler.config.get('SETTINGS', 'last_strategy', fallback='통합 전략')
            index = self.comboStg.findText(last_strategy)
            if index != -1:
                self.comboStg.setCurrentIndex(index)
            else:
                # 마지막 전략을 찾을 수 없으면 "통합 전략" 선택
                index = self.comboStg.findText("통합 전략")
                if index != -1:
                    self.comboStg.setCurrentIndex(index)
            
            self.comboStg.blockSignals(False)
            
            logging.info(f"✅ 전략 목록 로드 완료 ({len(existing_stgnames)}개)")

            # ===== 7. 선택된 전략 활성화 (stgChanged 호출) =====
            # 조건검색 리스트는 stgChanged()에서 필요할 때만 로드
            self.is_loading_strategy = True
            self.stgChanged()
            self.is_loading_strategy = False

        except Exception as ex:
            logging.error(f"load_strategy -> {ex}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "오류", f"전략 로드 중 오류:\n{str(ex)}")

    def load_strategy_async(self):
        """전략 로드 (백그라운드 비동기 - API 제한 회피)"""
        try:
            # ===== 1. 기본 설정만 먼저 로드 (즉시) =====
            self.dataStg = []
            self.data8537 = {}
            self.strategies = {}

            self.comboStg.clear()
            self.comboBuyStg.clear()
            self.buystgInputWidget.clear()

            # ===== 2. 설정 파일에서 전략 목록 읽기 (API 호출 없음) =====
            # ✅ 순서 보장을 위해 list 사용 (set → list 변경)
            if self.login_handler.config.has_section('STRATEGIES'):
                # STRATEGIES 섹션의 키를 정렬하여 순서 보장
                strategy_keys = sorted(self.login_handler.config['STRATEGIES'].keys())
                existing_stgnames = []
                seen = set()  # 중복 제거용
                for key in strategy_keys:
                    stg_value = self.login_handler.config['STRATEGIES'][key]
                    if stg_value not in seen:
                        existing_stgnames.append(stg_value)
                        seen.add(stg_value)
            else:
                existing_stgnames = []
                
            logging.debug(f"설정 파일에서 {len(existing_stgnames)}개 전략 로드 (순서 보장)")

            # ===== 3. 설정 파일에서 전략별 매수/매도 조건 읽기 =====
            for investment_strategy in existing_stgnames:
                if self.login_handler.config.has_section(investment_strategy):
                    self.strategies[investment_strategy] = []
                    
                    # 매수 전략 로드
                    buy_keys = sorted(
                        [k for k in self.login_handler.config[investment_strategy] if k.startswith('buy_stg_')],
                        key=lambda x: int(x.split('_')[-1])
                    )
                    for buy_key in buy_keys:
                        try:
                            buy_strategy = json.loads(self.login_handler.config.get(investment_strategy, buy_key))
                            buy_strategy['key'] = buy_key
                            self.strategies[investment_strategy].append(buy_strategy)
                        except json.JSONDecodeError as ex:
                            logging.warning(f"{investment_strategy} - {buy_key} 파싱 실패: {ex}")

                    # 매도 전략 로드
                    sell_keys = sorted(
                        [k for k in self.login_handler.config[investment_strategy] if k.startswith('sell_stg_')],
                        key=lambda x: int(x.split('_')[-1])
                    )
                    for sell_key in sell_keys:
                        try:
                            sell_strategy = json.loads(self.login_handler.config.get(investment_strategy, sell_key))
                            sell_strategy['key'] = sell_key
                            self.strategies[investment_strategy].append(sell_strategy)
                        except json.JSONDecodeError as ex:
                            logging.warning(f"{investment_strategy} - {sell_key} 파싱 실패: {ex}")

            # ===== 4. 콤보박스에 전략 추가 =====
            self.comboStg.blockSignals(True)
            for stgname in existing_stgnames:
                self.comboStg.addItem(stgname)
            
            # ===== 4-1. 백테스팅 탭 전략 콤보박스에도 추가 =====
            if hasattr(self, 'bt_strategy_combo'):
                self.bt_strategy_combo.clear()
                for stgname in existing_stgnames:
                    self.bt_strategy_combo.addItem(stgname)
                
                # 기본값: 통합 전략
                index = self.bt_strategy_combo.findText("통합 전략")
                if index != -1:
                    self.bt_strategy_combo.setCurrentIndex(index)
                logging.info(f"✅ 백테스팅 전략 콤보박스 초기화 완료 ({len(existing_stgnames)}개)")
            
            # 마지막 선택 전략 복원 (SETTINGS 섹션에서 로드)
            if self.login_handler.config.has_option('SETTINGS', 'last_strategy'):
                last_strategy = self.login_handler.config.get('SETTINGS', 'last_strategy')
                index = self.comboStg.findText(last_strategy)
                if index != -1:
                    self.comboStg.setCurrentIndex(index)
                    logging.info(f"✅ 마지막 전략 복원: {last_strategy}")
                else:
                    # 마지막 전략을 찾을 수 없으면 "통합 전략" 선택
                    index = self.comboStg.findText("통합 전략")
                    if index != -1:
                        self.comboStg.setCurrentIndex(index)
                        logging.info(f"⚠️ 마지막 전략 '{last_strategy}'을 찾을 수 없어 '통합 전략' 선택")
            else:
                # last_strategy 설정이 없으면 "통합 전략" 기본 선택
                index = self.comboStg.findText("통합 전략")
                if index != -1:
                    self.comboStg.setCurrentIndex(index)
                    logging.info(f"ℹ️ 저장된 전략 없음, 기본값 '통합 전략' 선택")
            
            self.comboStg.blockSignals(False)
            
            logging.info(f"✅ 전략 목록 로드 완료 ({len(existing_stgnames)}개)")

            # ===== 5. 초기 전략 설정 (백그라운드 로드 전에 먼저 실행) =====
            self.is_loading_strategy = True
            self.stgChanged()
            self.is_loading_strategy = False

            # ===== 6. 조건검색 리스트는 메인 스레드에서 로드 =====
            self._load_condition_list_background(existing_stgnames)
            
        except Exception as ex:
            logging.error(f"load_strategy_async -> {ex}\n{traceback.format_exc()}")

    def _load_condition_list_background(self, existing_stgnames):
        """조건검색 리스트 백그라운드 로드 (COM 초기화 포함)"""
        def worker():
            try:
                # COM 초기화 (백그라운드 스레드에서 필요)
                import pythoncom
                pythoncom.CoInitialize()
                
                logging.info("📋 조건검색 리스트 로드 중... (백그라운드)")
                
                # API 제한 확인
                if not self._check_api_limit_and_wait("조건검색 리스트 로드", 0):
                    logging.warning("❌ API 제한으로 조건검색 리스트 로드 거부")
                    return
                
                # 조건검색 리스트 로드
                self.data8537 = self.objstg.requestList()
                self._condition_list_loaded = True
                
                # 새로운 전략 추가
                for stgname, v in self.data8537.items():
                    if stgname not in existing_stgnames:
                        existing_keys = self.login_handler.config['STRATEGIES'].keys()
                        existing_numbers = []
                        for k in existing_keys:
                            match = re.match(r'stg_?(\d+)', k)
                            if match:
                                existing_numbers.append(int(match.group(1)))
                        next_number = max(existing_numbers, default=0) + 1
                        new_key = f'stg_{next_number}'
                        
                        self.login_handler.config.set('STRATEGIES', new_key, stgname)
                        self.comboStg.addItem(stgname)
                        existing_stgnames.add(stgname)
                
                # 설정 파일 저장
                if len(self.data8537) != len(existing_stgnames) - 1:  # "통합 전략" 제외
                    with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                        self.login_handler.config.write(configfile)
                
                logging.info("✅ 조건검색 리스트 로드 완료")
                
                # 백그라운드 로드 완료 후 새로운 전략만 추가 (중복 실행 방지)
                QTimer.singleShot(0, self._on_background_load_complete)
                
            except Exception as ex:
                logging.error(f"백그라운드 조건검색 로드 실패: {ex}")
            finally:
                # COM 정리
                try:
                    pythoncom.CoUninitialize()
                except:
                    pass
        
        # 메인 스레드에서 직접 실행 (COM 초기화 불필요)
        try:
            logging.info("📋 조건검색 리스트 로드 중... (메인 스레드)")
            
            # API 제한 확인
            if not self._check_api_limit_and_wait("조건검색 리스트 로드", 0):
                logging.warning("❌ API 제한으로 조건검색 리스트 로드 거부")
                return
            
            # 조건검색 리스트 로드
            self.data8537 = self.objstg.requestList()
            self._condition_list_loaded = True
            
            # 새로운 전략 추가
            for stgname, v in self.data8537.items():
                if stgname not in existing_stgnames:
                    existing_keys = self.login_handler.config['STRATEGIES'].keys()
                    existing_numbers = []
                    for k in existing_keys:
                        match = re.match(r'stg_?(\d+)', k)
                        if match:
                            existing_numbers.append(int(match.group(1)))
                    next_number = max(existing_numbers, default=0) + 1
                    new_key = f'stg_{next_number}'
                    
                    self.login_handler.config.set('STRATEGIES', new_key, stgname)
                    self.comboStg.addItem(stgname)
                    existing_stgnames.add(stgname)
            
            # 설정 파일 저장
            if len(self.data8537) != len(existing_stgnames) - 1:  # "통합 전략" 제외
                with open(self.login_handler.config_file, 'w', encoding='utf-8') as configfile:
                    self.login_handler.config.write(configfile)
            
            logging.info("✅ 조건검색 리스트 로드 완료")
            
            # 메인 스레드에서 직접 처리
            self._on_condition_list_loaded()
            
        except Exception as ex:
            logging.error(f"메인 스레드 조건검색 로드 실패: {ex}")

    def _on_condition_list_loaded(self):
        """조건검색 리스트 로드 완료 후 조건검색 시작"""
        try:
            current_stg = self.comboStg.currentText()
            
            # ===== 조건검색 리스트 로드 완료 후 처리 =====
            if not hasattr(self, 'data8537') or not self.data8537:
                logging.warning("⚠️ 조건검색 리스트가 비어있음")
                return
            
            logging.info(f"✅ 조건검색 리스트 로드 완료 - {current_stg} 조건검색 시작")
            
            # 전략별 조건검색 시작
            if current_stg == "통합 전략":
                # 통합 전략은 급등주, 갭상승 조건검색 시작
                self._start_condition_search("급등주")
                self._start_condition_search("갭상승")
                
            elif current_stg == "VI 발동":
                # VI 발동은 조건검색 없음 (실시간 VI 감시)
                logging.info(f"{current_stg} 전략은 조건검색 없음 (VI 실시간 감시)")
                
            else:
                # 기타 전략은 해당 전략 조건검색 시작
                if current_stg in self.data8537:
                    self._start_condition_search(current_stg)
                    
                    # static 전략인 경우 종목 로드도 시작
                    strategy_type = self.get_strategy_type(current_stg)
                    if strategy_type == 'static':
                        logging.info(f"🔍 static 전략 '{current_stg}' 종목 로드 시작")
                        item = self.data8537.get(current_stg)
                        if item:
                            id = item['ID']
                            name = item['전략명']
                            self._load_static_strategy(id, name)
                else:
                    logging.warning(f"⚠️ 조건검색 '{current_stg}'을 data8537에서 찾을 수 없음")
        except Exception as ex:
            logging.error(f"조건검색 시작 실패: {ex}")

    def _load_static_strategy(self, id, name):
        """static 전략 로드 (전일급등 등) - 메인 스레드에서 실행"""
        try:
            logging.info(f"📋 {name} 전략 로드 중... (메인 스레드)")
            
            # API 제한 확인
            if not self._check_api_limit_and_wait(f"{name} 전략 로드", 0):
                logging.warning(f"❌ {name}: API 제한으로 전략 로드 거부")
                return
            
            # static 전략 데이터 로드
            ret, dataStg = self.objstg.requestStgID(id)
            if ret and len(dataStg) > 0:
                max_load = self.get_max_static_load()
                stock_count = len(dataStg)
                
                logging.info(f"✅ {name} 전략 로드 완료 ({stock_count}개 종목)")
                
                # 메인 스레드에서 직접 처리
                self._load_static_stocks_complete(dataStg, max_load)
            else:
                logging.warning(f"{name} 전략 데이터 없음")
                
        except Exception as ex:
            logging.error(f"메인 스레드 {name} 전략 로드 실패: {ex}")

    def _load_static_stocks_complete_direct(self, stock_list, max_load):
        """static 전략 종목 로드 완료 (백그라운드 스레드에서 직접 실행)"""
        try:
            logging.info(f"🔍 _load_static_stocks_complete_direct 호출됨 - stock_list: {len(stock_list) if stock_list else 'None'}")
            
            if not stock_list:
                logging.warning("static 전략 종목 리스트가 비어있음")
                return
            
            stock_count = len(stock_list)
            logging.info(f"📦 static 전략 종목 로드 시작 ({stock_count}개)")
            
            # 종목 리스트 내용 확인
            for i, stock in enumerate(stock_list[:3]):  # 처음 3개만 로그
                logging.info(f"  종목 {i+1}: {stock}")
            
            if stock_count > max_load:
                logging.info(f"📦 제한 적용: {max_load}개만 로드")
                self._load_stocks_from_list_safely_with_limit_direct(stock_list, max_count=max_load)
            else:
                logging.info(f"📦 전체 로드: {stock_count}개 모두 로드")
                self._load_stocks_from_list_safely_direct(stock_list)
                
            logging.info(f"✅ static 전략 종목 로드 완료")
            
        except Exception as ex:
            logging.error(f"static 전략 종목 로드 실패: {ex}")
            import traceback
            logging.error(f"상세 오류: {traceback.format_exc()}")

    def _load_stocks_from_list_safely_direct(self, stock_list):
        """리스트에서 종목 안전하게 로드 (백그라운드에서 직접 실행)"""
        try:
            logging.info(f"🔍 _load_stocks_from_list_safely_direct 호출됨")
            
            if not stock_list:
                logging.warning("stock_list가 비어있음")
                return
                
            total_count = len(stock_list)
            logging.info(f"조건검색에서 {total_count}개 종목 로드 시작...")
            
            # stock_list 구조 확인
            if total_count > 0:
                logging.debug(f"첫 번째 종목 구조: {stock_list[0]}")
            
            loaded_count = 0
            failed_count = 0
            
            for idx, stock_item in enumerate(stock_list, 1):
                try:
                    # 다양한 키 시도
                    code = stock_item.get('code') or stock_item.get('Code') or stock_item.get('CODE')
                    if not code:
                        logging.warning(f"종목 {idx}: 코드가 없음 - 키들: {list(stock_item.keys())}")
                        failed_count += 1
                        continue
                    
                    stock_name = cpCodeMgr.CodeToName(code)
                    logging.debug(f"종목 {idx}/{total_count}: {stock_name}({code}) 로드 중...")
                    
                    # 종목 로드 (메인 스레드에서 안전하게)
                    if self._load_single_stock_safely(code):
                        loaded_count += 1
                        logging.info(f"✅ {stock_name}({code}) 로드 성공")
                    else:
                        failed_count += 1
                        logging.warning(f"❌ {stock_name}({code}) 로드 실패")
                    
                    time.sleep(1.5)  # API 제한 고려하여 증가
                    
                except Exception as ex:
                    logging.error(f"종목 {idx} 로드 실패: {ex}")
                    failed_count += 1
                    continue
            
            logging.info(
                f"조건검색 종목 로드 완료: 성공 {loaded_count}개, "
                f"실패 {failed_count}개 / 전체 {total_count}개"
            )
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_list_safely_direct: {ex}")
            import traceback
            logging.error(f"상세 오류: {traceback.format_exc()}")

    def _load_single_stock_safely_direct(self, code):
        """단일 종목 로드 (백그라운드에서 직접 실행)"""
        try:
            # 모니터링 시작
            if self.trader.daydata.monitor_code(code) and self.trader.mindata.monitor_code(code):
                self.trader.monistock_set.add(code)
                
                # 백그라운드에서 직접 UI 업데이트 시도
                logging.info(f"📋 {code} 모니터링 세트에 추가 완료")
                logging.info(f"📋 현재 모니터링 종목 수: {len(self.trader.monistock_set)}")
                
                # UI 업데이트를 위한 직접 호출
                try:
                    self.trader.stock_added_to_monitor.emit(code)
                    logging.info(f"📋 {code} UI 업데이트 시그널 발송 완료")
                except Exception as ui_ex:
                    logging.error(f"UI 업데이트 시그널 발송 실패: {ui_ex}")
                
                return True
            else:
                logging.warning(f"{code} 데이터 로드 실패")
                return False
        except Exception as ex:
            logging.error(f"{code} 로드 실패: {ex}")
            return False

    def _load_stocks_from_list_safely_with_limit_direct(self, stock_list, max_count=10):
        """리스트에서 종목 안전하게 로드 (개수 제한 + 메인 스레드 실행)"""
        try:
            if not stock_list:
                logging.info("로드할 종목 리스트 없음")
                return
            
            total_codes = len(stock_list)
            actual_count = min(total_codes, max_count)
            
            if total_codes > max_count:
                logging.info(f"종목 개수 제한: {total_codes}개 → {actual_count}개")
                limited_list = stock_list[:actual_count]
            else:
                limited_list = stock_list
            
            # 메인 스레드에서 직접 로드
            self._load_stocks_from_list_safely_sync(limited_list)
            
            if total_codes > max_count:
                remaining = total_codes - max_count
                logging.info(f"💡 나머지 {remaining}개 종목은 실시간 편입으로 추가됩니다.")
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_list_safely_with_limit_direct: {ex}")

    def _load_static_stocks_complete(self, stock_list, max_load):
        """static 전략 종목 로드 완료 (메인 스레드에서 실행)"""
        try:
            logging.debug(f"🔍 _load_static_stocks_complete 호출됨 - stock_list: {len(stock_list) if stock_list else 'None'}")
            
            if not stock_list:
                logging.warning("static 전략 종목 리스트가 비어있음")
                return
            
            stock_count = len(stock_list)
            logging.info(f"📦 static 전략 종목 로드 시작 ({stock_count}개)")
            
            # 종목 리스트 내용 확인
            for i, stock in enumerate(stock_list[:3]):  # 처음 3개만 로그
                logging.debug(f"  종목 {i+1}: {stock}")
            
            if stock_count > max_load:
                logging.info(f"📦 제한 적용: {max_load}개만 로드")
                self._load_stocks_from_list_safely_with_limit(stock_list, max_count=max_load)
            else:
                logging.info(f"📦 전체 로드: {stock_count}개 모두 로드")
                # 메인 스레드에서 안전하게 실행
                self._load_stocks_from_list_safely_sync(stock_list)
                
            logging.info(f"✅ static 전략 종목 로드 완료")
            
        except Exception as ex:
            logging.error(f"static 전략 종목 로드 실패: {ex}")
            import traceback
            logging.error(f"상세 오류: {traceback.format_exc()}")

    def stgChanged(self, *args):
        """전략 변경 시 처리"""
        try:
            stgName = self.comboStg.currentText()
            self.save_last_stg()

            if not self.is_loading_strategy:
                self.sell_all_item()
                self.trader.clear_list_db('mylist.db')
                
                # ===== 모니터링 데이터 초기화 =====
                for code in list(self.trader.monistock_set):
                    self.trader.tickdata.monitor_stop(code)
                    self.trader.mindata.monitor_stop(code)
                    self.trader.daydata.monitor_stop(code)
                self.trader.monistock_set.clear()
                
                # ===== UI 리스트 박스 초기화 =====
                self.firstListBox.clear()
                self.secondListBox.clear()
                logging.info(f"전략 변경: {stgName} - 모니터링 데이터 및 UI 초기화 완료")
            
            # ===== ✅ 조건검색 리스트는 백그라운드에서 로드됨 =====
            # stgChanged에서는 조건검색 리스트 로드하지 않음 (중복 방지)
            
            # ===== ❌ 큐 스레드 시작 제거 (이미 post_login_setup에서 시작됨) =====
            # self.objstg.start_processing_queue()  # ← 삭제
            
            if hasattr(self, 'momentum_scanner') and self.momentum_scanner:
                self.momentum_scanner = None
            
            # ===== VI 발동 전략 =====
            if stgName == 'VI 발동':
                self.objstg.Clear()
                logging.info(f"전략 초기화: VI 발동")
                
                self.trader.init_stock_balance()
                self._update_listboxes_from_balance()
                self._load_stocks_from_db_safely('mylist.db')
                
                # pb9619 중복 구독 방지
                if not hasattr(self, 'pb9619') or self.pb9619 is None:
                    self.pb9619 = CpPB9619()
                    self.pb9619.Subscribe("", self.trader)
                    logging.info("pb9619 구독 시작")
                else:
                    logging.info("pb9619 이미 구독 중")

            # ===== 통합 전략 =====
            elif stgName == "통합 전략":
                if hasattr(self, 'pb9619'):
                    self.pb9619.Unsubscribe()
                self.objstg.Clear()
                
                logging.info(f"=== 통합 전략 시작 ===")
                self.trader.init_stock_balance()
                self._update_listboxes_from_balance()
                
                # ✅ Scanner는 CpStrategy.__init__에서 이미 초기화됨 (중복 제거)
                # self.objstg.momentum_scanner = MomentumScanner(self.trader)
                # self.objstg.gap_scanner = GapUpScanner(self.trader)
                
                self._load_stocks_from_db_safely('mylist.db')                
                
                self.volatility_strategy = VolatilityBreakout(self.trader)
                self.trader_thread.set_volatility_strategy(self.volatility_strategy)
                
                self.gap_scanner = self.objstg.gap_scanner
                logging.info("✅ 통합 전략 초기화 완료")
                
                # ===== ✅ 조건검색 리스트 로드 및 시작 =====
                if hasattr(self, 'data8537') and self.data8537:
                    # 이미 로드된 경우 바로 시작
                    logging.info("📋 조건검색 리스트 이미 로드됨 - 급등주, 갭상승 시작")
                    self._start_condition_search("급등주")
                    self._start_condition_search("갭상승")
                else:
                    # 백그라운드에서 조건검색 리스트 로드
                    logging.info("📋 조건검색 리스트 백그라운드 로드 시작...")
                    QTimer.singleShot(100, self._load_condition_search_background)

            # ===== 기타 전략 =====
            else:
                if hasattr(self, 'pb9619'):
                    self.pb9619.Unsubscribe()
                
                # ===== ✅ 기존 조건검색 모두 정리 =====
                self.objstg.Clear()
                
                logging.info(f"전략 초기화: {stgName}")
                self.trader.init_stock_balance()
                self._update_listboxes_from_balance()
                
                self._load_stocks_from_db_safely('mylist.db')
                
                # 조건검색 리스트가 로드되었는지 확인
                if hasattr(self, 'data8537') and self.data8537:
                    item = self.data8537.get(stgName)
                    if item:
                        id = item['ID']
                        name = item['전략명']
                        strategy_type = self.get_strategy_type(name)
                        
                        if strategy_type == 'static':
                            # static 전략은 백그라운드에서 로드
                            self._load_static_strategy(id, name)
                        
                        self._start_condition_search(stgName)
                    else:
                        logging.warning(f"전략 '{stgName}'을 조건검색 리스트에서 찾을 수 없음")
                else:
                    logging.info(f"조건검색 리스트 로드 대기 중... ({stgName})")
                    # 백그라운드에서 조건검색 리스트 로드
                    logging.info("📋 조건검색 리스트 백그라운드 로드 시작...")
                    QTimer.singleShot(100, self._load_condition_search_background)
            
            logging.info(f"{stgName} 전략 감시 시작")
            
            # 콤보박스 업데이트
            self.comboBuyStg.clear()
            self.comboSellStg.clear()
            if stgName in self.strategies:
                for strategy in self.strategies[stgName]:
                    strategy_name = strategy.get('name')
                    strategy_key = strategy.get('key')
                    if strategy_key.startswith('buy'):
                        self.comboBuyStg.addItem(strategy_name, strategy_key)
                    elif strategy_key.startswith('sell'):
                        self.comboSellStg.addItem(strategy_name, strategy_key)
                
                if self.comboBuyStg.count() > 0:
                    self.comboBuyStg.setCurrentIndex(0)
                    self.buyStgChanged()
                
                if self.comboSellStg.count() > 0:
                    self.comboSellStg.setCurrentIndex(0)
                    self.sellStgChanged()
            
        except Exception as ex:
            logging.error(f"stgChanged: {ex}\n{traceback.format_exc()}")
            
    def _load_stocks_from_list_safely_with_limit(self, stock_list, max_count=10):
        """리스트에서 종목 안전하게 로드 (개수 제한 + 백그라운드)"""
        try:
            if not stock_list:
                logging.info("로드할 종목 리스트 없음")
                return
            
            total_codes = len(stock_list)
            actual_count = min(total_codes, max_count)
            
            if total_codes > max_count:
                logging.warning(
                    f"⚠️ 종목 수({total_codes}개)가 제한({max_count}개)을 초과합니다. "
                    f"상위 {max_count}개만 로드합니다."
                )
            
            # 제한된 리스트로 백그라운드 로드
            limited_list = stock_list[:actual_count]
            
            if actual_count > 3:
                self._load_stocks_in_background(limited_list)
            else:
                self._load_stocks_from_list_safely_sync(limited_list)
            
            if total_codes > max_count:
                remaining = total_codes - max_count
                logging.info(f"💡 나머지 {remaining}개 종목은 실시간 편입으로 추가됩니다.")
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_list_safely_with_limit: {ex}")
            
    def _start_condition_search(self, strategy_name):
        """조건검색 시작 (안전)"""
        try:
            stg_item = self.data8537.get(strategy_name)
            if not stg_item:
                logging.warning(f"⚠️ 조건검색 '{strategy_name}'을 찾을 수 없습니다. HTS에서 생성하세요.")
                return False
            
            id = stg_item['ID']
            name = stg_item['전략명']
            
            # ===== ✅ 200종목 제한 체크 (실시간 감시 시작 전) =====
            ret, stockList = self.objstg.requestStgID(id)
            if not ret:
                logging.warning(f"❌ 조건검색 '{name}' 종목 조회 실패")
                return False
            
            stock_count = len(stockList)
            logging.info(f"📊 조건검색 '{name}': 현재 {stock_count}개 종목 검색됨")
            
            
            if stock_count >= 200:
                logging.warning(f"{'='*40}")
                logging.warning(f"⚠️  경고: 조건검색 '{name}' 종목 수 초과!")
                logging.warning(f"⚠️  검색된 종목: {stock_count}개 (제한: 200개)")
                logging.warning(f"⚠️  실시간 감시가 불가능합니다.")
                logging.warning(f"⚠️  대신증권 HTS에서 조건을 더 엄격하게 수정하세요.")
                logging.warning(f"{'='*40}")
                return False
            
            # ===== ✅ 기존 편입 종목 검증 처리 (모니터 ID 획득 성공/실패와 무관) =====
            if stock_count > 0 and stockList:
                logging.info(f"🔍 기존 편입 종목 {stock_count}개 검증 시작...")
                
                # 큐 처리 스레드가 시작되었는지 확인
                if not self.objstg.is_thread_started:
                    self.objstg.start_processing_queue()
                
                # 각 종목을 처리 큐에 추가
                added_count = 0
                for stock_info in stockList:
                    try:
                        code = stock_info.get('code', '')
                        price = stock_info.get('price', 0)
                        
                        if code:
                            # 현재 시간 생성
                            time_str = datetime.now().strftime('%m/%d %H:%M:%S')
                            
                            # 큐에 직접 추가 (모니터 ID 없이도 처리 가능하도록 수정)
                            self.objstg.stock_queue.put({
                                'stgid': id,
                                'stgmonid': 0,  # 모니터 ID 없음
                                'code': code,
                                'stgprice': price,
                                'time': time_str,
                                'stgname': name
                            })
                            added_count += 1
                            
                            # API 제한 고려하여 짧은 대기
                            time.sleep(0.05)
                            
                    except Exception as ex:
                        logging.error(f"기존 종목 {code} 큐 추가 실패: {ex}")
                        continue
                
                logging.info(f"✅ 기존 편입 종목 {added_count}개를 처리 큐에 추가 완료")
            
            # ===== ✅ 모니터 ID 획득 시도 =====
            ret, monid = self.objstg.requestMonitorID(id)
            if not ret:
                logging.info(f"ℹ️ 조건검색 '{name}' 모니터 ID 획득 실패 - 다음날 대신증권에서 자동 초기화됩니다")
                # 기존 편입 종목은 이미 처리했으므로 True 반환
                return True
            
            ret, status = self.objstg.requestStgControl(id, monid, True, name)
            if ret:
                logging.info(f"✅ 조건검색 감시 시작: [{name}] ({stock_count}개 종목)")
                return True
            else:
                logging.info(f"ℹ️ 조건검색 '{name}' 시작 실패 - 다음날 대신증권에서 자동 초기화됩니다")
                # 기존 편입 종목은 이미 처리했으므로 True 반환
                return True
                
        except Exception as ex:
            logging.error(f"_start_condition_search({strategy_name}) -> {ex}")
            return False

    def _load_condition_search_background(self):
        """조건검색 리스트를 백그라운드에서 로드"""
        try:
            logging.info("📋 조건검색 리스트 로드 시도...")
            # 로그 즉시 플러시
            for handler in logging.getLogger().handlers:
                if hasattr(handler, 'stream'):
                    handler.stream.flush()
            
            # 조건검색 리스트 로드
            self.data8537 = self.objstg.requestList()
            self._condition_list_loaded = True
            logging.info("✅ 조건검색 리스트 로드 완료")
            # 로그 즉시 플러시
            for handler in logging.getLogger().handlers:
                if hasattr(handler, 'stream'):
                    handler.stream.flush()
            
            # 로드 후 바로 시작
            self._start_condition_search("급등주")
            self._start_condition_search("갭상승")
            
        except Exception as ex:
            logging.error(f"조건검색 리스트 로드 실패: {ex}")

    def _load_stocks_from_db_safely(self, db_file='mylist.db'):
        """DB에서 종목 안전하게 로드 (백그라운드)"""
        try:
            self.trader.load_from_list_db(db_file)
            
            codes_to_load = list(self.trader.database_set)
            total_codes = len(codes_to_load)
            
            if total_codes == 0:
                logging.info("mylist.db에 저장된 종목 없음")
                return
            
            logging.info(f"DB에서 {total_codes}개 종목 로드...")
            logging.info(f"로드할 종목 목록: {codes_to_load}")
            
            # ===== ✅ 모든 종목을 리스트박스에 추가 (중복 방지) =====
            for code in codes_to_load:
                self._add_to_listbox_if_not_exists(self.firstListBox, code, "DB")
            
            # ===== ✅ 종목이 많으면 백그라운드로 =====
            if total_codes > 3:
                # 종목 리스트를 딕셔너리 형태로 변환
                stock_list = [{'code': code} for code in codes_to_load]
                self._load_stocks_in_background(stock_list)
            else:
                # 동기 처리
                self._load_stocks_from_db_safely_sync(codes_to_load)
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_db_safely: {ex}")

    def _update_listboxes_from_balance(self):
        """잔고에 있는 종목들을 리스트박스에 추가 (중복 방지)"""
        try:
            # 투자대상종목 리스트박스에 추가
            for code in self.trader.monistock_set:
                self._add_to_listbox_if_not_exists(self.firstListBox, code, "잔고")
            
            # 투자종목 리스트박스에 추가 (매수된 종목들)
            for code in self.trader.bought_set:
                self._add_to_listbox_if_not_exists(self.secondListBox, code, "잔고")
            
            logging.info(f"📋 잔고 종목 리스트박스 업데이트 완료: 모니터링 {len(self.trader.monistock_set)}개, 매수 {len(self.trader.bought_set)}개")
            
        except Exception as ex:
            logging.error(f"_update_listboxes_from_balance: {ex}")

    def _add_to_listbox_if_not_exists(self, listbox, code, source="알 수 없음"):
        """리스트박스에 종목을 중복 없이 추가"""
        try:
            existing_items = [listbox.item(i).text() for i in range(listbox.count())]
            if code not in existing_items:
                listbox.addItem(code)
                logging.info(f"📋 {code} {source}에서 리스트박스에 추가됨")
                return True
            else:
                logging.debug(f"📋 {code} 이미 리스트박스에 존재함 ({source}, 중복 방지)")
                return False
        except Exception as ex:
            logging.error(f"_add_to_listbox_if_not_exists: {ex}")
            return False

    def _load_stocks_from_db_safely_sync(self, codes_to_load):
        """DB에서 종목 동기 로드 (적은 종목용)"""
        try:
            total_codes = len(codes_to_load)
            loaded_count = 0
            failed_count = 0
            
            for idx, code in enumerate(codes_to_load, 1):
                try:
                    if code in self.trader.monistock_set:
                        logging.debug(f"{code}: 이미 모니터링 중")
                        # 이미 모니터링 중이어도 리스트박스에 직접 추가 (중복 방지)
                        self._add_to_listbox_if_not_exists(self.firstListBox, code, "이미 모니터링 중")
                        loaded_count += 1
                        continue
                    
                    if self._load_single_stock_safely(code):
                        loaded_count += 1
                    else:
                        failed_count += 1
                    
                    time.sleep(1.5)  # API 제한 고려하여 증가
                    
                except Exception as ex:
                    logging.error(f"{code} 로드 실패: {ex}")
                    failed_count += 1
                    continue
            
            logging.info(
                f"DB 종목 로드 완료: 성공 {loaded_count}개, "
                f"실패 {failed_count}개 / 전체 {total_codes}개"
            )
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_db_safely_sync: {ex}")

    def _load_stocks_from_list_safely(self, stock_list):
        """리스트에서 종목 안전하게 로드 (백그라운드)"""
        try:
            if not stock_list:
                logging.info("로드할 종목 리스트 없음")
                return
            
            total_count = len(stock_list)
            
            # ===== ✅ 종목이 많으면 백그라운드로 처리 =====
            if total_count > 3:
                self._load_stocks_in_background(stock_list)
            else:
                # 종목이 적으면 동기 처리
                self._load_stocks_from_list_safely_sync(stock_list)
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_list_safely: {ex}")

    def _load_stocks_from_list_safely_sync(self, stock_list):
        """리스트에서 종목 동기 로드 (적은 종목용)"""
        try:
            logging.debug(f"🔍 _load_stocks_from_list_safely_sync 호출됨")
            
            if not stock_list:
                logging.warning("stock_list가 비어있음")
                return
                
            total_codes = len(stock_list)
            logging.info(f"조건검색에서 {total_codes}개 종목 로드 시작...")
            
            # stock_list 구조 확인
            if total_codes > 0:
                logging.info(f"첫 번째 종목 구조: {stock_list[0]}")
            
            loaded_count = 0
            failed_count = 0
            
            for idx, stock_item in enumerate(stock_list, 1):
                try:
                    # 다양한 키 시도
                    code = stock_item.get('code') or stock_item.get('Code') or stock_item.get('CODE')
                    if not code:
                        logging.warning(f"종목 {idx}: 코드가 없음 - 키들: {list(stock_item.keys())}")
                        failed_count += 1
                        continue
                    
                    stock_name = cpCodeMgr.CodeToName(code)
                    logging.info(f"종목 {idx}/{total_codes}: {stock_name}({code}) 로드 중...")
                    
                    # 종목 로드
                    if self._load_single_stock_safely(code):
                        loaded_count += 1
                        logging.info(f"✅ {stock_name}({code}) 로드 성공")
                    else:
                        failed_count += 1
                        logging.warning(f"❌ {stock_name}({code}) 로드 실패")
                    
                    time.sleep(1.5)  # API 제한 고려하여 증가
                    
                except Exception as ex:
                    logging.error(f"종목 {idx} 로드 실패: {ex}")
                    failed_count += 1
                    continue
            
            logging.info(
                f"조건검색 종목 로드 완료: 성공 {loaded_count}개, "
                f"실패 {failed_count}개 / 전체 {total_codes}개"
            )
            
        except Exception as ex:
            logging.error(f"_load_stocks_from_list_safely_sync: {ex}")
            import traceback
            logging.error(f"상세 오류: {traceback.format_exc()}")

    def _load_stocks_in_background(self, stock_list):
        """백그라운드에서 종목 로드"""
        try:
            total_count = len(stock_list)
            logging.info(f"📦 {total_count}개 종목을 백그라운드에서 로드합니다...")
            
            # 프로그레스 다이얼로그 생성
            self.progress_dialog = StockLoadingProgressDialog(self, total_count)
            
            # 로딩 스레드 생성
            self.loader_thread = StockLoaderThread(
                self,
                stock_list,
                self._load_single_stock_safely
            )
            
            # 시그널 연결
            self.loader_thread.progress_updated.connect(self.progress_dialog.update_progress)
            self.loader_thread.stock_loaded.connect(self.progress_dialog.on_stock_loaded)
            self.loader_thread.loading_completed.connect(self.on_background_loading_completed)
            self.loader_thread.error_occurred.connect(self.on_background_loading_error)
            
            # 취소 버튼 연결
            self.progress_dialog.cancel_button.clicked.connect(self.loader_thread.cancel)
            self.progress_dialog.close_button.clicked.connect(self.progress_dialog.accept)
            
            # 스레드 시작
            self.loader_thread.start()
            
            # 다이얼로그 표시 (모달)
            self.progress_dialog.exec_()
            
        except Exception as ex:
            logging.error(f"_load_stocks_in_background: {ex}")

    def _load_single_stock_safely(self, code, max_retries=2):
        """단일 종목 안전하게 로드"""
        try:
            stock_name = cpCodeMgr.CodeToName(code)
            
            for attempt in range(max_retries):
                try:                    
                    # 일봉 로드
                    if not self.trader.daydata.select_code(code):
                        if attempt < max_retries - 1:
                            logging.debug(f"{code}: 일봉 로드 실패, 재시도 {attempt+1}/{max_retries}")
                            time.sleep(0.3)
                            continue
                        else:
                            logging.warning(f"{stock_name}({code}): 일봉 로드 최종 실패")
                            return False
                    
                    # 틱/분 데이터 모니터링 시작
                    tick_ok = self.trader.tickdata.monitor_code(code)
                    min_ok = self.trader.mindata.monitor_code(code)
                    
                    if not (tick_ok and min_ok):
                        if attempt < max_retries - 1:
                            logging.debug(f"{code}: 틱/분 로드 실패, 재시도 {attempt+1}/{max_retries}")
                            self.trader.daydata.monitor_stop(code)
                            time.sleep(0.3)
                            continue
                        else:
                            logging.warning(f"{stock_name}({code}): 틱/분 로드 최종 실패")
                            self.trader.daydata.monitor_stop(code)
                            return False
                    
                    # 성공
                    if code not in self.trader.starting_time:
                        self.trader.starting_time[code] = datetime.now().strftime('%m/%d 09:00:00')
                    
                    self.trader.monistock_set.add(code)
                    
                    # 리스트박스에 직접 추가 (시그널 방식 대신, 중복 방지)
                    self._add_to_listbox_if_not_exists(self.firstListBox, code, f"{stock_name}({code}) 신규 로드")
                    
                    logging.debug(f"✅ {stock_name}({code}) 로드 성공")
                    return True
                    
                except Exception as ex:
                    if attempt < max_retries - 1:
                        logging.debug(f"{code}: 로드 중 오류, 재시도 {attempt+1}/{max_retries}: {ex}")
                        time.sleep(0.3)
                    else:
                        logging.error(f"{stock_name}({code}): 로드 중 오류: {ex}")
                        self.trader.daydata.monitor_stop(code)
                        self.trader.tickdata.monitor_stop(code)
                        self.trader.mindata.monitor_stop(code)
                        return False
            
            return False
            
        except Exception as ex:
            logging.error(f"_load_single_stock_safely({code}) -> {ex}")
            return False
        
    def on_background_loading_completed(self, success, failed):
        """백그라운드 로딩 완료"""
        try:
            total = success + failed
            logging.info(f"✅ 백그라운드 종목 로딩 완료: 성공 {success}개, 실패 {failed}개 / 전체 {total}개")
            
            # 프로그레스 다이얼로그 업데이트
            if hasattr(self, 'progress_dialog'):
                self.progress_dialog.on_loading_completed(success, failed)
            
        except Exception as ex:
            logging.error(f"on_background_loading_completed: {ex}")

    def on_background_loading_error(self, error_msg):
        """백그라운드 로딩 에러"""
        logging.error(f"백그라운드 로딩 에러: {error_msg}")
        QMessageBox.critical(self, "로딩 오류", f"종목 로딩 중 오류 발생:\n{error_msg}")
        
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.reject()
        
    @pyqtSlot(dict)
    def on_momentum_stock_found(self, stock):
        """급등주 발견 시 처리"""
        logging.info(f"급등주 발견: {stock['name']}({stock['code']}) - 점수: {stock['score']}")

    def buyStgChanged(self):
        try:
            investment_strategy = self.comboStg.currentText()
            buy_strategy = self.comboBuyStg.currentText()

            if not investment_strategy or not buy_strategy:
                self.buystgInputWidget.clear()
                return

            if investment_strategy in self.strategies:
                selected_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == buy_strategy), None)
                if selected_strategy:
                    strategy_content = selected_strategy.get('content', '')
                    # ✅ 이스케이프된 \n을 실제 줄바꿈으로 변환
                    strategy_content = strategy_content.replace('\\n', '\n')
                    self.buystgInputWidget.setPlainText(strategy_content)
                else:
                    self.buystgInputWidget.clear()
            else:
                self.buystgInputWidget.clear()
        except Exception as ex:
            logging.error(f"buyStgChanged -> {ex}")

    def sellStgChanged(self):
        try:
            investment_strategy = self.comboStg.currentText()
            sell_strategy = self.comboSellStg.currentText()

            if not investment_strategy or not sell_strategy:
                self.sellstgInputWidget.clear()
                return

            if investment_strategy in self.strategies:
                selected_strategy = next((stg for stg in self.strategies[investment_strategy] if stg['name'] == sell_strategy), None)
                if selected_strategy:
                    strategy_content = selected_strategy.get('content', '')
                    # ✅ 이스케이프된 \n을 실제 줄바꿈으로 변환
                    strategy_content = strategy_content.replace('\\n', '\n')
                    self.sellstgInputWidget.setPlainText(strategy_content)
                else:
                    self.sellstgInputWidget.clear()
            else:
                self.sellstgInputWidget.clear()
        except Exception as ex:
            logging.error(f"sellStgChanged -> {ex}")

    def close_external_popup(self):
        try:
            windows = gw.getWindowsWithTitle('공지사항')
            for window in windows:
                window.close()
        except Exception as ex:
            logging.error(f"close_external_popup -> {ex}")

    def listBoxChanged(self, current):
        if current:
            self.chartdrawer.set_code(current.text())
        else:
            self.chartdrawer.set_code(None)

    def delete_select_item(self):
        selected_items = self.firstListBox.selectedItems()
        if selected_items:
            for item in selected_items:
                self.firstListBox.takeItem(self.firstListBox.row(item))
                code = item.text()
                if code in self.trader.monistock_set:
                    self.trader.monistock_set.remove(code)
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                self.trader.daydata.monitor_stop(code)
                self.trader.delete_list_db(code)

                if code == self.chartdrawer.code:
                    self.chartdrawer.set_code(None)
                logging.info(f"{cpCodeMgr.CodeToName(code)}({code}) -> 투자 대상 종목 삭제")

    def buy_item(self):
        selected_items = self.firstListBox.selectedItems()
        if selected_items:
            for item in selected_items:
                self.trader.buy_stock(item.text(), '기타', '0', '03')

    def sell_item(self):
        selected_items = self.secondListBox.selectedItems()
        if selected_items:
            for item in selected_items:
                self.trader.sell_stock(item.text(), '직접 매도')

    def sell_all_item(self):
        if self.trader.sell_all():
            for code in list(self.trader.monistock_set):
                self.on_stock_removed(code)

    def print_chart(self):
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPrinter.A4)
        printer.setOrientation(QPrinter.Portrait)

        printDialog = QPrintDialog(printer, self)
        if printDialog.exec_() == QPrintDialog.Accepted:
            painter = QPainter(printer)
            pageRect = printer.pageRect()
            canvasSize = self.canvas.size()
            xScale = pageRect.width() / canvasSize.width()
            yScale = pageRect.height() / canvasSize.height()
            scale = min(xScale, yScale)
            painter.scale(scale, scale)
            self.canvas.render(painter)
            painter.end()

    def output_current_data(self):
        try:
            tick_data_all = self.trader.tickdata.stockdata
            min_data_all = self.trader.mindata.stockdata
            day_data_all = self.trader.daydata.stockdata

            if not tick_data_all and not min_data_all and not day_data_all:
                QMessageBox.warning(self, "데이터 없음", "현재 모니터링 중인 종목의 데이터가 없습니다.")
                return

            wb = Workbook()
            
            ws_tick = wb.active
            ws_tick.title = "Tick Data"
            for code, tick_data in tick_data_all.items():
                ws_tick.append([f"종목코드: {code}"])
                tick_keys = list(tick_data.keys())
                ws_tick.append(tick_keys)

                max_len_tick = max(len(v) for v in tick_data.values())

                for i in range(max_len_tick):
                    row = []
                    for key in tick_keys:
                        if i < len(tick_data[key]):
                            row.append(tick_data[key][i])
                        else:
                            row.append(None)
                    ws_tick.append(row)
                ws_tick.append([])

            ws_min = wb.create_sheet(title="Minute Data")
            for code, min_data in min_data_all.items():
                ws_min.append([f"종목코드: {code}"])
                min_keys = list(min_data.keys())
                ws_min.append(min_keys)

                max_len_min = max(len(v) for v in min_data.values())

                for i in range(max_len_min):
                    row = []
                    for key in min_keys:
                        if i < len(min_data[key]):
                            row.append(min_data[key][i])
                        else:
                            row.append(None)
                    ws_min.append(row)
                ws_min.append([])

            ws_day = wb.create_sheet(title="Day Data")
            for code, day_data in day_data_all.items():
                ws_day.append([f"종목코드: {code}"])
                day_keys = list(day_data.keys())
                ws_day.append(day_keys)

                max_len_day = max(len(v) for v in day_data.values())

                for i in range(max_len_day):
                    row = []
                    for key in day_keys:
                        if i < len(day_data[key]):
                            row.append(day_data[key][i])
                        else:
                            row.append(None)
                    ws_day.append(row)
                ws_day.append([])

            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            filename, _ = QFileDialog.getSaveFileName(self, "Save Excel File",
                                                    f"stock_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                                    "Excel Files (*.xlsx);;All Files (*)", options=options)
            if not filename:
                QMessageBox.warning(self, "저장 취소", "파일 저장이 취소되었습니다.")
                return

            wb.save(filename)
            QMessageBox.information(self, "성공", f"데이터가 '{filename}'에 저장되었습니다.")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"데이터 저장 중 오류가 발생했습니다:\n{str(e)}")

    def print_terminal(self):
        printer = QPrinter()
        printDialog = QPrintDialog(printer, self)
        if printDialog.exec_() == QPrintDialog.Accepted:
            self.terminalOutput.print_(printer)

    def closeEvent(self, event):
        """프로그램 종료 처리"""
        reply = QMessageBox.question(self, 'Message', "Are you sure you want to quit?", 
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.save_last_stg()

            # 전략 객체 정리
            if self.momentum_scanner:
                self.momentum_scanner.stop_screening()            
            
            # 차트 업데이트 타이머 정지
            if getattr(self.trader, 'tickdata', None):
                self.trader.tickdata.update_data_timer.stop()
            if getattr(self.trader, 'mindata', None):
                self.trader.mindata.update_data_timer.stop()
            if getattr(self.trader, 'daydata', None):
                self.trader.daydata.update_data_timer.stop()
            if getattr(self, 'ui_update_timer', None):
                self.ui_update_timer.stop()

            if self.chartdrawer.chart_thread:
                self.chartdrawer.chart_thread.stop()

            # CpStrategy QThread 종료
            if self.objstg:
                self.objstg.stop_processing_queue()

            if self.trader_thread:
                self.trader_thread.stop()

            for code in list(self.trader.monistock_set):
                self.trader.tickdata.monitor_stop(code)
                self.trader.mindata.monitor_stop(code)
                self.trader.daydata.monitor_stop(code)
            self.trader.monistock_set.clear()

            self.objstg.Clear()

            self.clean_up_processes()

            logger = logging.getLogger()
            for handler in logger.handlers[:]:
                handler.close()
                logger.removeHandler(handler)

            cpStatus.PlusDisconnect()

            event.accept()
        else:
            event.ignore()

    def clean_up_processes(self):
        process_names = ['coStarter', 'CpStart', 'DibServer']
        for process_name in process_names:
            for proc in psutil.process_iter(attrs=['pid', 'name']):
                if process_name.lower() in proc.info['name'].lower():
                    proc.kill()
                    logging.warning(f"Killed  {proc.info['name']} (PID: {proc.info['pid']})")

    def init_ui(self):
        """UI 초기화 (탭 구조)"""
        self.setWindowTitle("초단타 매매 프로그램 v3.0 - 백테스팅")
        self.setGeometry(0, 0, 1900, 980)

        # ===== 메인 탭 위젯 생성 =====
        self.tab_widget = QTabWidget()
        
        # 탭 1: 실시간 매매
        self.trading_tab = QWidget()
        self.init_trading_tab()
        self.tab_widget.addTab(self.trading_tab, "실시간 매매")
        
        # 탭 2: 백테스팅
        self.backtest_tab = QWidget()
        self.init_backtest_tab()
        self.tab_widget.addTab(self.backtest_tab, "백테스팅")
        
        # 메인 레이아웃
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tab_widget)
        self.setLayout(main_layout)

    def init_trading_tab(self):
        """실시간 매매 탭 초기화"""
        
        # ===== 로그인 영역 =====
        loginLayout = QVBoxLayout()

        loginLabelLayout = QHBoxLayout()
        loginLabel = QLabel("아이디:")
        loginLabelLayout.addWidget(loginLabel)
        self.loginEdit = QLineEdit()
        loginLabelLayout.addWidget(self.loginEdit)

        passwordLabelLayout = QHBoxLayout()
        passwordLabel = QLabel("비밀번호:")
        passwordLabelLayout.addWidget(passwordLabel)
        self.passwordEdit = QLineEdit()
        self.passwordEdit.setEchoMode(QLineEdit.Password)
        passwordLabelLayout.addWidget(self.passwordEdit)

        certpasswordLabelLayout = QHBoxLayout()
        certpasswordLabel = QLabel("인증번호:")
        certpasswordLabelLayout.addWidget(certpasswordLabel)
        self.certpasswordEdit = QLineEdit()
        self.certpasswordEdit.setEchoMode(QLineEdit.Password)
        certpasswordLabelLayout.addWidget(self.certpasswordEdit)

        label_width = 70
        loginLabel.setFixedWidth(label_width)
        passwordLabel.setFixedWidth(label_width)
        certpasswordLabel.setFixedWidth(label_width)

        loginButtonLayout = QHBoxLayout()
        self.autoLoginCheckBox = QCheckBox("자동 로그인")
        loginButtonLayout.addWidget(self.autoLoginCheckBox)
        self.loginButton = QPushButton("로그인")
        loginButtonLayout.addWidget(self.loginButton)

        loginLayout.addLayout(loginLabelLayout)
        loginLayout.addLayout(passwordLabelLayout)
        loginLayout.addLayout(certpasswordLabelLayout)
        loginLayout.addLayout(loginButtonLayout)

        # ===== 투자 설정 =====
        buycountLayout = QHBoxLayout()
        buycountLabel = QLabel("최대투자 종목수 :")
        buycountLayout.addWidget(buycountLabel)
        self.buycountEdit = QLineEdit()
        buycountLayout.addWidget(self.buycountEdit)
        self.buycountButton = QPushButton("설정")
        self.buycountButton.setFixedWidth(70)
        buycountLayout.addWidget(self.buycountButton)

        # ===== 투자 대상 종목 리스트 =====
        firstListBoxLayout = QVBoxLayout()
        listBoxLabel = QLabel("투자 대상 종목 :")
        firstListBoxLayout.addWidget(listBoxLabel)
        self.firstListBox = QListWidget()
        firstListBoxLayout.addWidget(self.firstListBox, 1)
        firstButtonLayout = QHBoxLayout()
        self.buyButton = QPushButton("매입")
        firstButtonLayout.addWidget(self.buyButton)
        self.deleteFirstButton = QPushButton("삭제")        
        firstButtonLayout.addWidget(self.deleteFirstButton)        
        firstListBoxLayout.addLayout(firstButtonLayout)

        # ===== 투자 종목 리스트 =====
        secondListBoxLayout = QVBoxLayout()
        secondListBoxLabel = QLabel("투자 종목 :")
        secondListBoxLayout.addWidget(secondListBoxLabel)
        self.secondListBox = QListWidget()        
        secondListBoxLayout.addWidget(self.secondListBox, 1)
        secondButtonLayout = QHBoxLayout()
        self.sellButton = QPushButton("매도")
        secondButtonLayout.addWidget(self.sellButton)
        self.sellAllButton = QPushButton("전부 매도")
        secondButtonLayout.addWidget(self.sellAllButton)     
        secondListBoxLayout.addLayout(secondButtonLayout)

        # ===== 출력 버튼 =====
        printLayout = QHBoxLayout()
        self.printChartButton = QPushButton("차트 출력")
        printLayout.addWidget(self.printChartButton)
        self.dataOutputButton2 = QPushButton("차트데이터 저장")
        printLayout.addWidget(self.dataOutputButton2)

        # ===== 왼쪽 영역 통합 =====
        listBoxesLayout = QVBoxLayout()
        listBoxesLayout.addLayout(loginLayout)
        listBoxesLayout.addLayout(buycountLayout)
        listBoxesLayout.addLayout(firstListBoxLayout, 6)
        listBoxesLayout.addLayout(secondListBoxLayout, 4)
        listBoxesLayout.addLayout(printLayout)

        # ===== 차트 영역 =====
        chartLayout = QVBoxLayout()
        self.fig = Figure(figsize=(12, 8))
        self.canvas = FigureCanvas(self.fig)
        chartLayout.addWidget(self.canvas)

        # ===== 차트와 리스트 통합 =====
        chartAndListLayout = QHBoxLayout()
        chartAndListLayout.addLayout(listBoxesLayout, 1)
        chartAndListLayout.addLayout(chartLayout, 4)

        # ===== 전략 및 거래 정보 영역 =====
        strategyAndTradeLayout = QVBoxLayout()

        # 투자 전략
        strategyLayout = QHBoxLayout()
        strategyLabel = QLabel("투자전략:")
        strategyLabel.setFixedWidth(label_width)
        strategyLayout.addWidget(strategyLabel, Qt.AlignLeft)
        self.comboStg = QComboBox()
        self.comboStg.setFixedWidth(200)
        strategyLayout.addWidget(self.comboStg, Qt.AlignLeft)
        strategyLayout.addStretch()
        self.counterlabel = QLabel('타이머: 0')
        self.counterlabel.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        strategyLayout.addWidget(self.counterlabel)
        self.chart_status_label = QLabel("Chart: None")
        self.chart_status_label.setStyleSheet("color: red")
        self.chart_status_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        strategyLayout.addWidget(self.chart_status_label)

        # 매수 전략
        buyStrategyLayout = QHBoxLayout()
        buyStgLabel = QLabel("매수전략:")
        buyStgLabel.setFixedWidth(label_width)
        buyStrategyLayout.addWidget(buyStgLabel, alignment=Qt.AlignLeft)
        self.comboBuyStg = QComboBox()
        self.comboBuyStg.setFixedWidth(200)
        buyStrategyLayout.addWidget(self.comboBuyStg, alignment=Qt.AlignLeft)
        buyStrategyLayout.addStretch()
        self.saveBuyStgButton = QPushButton("수정")
        self.saveBuyStgButton.setFixedWidth(100)
        buyStrategyLayout.addWidget(self.saveBuyStgButton, alignment=Qt.AlignRight)
        self.buystgInputWidget = QTextEdit()
        self.buystgInputWidget.setPlaceholderText("매수전략의 내용을 입력하세요...")
        self.buystgInputWidget.setFixedHeight(80)

        # 매도 전략
        sellStrategyLayout = QHBoxLayout()
        sellStgLabel = QLabel("매도전략:")
        sellStgLabel.setFixedWidth(label_width)
        sellStrategyLayout.addWidget(sellStgLabel, alignment=Qt.AlignLeft)
        self.comboSellStg = QComboBox()
        self.comboSellStg.setFixedWidth(200)
        sellStrategyLayout.addWidget(self.comboSellStg, alignment=Qt.AlignLeft)
        sellStrategyLayout.addStretch()
        self.saveSellStgButton = QPushButton("수정")
        self.saveSellStgButton.setFixedWidth(100)
        sellStrategyLayout.addWidget(self.saveSellStgButton, alignment=Qt.AlignRight)
        self.sellstgInputWidget = QTextEdit()
        self.sellstgInputWidget.setPlaceholderText("매도전략의 내용을 입력하세요...")
        self.sellstgInputWidget.setFixedHeight(63)

        # 주식 현황 테이블
        self.stock_table = QTableWidget()
        self.stock_table.setRowCount(0)
        self.stock_table.setColumnCount(6)
        self.stock_table.setHorizontalHeaderLabels(["종목코드", "현재가", "상승확률(%)", "매수가", "평가손익", "수익률(%)"])
        self.stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.stock_table.setFixedHeight(220)
        self.stock_table.verticalHeader().setDefaultSectionSize(20)

        strategyAndTradeLayout.addLayout(strategyLayout)
        strategyAndTradeLayout.addLayout(buyStrategyLayout)
        strategyAndTradeLayout.addWidget(self.buystgInputWidget)
        strategyAndTradeLayout.addLayout(sellStrategyLayout)
        strategyAndTradeLayout.addWidget(self.sellstgInputWidget)
        strategyAndTradeLayout.addWidget(self.stock_table)

        # ===== 터미널 출력 =====
        self.terminalOutput = QTextEdit()
        self.terminalOutput.setReadOnly(True)

        counterAndterminalLayout = QVBoxLayout()
        counterAndterminalLayout.addLayout(strategyAndTradeLayout)
        counterAndterminalLayout.addWidget(self.terminalOutput)

        # ===== 메인 레이아웃 =====
        mainLayout = QHBoxLayout()
        mainLayout.addLayout(chartAndListLayout, 70)
        mainLayout.addLayout(counterAndterminalLayout, 30)
        self.trading_tab.setLayout(mainLayout)

        # ===== 이벤트 연결 =====
        self.loginButton.clicked.connect(self.login_handler.handle_login)
        self.buycountButton.clicked.connect(self.login_handler.buycount_setting)

        self.buyButton.clicked.connect(self.buy_item)
        self.deleteFirstButton.clicked.connect(self.delete_select_item)
        self.sellButton.clicked.connect(self.sell_item)
        self.sellAllButton.clicked.connect(self.sell_all_item)

        self.firstListBox.currentItemChanged.connect(self.listBoxChanged)
        self.firstListBox.itemClicked.connect(self.listBoxChanged)
        self.secondListBox.currentItemChanged.connect(self.listBoxChanged)
        self.secondListBox.itemClicked.connect(self.listBoxChanged)

        self.printChartButton.clicked.connect(self.print_chart)
        self.dataOutputButton2.clicked.connect(self.output_current_data)

        self.comboStg.currentIndexChanged.connect(self.stgChanged)
        self.comboBuyStg.currentIndexChanged.connect(self.buyStgChanged)
        self.comboSellStg.currentIndexChanged.connect(self.sellStgChanged)
        self.saveBuyStgButton.clicked.connect(self.save_buystrategy)
        self.saveSellStgButton.clicked.connect(self.save_sellstrategy)

    def init_backtest_tab(self):
        """백테스팅 탭 초기화"""
        
        layout = QVBoxLayout()
        
        # ===== 설정 영역 =====
        settings_group = QGroupBox("백테스팅 설정")
        settings_layout = QGridLayout()
        
        # 기간 선택
        settings_layout.addWidget(QLabel("시작일:"), 0, 0)
        self.bt_start_date = QLineEdit()
        self.bt_start_date.setPlaceholderText("YYYYMMDD (예: 20250101)")
        self.bt_start_date.setFixedWidth(150)
        settings_layout.addWidget(self.bt_start_date, 0, 1)
        
        settings_layout.addWidget(QLabel("종료일:"), 0, 2)
        self.bt_end_date = QLineEdit()
        self.bt_end_date.setPlaceholderText("YYYYMMDD (예: 20250131)")
        self.bt_end_date.setFixedWidth(150)
        settings_layout.addWidget(self.bt_end_date, 0, 3)
        
        # DB 기간 불러오기 버튼
        self.bt_load_period_button = QPushButton("DB 기간 불러오기")
        self.bt_load_period_button.setFixedWidth(130)
        self.bt_load_period_button.clicked.connect(self.load_db_period)
        settings_layout.addWidget(self.bt_load_period_button, 0, 4)
        
        # 초기 자금
        settings_layout.addWidget(QLabel("초기 자금:"), 1, 0)
        self.bt_initial_cash = QLineEdit("10000000")
        self.bt_initial_cash.setFixedWidth(150)
        settings_layout.addWidget(self.bt_initial_cash, 1, 1)
        
        # 전략 선택
        settings_layout.addWidget(QLabel("투자 전략:"), 2, 0)
        self.bt_strategy_combo = QComboBox()
        self.bt_strategy_combo.setFixedWidth(150)
        settings_layout.addWidget(self.bt_strategy_combo, 2, 1)
        
        # 실행 버튼
        self.bt_run_button = QPushButton("백테스팅 실행")
        self.bt_run_button.setFixedWidth(150)
        self.bt_run_button.clicked.connect(self.run_backtest)
        settings_layout.addWidget(self.bt_run_button, 2, 2)
        
        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)
        
        # ===== 결과 영역 (탭 구조) =====
        results_tab_widget = QTabWidget()
        
        # 탭 1: 전체 결과
        overall_tab = QWidget()
        overall_layout = QHBoxLayout()
        
        # 왼쪽: 결과 요약
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        
        left_layout.addWidget(QLabel("백테스팅 결과:"))
        self.bt_results_text = QTextEdit()
        self.bt_results_text.setReadOnly(True)
        self.bt_results_text.setMaximumWidth(450)
        left_layout.addWidget(self.bt_results_text)
        
        left_widget.setLayout(left_layout)
        
        # 오른쪽: 차트
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        
        self.bt_fig = Figure(figsize=(10, 8))
        self.bt_canvas = FigureCanvas(self.bt_fig)
        right_layout.addWidget(self.bt_canvas)
        
        right_widget.setLayout(right_layout)
        
        overall_layout.addWidget(left_widget, 1)
        overall_layout.addWidget(right_widget, 2)
        overall_tab.setLayout(overall_layout)
        
        # 탭 2: 일별 성과
        daily_tab = QWidget()
        daily_layout = QHBoxLayout()
        
        # 왼쪽: 일별 성과 테이블
        daily_left_widget = QWidget()
        daily_left_layout = QVBoxLayout()
        
        daily_left_layout.addWidget(QLabel("일별 성과 내역:"))
        self.bt_daily_table = QTableWidget()
        self.bt_daily_table.setColumnCount(8)
        self.bt_daily_table.setHorizontalHeaderLabels([
            "날짜", "일손익", "수익률(%)", "거래수", "승", "패", "누적손익", "포트폴리오"
        ])
        self.bt_daily_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.bt_daily_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.bt_daily_table.setMaximumWidth(600)
        daily_left_layout.addWidget(self.bt_daily_table)
        
        daily_left_widget.setLayout(daily_left_layout)
        
        # 오른쪽: 일별 차트
        daily_right_widget = QWidget()
        daily_right_layout = QVBoxLayout()
        
        self.bt_daily_fig = Figure(figsize=(10, 8))
        self.bt_daily_canvas = FigureCanvas(self.bt_daily_fig)
        daily_right_layout.addWidget(self.bt_daily_canvas)
        
        daily_right_widget.setLayout(daily_right_layout)
        
        daily_layout.addWidget(daily_left_widget, 1)
        daily_layout.addWidget(daily_right_widget, 2)
        daily_tab.setLayout(daily_layout)
        
        # 탭 추가
        results_tab_widget.addTab(overall_tab, "전체 성과")
        results_tab_widget.addTab(daily_tab, "일별 성과")
        
        layout.addWidget(results_tab_widget)
        
        self.backtest_tab.setLayout(layout)
        
        # 초기화 시 DB 기간 자동 로드
        QTimer.singleShot(100, self.load_db_period)

    def run_backtest(self):
        """백테스팅 실행"""
        
        try:
            from backtester import Backtester
            
            start_date = self.bt_start_date.text()
            end_date = self.bt_end_date.text()
            initial_cash = int(self.bt_initial_cash.text())
            
            # 입력 검증
            if len(start_date) != 8 or len(end_date) != 8:
                QMessageBox.warning(self, "입력 오류", "날짜 형식: YYYYMMDD (예: 20250101)")
                return
            
            # 전략 콤보박스가 비어있으면 전략 로드 시도
            if self.bt_strategy_combo.count() == 0:
                self.load_strategies_for_backtest()
                if self.bt_strategy_combo.count() == 0:
                    QMessageBox.warning(self, "오류", "전략을 불러올 수 없습니다.\nsettings.ini 파일을 확인해주세요.")
                    return
            
            # DB 파일 경로 확인
            if not hasattr(self, 'trader'):
                # 로그인하지 않은 경우 기본 DB 경로 사용
                import os
                db_path = 'vi_stock_data.db'
                if not os.path.exists(db_path):
                    QMessageBox.warning(self, "오류", f"데이터베이스 파일을 찾을 수 없습니다.\n경로: {db_path}")
                return
            else:
                db_path = self.trader.db_name
            
            self.bt_results_text.clear()
            self.bt_results_text.append(f"백테스팅 시작: {start_date} ~ {end_date}")
            self.bt_results_text.append(f"초기 자금: {initial_cash:,}원\n")
            self.bt_results_text.append("처리 중...\n")
            
            QApplication.processEvents()
            
            # 백테스팅 실행 (settings.ini 포함)
            bt = Backtester(
                db_path=db_path,
                config_file='settings.ini',
                initial_cash=initial_cash
            )
            
            # 백테스팅 탭의 전략 선택 사용
            strategy_name = self.bt_strategy_combo.currentText() if self.bt_strategy_combo.currentText() else '통합 전략'
            logging.info(f"선택된 전략: {strategy_name}")
            results = bt.run(start_date, end_date, strategy_name=strategy_name)
            
            # 결과 표시
            result_text = f"""
=== 백테스팅 결과 ===

【 기본 정보 】
전략: {results['strategy']}
기간: {results['start_date']} ~ {results['end_date']}

【 수익 성과 】
초기 자금: {results['initial_cash']:,}원
최종 자금: {results['final_cash']:,}원
총 수익: {results['total_profit']:,.0f}원
수익률: {results['total_return_pct']:.2f}%

【 거래 통계 】
총 거래: {results['total_trades']}회
승리: {results['win_trades']}회
패배: {results['lose_trades']}회
승률: {results['win_rate']:.1f}%

【 손익 분석 】
평균 수익률: {results['avg_profit_pct']:.2f}%
최대 수익: {results['max_profit_pct']:.2f}%
최대 손실: {results['max_loss_pct']:.2f}%
MDD (최대 낙폭): {results['mdd']:.2f}%

【 기타 지표 】
샤프 비율: {results['sharpe_ratio']:.2f}
평균 보유 시간: {results['avg_hold_minutes']:.0f}분

※ 백테스팅 결과는 참고용이며, 실제 매매 결과와 다를 수 있습니다.
"""
            
            self.bt_results_text.setPlainText(result_text)
            
            # 전체 성과 차트 그리기
            bt.plot_results(self.bt_fig)
            self.bt_canvas.draw()
            
            # 일별 성과 테이블 업데이트
            self.update_daily_results_table(results.get('daily_results', []))
            
            # 일별 성과 차트 그리기
            if len(results.get('daily_results', [])) > 0:
                bt.plot_daily_results(self.bt_daily_fig)
                self.bt_daily_canvas.draw()
            
            QMessageBox.information(self, "완료", "백테스팅이 완료되었습니다!")
            
        except FileNotFoundError:
            QMessageBox.critical(self, "오류", "backtester.py 파일을 찾을 수 없습니다.\n같은 폴더에 backtester.py가 있는지 확인하세요.")
        except Exception as ex:
            logging.error(f"run_backtest -> {ex}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "오류", f"백테스팅 실패:\n{str(ex)}")
    
    def update_daily_results_table(self, daily_results):
        """일별 성과 테이블 업데이트"""
        try:
            self.bt_daily_table.setRowCount(0)
            
            if not daily_results:
                return
            
            self.bt_daily_table.setRowCount(len(daily_results))
            
            for row_idx, daily in enumerate(daily_results):
                # 날짜 (YYYYMMDD -> YYYY-MM-DD)
                date_str = daily['date']
                formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                self.bt_daily_table.setItem(row_idx, 0, QTableWidgetItem(formatted_date))
                
                # 일손익
                daily_profit = daily['daily_profit']
                profit_item = QTableWidgetItem(f"{daily_profit:,.0f}")
                profit_item.setForeground(QColor('green') if daily_profit > 0 else QColor('red'))
                profit_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.bt_daily_table.setItem(row_idx, 1, profit_item)
                
                # 수익률
                daily_return = daily['daily_return_pct']
                return_item = QTableWidgetItem(f"{daily_return:.2f}")
                return_item.setForeground(QColor('green') if daily_return > 0 else QColor('red'))
                return_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.bt_daily_table.setItem(row_idx, 2, return_item)
                
                # 거래수
                trades_item = QTableWidgetItem(f"{daily['total_trades']}")
                trades_item.setTextAlignment(Qt.AlignCenter)
                self.bt_daily_table.setItem(row_idx, 3, trades_item)
                
                # 승
                win_item = QTableWidgetItem(f"{daily['win_trades']}")
                win_item.setTextAlignment(Qt.AlignCenter)
                self.bt_daily_table.setItem(row_idx, 4, win_item)
                
                # 패
                lose_item = QTableWidgetItem(f"{daily['lose_trades']}")
                lose_item.setTextAlignment(Qt.AlignCenter)
                self.bt_daily_table.setItem(row_idx, 5, lose_item)
                
                # 누적손익
                cumulative = daily['cumulative_profit']
                cumulative_item = QTableWidgetItem(f"{cumulative:,.0f}")
                cumulative_item.setForeground(QColor('blue'))
                cumulative_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.bt_daily_table.setItem(row_idx, 6, cumulative_item)
                
                # 포트폴리오 가치
                portfolio_item = QTableWidgetItem(f"{daily['portfolio_value']:,.0f}")
                portfolio_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.bt_daily_table.setItem(row_idx, 7, portfolio_item)
            
            logging.info(f"일별 성과 테이블 업데이트 완료: {len(daily_results)}건")
            
        except Exception as ex:
            logging.error(f"update_daily_results_table -> {ex}\n{traceback.format_exc()}")
    
    def load_strategies_for_backtest(self):
        """백테스팅용 전략 로드 (로그인 없이도 가능)"""
        try:
            import configparser
            import os
            
            # ✅ RawConfigParser 사용 (% 문자 이슈 완전 해결)
            config = configparser.RawConfigParser()
            if not os.path.exists('settings.ini'):
                logging.warning("settings.ini 파일이 없습니다.")
                return
            
            config.read('settings.ini', encoding='utf-8')
            
            # STRATEGIES 섹션에서 전략 목록 읽기
            if config.has_section('STRATEGIES'):
                strategy_keys = sorted(config['STRATEGIES'].keys())
                existing_stgnames = []
                seen = set()
                for key in strategy_keys:
                    stg_value = config['STRATEGIES'][key]
                    if stg_value not in seen:
                        existing_stgnames.append(stg_value)
                        seen.add(stg_value)
            else:
                existing_stgnames = []
            
            # 백테스팅 콤보박스에 추가
            self.bt_strategy_combo.clear()
            for stgname in existing_stgnames:
                self.bt_strategy_combo.addItem(stgname)
            
            # 기본값: 통합 전략
            index = self.bt_strategy_combo.findText("통합 전략")
            if index != -1:
                self.bt_strategy_combo.setCurrentIndex(index)
            
            logging.info(f"✅ 백테스팅 전략 로드 완료: {len(existing_stgnames)}개")
            
        except Exception as ex:
            logging.error(f"load_strategies_for_backtest -> {ex}\n{traceback.format_exc()}")
    
    def load_db_period(self):
        """DB에서 사용 가능한 데이터 기간 조회 및 자동 입력"""
        try:
            import sqlite3
            import os
            
            # DB 파일 경로 확인
            if hasattr(self, 'trader') and hasattr(self.trader, 'db_name'):
                db_path = self.trader.db_name
            else:
                db_path = 'vi_stock_data.db'
            
            if not os.path.exists(db_path):
                logging.debug(f"DB 파일 없음: {db_path}")
                return
            
            # DB 연결 및 기간 조회
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # combined_tick_data 테이블에서 최소/최대 날짜 조회
            cursor.execute('''
                SELECT MIN(date), MAX(date)
                FROM combined_tick_data
                WHERE date IS NOT NULL AND date != ''
            ''')
            
            result = cursor.fetchone()
            conn.close()
            
            if result and result[0] and result[1]:
                start_date = result[0]
                end_date = result[1]
                
                # 입력 필드에 자동 입력
                self.bt_start_date.setText(start_date)
                self.bt_end_date.setText(end_date)
                
                logging.info(f"✅ DB 기간 로드 완료: {start_date} ~ {end_date}")
                
                # 상태 메시지 표시 (선택사항)
                if hasattr(self, 'bt_results_text'):
                    self.bt_results_text.clear()
                    self.bt_results_text.append(f"📅 DB 데이터 기간: {start_date} ~ {end_date}")
                    self.bt_results_text.append(f"\n백테스팅 기간이 자동으로 설정되었습니다.")
            else:
                logging.warning("DB에 데이터가 없습니다.")
                if hasattr(self, 'bt_results_text'):
                    self.bt_results_text.clear()
                    self.bt_results_text.append("⚠️ DB에 데이터가 없습니다.")
            
        except sqlite3.Error as ex:
            logging.error(f"load_db_period (DB 오류) -> {ex}")
        except Exception as ex:
            logging.error(f"load_db_period -> {ex}\n{traceback.format_exc()}")

# ==================== QTextEditLogger ====================
class QTextEditLogger(QObject, logging.Handler):
    log_signal = pyqtSignal(str)

    def __init__(self, text_edit):
        super().__init__()
        self.text_edit = text_edit
        self.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
        self.log_signal.connect(self.append_log)

    @pyqtSlot(str)
    def append_log(self, msg):
        self.text_edit.append(msg)

    def emit(self, record):
        # 빈 로그 메시지 필터링 (더 강력한 필터)
        message = record.getMessage()
        if not message or not message.strip() or message.strip() == "":
            return
        
        msg = self.format(record)
        
        # 포맷팅된 메시지에서도 확인 (타임스탬프 제거 후)
        # 형식: "2025-10-13 10:14:04,278 - MESSAGE"
        if " - " in msg:
            content = msg.split(" - ", 1)[1].strip() if len(msg.split(" - ", 1)) > 1 else ""
            if not content or content == "":
                return
        
        if '매매이익' in msg:
            msg = f"<span style='color:green;'>{msg}</span>"
        elif '매매손실' in msg:
            msg = f"<span style='color:red;'>{msg}</span>"
        elif '매매실현손익' in msg:
            msg = f"<span style='font-weight:bold;'>{msg}</span>"
        else:
            msg = f"<span>{msg}</span>"

        self.log_signal.emit(msg)

class SplashScreen(QWidget):
    """로딩 스플래시 스크린"""
    
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        
        # 레이아웃
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        
        # 로고/타이틀
        title_label = QLabel("초단타 매매 프로그램")
        title_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 24px;
                font-weight: bold;
                background-color: rgba(0, 0, 0, 180);
                padding: 20px;
                border-radius: 10px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 로딩 메시지
        self.message_label = QLabel("초기화 중...")
        self.message_label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 14px;
                background-color: rgba(0, 0, 0, 180);
                padding: 10px;
                border-radius: 5px;
            }
        """)
        self.message_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.message_label)
        
        # 프로그레스 바
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid grey;
                border-radius: 5px;
                text-align: center;
                background-color: rgba(255, 255, 255, 200);
            }
            QProgressBar::chunk {
                background-color: #05B8CC;
                width: 10px;
                margin: 0.5px;
            }
        """)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        layout.addWidget(self.progress_bar)
        
        self.setLayout(layout)
        self.resize(400, 200)
        
        # 화면 중앙 배치
        self.center()
    
    def center(self):
        """화면 중앙 정렬"""
        qr = self.frameGeometry()
        cp = QApplication.desktop().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
    
    def update_progress(self, value, message=""):
        """진행 상황 업데이트"""
        self.progress_bar.setValue(value)
        if message:
            self.message_label.setText(message)
        QApplication.processEvents()

# ==================== Main ====================
if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
        os.chdir(application_path)
    
    try:
        # 로그 초기화 (빠름)
        setup_logging()
        logging.info("=" * 40)
        logging.info("=== 초단타 매매 프로그램 시작 ===")
        logging.info(f"실행 경로: {os.getcwd()}")
        logging.info("=" * 40)

        # QApplication 생성 (빠름)
        app = QApplication(sys.argv)
        
        # 폰트 설정 (빠름)
        try:
            app.setFont(QFont("Malgun Gothic", 9))
        except Exception as ex:
            logging.warning(f"폰트 설정 실패: {ex}")
        
        # ===== ✅ 메인 윈도우 즉시 생성 및 표시 =====
        logging.info("메인 윈도우 생성 중...")
        myWindow = MyWindow()
        
        # 아이콘 설정 (빠름)
        try:
            icon_path = 'stock_trader.ico'
            if getattr(sys, 'frozen', False):
                icon_path = os.path.join(application_path, 'stock_trader.ico')
            
            if os.path.exists(icon_path):
                myWindow.setWindowIcon(QIcon(icon_path))
        except Exception as ex:
            logging.warning(f"아이콘 설정 실패: {ex}")
        
        # ===== ✅ 창 즉시 표시 =====
        myWindow.showMaximized()
        logging.info("GUI 표시 완료")
        
        # 이벤트 루프 실행
        exit_code = app.exec_()
        logging.info(f"프로그램 종료 (exit code: {exit_code})")
        sys.exit(exit_code)
        
    except Exception as ex:
        # 최상위 예외 처리
        error_msg = (
            f"프로그램 실행 중 치명적 오류 발생:\n\n"
            f"{type(ex).__name__}: {ex}\n\n"
            f"상세 정보:\n{traceback.format_exc()}"
        )
        
        # 로그 파일에 기록
        try:
            logging.critical(error_msg)
        except:
            pass
        
        # 오류 파일 생성
        try:
            error_file = os.path.join(os.getcwd(), 'error.txt')
            with open(error_file, 'w', encoding='utf-8') as f:
                f.write(f"발생 시간: {datetime.now()}\n")
                f.write(f"실행 경로: {os.getcwd()}\n")
                f.write(f"Python: {sys.version}\n\n")
                f.write(error_msg)
            print(f"\n오류 정보가 저장되었습니다: {error_file}\n")
        except Exception as e:
            print(f"오류 파일 저장 실패: {e}")
        
        # 메시지 박스 표시
        try:
            from PyQt5.QtWidgets import QMessageBox, QApplication
            app = QApplication.instance()
            if app is None:
                app = QApplication(sys.argv)
            
            QMessageBox.critical(
                None, 
                "프로그램 오류", 
                f"프로그램 실행 중 오류가 발생했습니다.\n\n"
                f"{type(ex).__name__}: {ex}\n\n"
                f"자세한 내용은 error.txt 파일을 확인하세요."
            )
        except:
            # 메시지 박스도 실패하면 콘솔에 출력
            print("\n" + "=" * 60)
            print(error_msg)
            print("=" * 60)
            input("\nEnter 키를 눌러 종료하세요...")
        
        sys.exit(1)